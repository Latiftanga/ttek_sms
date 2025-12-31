from collections import defaultdict, Counter
from decimal import Decimal
import logging
import json
from django.utils import timezone

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.db.models import Prefetch, Sum, Avg, Count, Q, F
from django.db import models, transaction

from django.contrib import messages

from .models import (
    GradingSystem, GradeScale, AssessmentCategory,
    Assignment, Score, SubjectTermGrade, TermReport, ScoreAuditLog,
    RemarkTemplate, ReportDistributionLog
)
from .signals import signals_disabled
from . import config
from .utils import (
    check_transcript_permission,
    get_transcript_data,
    build_academic_history,
    get_school_context,
)
from academics.models import Class, Subject, ClassSubject
from students.models import Student
from core.models import Term
from teachers.models import Teacher

logger = logging.getLogger(__name__)


def get_client_ip(request):
    """Get client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def is_school_admin(user):
    """Check if user is a school admin or superuser."""
    return user.is_superuser or getattr(user, 'is_school_admin', False)


def is_teacher_or_admin(user):
    """Check if user is a teacher, school admin, or superuser."""
    return (user.is_superuser or
            getattr(user, 'is_school_admin', False) or
            getattr(user, 'is_teacher', False))


def admin_required(view_func):
    """Decorator to require school admin or superuser."""
    return user_passes_test(is_school_admin, login_url='/')(view_func)


def teacher_or_admin_required(view_func):
    """Decorator to require teacher, school admin, or superuser."""
    return user_passes_test(is_teacher_or_admin, login_url='/')(view_func)


def can_edit_scores(user, class_obj, subject):
    """
    Check if a user can edit scores for a specific class/subject.

    Returns True if:
    - User is superuser or school admin
    - User is the teacher assigned to this subject for this class
    """
    # Admins can always edit
    if user.is_superuser or getattr(user, 'is_school_admin', False):
        return True

    # Check if user has a teacher profile
    if not hasattr(user, 'teacher_profile') or not user.teacher_profile:
        return False

    teacher = user.teacher_profile

    # Check if this teacher is assigned to teach this subject to this class
    return ClassSubject.objects.filter(
        class_assigned=class_obj,
        subject=subject,
        teacher=teacher
    ).exists()


def get_teacher_subjects(user, class_obj):
    """
    Get subjects a teacher can edit for a specific class.

    Returns all subjects if admin, otherwise only assigned subjects.
    """
    if user.is_superuser or getattr(user, 'is_school_admin', False):
        # Admins see all subjects for the class
        return Subject.objects.filter(
            class_allocations__class_assigned=class_obj
        ).distinct()

    # Teachers only see their assigned subjects
    if not hasattr(user, 'teacher_profile') or not user.teacher_profile:
        return Subject.objects.none()

    return Subject.objects.filter(
        class_allocations__class_assigned=class_obj,
        class_allocations__teacher=user.teacher_profile
    ).distinct()


def htmx_render(request, full_template, partial_template, context=None):
    """Render full template for regular requests, partial for HTMX requests."""
    context = context or {}
    template = partial_template if request.htmx else full_template
    return render(request, template, context)


@login_required
@admin_required
def index(request):
    """Gradebook dashboard (Admin only)."""
    current_term = Term.get_current()
    classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')

    # Get grading systems
    grading_systems = GradingSystem.objects.filter(is_active=True)
    categories = AssessmentCategory.objects.filter(is_active=True)

    # Enhanced Stats - optimized with single queries
    total_students = Student.objects.filter(status='active').count()
    assignments_this_term = Assignment.objects.filter(term=current_term).count() if current_term else 0
    scores_entered = Score.objects.filter(assignment__term=current_term).count() if current_term else 0
    reports_generated = TermReport.objects.filter(term=current_term).count() if current_term else 0

    # Calculate score entry progress using optimized single query
    # Count expected scores: for each assignment, count students in classes that have that subject
    total_possible_scores = 0
    if current_term:
        # Get all assignments for current term with their subjects
        term_assignments = Assignment.objects.filter(term=current_term).values_list('subject_id', flat=True)
        if term_assignments:
            # Count unique (assignment, student) pairs where student's class has that subject
            # Using aggregation instead of nested loops
            total_possible_scores = Score.objects.filter(
                assignment__term=current_term
            ).values('student_id', 'assignment_id').distinct().count()

            # For a more accurate "expected" count, calculate based on class-subject assignments
            # This counts: sum of (students per class * assignments for subjects in that class)
            class_subject_data = ClassSubject.objects.filter(
                subject_id__in=term_assignments
            ).select_related('class_assigned').values(
                'subject_id', 'class_assigned_id'
            )

            # Build a mapping of subject -> list of class IDs
            subject_classes = {}
            for cs in class_subject_data:
                subject_classes.setdefault(cs['subject_id'], set()).add(cs['class_assigned_id'])

            # Get student counts per class (single query)
            class_student_counts = dict(
                Student.objects.filter(
                    status='active',
                    current_class__isnull=False
                ).values('current_class_id').annotate(
                    count=Count('id')
                ).values_list('current_class_id', 'count')
            )

            # Get assignment counts per subject (single query)
            subject_assignment_counts = dict(
                Assignment.objects.filter(
                    term=current_term
                ).values('subject_id').annotate(
                    count=Count('id')
                ).values_list('subject_id', 'count')
            )

            # Calculate total expected scores
            total_possible_scores = 0
            for subject_id, class_ids in subject_classes.items():
                assignment_count = subject_assignment_counts.get(subject_id, 0)
                for class_id in class_ids:
                    student_count = class_student_counts.get(class_id, 0)
                    total_possible_scores += assignment_count * student_count

    score_progress = round((scores_entered / total_possible_scores * 100) if total_possible_scores > 0 else 0, 1)

    # Recent activity - get latest score entries
    recent_scores = Score.objects.filter(
        assignment__term=current_term
    ).select_related(
        'student', 'assignment__subject'
    ).order_by('-updated_at')[:5] if current_term else []

    # Classes with incomplete scores - optimized with annotations
    classes_needing_scores = []
    if current_term:
        # Get top 6 active classes with student counts in a single query
        top_classes = classes[:6]
        top_class_ids = [c.id for c in top_classes]

        # Get student counts per class
        class_student_counts = dict(
            Student.objects.filter(
                status='active',
                current_class_id__in=top_class_ids
            ).values('current_class_id').annotate(
                count=Count('id')
            ).values_list('current_class_id', 'count')
        )

        # Get subjects per class
        class_subjects_map = {}
        for cs in ClassSubject.objects.filter(class_assigned_id__in=top_class_ids).values('class_assigned_id', 'subject_id'):
            class_subjects_map.setdefault(cs['class_assigned_id'], set()).add(cs['subject_id'])

        # Get assignment counts per subject for current term
        subject_assignment_counts = dict(
            Assignment.objects.filter(term=current_term).values('subject_id').annotate(
                count=Count('id')
            ).values_list('subject_id', 'count')
        )

        # Get actual score counts per class
        class_score_counts = dict(
            Score.objects.filter(
                assignment__term=current_term,
                student__current_class_id__in=top_class_ids
            ).values('student__current_class_id').annotate(
                count=Count('id')
            ).values_list('student__current_class_id', 'count')
        )

        for cls in top_classes:
            student_count = class_student_counts.get(cls.id, 0)
            if student_count > 0:
                # Calculate expected scores for this class
                class_subject_ids = class_subjects_map.get(cls.id, set())
                total_assignments = sum(
                    subject_assignment_counts.get(subj_id, 0)
                    for subj_id in class_subject_ids
                )
                if total_assignments > 0:
                    expected_scores = total_assignments * student_count
                    actual_scores = class_score_counts.get(cls.id, 0)
                    progress = round((actual_scores / expected_scores * 100) if expected_scores > 0 else 0)
                    classes_needing_scores.append({
                        'class': cls,
                        'student_count': student_count,
                        'progress': progress,
                        'assignments': total_assignments,
                    })

    stats = {
        'classes': classes.count(),
        'students': total_students,
        'grading_systems': grading_systems.count(),
        'categories': categories.count(),
        'assignments': assignments_this_term,
        'scores_entered': scores_entered,
        'reports_generated': reports_generated,
        'score_progress': score_progress,
    }

    context = {
        'current_term': current_term,
        'classes': classes,
        'grading_systems': grading_systems,
        'categories': categories,
        'stats': stats,
        'recent_scores': recent_scores,
        'classes_needing_scores': classes_needing_scores,
    }

    return htmx_render(
        request,
        'gradebook/index.html',
        'gradebook/partials/index_content.html',
        context
    )


@login_required
@admin_required
def settings(request):
    """Gradebook settings - grading systems and categories (Admin only)."""
    grading_systems = GradingSystem.objects.prefetch_related('scales').all()
    categories = AssessmentCategory.objects.all()

    # Check if percentages sum to 100
    total_percentage = sum(c.percentage for c in categories if c.is_active)

    context = {
        'grading_systems': grading_systems,
        'categories': categories,
        'total_percentage': total_percentage,
    }

    return htmx_render(
        request,
        'gradebook/settings.html',
        'gradebook/partials/settings_content.html',
        context
    )


# ============ Grading System CRUD ============

@login_required
@admin_required
def grading_systems(request):
    """List all grading systems (Admin only)."""
    systems = GradingSystem.objects.prefetch_related('scales').all()
    return render(request, 'gradebook/partials/grading_systems_list.html', {
        'grading_systems': systems,
    })


@login_required
@admin_required
def grading_system_create(request):
    """Create a new grading system (Admin only)."""
    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_grading_system.html', {
            'levels': GradingSystem.SCHOOL_LEVELS,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    name = request.POST.get('name', '').strip()
    level = request.POST.get('level', 'BASIC')
    description = request.POST.get('description', '').strip()

    if not name:
        return render(request, 'gradebook/includes/modal_grading_system.html', {
            'error': 'Name is required.',
            'levels': GradingSystem.SCHOOL_LEVELS,
        })

    GradingSystem.objects.create(
        name=name,
        level=level,
        description=description,
    )

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'closeModal, refreshSettings'
    return response


@login_required
@admin_required
def grading_system_edit(request, pk):
    """Edit a grading system (Admin only)."""
    system = get_object_or_404(GradingSystem, pk=pk)

    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_grading_system.html', {
            'system': system,
            'levels': GradingSystem.SCHOOL_LEVELS,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    system.name = request.POST.get('name', '').strip()
    system.level = request.POST.get('level', 'BASIC')
    system.description = request.POST.get('description', '').strip()
    system.is_active = request.POST.get('is_active') == 'on'
    system.save()

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'closeModal, refreshSettings'
    return response


@login_required
@admin_required
def grading_system_delete(request, pk):
    """Delete a grading system (Admin only)."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    system = get_object_or_404(GradingSystem, pk=pk)
    system.delete()

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'refreshSettings'
    return response


# ============ Grade Scale CRUD ============

@login_required
@admin_required
def grade_scales(request, system_id):
    """List grades for a grading system (Admin only)."""
    system = get_object_or_404(GradingSystem, pk=system_id)
    scales = system.scales.all()

    return render(request, 'gradebook/partials/grade_scales_list.html', {
        'system': system,
        'scales': scales,
    })


@login_required
@admin_required
def grade_scale_create(request, system_id):
    """Create a new grade scale (Admin only)."""
    system = get_object_or_404(GradingSystem, pk=system_id)

    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_grade_scale.html', {
            'system': system,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    try:
        GradeScale.objects.create(
            grading_system=system,
            grade_label=request.POST.get('grade_label', '').strip(),
            min_percentage=Decimal(request.POST.get('min_percentage', '0')),
            max_percentage=Decimal(request.POST.get('max_percentage', '0')),
            aggregate_points=int(request.POST.get('aggregate_points') or 0) or None,
            interpretation=request.POST.get('interpretation', '').strip(),
            is_pass=request.POST.get('is_pass') == 'on',
            is_credit=request.POST.get('is_credit') == 'on',
            order=int(request.POST.get('order') or 0),
        )
    except Exception as e:
        return render(request, 'gradebook/includes/modal_grade_scale.html', {
            'system': system,
            'error': str(e),
        })

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'closeModal, refreshSettings'
    return response


@login_required
@admin_required
def grade_scale_edit(request, pk):
    """Edit a grade scale (Admin only)."""
    scale = get_object_or_404(GradeScale, pk=pk)

    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_grade_scale.html', {
            'system': scale.grading_system,
            'scale': scale,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    try:
        scale.grade_label = request.POST.get('grade_label', '').strip()
        scale.min_percentage = Decimal(request.POST.get('min_percentage', '0'))
        scale.max_percentage = Decimal(request.POST.get('max_percentage', '0'))
        scale.aggregate_points = int(request.POST.get('aggregate_points') or 0) or None
        scale.interpretation = request.POST.get('interpretation', '').strip()
        scale.is_pass = request.POST.get('is_pass') == 'on'
        scale.is_credit = request.POST.get('is_credit') == 'on'
        scale.order = int(request.POST.get('order') or 0)
        scale.save()
    except Exception as e:
        return render(request, 'gradebook/includes/modal_grade_scale.html', {
            'system': scale.grading_system,
            'scale': scale,
            'error': str(e),
        })

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'closeModal, refreshSettings'
    return response


@login_required
@admin_required
def grade_scale_delete(request, pk):
    """Delete a grade scale (Admin only)."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    scale = get_object_or_404(GradeScale, pk=pk)
    scale.delete()

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'refreshSettings'
    return response


# ============ Assessment Category CRUD ============

@login_required
@admin_required
def categories(request):
    """List all assessment categories (Admin only)."""
    cats = AssessmentCategory.objects.all()
    total = sum(c.percentage for c in cats if c.is_active)

    return render(request, 'gradebook/partials/categories_list.html', {
        'categories': cats,
        'total_percentage': total,
    })


@login_required
@admin_required
def category_create(request):
    """Create a new assessment category (Admin only)."""
    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_category.html', {})

    if request.method != 'POST':
        return HttpResponse(status=405)

    name = request.POST.get('name', '').strip()
    short_name = request.POST.get('short_name', '').strip().upper()
    percentage = int(request.POST.get('percentage', 0))

    if not name or not short_name:
        return render(request, 'gradebook/includes/modal_category.html', {
            'error': 'Name and short name are required.',
        })

    # Check total won't exceed 100%
    current_total = AssessmentCategory.objects.filter(
        is_active=True
    ).aggregate(total=models.Sum('percentage'))['total'] or 0

    if current_total + percentage > 100:
        return render(request, 'gradebook/includes/modal_category.html', {
            'error': f'Total percentage would exceed 100%. Current: {current_total}%',
        })

    AssessmentCategory.objects.create(
        name=name,
        short_name=short_name,
        percentage=percentage,
        order=int(request.POST.get('order', 0)),
    )

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'closeModal, refreshSettings'
    return response


@login_required
@admin_required
def category_edit(request, pk):
    """Edit an assessment category (Admin only)."""
    category = get_object_or_404(AssessmentCategory, pk=pk)

    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_category.html', {
            'category': category,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    category.name = request.POST.get('name', '').strip()
    category.short_name = request.POST.get('short_name', '').strip().upper()
    category.percentage = int(request.POST.get('percentage', 0))
    category.order = int(request.POST.get('order', 0))
    category.is_active = request.POST.get('is_active') == 'on'
    category.save()

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'closeModal, refreshSettings'
    return response


@login_required
@admin_required
def category_delete(request, pk):
    """Delete an assessment category (Admin only)."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    category = get_object_or_404(AssessmentCategory, pk=pk)
    category.delete()

    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'refreshSettings'
    return response


# ============ Score Entry ============

@login_required
def score_entry(request):
    """Score entry page - select class and subject."""
    current_term = Term.get_current()
    user = request.user

    # Filter classes based on user role
    if is_school_admin(user):
        # Admins see all classes
        classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')
    elif getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile'):
        # Teachers see only classes they're assigned to
        teacher = user.teacher_profile
        assigned_class_ids = ClassSubject.objects.filter(
            teacher=teacher
        ).values_list('class_assigned_id', flat=True).distinct()
        classes = Class.objects.filter(
            id__in=assigned_class_ids,
            is_active=True
        ).order_by('level_number', 'name')
    else:
        classes = Class.objects.none()

    context = {
        'current_term': current_term,
        'classes': classes,
        'is_admin': is_school_admin(request.user),
    }

    return htmx_render(
        request,
        'gradebook/score_entry.html',
        'gradebook/partials/score_entry_content.html',
        context
    )


@login_required
@teacher_or_admin_required
def score_entry_form(request, class_id, subject_id):
    """Score entry form for a specific class/subject.

    OPTIMIZED: Uses select_related and builds lookup dict for O(1) score access.
    """
    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    grades_locked = current_term.grades_locked if current_term else False

    # Check if user can edit scores for this subject/class
    can_edit = can_edit_scores(request.user, class_obj, subject)

    # Get students - only fetch needed fields
    students = list(Student.objects.filter(
        current_class=class_obj
    ).only(
        'id', 'first_name', 'last_name', 'admission_number'
    ).order_by('last_name', 'first_name'))

    # Get assignments for this subject/term
    assignments = list(Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name'))

    # Get existing scores in single query - build nested dict for O(1) template lookup
    # Structure: {student_id: {assignment_id: points}}
    scores_dict = defaultdict(dict)
    if students and assignments:
        student_ids = [s.id for s in students]
        assignment_ids = [a.id for a in assignments]
        for score in Score.objects.filter(
            student_id__in=student_ids,
            assignment_id__in=assignment_ids
        ).only('student_id', 'assignment_id', 'points'):
            scores_dict[score.student_id][score.assignment_id] = score.points

    # Get categories (small table, usually cached)
    categories = list(AssessmentCategory.objects.filter(is_active=True).order_by('order'))

    # Determine if editing is allowed (not locked AND authorized)
    editing_allowed = can_edit and not grades_locked

    context = {
        'class_obj': class_obj,
        'subject': subject,
        'current_term': current_term,
        'students': students,
        'assignments': assignments,
        'categories': categories,
        'scores_dict': dict(scores_dict),  # Convert to regular dict for template
        'grades_locked': grades_locked,
        'can_edit': can_edit,
        'editing_allowed': editing_allowed,
    }

    return render(request, 'gradebook/partials/score_form.html', context)


@login_required
@teacher_or_admin_required
def score_save(request):
    """Save scores via HTMX with audit logging."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    student_id = request.POST.get('student_id')
    assignment_id = request.POST.get('assignment_id')
    points = request.POST.get('points', '').strip()

    if not all([student_id, assignment_id]):
        response = HttpResponse(status=200)
        response['HX-Trigger'] = '{"showToast": {"message": "Missing data", "type": "error"}}'
        return response

    student = get_object_or_404(Student.objects.select_related('current_class'), pk=student_id)
    assignment = get_object_or_404(Assignment.objects.select_related('term', 'subject'), pk=assignment_id)

    # Early check for authorization (before transaction)
    if not can_edit_scores(request.user, student.current_class, assignment.subject):
        response = HttpResponse(status=200)
        response['HX-Trigger'] = '{"showToast": {"message": "You are not authorized to edit scores for this subject", "type": "error"}}'
        return response

    # Get audit context
    client_ip = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')[:255]

    if points == '' or points is None:
        # Delete score if empty
        try:
            with transaction.atomic():
                # Re-check grade lock inside transaction with row lock
                term = Term.objects.select_for_update().get(pk=assignment.term_id)
                if term.grades_locked:
                    response = HttpResponse(status=200)
                    response['HX-Trigger'] = '{"showToast": {"message": "Grades are locked for this term", "type": "error"}}'
                    return response

                existing_score = Score.objects.filter(student=student, assignment=assignment).first()
                if existing_score:
                    old_value = existing_score.points
                    # Log deletion first
                    ScoreAuditLog.objects.create(
                        score=None,
                        student=student,
                        assignment=assignment,
                        user=request.user,
                        action='DELETE',
                        old_value=old_value,
                        new_value=None,
                        ip_address=client_ip,
                        user_agent=user_agent
                    )
                    existing_score.delete()
        except Exception as e:
            logger.error(f"Error deleting score: {e}")
            response = HttpResponse(status=200)
            response['HX-Trigger'] = '{"showToast": {"message": "Error deleting score", "type": "error"}}'
            return response
        return HttpResponse(status=200)

    # Get existing score for reverting on error
    existing_score = Score.objects.filter(student=student, assignment=assignment).first()
    old_value = existing_score.points if existing_score else None
    old_value_str = str(old_value) if old_value is not None else ''

    try:
        points_decimal = Decimal(points)
    except Exception:
        response = HttpResponse(status=200)
        response['HX-Trigger'] = (
            f'{{"showToast": {{"message": "Invalid number", "type": "error"}}, '
            f'"revertScore": {{"student": {student_id}, "assignment": {assignment_id}, "value": "{old_value_str}"}}}}'
        )
        return response

    # Validate range
    max_points = float(assignment.points_possible)
    if points_decimal < 0:
        response = HttpResponse(status=200)
        response['HX-Trigger'] = (
            f'{{"showToast": {{"message": "Score cannot be negative", "type": "error"}}, '
            f'"revertScore": {{"student": {student_id}, "assignment": {assignment_id}, "value": "{old_value_str}"}}}}'
        )
        return response

    if points_decimal > assignment.points_possible:
        response = HttpResponse(status=200)
        response['HX-Trigger'] = (
            f'{{"showToast": {{"message": "Maximum score is {max_points:.0f}", "type": "error"}}, '
            f'"revertScore": {{"student": {student_id}, "assignment": {assignment_id}, "value": "{old_value_str}"}}}}'
        )
        return response

    # Save score with transaction handling and race condition protection
    try:
        with transaction.atomic():
            # Re-check grade lock inside transaction with row lock to prevent race condition
            term = Term.objects.select_for_update().get(pk=assignment.term_id)
            if term.grades_locked:
                response = HttpResponse(status=200)
                response['HX-Trigger'] = '{"showToast": {"message": "Grades are locked for this term", "type": "error"}}'
                return response

            score, created = Score.objects.update_or_create(
                student=student,
                assignment=assignment,
                defaults={'points': points_decimal}
            )

            # Log the change
            ScoreAuditLog.objects.create(
                score=score,
                student=student,
                assignment=assignment,
                user=request.user,
                action='CREATE' if created else 'UPDATE',
                old_value=old_value,
                new_value=points_decimal,
                ip_address=client_ip,
                user_agent=user_agent
            )
    except Exception as e:
        logger.error(f"Error saving score: {e}")
        response = HttpResponse(status=200)
        response['HX-Trigger'] = (
            f'{{"showToast": {{"message": "Error saving score", "type": "error"}}, '
            f'"revertScore": {{"student": {student_id}, "assignment": "{assignment_id}", "value": "{old_value_str}"}}}}'
        )
        return response

    return HttpResponse(status=200)


# ============ Score Audit History ============

@login_required
@admin_required
def score_audit_history(request, student_id, assignment_id):
    """View audit history for a specific score."""
    student = get_object_or_404(Student, pk=student_id)
    assignment = get_object_or_404(Assignment, pk=assignment_id)

    logs = ScoreAuditLog.objects.filter(
        student=student,
        assignment=assignment
    ).select_related('user').order_by('-created_at')[:config.AUDIT_LOG_DISPLAY_LIMIT]

    return render(request, 'gradebook/partials/score_audit_history.html', {
        'student': student,
        'assignment': assignment,
        'logs': logs,
    })


# ============ Assignments ============

@login_required
def assignments(request, subject_id):
    """List assignments for a subject in current term."""
    current_term = Term.get_current()
    subject = get_object_or_404(Subject, pk=subject_id)

    assigns = Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name')

    categories = AssessmentCategory.objects.filter(is_active=True)

    return render(request, 'gradebook/partials/assignments_list.html', {
        'subject': subject,
        'assignments': assigns,
        'categories': categories,
        'current_term': current_term,
    })


@login_required
def assignment_create(request):
    """Create a new assignment."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    current_term = Term.get_current()
    if not current_term:
        return HttpResponse('No current term set', status=400)

    subject_id = request.POST.get('subject_id')
    category_id = request.POST.get('category_id')
    name = request.POST.get('name', '').strip()
    points_possible = request.POST.get('points_possible', '100')

    if not all([subject_id, category_id, name]):
        return HttpResponse('Missing required fields', status=400)

    subject = get_object_or_404(Subject, pk=subject_id)
    category = get_object_or_404(AssessmentCategory, pk=category_id)

    Assignment.objects.create(
        assessment_category=category,
        subject=subject,
        term=current_term,
        name=name,
        points_possible=int(points_possible),
    )

    # Return updated assignments list
    assigns = Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name')

    categories = AssessmentCategory.objects.filter(is_active=True)

    response = render(request, 'gradebook/partials/assignments_list.html', {
        'subject': subject,
        'assignments': assigns,
        'categories': categories,
        'current_term': current_term,
    })
    # Trigger score form refresh
    response['HX-Trigger'] = '{"assignmentsChanged": {"subject_id": %d}}' % subject.pk
    return response


@login_required
@teacher_or_admin_required
def assignment_edit(request, pk):
    """Edit an assignment."""
    assignment = get_object_or_404(Assignment.objects.select_related('subject', 'term', 'assessment_category'), pk=pk)
    subject = assignment.subject
    current_term = assignment.term

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        points_possible = request.POST.get('points_possible', '100')
        category_id = request.POST.get('category_id')

        if not name:
            return HttpResponse('Name is required', status=400)

        try:
            points = Decimal(points_possible)
            if points < 1:
                return HttpResponse('Points must be at least 1', status=400)
        except Exception:
            return HttpResponse('Invalid points value', status=400)

        # Update assignment
        assignment.name = name
        assignment.points_possible = points
        if category_id:
            category = get_object_or_404(AssessmentCategory, pk=category_id)
            assignment.assessment_category = category
        assignment.save()

        # Return updated assignments list
        assigns = Assignment.objects.filter(
            subject=subject,
            term=current_term
        ).select_related('assessment_category').order_by('assessment_category__order', 'name')

        categories = AssessmentCategory.objects.filter(is_active=True)

        response = render(request, 'gradebook/partials/assignments_list.html', {
            'subject': subject,
            'assignments': assigns,
            'categories': categories,
            'current_term': current_term,
        })
        response['HX-Trigger'] = '{"assignmentsChanged": {"subject_id": %d}}' % subject.pk
        return response

    # GET request - return edit form
    categories = AssessmentCategory.objects.filter(is_active=True)
    return render(request, 'gradebook/partials/assignment_edit_form.html', {
        'assignment': assignment,
        'categories': categories,
    })


@login_required
@teacher_or_admin_required
def assignment_delete(request, pk):
    """Delete an assignment and all associated scores."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    assignment = get_object_or_404(Assignment, pk=pk)
    subject = assignment.subject
    current_term = assignment.term

    # Check for existing scores (they will be cascade deleted)
    score_count = Score.objects.filter(assignment=assignment).count()
    if score_count > 0:
        # Log the deletion for audit purposes
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(
            f"Assignment '{assignment.name}' deleted by {request.user}. "
            f"{score_count} scores were cascade deleted."
        )

    assignment.delete()

    # Return updated assignments list
    assigns = Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name')

    categories = AssessmentCategory.objects.filter(is_active=True)

    response = render(request, 'gradebook/partials/assignments_list.html', {
        'subject': subject,
        'assignments': assigns,
        'categories': categories,
        'current_term': current_term,
    })
    # Trigger score form refresh
    response['HX-Trigger'] = '{"assignmentsChanged": {"subject_id": %d}}' % subject.pk
    return response


# ============ Bulk Score Import ============

@login_required
@teacher_or_admin_required
def score_import_template(request, class_id, subject_id):
    """Download Excel template for score import."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as DjangoHttpResponse

    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    # Check authorization
    if not can_edit_scores(request.user, class_obj, subject):
        return HttpResponse("Not authorized", status=403)

    # Get students and assignments
    students = Student.objects.filter(
        current_class=class_obj
    ).order_by('last_name', 'first_name')

    assignments = Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=config.EXCEL_HEADER_COLOR, end_color=config.EXCEL_HEADER_COLOR, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ["Student ID", "Student Name"]
    for assign in assignments:
        headers.append(f"{assign.assessment_category.short_name}: {assign.name} (/{assign.points_possible})")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Get existing scores
    existing_scores = {}
    for score in Score.objects.filter(
        student__in=students,
        assignment__in=assignments
    ).select_related('student', 'assignment'):
        key = (score.student_id, score.assignment_id)
        existing_scores[key] = score.points

    # Data rows
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.admission_number).border = thin_border
        ws.cell(row=row, column=2, value=f"{student.last_name}, {student.first_name}").border = thin_border

        for col, assign in enumerate(assignments, 3):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            # Pre-fill existing scores
            existing = existing_scores.get((student.id, assign.id))
            if existing is not None:
                cell.value = float(existing)

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Add metadata sheet for import validation
    meta_ws = wb.create_sheet("_metadata")
    meta_ws.cell(row=1, column=1, value="class_id")
    meta_ws.cell(row=1, column=2, value=class_id)
    meta_ws.cell(row=2, column=1, value="subject_id")
    meta_ws.cell(row=2, column=2, value=subject_id)
    meta_ws.cell(row=3, column=1, value="term_id")
    meta_ws.cell(row=3, column=2, value=current_term.id if current_term else "")

    # Assignment IDs in order
    for col, assign in enumerate(assignments, 1):
        meta_ws.cell(row=4, column=col, value=assign.id)

    # Hide metadata sheet
    meta_ws.sheet_state = 'hidden'

    # Create response
    response = DjangoHttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"scores_{class_obj.name}_{subject.short_name}_{current_term.name if current_term else 'noterm'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


@login_required
@teacher_or_admin_required
def score_import_upload(request, class_id, subject_id):
    """Handle score import file upload and show preview."""
    import openpyxl
    from decimal import InvalidOperation

    if request.method != 'POST':
        return HttpResponse(status=405)

    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    # Check authorization
    if not can_edit_scores(request.user, class_obj, subject):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'You are not authorized to import scores for this subject.'
        })

    # Check if grades are locked
    if current_term and current_term.grades_locked:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Grades are locked for this term.'
        })

    file = request.FILES.get('file')
    if not file:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'No file uploaded.'
        })

    if not file.name.endswith('.xlsx'):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Please upload an Excel file (.xlsx).'
        })

    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active

        # Get assignments for validation
        assignments = list(Assignment.objects.filter(
            subject=subject,
            term=current_term
        ).select_related('assessment_category').order_by('assessment_category__order', 'name'))

        # Get students lookup
        students_by_id = {
            s.admission_number: s for s in Student.objects.filter(current_class=class_obj)
        }

        # Parse data
        preview_data = []
        errors = []
        row_num = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or not row[0]:  # Skip empty rows
                continue

            student_id = str(row[0]).strip()
            student = students_by_id.get(student_id)

            row_data = {
                'row_num': row_num + 1,
                'student_id': student_id,
                'student_name': row[1] if len(row) > 1 else '',
                'student': student,
                'scores': [],
                'has_error': False,
            }

            if not student:
                row_data['has_error'] = True
                errors.append(f"Row {row_num + 1}: Student ID '{student_id}' not found in this class.")

            # Parse scores
            for col, assign in enumerate(assignments, 2):
                value = row[col] if len(row) > col else None
                score_data = {
                    'assignment': assign,
                    'value': value,
                    'error': None,
                }

                if value is not None and value != '':
                    try:
                        points = Decimal(str(value))
                        if points < 0:
                            score_data['error'] = 'Negative value'
                            row_data['has_error'] = True
                            errors.append(f"Row {row_num + 1}, {assign.name}: Negative value not allowed.")
                        elif points > assign.points_possible:
                            score_data['error'] = f'Exceeds max ({assign.points_possible})'
                            row_data['has_error'] = True
                            errors.append(f"Row {row_num + 1}, {assign.name}: Value {points} exceeds maximum {assign.points_possible}.")
                        else:
                            score_data['value'] = points
                    except (InvalidOperation, ValueError):
                        score_data['error'] = 'Invalid number'
                        row_data['has_error'] = True
                        errors.append(f"Row {row_num + 1}, {assign.name}: Invalid number '{value}'.")

                row_data['scores'].append(score_data)

            preview_data.append(row_data)

        wb.close()

        # Store data in session for confirmation
        import json
        import_data = []
        for row in preview_data:
            if row['student'] and not row['has_error']:
                for score in row['scores']:
                    if score['value'] is not None and score['value'] != '' and not score['error']:
                        import_data.append({
                            'student_id': row['student'].id,
                            'assignment_id': score['assignment'].id,
                            'points': str(score['value']),
                        })

        request.session['import_data'] = json.dumps(import_data)
        request.session['import_class_id'] = class_id
        request.session['import_subject_id'] = subject_id

        return render(request, 'gradebook/partials/import_preview.html', {
            'class_obj': class_obj,
            'subject': subject,
            'assignments': assignments,
            'preview_data': preview_data,
            'errors': errors,
            'total_scores': len(import_data),
            'has_errors': len(errors) > 0,
        })

    except Exception as e:
        logger.exception("Error parsing import file")
        return render(request, 'gradebook/partials/import_error.html', {
            'error': f'Error reading file: {str(e)}'
        })


@login_required
@teacher_or_admin_required
def score_import_confirm(request, class_id, subject_id):
    """Confirm and execute score import."""
    import json

    if request.method != 'POST':
        return HttpResponse(status=405)

    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    current_term = Term.get_current()

    # Check authorization
    if not can_edit_scores(request.user, class_obj, subject):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'You are not authorized to import scores for this subject.'
        })

    # Check if grades are locked
    if current_term and current_term.grades_locked:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Grades are locked for this term.'
        })

    # Get data from session
    import_data_json = request.session.get('import_data')
    if not import_data_json:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'No import data found. Please upload the file again.'
        })

    # Validate session data matches current request
    if (request.session.get('import_class_id') != class_id or
        request.session.get('import_subject_id') != subject_id):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Import data mismatch. Please upload the file again.'
        })

    try:
        import_data = json.loads(import_data_json)
    except json.JSONDecodeError:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Invalid import data. Please upload the file again.'
        })

    # Get audit context
    client_ip = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')[:255]

    # Import scores
    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for item in import_data:
            student_id = item['student_id']
            assignment_id = item['assignment_id']
            points = Decimal(item['points'])

            existing = Score.objects.filter(
                student_id=student_id,
                assignment_id=assignment_id
            ).first()

            old_value = existing.points if existing else None

            score, created = Score.objects.update_or_create(
                student_id=student_id,
                assignment_id=assignment_id,
                defaults={'points': points}
            )

            # Audit log
            ScoreAuditLog.objects.create(
                score=score,
                student_id=student_id,
                assignment_id=assignment_id,
                user=request.user,
                action='CREATE' if created else 'UPDATE',
                old_value=old_value,
                new_value=points,
                ip_address=client_ip,
                user_agent=f"BULK_IMPORT: {user_agent[:240]}"
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

    # Clear session data
    request.session.pop('import_data', None)
    request.session.pop('import_class_id', None)
    request.session.pop('import_subject_id', None)

    return render(request, 'gradebook/partials/import_success.html', {
        'created_count': created_count,
        'updated_count': updated_count,
        'total_count': created_count + updated_count,
        'class_obj': class_obj,
        'subject': subject,
    })


# ============ Grade Calculation ============

@login_required
@admin_required
def calculate_grades(request):
    """Calculate grades page - select class (Admin only)."""
    current_term = Term.get_current()
    classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')
    grading_systems = GradingSystem.objects.filter(is_active=True)

    context = {
        'current_term': current_term,
        'classes': classes,
        'grading_systems': grading_systems,
    }

    return htmx_render(
        request,
        'gradebook/calculate.html',
        'gradebook/partials/calculate_content.html',
        context
    )


@login_required
@admin_required
def calculate_class_grades(request, class_id):
    """
    Calculate grades for all students in a class.
    Uses Ghana grading standards with configurable pass marks,
    WAEC aggregate calculation, and promotion eligibility checks.

    OPTIMIZED: Uses bulk prefetching and bulk_update to minimize queries.
    Before: ~3000+ queries for 40 students x 12 subjects
    After: ~20 queries total
    """
    if request.method != 'POST':
        return HttpResponse(status=405)

    current_term = Term.get_current()
    if not current_term:
        return HttpResponse('No current term set', status=400)

    class_obj = get_object_or_404(Class, pk=class_id)
    grading_system_id = request.POST.get('grading_system_id')
    grading_system = get_object_or_404(GradingSystem, pk=grading_system_id) if grading_system_id else None

    # Check if this is the final term (Term 3) for promotion decisions
    is_final_term = current_term.term_number == 3 if hasattr(current_term, 'term_number') else False

    try:
        # Disable auto-calculation signals during bulk operation
        with signals_disabled(), transaction.atomic():
            # ========== PHASE 1: Bulk prefetch all data ==========

            # Get students
            students = list(Student.objects.filter(
                current_class=class_obj
            ).order_by('last_name', 'first_name'))

            if not students:
                return HttpResponse('No students in this class', status=400)

            student_ids = [s.id for s in students]

            # Get subjects for this class
            class_subjects = ClassSubject.objects.filter(
                class_assigned=class_obj
            ).select_related('subject')
            subjects = [cs.subject for cs in class_subjects]

            if not subjects:
                return HttpResponse('No subjects assigned to this class', status=400)

            subject_ids = [s.id for s in subjects]

            # Prefetch all categories (small table, usually 2-3 rows)
            categories = list(AssessmentCategory.objects.filter(is_active=True))

            # Prefetch all assignments for these subjects in this term
            assignments = list(Assignment.objects.filter(
                subject_id__in=subject_ids,
                term=current_term
            ).select_related('assessment_category'))

            # Build assignment lookup: {(subject_id, category_id): [assignments]}
            assignments_by_subject_category = defaultdict(list)
            assignments_by_subject = defaultdict(list)
            for assign in assignments:
                key = (assign.subject_id, assign.assessment_category_id)
                assignments_by_subject_category[key].append(assign)
                assignments_by_subject[assign.subject_id].append(assign)

            # Prefetch all scores for these students and assignments
            assignment_ids = [a.id for a in assignments]
            scores = Score.objects.filter(
                student_id__in=student_ids,
                assignment_id__in=assignment_ids
            ).select_related('assignment')

            # Build score lookup: {(student_id, assignment_id): score}
            scores_lookup = {
                (s.student_id, s.assignment_id): s for s in scores
            }

            # Prefetch grade scales for the grading system
            grade_scales = []
            if grading_system:
                grade_scales = list(grading_system.scales.all().order_by('-min_percentage'))

            # ========== PHASE 2: Calculate grades in memory ==========

            # Get or create SubjectTermGrade objects in bulk
            existing_grades = {
                (g.student_id, g.subject_id): g
                for g in SubjectTermGrade.objects.filter(
                    student_id__in=student_ids,
                    subject_id__in=subject_ids,
                    term=current_term
                )
            }

            grades_to_create = []
            grades_to_update = []

            for student in students:
                for subject in subjects:
                    key = (student.id, subject.id)

                    if key in existing_grades:
                        grade = existing_grades[key]
                    else:
                        grade = SubjectTermGrade(
                            student=student,
                            subject=subject,
                            term=current_term
                        )
                        grades_to_create.append(grade)
                        existing_grades[key] = grade

            # Bulk create new grades
            if grades_to_create:
                SubjectTermGrade.objects.bulk_create(grades_to_create)
                # Refresh to get IDs
                for grade in grades_to_create:
                    existing_grades[(grade.student_id, grade.subject_id)] = grade

            # Calculate scores for each grade using prefetched data
            for student in students:
                for subject in subjects:
                    grade = existing_grades[(student.id, subject.id)]

                    # Calculate scores using prefetched data (no DB queries)
                    category_totals = {}
                    total = Decimal('0.0')

                    for category in categories:
                        cat_assignments = assignments_by_subject_category.get(
                            (subject.id, category.id), []
                        )

                        if not cat_assignments:
                            continue

                        # Calculate weight per assignment
                        weight_per_assignment = (
                            Decimal(str(category.percentage)) / Decimal(str(len(cat_assignments)))
                        )
                        category_total = Decimal('0.0')

                        for assign in cat_assignments:
                            score = scores_lookup.get((student.id, assign.id))
                            if score:
                                score_pct = Decimal(str(score.points)) / Decimal(str(assign.points_possible))
                                category_total += score_pct * weight_per_assignment

                        # Store by both short_name and category_type for flexibility
                        category_totals[category.short_name] = round(category_total, 2)
                        category_totals[f'_type_{category.category_type}'] = round(category_total, 2)
                        total += category_total

                    # Update grade object using category_type (with fallback to short_name)
                    grade.class_score = category_totals.get('_type_CLASS_SCORE', category_totals.get('CA', Decimal('0.0')))
                    grade.exam_score = category_totals.get('_type_EXAM', category_totals.get('EXAM', Decimal('0.0')))
                    grade.total_score = round(total, 2)

                    # Determine grade from scale (no DB query - uses prefetched scales)
                    grade.is_passing = False  # Default to not passing
                    if grading_system and grade.total_score is not None:
                        for scale in grade_scales:
                            if scale.min_percentage <= grade.total_score <= scale.max_percentage:
                                grade.grade = scale.grade_label
                                grade.grade_remark = scale.interpretation
                                grade.is_passing = scale.is_pass
                                break

                    grades_to_update.append(grade)

            # Bulk update all grades
            SubjectTermGrade.objects.bulk_update(
                grades_to_update,
                ['class_score', 'exam_score', 'total_score', 'grade', 'grade_remark', 'is_passing'],
                batch_size=config.BULK_UPDATE_BATCH_SIZE
            )

            # ========== PHASE 3: Calculate positions per subject ==========

            # Refresh grades for position calculation
            all_grades = list(SubjectTermGrade.objects.filter(
                student_id__in=student_ids,
                subject_id__in=subject_ids,
                term=current_term,
                total_score__isnull=False
            ))

            # Group by subject and calculate positions
            grades_by_subject = defaultdict(list)
            for grade in all_grades:
                grades_by_subject[grade.subject_id].append(grade)

            position_updates = []
            for subject_id, subject_grades in grades_by_subject.items():
                # Sort by total_score descending
                subject_grades.sort(key=lambda g: g.total_score or Decimal('0'), reverse=True)

                position = 0
                last_score = None
                for i, grade in enumerate(subject_grades, 1):
                    if grade.total_score != last_score:
                        position = i
                    grade.position = position
                    position_updates.append(grade)
                    last_score = grade.total_score

            # Bulk update positions
            SubjectTermGrade.objects.bulk_update(
                position_updates,
                ['position'],
                batch_size=config.BULK_UPDATE_BATCH_SIZE
            )

            # ========== PHASE 4: Calculate term reports ==========

            # Get or create TermReport objects
            existing_reports = {
                r.student_id: r
                for r in TermReport.objects.filter(
                    student_id__in=student_ids,
                    term=current_term
                )
            }

            reports_to_create = []
            for student in students:
                if student.id not in existing_reports:
                    report = TermReport(student=student, term=current_term)
                    reports_to_create.append(report)
                    existing_reports[student.id] = report

            if reports_to_create:
                TermReport.objects.bulk_create(reports_to_create)

            # Build grade lookup for aggregates
            grades_by_student = defaultdict(list)
            for grade in all_grades:
                grades_by_student[grade.student_id].append(grade)

            # Get subjects for core check
            subjects_dict = {s.id: s for s in subjects}

            # Calculate aggregates for each report
            reports_to_update = []
            pass_mark = grading_system.pass_mark if grading_system else config.DEFAULT_PASS_MARK
            credit_mark = grading_system.credit_mark if grading_system else config.DEFAULT_CREDIT_MARK

            for student in students:
                report = existing_reports[student.id]
                student_grades = grades_by_student.get(student.id, [])

                if student_grades:
                    total = sum(g.total_score for g in student_grades if g.total_score)
                    count = len([g for g in student_grades if g.total_score is not None])

                    report.total_marks = total
                    report.average = round(total / count, 2) if count > 0 else Decimal('0.0')
                    report.subjects_taken = count

                    # Count passed/failed
                    passed = [g for g in student_grades if g.total_score and g.total_score >= pass_mark]
                    report.subjects_passed = len(passed)
                    report.subjects_failed = count - len(passed)

                    # Count credits
                    report.credits_count = len([
                        g for g in student_grades if g.total_score and g.total_score >= credit_mark
                    ])

                    # Core subjects
                    core_grades = [
                        g for g in student_grades
                        if subjects_dict.get(g.subject_id) and subjects_dict[g.subject_id].is_core
                    ]
                    report.core_subjects_total = len(core_grades)
                    report.core_subjects_passed = len([
                        g for g in core_grades if g.total_score and g.total_score >= pass_mark
                    ])

                    # Calculate aggregate if grading system provided
                    if grading_system:
                        grade_points = []
                        for g in student_grades:
                            if g.total_score is not None:
                                for scale in grade_scales:
                                    if scale.min_percentage <= g.total_score <= scale.max_percentage:
                                        if scale.aggregate_points:
                                            grade_points.append(scale.aggregate_points)
                                        break

                        if grade_points:
                            grade_points.sort()
                            best_n = grade_points[:grading_system.aggregate_subjects_count]
                            report.aggregate = sum(best_n)

                report.out_of = len(students)

                # Check promotion eligibility for final term
                if is_final_term and grading_system:
                    is_eligible, reasons = grading_system.check_promotion_eligibility(report)
                    report.promoted = is_eligible
                    report.promotion_remarks = '; '.join(reasons) if reasons else 'Meets all requirements'

                reports_to_update.append(report)

            # Bulk update reports
            TermReport.objects.bulk_update(
                reports_to_update,
                ['total_marks', 'average', 'subjects_taken', 'subjects_passed',
                 'subjects_failed', 'credits_count', 'core_subjects_total',
                 'core_subjects_passed', 'aggregate', 'out_of', 'promoted',
                 'promotion_remarks'],
                batch_size=config.BULK_UPDATE_BATCH_SIZE
            )

            # ========== PHASE 5: Calculate overall positions ==========

            # SHS uses aggregate (lower is better), Basic uses average (higher is better)
            if grading_system and grading_system.level == 'SHS':
                reports_to_update.sort(
                    key=lambda r: (r.aggregate is None, r.aggregate or 999, -(r.average or 0))
                )
            else:
                reports_to_update.sort(key=lambda r: -(r.average or 0))

            position = 0
            last_value = None
            for i, report in enumerate(reports_to_update, 1):
                if grading_system and grading_system.level == 'SHS':
                    current_value = report.aggregate
                else:
                    current_value = report.average

                if current_value != last_value:
                    position = i
                report.position = position
                last_value = current_value

            # Bulk update positions
            TermReport.objects.bulk_update(
                reports_to_update,
                ['position'],
                batch_size=config.BULK_UPDATE_BATCH_SIZE
            )

            logger.info(
                f'Calculated grades for {len(students)} students in {class_obj.name} '
                f'using {grading_system.name if grading_system else "default"} grading system'
            )

    except Exception as e:
        logger.error(f'Error calculating grades for class {class_id}: {str(e)}')
        return HttpResponse(f'Error: {str(e)}', status=500)

    # Return success HTML
    return HttpResponse(f'''
        <div class="alert alert-success mt-2">
            <i class="fa-solid fa-check-circle"></i>
            <div>
                <div class="font-bold">Grades Calculated Successfully!</div>
                <div class="text-sm">{len(students)} students in {class_obj.name} using {grading_system.name if grading_system else "default"} grading system</div>
            </div>
            <a href="/gradebook/reports/?class={class_id}" class="btn btn-sm btn-ghost">View Reports</a>
        </div>
    ''')


# ============ Grade Locking ============

@login_required
@admin_required
def toggle_grade_lock(request, term_id):
    """Toggle grade lock status for a term."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    term = get_object_or_404(Term, pk=term_id)

    if term.grades_locked:
        term.unlock_grades()
        message = f"Grades unlocked for {term.name}"
    else:
        term.lock_grades(request.user)
        message = f"Grades locked for {term.name}"

    response = HttpResponse(status=204)
    response['HX-Trigger'] = f'{{"showToast": {{"message": "{message}", "type": "success"}}, "refreshLockStatus": true}}'
    return response


@login_required
def grade_lock_status(request):
    """Get current term's lock status (for HTMX refresh)."""
    current_term = Term.get_current()
    return render(request, 'gradebook/partials/grade_lock_status.html', {
        'current_term': current_term,
        'is_admin': is_school_admin(request.user),
    })


# ============ Report Cards ============

@login_required
def report_cards(request):
    """Report cards page - select class/term.

    OPTIMIZED: Uses dict lookup for O(1) report access per student.

    For teachers: Only show classes where they are the form master (class_teacher).
    For admins: Show all classes. Admins can also filter by student status to view
    transcripts for past students (graduated, withdrawn, etc.).
    """
    from students.models import Enrollment

    current_term = Term.get_current()
    user = request.user
    is_admin = is_school_admin(user)

    # Filter classes based on user role
    if is_admin:
        # Admins see all classes
        classes = Class.objects.filter(is_active=True).only(
            'id', 'name', 'level_number'
        ).order_by('level_number', 'name')
    elif getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile'):
        # Teachers only see classes where they are the form master
        teacher = user.teacher_profile
        classes = Class.objects.filter(
            class_teacher=teacher,
            is_active=True
        ).only(
            'id', 'name', 'level_number'
        ).order_by('level_number', 'name')
    else:
        classes = Class.objects.none()

    # Get filters
    class_id = request.GET.get('class')
    status_filter = request.GET.get('status', 'active')  # Default to active
    students = []
    class_obj = None

    # Status choices for admins (to filter past students)
    status_choices = Student.Status.choices if is_admin else []

    if class_id:
        class_obj = get_object_or_404(Class, pk=class_id)

        # Verify teacher has permission to view this class
        if not is_admin:
            if getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile'):
                teacher = user.teacher_profile
                if class_obj.class_teacher != teacher:
                    from django.contrib import messages
                    messages.error(request, 'You can only view reports for classes you are the form master of.')
                    return redirect('gradebook:reports')
            else:
                from django.contrib import messages
                messages.error(request, 'You do not have permission to view reports.')
                return redirect('core:index')

        if status_filter == 'active':
            # Active students: filter by current_class (existing behavior)
            students = list(Student.objects.filter(
                current_class=class_obj
            ).only(
                'id', 'first_name', 'last_name', 'admission_number', 'status'
            ).order_by('last_name', 'first_name'))
        elif is_admin:
            # Non-active students (graduated, etc.): find via enrollment history
            # Get student IDs who were enrolled in this class with the selected status
            student_ids = Enrollment.objects.filter(
                class_assigned=class_obj,
                student__status=status_filter
            ).values_list('student_id', flat=True).distinct()

            students = list(Student.objects.filter(
                pk__in=student_ids,
                status=status_filter
            ).only(
                'id', 'first_name', 'last_name', 'admission_number', 'status'
            ).order_by('last_name', 'first_name'))

        if students:
            # Get term reports in single query - O(1) lookup per student
            student_ids = [s.id for s in students]
            reports = {
                r.student_id: r
                for r in TermReport.objects.filter(
                    student_id__in=student_ids,
                    term=current_term
                ).only(
                    'id', 'student_id', 'average', 'position', 'out_of',
                    'subjects_passed', 'subjects_failed', 'aggregate'
                )
            }
            for s in students:
                s.term_report = reports.get(s.id)

    context = {
        'current_term': current_term,
        'classes': classes,
        'selected_class': class_obj,
        'students': students,
        'is_admin': is_admin,
        'status_choices': status_choices,
        'status_filter': status_filter,
    }

    return htmx_render(
        request,
        'gradebook/reports.html',
        'gradebook/partials/reports_content.html',
        context
    )


@login_required
def student_report(request, student_id):
    """View individual student report card with Ghana-specific data.

    OPTIMIZED: Uses select_related and computes grade summary in-memory.

    For teachers: Only allow viewing reports for students in their homeroom classes.
    For admins: Allow viewing all reports.
    """
    current_term = Term.get_current()
    student = get_object_or_404(Student.objects.select_related('current_class'), pk=student_id)
    user = request.user

    # Permission check for teachers
    if not is_school_admin(user):
        if getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile'):
            teacher = user.teacher_profile
            # Teacher must be the form master of the student's class
            if not student.current_class or student.current_class.class_teacher != teacher:
                from django.contrib import messages
                messages.error(request, 'You can only view reports for students in your homeroom class.')
                return redirect('gradebook:reports')
        else:
            from django.contrib import messages
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('core:index')

    # Get subject grades with core/elective distinction - single query
    subject_grades = list(SubjectTermGrade.objects.filter(
        student=student,
        term=current_term
    ).select_related('subject').order_by('-subject__is_core', 'subject__name'))

    # Separate core and elective subjects in memory (no extra queries)
    core_grades = [sg for sg in subject_grades if sg.subject.is_core]
    elective_grades = [sg for sg in subject_grades if not sg.subject.is_core]

    # Get term report
    term_report = TermReport.objects.filter(
        student=student,
        term=current_term
    ).first()

    # Compute grade summary in memory (no extra query)
    from collections import Counter
    grade_summary = dict(Counter(
        sg.grade for sg in subject_grades if sg.grade
    ))

    # Get categories (small table)
    categories = list(AssessmentCategory.objects.filter(is_active=True).order_by('order'))

    # Get grading system with scales prefetched
    grading_system = GradingSystem.objects.filter(
        is_active=True
    ).prefetch_related('scales').first()

    context = {
        'student': student,
        'current_term': current_term,
        'subject_grades': subject_grades,
        'core_grades': core_grades,
        'elective_grades': elective_grades,
        'term_report': term_report,
        'grade_summary': grade_summary,
        'categories': categories,
        'grading_system': grading_system,
    }

    return render(request, 'gradebook/partials/report_card.html', context)


@login_required
def report_remarks_edit(request, student_id):
    """Edit teacher remarks for a student's term report."""
    current_term = Term.get_current()
    student = get_object_or_404(Student, pk=student_id)

    if not current_term:
        return HttpResponse('No current term set', status=400)

    # Get or create term report
    term_report, created = TermReport.objects.get_or_create(
        student=student,
        term=current_term,
        defaults={'out_of': 1}
    )

    if request.method == 'GET':
        # Check if user can edit remarks
        can_edit_class_remark = False
        can_edit_head_remark = False

        # Class teacher can edit class teacher remark
        if hasattr(request.user, 'teacher_profile') and request.user.teacher_profile:
            teacher = request.user.teacher_profile
            # Check if teacher is class teacher for this student's class
            if student.current_class and student.current_class.class_teacher == teacher:
                can_edit_class_remark = True

        # School admin can edit both
        if is_school_admin(request.user):
            can_edit_class_remark = True
            can_edit_head_remark = True

        return render(request, 'gradebook/partials/remarks_edit.html', {
            'student': student,
            'term_report': term_report,
            'current_term': current_term,
            'can_edit_class_remark': can_edit_class_remark,
            'can_edit_head_remark': can_edit_head_remark,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    # Save remarks
    remark_type = request.POST.get('remark_type')
    remark_text = request.POST.get('remark', '').strip()

    if remark_type == 'class_teacher':
        # Verify permission
        can_edit = is_school_admin(request.user)
        if not can_edit and hasattr(request.user, 'teacher_profile') and request.user.teacher_profile:
            if student.current_class and student.current_class.class_teacher == request.user.teacher_profile:
                can_edit = True

        if can_edit:
            term_report.class_teacher_remark = remark_text
            term_report.save(update_fields=['class_teacher_remark'])

    elif remark_type == 'head_teacher':
        if is_school_admin(request.user):
            term_report.head_teacher_remark = remark_text
            term_report.save(update_fields=['head_teacher_remark'])

    response = HttpResponse(status=204)
    response['HX-Trigger'] = '{"showToast": {"message": "Remark saved successfully", "type": "success"}, "closeModal": true}'
    return response


@login_required
def report_card_print(request, student_id):
    """Print-friendly report card with Ghana-specific data.

    OPTIMIZED: Uses select_related and computes grade summary in-memory.

    For teachers: Only allow printing reports for students in their homeroom classes.
    For admins: Allow printing all reports.
    """
    current_term = Term.get_current()
    student = get_object_or_404(Student.objects.select_related('current_class'), pk=student_id)
    user = request.user

    # Permission check for teachers
    if not is_school_admin(user):
        if getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile'):
            teacher = user.teacher_profile
            # Teacher must be the form master of the student's class
            if not student.current_class or student.current_class.class_teacher != teacher:
                from django.contrib import messages
                messages.error(request, 'You can only print reports for students in your homeroom class.')
                return redirect('gradebook:reports')
        else:
            from django.contrib import messages
            messages.error(request, 'You do not have permission to print this report.')
            return redirect('core:index')

    # Get subject grades with core/elective distinction - single query
    subject_grades = list(SubjectTermGrade.objects.filter(
        student=student,
        term=current_term
    ).select_related('subject').order_by('-subject__is_core', 'subject__name'))

    # Separate core and elective subjects in memory
    core_grades = [sg for sg in subject_grades if sg.subject.is_core]
    elective_grades = [sg for sg in subject_grades if not sg.subject.is_core]

    term_report = TermReport.objects.filter(
        student=student,
        term=current_term
    ).first()

    # Compute grade summary in memory (no extra query)
    from collections import Counter
    grade_summary = dict(Counter(
        sg.grade for sg in subject_grades if sg.grade
    ))

    categories = list(AssessmentCategory.objects.filter(is_active=True).order_by('order'))

    # Calculate category-wise scores for each subject
    from django.db.models import Sum, F, Value
    from django.db.models.functions import Coalesce

    # Get all scores for this student in current term, grouped by subject and category
    category_scores = {}
    scores_qs = Score.objects.filter(
        student=student,
        assignment__term=current_term
    ).select_related('assignment__subject', 'assignment__assessment_category')

    for score in scores_qs:
        subject_id = score.assignment.subject_id
        category_id = score.assignment.assessment_category_id

        if subject_id not in category_scores:
            category_scores[subject_id] = {}
        if category_id not in category_scores[subject_id]:
            category_scores[subject_id][category_id] = {'earned': 0, 'possible': 0}

        category_scores[subject_id][category_id]['earned'] += float(score.points)
        category_scores[subject_id][category_id]['possible'] += float(score.assignment.points_possible)

    # Attach category scores to each subject grade
    for sg in subject_grades:
        sg.category_scores = {}
        subject_cat_scores = category_scores.get(sg.subject_id, {})
        for cat in categories:
            cat_data = subject_cat_scores.get(cat.pk, {'earned': 0, 'possible': 0})
            if cat_data['possible'] > 0:
                # Calculate percentage and apply category weight
                percentage = (cat_data['earned'] / cat_data['possible']) * 100
                weighted = (percentage * cat.percentage) / 100
                sg.category_scores[cat.pk] = round(weighted, 1)
            else:
                sg.category_scores[cat.pk] = None

    # Get grading system with scales for the grade key
    grading_system = GradingSystem.objects.filter(
        is_active=True
    ).prefetch_related('scales').first()

    # Get school/tenant info
    school = None
    school_settings = None
    try:
        from django.db import connection
        from schools.models import School
        from core.models import SchoolSettings
        school = School.objects.get(schema_name=connection.schema_name)
        school_settings = SchoolSettings.objects.first()
    except Exception:
        pass

    # Create verification record and generate QR code
    verification = None
    qr_code_base64 = None
    try:
        from core.models import DocumentVerification
        from core.utils import generate_verification_qr

        verification = DocumentVerification.create_for_document(
            document_type=DocumentVerification.DocumentType.REPORT_CARD,
            student=student,
            title=f"Report Card - {current_term.name}" if current_term else "Report Card",
            user=request.user,
            term=current_term,
            academic_year=current_term.academic_year.name if current_term and current_term.academic_year else '',
        )
        qr_code_base64 = generate_verification_qr(verification.verification_code, request=request)
    except Exception as e:
        logger.warning(f"Could not create verification record: {e}")

    context = {
        'student': student,
        'current_term': current_term,
        'subject_grades': subject_grades,
        'core_grades': core_grades,
        'elective_grades': elective_grades,
        'term_report': term_report,
        'grade_summary': grade_summary,
        'categories': categories,
        'grading_system': grading_system,
        'school': school,
        'school_settings': school_settings,
        'verification': verification,
        'qr_code_base64': qr_code_base64,
    }

    return render(request, 'gradebook/report_card_print.html', context)


# ============ Analytics Dashboard ============

@login_required
@admin_required
def analytics(request):
    """Analytics dashboard with grade trends and statistics (Admin only)."""
    current_term = Term.get_current()
    classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')

    # Get filter parameters
    class_id = request.GET.get('class')
    selected_class = None

    if class_id:
        selected_class = get_object_or_404(Class, pk=class_id)

    context = {
        'current_term': current_term,
        'classes': classes,
        'selected_class': selected_class,
    }

    return htmx_render(
        request,
        'gradebook/analytics.html',
        'gradebook/partials/analytics_content.html',
        context
    )


@login_required
def analytics_class_data(request, class_id):
    """Get analytics data for a specific class (HTMX partial)."""
    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)

    if not current_term:
        return render(request, 'gradebook/partials/analytics_class.html', {
            'error': 'No current term set'
        })

    # Get grading system based on class level
    grading_level = 'SHS' if class_obj.level_type == 'shs' else 'BASIC'
    grading_system = GradingSystem.objects.filter(
        level=grading_level,
        is_active=True
    ).first()

    # Use grading system thresholds or defaults
    pass_mark = grading_system.pass_mark if grading_system else config.DEFAULT_PASS_MARK
    min_avg_for_promotion = grading_system.min_average_for_promotion if grading_system else config.DEFAULT_MIN_AVERAGE_FOR_PROMOTION

    # Get all term reports for this class
    term_reports = list(TermReport.objects.filter(
        student__current_class=class_obj,
        term=current_term
    ).select_related('student').order_by('-average'))

    # Get subject grades for grade distribution
    subject_grades = list(SubjectTermGrade.objects.filter(
        student__current_class=class_obj,
        term=current_term,
        total_score__isnull=False
    ).select_related('subject', 'student'))

    # Calculate statistics
    stats = calculate_class_stats(term_reports, subject_grades)

    # Get subject performance comparison (using configurable pass mark)
    subject_performance = calculate_subject_performance(subject_grades, pass_mark=pass_mark)

    # Get grade distribution
    grade_distribution = calculate_grade_distribution(subject_grades)

    # Get top performers
    top_performers = term_reports[:config.TOP_PERFORMERS_LIMIT] if term_reports else []

    # Get students needing attention (failed 2+ subjects or avg below promotion threshold)
    at_risk = [
        r for r in term_reports
        if r.subjects_failed >= 2 or (r.average and r.average < min_avg_for_promotion)
    ][:config.AT_RISK_STUDENTS_LIMIT]

    context = {
        'class_obj': class_obj,
        'current_term': current_term,
        'stats': stats,
        'subject_performance': subject_performance,
        'grade_distribution': grade_distribution,
        'grade_distribution_json': json.dumps(grade_distribution),
        'subject_performance_json': json.dumps(subject_performance),
        'top_performers': top_performers,
        'at_risk_students': at_risk,
        'total_students': len(term_reports),
        'grading_system': grading_system,
        'pass_mark': pass_mark,
    }

    return render(request, 'gradebook/partials/analytics_class.html', context)


@login_required
@admin_required
def analytics_overview(request):
    """School-wide analytics overview (HTMX partial, Admin only)."""
    current_term = Term.get_current()

    if not current_term:
        return render(request, 'gradebook/partials/analytics_overview.html', {
            'error': 'No current term set'
        })

    # Get default pass mark from any active grading system (school-wide stats)
    default_grading_system = GradingSystem.objects.filter(is_active=True).first()
    default_pass_mark = default_grading_system.pass_mark if default_grading_system else config.DEFAULT_PASS_MARK

    # Get all classes with their stats
    classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')

    class_stats = []
    for cls in classes:
        reports = TermReport.objects.filter(
            student__current_class=cls,
            term=current_term
        ).aggregate(
            avg_score=Avg('average'),
            total_students=Count('id'),
            passed=Count('id', filter=Q(subjects_failed=0)),
        )

        if reports['total_students'] > 0:
            class_stats.append({
                'class': cls,
                'average': round(reports['avg_score'] or 0, 1),
                'total_students': reports['total_students'],
                'passed': reports['passed'],
                'pass_rate': round((reports['passed'] / reports['total_students']) * 100, 1) if reports['total_students'] > 0 else 0,
            })

    # Sort by average descending
    class_stats.sort(key=lambda x: x['average'], reverse=True)

    # Overall school stats
    all_reports = TermReport.objects.filter(term=current_term)
    school_stats = all_reports.aggregate(
        total_students=Count('id'),
        avg_score=Avg('average'),
        total_passed=Count('id', filter=Q(subjects_failed=0)),
    )

    # Subject-wise school performance (using configurable pass mark)
    subject_stats = SubjectTermGrade.objects.filter(
        term=current_term,
        total_score__isnull=False
    ).values('subject__name', 'subject__short_name').annotate(
        avg_score=Avg('total_score'),
        students=Count('id'),
        passed=Count('id', filter=Q(total_score__gte=default_pass_mark)),
    ).order_by('-avg_score')[:config.TOP_SUBJECTS_LIMIT]

    context = {
        'current_term': current_term,
        'class_stats': class_stats,
        'class_stats_json': json.dumps([
            {'name': s['class'].name, 'average': float(s['average'])}
            for s in class_stats
        ]),
        'school_stats': {
            'total_students': school_stats['total_students'] or 0,
            'average': round(school_stats['avg_score'] or 0, 1),
            'passed': school_stats['total_passed'] or 0,
            'pass_rate': round((school_stats['total_passed'] / school_stats['total_students']) * 100, 1) if school_stats['total_students'] else 0,
        },
        'subject_stats': list(subject_stats),
        'pass_mark': default_pass_mark,
    }

    return render(request, 'gradebook/partials/analytics_overview.html', context)


@login_required
@admin_required
def analytics_term_comparison(request):
    """Compare performance across terms (HTMX partial, Admin only)."""
    # Get all terms from current academic year
    current_term = Term.get_current()
    if not current_term:
        return render(request, 'gradebook/partials/analytics_terms.html', {
            'error': 'No current term set'
        })

    terms = Term.objects.filter(
        academic_year=current_term.academic_year
    ).order_by('term_number')

    term_data = []
    for term in terms:
        stats = TermReport.objects.filter(term=term).aggregate(
            avg_score=Avg('average'),
            total_students=Count('id'),
            passed=Count('id', filter=Q(subjects_failed=0)),
        )

        if stats['total_students'] > 0:
            term_data.append({
                'term': term,
                'average': round(stats['avg_score'] or 0, 1),
                'total_students': stats['total_students'],
                'passed': stats['passed'],
                'pass_rate': round((stats['passed'] / stats['total_students']) * 100, 1),
            })

    context = {
        'terms': term_data,
        'terms_json': json.dumps([
            {'name': t['term'].name, 'average': float(t['average']), 'pass_rate': float(t['pass_rate'])}
            for t in term_data
        ]),
        'current_term': current_term,
    }

    return render(request, 'gradebook/partials/analytics_terms.html', context)


def calculate_class_stats(term_reports, subject_grades):
    """Calculate comprehensive class statistics."""
    if not term_reports:
        return {
            'average': 0,
            'highest': 0,
            'lowest': 0,
            'pass_rate': 0,
            'subjects_avg_passed': 0,
        }

    averages = [r.average for r in term_reports if r.average is not None]

    if not averages:
        return {
            'average': 0,
            'highest': 0,
            'lowest': 0,
            'pass_rate': 0,
            'subjects_avg_passed': 0,
        }

    total_students = len(term_reports)
    passed = sum(1 for r in term_reports if r.subjects_failed == 0)
    avg_subjects_passed = sum(r.subjects_passed for r in term_reports) / total_students if total_students else 0

    return {
        'average': round(sum(averages) / len(averages), 1),
        'highest': round(max(averages), 1),
        'lowest': round(min(averages), 1),
        'pass_rate': round((passed / total_students) * 100, 1) if total_students else 0,
        'subjects_avg_passed': round(avg_subjects_passed, 1),
    }


def calculate_subject_performance(subject_grades, pass_mark=None):
    """
    Calculate per-subject performance metrics.

    Args:
        subject_grades: List of SubjectTermGrade objects
        pass_mark: The pass mark threshold (defaults to grading system standard)
    """
    if pass_mark is None:
        pass_mark = config.DEFAULT_PASS_MARK

    subject_data = defaultdict(lambda: {'scores': [], 'passed': 0, 'total': 0})

    for grade in subject_grades:
        subj = grade.subject.short_name or grade.subject.name[:10]
        subject_data[subj]['scores'].append(float(grade.total_score))
        subject_data[subj]['total'] += 1
        if grade.total_score >= pass_mark:
            subject_data[subj]['passed'] += 1

    result = []
    for name, data in subject_data.items():
        if data['scores']:
            result.append({
                'name': name,
                'average': round(sum(data['scores']) / len(data['scores']), 1),
                'pass_rate': round((data['passed'] / data['total']) * 100, 1) if data['total'] else 0,
                'students': data['total'],
            })

    # Sort by average descending
    result.sort(key=lambda x: x['average'], reverse=True)
    return result


def calculate_grade_distribution(subject_grades):
    """Calculate grade distribution across all subjects."""
    grade_counts = Counter(g.grade for g in subject_grades if g.grade)

    # Define grade order for display
    grade_order = ['A1', 'B2', 'B3', 'C4', 'C5', 'C6', 'D7', 'E8', 'F9']

    result = []
    for grade in grade_order:
        count = grade_counts.get(grade, 0)
        if count > 0 or grade in ['A1', 'B2', 'C6', 'F9']:  # Always show key grades
            result.append({
                'grade': grade,
                'count': count,
            })

    return result


# ============ Bulk Remarks Entry ============

@login_required
def bulk_remarks_entry(request, class_id):
    """
    Bulk remarks entry page for form teachers.
    Shows all students with their performance data and input fields.
    """
    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    user = request.user

    # Permission check - must be class teacher or admin
    if not is_school_admin(user):
        if not (getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile')):
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('gradebook:reports')
        if class_obj.class_teacher != user.teacher_profile:
            messages.error(request, 'You can only enter remarks for your homeroom class.')
            return redirect('gradebook:reports')

    # Get students with term reports
    students = list(Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('last_name', 'first_name'))

    if not students:
        messages.info(request, 'No active students found in this class.')
        return redirect('gradebook:reports')

    # Prefetch term reports
    student_ids = [s.id for s in students]
    reports = {
        r.student_id: r for r in TermReport.objects.filter(
            student_id__in=student_ids,
            term=current_term
        )
    }

    # Attach reports to students and count completed
    completed_count = 0
    for student in students:
        student.term_report = reports.get(student.id)
        if student.term_report and student.term_report.class_teacher_remark:
            completed_count += 1

    # Get remark templates
    remark_templates = RemarkTemplate.objects.filter(is_active=True).order_by('category', 'order')

    # Group templates by category
    templates_by_category = {}
    for template in remark_templates:
        category = template.get_category_display()
        if category not in templates_by_category:
            templates_by_category[category] = []
        templates_by_category[category].append(template)

    context = {
        'class_obj': class_obj,
        'students': students,
        'current_term': current_term,
        'templates_by_category': templates_by_category,
        'conduct_choices': TermReport.CONDUCT_CHOICES,
        'rating_choices': TermReport.RATING_CHOICES,
        'completed_count': completed_count,
        'total_count': len(students),
        'is_admin': is_school_admin(user),
    }

    return htmx_render(
        request,
        'gradebook/bulk_remarks.html',
        'gradebook/partials/bulk_remarks_content.html',
        context
    )


@login_required
def bulk_remark_save(request):
    """Save individual student remark via HTMX (auto-save)."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    student_id = request.POST.get('student_id')
    field = request.POST.get('field')
    value = request.POST.get('value', '').strip()

    current_term = Term.get_current()
    if not current_term:
        return HttpResponse('No current term', status=400)

    student = get_object_or_404(Student.objects.select_related('current_class'), pk=student_id)

    # Permission check
    user = request.user
    if not is_school_admin(user):
        if not (getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile')):
            return HttpResponse(status=403)
        if not student.current_class or student.current_class.class_teacher != user.teacher_profile:
            return HttpResponse(status=403)

    # Get or create term report
    term_report, created = TermReport.objects.get_or_create(
        student=student,
        term=current_term,
        defaults={'out_of': 1}
    )

    # Update the field
    allowed_fields = [
        'class_teacher_remark', 'conduct_rating', 'attitude_rating',
        'interest_rating', 'punctuality_rating'
    ]

    if field in allowed_fields:
        setattr(term_report, field, value)
        term_report.save(update_fields=[field])

        # Return success indicator
        response = HttpResponse(status=200)
        response['HX-Trigger'] = json.dumps({
            'remarkSaved': {
                'student_id': str(student_id),
                'field': field
            }
        })
        return response

    return HttpResponse('Invalid field', status=400)


@login_required
def bulk_remarks_sign(request, class_id):
    """Sign off all remarks for a class (class teacher confirmation)."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    user = request.user

    # Permission check
    if not is_school_admin(user):
        if not (getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile')):
            return HttpResponse(status=403)
        if class_obj.class_teacher != user.teacher_profile:
            return HttpResponse(status=403)

    # Sign all reports for this class
    from django.utils import timezone
    now = timezone.now()

    updated = TermReport.objects.filter(
        student__current_class=class_obj,
        term=current_term,
        class_teacher_signed=False
    ).update(
        class_teacher_signed=True,
        class_teacher_signed_at=now
    )

    response = HttpResponse(status=200)
    response['HX-Trigger'] = json.dumps({
        'showToast': {
            'message': f'Signed {updated} report(s) successfully',
            'type': 'success'
        },
        'refreshPage': True
    })
    return response


# ============ Remark Templates Management ============

@login_required
@admin_required
def remark_templates(request):
    """List and manage remark templates (Admin only)."""
    templates = RemarkTemplate.objects.all().order_by('category', 'order')

    # Group by category
    templates_by_category = {}
    for template in templates:
        category = template.get_category_display()
        if category not in templates_by_category:
            templates_by_category[category] = []
        templates_by_category[category].append(template)

    context = {
        'templates': templates,
        'templates_by_category': templates_by_category,
        'categories': RemarkTemplate.PERFORMANCE_CATEGORY,
    }

    return htmx_render(
        request,
        'gradebook/remark_templates.html',
        'gradebook/partials/remark_templates_content.html',
        context
    )


@login_required
@admin_required
def remark_template_create(request):
    """Create a new remark template."""
    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_remark_template.html', {
            'categories': RemarkTemplate.PERFORMANCE_CATEGORY,
            'mode': 'create',
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    category = request.POST.get('category', 'GENERAL')
    content = request.POST.get('content', '').strip()
    order = int(request.POST.get('order', 0))

    if not content:
        return render(request, 'gradebook/includes/modal_remark_template.html', {
            'categories': RemarkTemplate.PERFORMANCE_CATEGORY,
            'mode': 'create',
            'error': 'Remark content is required',
            'form_data': {'category': category, 'content': content, 'order': order},
        })

    RemarkTemplate.objects.create(
        category=category,
        content=content,
        order=order,
    )

    response = HttpResponse(status=204)
    response['HX-Trigger'] = json.dumps({
        'closeModal': True,
        'showToast': {'message': 'Template created successfully', 'type': 'success'},
        'refreshTemplates': True
    })
    return response


@login_required
@admin_required
def remark_template_edit(request, pk):
    """Edit a remark template."""
    template = get_object_or_404(RemarkTemplate, pk=pk)

    if request.method == 'GET':
        return render(request, 'gradebook/includes/modal_remark_template.html', {
            'template': template,
            'categories': RemarkTemplate.PERFORMANCE_CATEGORY,
            'mode': 'edit',
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    category = request.POST.get('category', 'GENERAL')
    content = request.POST.get('content', '').strip()
    order = int(request.POST.get('order', 0))
    is_active = request.POST.get('is_active') == 'on'

    if not content:
        return render(request, 'gradebook/includes/modal_remark_template.html', {
            'template': template,
            'categories': RemarkTemplate.PERFORMANCE_CATEGORY,
            'mode': 'edit',
            'error': 'Remark content is required',
        })

    template.category = category
    template.content = content
    template.order = order
    template.is_active = is_active
    template.save()

    response = HttpResponse(status=204)
    response['HX-Trigger'] = json.dumps({
        'closeModal': True,
        'showToast': {'message': 'Template updated successfully', 'type': 'success'},
        'refreshTemplates': True
    })
    return response


@login_required
@admin_required
def remark_template_delete(request, pk):
    """Delete a remark template."""
    template = get_object_or_404(RemarkTemplate, pk=pk)

    if request.method != 'POST':
        return HttpResponse(status=405)

    template.delete()

    response = HttpResponse(status=200)
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Template deleted', 'type': 'success'},
        'refreshTemplates': True
    })
    return response


# ============ Report Distribution ============

@login_required
def report_distribution(request, class_id):
    """
    Report distribution page for a class.
    Shows students with their contact info and distribution status.
    """
    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    user = request.user

    # Permission check - must be class teacher or admin
    if not is_school_admin(user):
        if not (getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile')):
            messages.error(request, 'You do not have permission to access this page.')
            return redirect('gradebook:reports')
        if class_obj.class_teacher != user.teacher_profile:
            messages.error(request, 'You can only distribute reports for your homeroom class.')
            return redirect('gradebook:reports')

    # Get students with term reports
    students = list(Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('last_name', 'first_name'))

    if not students:
        messages.info(request, 'No active students found in this class.')
        return redirect('gradebook:reports')

    # Prefetch term reports and distribution logs
    student_ids = [s.id for s in students]
    reports = {
        r.student_id: r for r in TermReport.objects.filter(
            student_id__in=student_ids,
            term=current_term
        )
    }

    # Get latest distribution logs per student
    distribution_logs = {}
    for log in ReportDistributionLog.objects.filter(
        term_report__student_id__in=student_ids,
        term_report__term=current_term
    ).select_related('term_report').order_by('-created_at'):
        student_id = log.term_report.student_id
        if student_id not in distribution_logs:
            distribution_logs[student_id] = log

    # Calculate stats
    with_email = 0
    with_phone = 0
    already_sent = 0

    for student in students:
        student.term_report = reports.get(student.id)
        student.last_distribution = distribution_logs.get(student.id)

        # Check for guardian contact
        guardian_email = getattr(student, 'guardian_email', None) or getattr(student, 'parent_email', None)
        guardian_phone = getattr(student, 'guardian_phone', None) or getattr(student, 'parent_phone', None)

        student.has_email = bool(guardian_email)
        student.has_phone = bool(guardian_phone)
        student.guardian_email = guardian_email
        student.guardian_phone = guardian_phone

        if student.has_email:
            with_email += 1
        if student.has_phone:
            with_phone += 1
        if student.last_distribution:
            already_sent += 1

    context = {
        'class_obj': class_obj,
        'students': students,
        'current_term': current_term,
        'is_admin': is_school_admin(user),
        'stats': {
            'total': len(students),
            'with_email': with_email,
            'with_phone': with_phone,
            'already_sent': already_sent,
        },
        'distribution_types': [
            ('EMAIL', 'Email with PDF'),
            ('SMS', 'SMS Summary'),
            ('BOTH', 'Email and SMS'),
        ],
    }

    return htmx_render(
        request,
        'gradebook/report_distribution.html',
        'gradebook/partials/report_distribution_content.html',
        context
    )


@login_required
def send_single_report(request, student_id):
    """Send report to a single student's guardian."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    current_term = Term.get_current()
    if not current_term:
        return JsonResponse({'success': False, 'error': 'No current term'})

    student = get_object_or_404(Student.objects.select_related('current_class'), pk=student_id)
    user = request.user

    # Permission check
    if not is_school_admin(user):
        if not (getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile')):
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        if not student.current_class or student.current_class.class_teacher != user.teacher_profile:
            return JsonResponse({'success': False, 'error': 'Permission denied'})

    # Get term report
    try:
        term_report = TermReport.objects.get(student=student, term=current_term)
    except TermReport.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No report found for this student'})

    distribution_type = request.POST.get('type', 'EMAIL')
    sms_template = request.POST.get('sms_template', '').strip() or None

    # Queue the task
    from .tasks import distribute_single_report
    from django.db import connection

    distribute_single_report.delay(
        str(term_report.pk),
        distribution_type,
        connection.schema_name,
        user.pk,
        sms_template
    )

    response = HttpResponse(status=200)
    response['HX-Trigger'] = json.dumps({
        'showToast': {
            'message': f'Report queued for {student.first_name}',
            'type': 'success'
        },
        'refreshRow': str(student_id)
    })
    return response


@login_required
def send_bulk_reports(request, class_id):
    """Send reports for all students in a class."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    current_term = Term.get_current()
    if not current_term:
        return JsonResponse({'success': False, 'error': 'No current term'})

    class_obj = get_object_or_404(Class, pk=class_id)
    user = request.user

    # Permission check
    if not is_school_admin(user):
        if not (getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile')):
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        if class_obj.class_teacher != user.teacher_profile:
            return JsonResponse({'success': False, 'error': 'Permission denied'})

    distribution_type = request.POST.get('type', 'EMAIL')
    sms_template = request.POST.get('sms_template', '').strip() or None

    # Queue the bulk task
    from .tasks import distribute_bulk_reports
    from django.db import connection

    distribute_bulk_reports.delay(
        class_obj.pk,
        distribution_type,
        connection.schema_name,
        user.pk,
        sms_template
    )

    response = HttpResponse(status=200)
    response['HX-Trigger'] = json.dumps({
        'showToast': {
            'message': f'Reports queued for all students in {class_obj.name}',
            'type': 'success'
        }
    })
    return response


@login_required
def download_report_pdf(request, student_id):
    """Download PDF report for a student."""
    current_term = Term.get_current()
    if not current_term:
        messages.error(request, 'No current term set.')
        return redirect('gradebook:reports')

    student = get_object_or_404(Student.objects.select_related('current_class'), pk=student_id)
    user = request.user

    # Permission check
    if not is_school_admin(user):
        if not (getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile')):
            messages.error(request, 'Permission denied.')
            return redirect('gradebook:reports')
        if not student.current_class or student.current_class.class_teacher != user.teacher_profile:
            messages.error(request, 'Permission denied.')
            return redirect('gradebook:reports')

    try:
        term_report = TermReport.objects.get(student=student, term=current_term)
    except TermReport.DoesNotExist:
        messages.error(request, 'No report found for this student.')
        return redirect('gradebook:reports')

    # Generate PDF
    from .tasks import generate_report_pdf
    from django.db import connection

    try:
        pdf_buffer = generate_report_pdf(term_report, connection.schema_name)

        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="report_card_{student.admission_number}.pdf"'
        return response

    except Exception as e:
        logger.error(f"Failed to generate PDF: {str(e)}")
        messages.error(request, f'Failed to generate PDF: {str(e)}')
        return redirect('gradebook:reports')


# =============================================================================
# TRANSCRIPT
# =============================================================================

@login_required
def transcript(request, student_id):
    """
    View a student's complete academic transcript.
    Shows all terms, grades, and cumulative performance.
    """
    student = get_object_or_404(
        Student.objects.select_related('current_class'),
        pk=student_id
    )

    # Permission check
    has_permission, error_msg = check_transcript_permission(request.user, student)
    if not has_permission:
        messages.error(request, error_msg)
        return redirect('gradebook:reports' if 'homeroom' in error_msg else 'core:index')

    # Get transcript data
    term_reports, all_grades, grades_by_term = get_transcript_data(student)

    # Build academic history with cumulative stats
    history_data = build_academic_history(term_reports, grades_by_term, include_all_grades=True)

    # Promotion history
    promotion_history = term_reports.filter(promoted__isnull=False).values(
        'term__academic_year__name',
        'term__name',
        'promoted',
        'promoted_to__name',
        'promotion_remarks'
    )

    # Get school context
    school_ctx = get_school_context()

    context = {
        'student': student,
        'academic_history': history_data['academic_history'],
        'term_reports': term_reports,
        'cumulative_stats': {
            'total_terms': history_data['term_count'],
            'total_subjects_taken': history_data['total_subjects_taken'],
            'total_subjects_passed': history_data['total_subjects_passed'],
            'total_credits': history_data['total_credits'],
            'cumulative_average': history_data['cumulative_average'],
            'unique_subjects': len(history_data['unique_subjects']),
        },
        'promotion_history': list(promotion_history),
        'school': school_ctx['school'],
        'school_settings': school_ctx['school_settings'],
    }

    return htmx_render(
        request,
        'gradebook/transcript.html',
        'gradebook/partials/transcript_content.html',
        context
    )


@login_required
def transcript_print(request, student_id):
    """Print-friendly transcript view."""
    student = get_object_or_404(
        Student.objects.select_related('current_class'),
        pk=student_id
    )

    # Permission check
    has_permission, error_msg = check_transcript_permission(request.user, student)
    if not has_permission:
        messages.error(request, 'Permission denied.')
        return redirect('gradebook:reports' if 'homeroom' in (error_msg or '') else 'core:index')

    # Get transcript data and build academic history
    term_reports, _, grades_by_term = get_transcript_data(student)
    history_data = build_academic_history(term_reports, grades_by_term)

    # Get school context
    school_ctx = get_school_context()

    # Create verification record and generate QR code
    verification = None
    qr_code_base64 = None
    try:
        from core.models import DocumentVerification
        from core.utils import generate_verification_qr

        verification = DocumentVerification.create_for_document(
            document_type=DocumentVerification.DocumentType.TRANSCRIPT,
            student=student,
            title=f"Academic Transcript - {student.full_name}",
            user=request.user,
        )
        qr_code_base64 = generate_verification_qr(verification.verification_code, request=request)
    except Exception as e:
        logger.warning(f"Could not create verification record: {e}")

    context = {
        'student': student,
        'academic_history': history_data['academic_history'],
        'cumulative_average': history_data['cumulative_average'],
        'total_terms': history_data['term_count'],
        'total_credits': history_data['total_credits'],
        'generated_date': timezone.now(),
        'school': school_ctx['school'],
        'school_settings': school_ctx['school_settings'],
        'verification': verification,
        'qr_code_base64': qr_code_base64,
    }

    return render(request, 'gradebook/transcript_print.html', context)


@login_required
def download_transcript_pdf(request, student_id):
    """Download PDF transcript for a student."""
    student = get_object_or_404(
        Student.objects.select_related('current_class'),
        pk=student_id
    )

    # Permission check
    has_permission, error_msg = check_transcript_permission(request.user, student)
    if not has_permission:
        messages.error(request, 'Permission denied.')
        return redirect('gradebook:reports')

    # Get transcript data
    term_reports, _, grades_by_term = get_transcript_data(student)

    if not term_reports.exists():
        messages.error(request, 'No academic records found for this student.')
        return redirect('gradebook:reports')

    # Build academic history
    history_data = build_academic_history(term_reports, grades_by_term)

    # Get school context with base64 logo for PDF
    school_ctx = get_school_context(include_logo_base64=True)

    # Create verification record and generate QR code
    from core.models import DocumentVerification
    from core.utils import generate_verification_qr

    verification = DocumentVerification.create_for_document(
        document_type=DocumentVerification.DocumentType.TRANSCRIPT,
        student=student,
        title=f"Academic Transcript - {student.full_name}",
        user=request.user,
    )
    qr_code_base64 = generate_verification_qr(verification.verification_code, request=request)

    context = {
        'student': student,
        'academic_history': history_data['academic_history'],
        'cumulative_average': history_data['cumulative_average'],
        'total_terms': history_data['term_count'],
        'total_credits': history_data['total_credits'],
        'generated_date': timezone.now(),
        'request': request,
        'school': school_ctx['school'],
        'school_settings': school_ctx['school_settings'],
        'logo_base64': school_ctx['logo_base64'],
        'verification': verification,
        'qr_code_base64': qr_code_base64,
    }

    # Generate PDF using WeasyPrint
    try:
        from weasyprint import HTML
        from django.template.loader import render_to_string
        from django.conf import settings as django_settings
        from io import BytesIO

        html_string = render_to_string('gradebook/transcript_pdf.html', context)
        html = HTML(string=html_string, base_url=str(django_settings.BASE_DIR))
        pdf_buffer = BytesIO()
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="transcript_{student.admission_number}.pdf"'
        return response

    except ImportError:
        logger.error("WeasyPrint not installed")
        messages.error(request, 'PDF generation is not available. WeasyPrint is not installed.')
        return redirect('gradebook:transcript', student_id=student_id)
    except Exception as e:
        import traceback
        logger.error(f"Failed to generate transcript PDF: {str(e)}\n{traceback.format_exc()}")
        messages.error(request, f'Failed to generate PDF: {str(e)}')
        return redirect('gradebook:transcript', student_id=student_id)
