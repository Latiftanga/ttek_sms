import zipfile

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse as DjangoHttpResponse
from decimal import Decimal, InvalidOperation
import logging
import json
import uuid

from django.core.cache import cache as django_cache
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import transaction


from .base import (
    teacher_or_admin_required, admin_required, can_edit_scores, get_client_ip
)
from ..models import (
    Assignment, Score, ScoreAuditLog, SubjectTermGrade, TermReport
)
from ..signals import signals_disabled
from .. import config
from academics.models import Class, Subject
from students.models import Student
from core.models import Term

logger = logging.getLogger(__name__)

# ============ Bulk Score Import ============

@login_required
@teacher_or_admin_required
def score_import_template(request, class_id, subject_id):
    """Download Excel template for score import."""
    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    # Check authorization
    if not can_edit_scores(request.user, class_obj, subject):
        return HttpResponse("Not authorized", status=403)

    # Get students and assignments
    students = Student.objects.filter(
        current_class=class_obj
    ).order_by('last_name', 'first_name')

    assignments = Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=config.EXCEL_HEADER_COLOR, end_color=config.EXCEL_HEADER_COLOR, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ["Student ID", "Student Name"]
    for assign in assignments:
        headers.append(f"{assign.assessment_category.short_name}: {assign.name} (/{assign.points_possible})")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Get existing scores
    existing_scores = {}
    for score in Score.objects.filter(
        student__in=students,
        assignment__in=assignments
    ).select_related('student', 'assignment'):
        key = (score.student_id, score.assignment_id)
        existing_scores[key] = score.points

    # Data rows
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.admission_number).border = thin_border
        ws.cell(row=row, column=2, value=f"{student.last_name}, {student.first_name}").border = thin_border

        for col, assign in enumerate(assignments, 3):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            # Pre-fill existing scores
            existing = existing_scores.get((student.id, assign.id))
            if existing is not None:
                cell.value = float(existing)

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Add metadata sheet for import validation
    meta_ws = wb.create_sheet("_metadata")
    meta_ws.cell(row=1, column=1, value="class_id")
    meta_ws.cell(row=1, column=2, value=str(class_id))
    meta_ws.cell(row=2, column=1, value="subject_id")
    meta_ws.cell(row=2, column=2, value=str(subject_id))
    meta_ws.cell(row=3, column=1, value="term_id")
    meta_ws.cell(row=3, column=2, value=str(current_term.id) if current_term else "")

    # Assignment IDs in order
    for col, assign in enumerate(assignments, 1):
        meta_ws.cell(row=4, column=col, value=str(assign.id))

    # Hide metadata sheet
    meta_ws.sheet_state = 'hidden'

    # Create response
    response = DjangoHttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"scores_{class_obj.name}_{subject.short_name}_{current_term.name if current_term else 'noterm'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


@login_required
@teacher_or_admin_required
def score_import_upload(request, class_id, subject_id):
    """Handle score import file upload and show preview."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    current_term = Term.get_current()
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    # Check authorization
    if not can_edit_scores(request.user, class_obj, subject):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'You are not authorized to import scores for this subject.'
        })

    # Check if grades are locked
    if current_term and current_term.grades_locked:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Grades are locked for this term.'
        })

    file = request.FILES.get('file')
    if not file:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'No file uploaded.'
        })

    # File size validation
    if file.size > config.MAX_FILE_SIZE:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'File too large. Maximum size is 5MB.'
        })

    if not file.name.endswith('.xlsx'):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Please upload an Excel file (.xlsx).'
        })

    try:
        wb = openpyxl.load_workbook(file, read_only=True)
    except (openpyxl.utils.exceptions.InvalidFileException, KeyError, ValueError, zipfile.BadZipFile):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Could not read the Excel file. Please ensure it is a valid .xlsx file.'
        })

    try:
        ws = wb.active

        # Get assignments for validation
        assignments = list(Assignment.objects.filter(
            subject=subject,
            term=current_term
        ).select_related('assessment_category').order_by('assessment_category__order', 'name'))

        # Get students lookup
        students_by_id = {
            s.admission_number: s for s in Student.objects.filter(current_class=class_obj)
        }

        # Parse data
        preview_data = []
        errors = []
        row_num = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or not row[0]:  # Skip empty rows
                continue

            student_id = str(row[0]).strip()
            student = students_by_id.get(student_id)

            row_data = {
                'row_num': row_num + 1,
                'student_id': student_id,
                'student_name': row[1] if len(row) > 1 else '',
                'student': student,
                'scores': [],
                'has_error': False,
            }

            if not student:
                row_data['has_error'] = True
                errors.append(f"Row {row_num + 1}: Student ID '{student_id}' not found in this class.")

            # Parse scores
            for col, assign in enumerate(assignments, 2):
                value = row[col] if len(row) > col else None
                score_data = {
                    'assignment': assign,
                    'value': value,
                    'error': None,
                }

                if value is not None and value != '':
                    try:
                        points = Decimal(str(value))
                        if points < 0:
                            score_data['error'] = 'Negative value'
                            row_data['has_error'] = True
                            errors.append(f"Row {row_num + 1}, {assign.name}: Negative value not allowed.")
                        elif points > assign.points_possible:
                            score_data['error'] = f'Exceeds max ({assign.points_possible})'
                            row_data['has_error'] = True
                            errors.append(f"Row {row_num + 1}, {assign.name}: Value {points} exceeds maximum {assign.points_possible}.")
                        else:
                            score_data['value'] = points
                    except (InvalidOperation, ValueError):
                        score_data['error'] = 'Invalid number'
                        row_data['has_error'] = True
                        errors.append(f"Row {row_num + 1}, {assign.name}: Invalid number '{value}'.")

                row_data['scores'].append(score_data)

            preview_data.append(row_data)

        # Store data in session for confirmation
        import_data = []
        for row in preview_data:
            if row['student'] and not row['has_error']:
                for score in row['scores']:
                    if score['value'] is not None and score['value'] != '' and not score['error']:
                        import_data.append({
                            'student_id': str(row['student'].id),
                            'assignment_id': str(score['assignment'].id),
                            'points': str(score['value']),
                        })

        nonce = uuid.uuid4().hex[:12]
        cache_key = f'score_import:{request.user.pk}:{nonce}'
        django_cache.set(cache_key, json.dumps({
            'data': import_data,
            'class_id': str(class_id),
            'subject_id': str(subject_id),
        }), 1800)  # 30 min TTL
        request.session['score_import_cache_key'] = cache_key

        return render(request, 'gradebook/partials/import_preview.html', {
            'class_obj': class_obj,
            'subject': subject,
            'assignments': assignments,
            'preview_data': preview_data,
            'errors': errors,
            'total_scores': len(import_data),
            'has_errors': len(errors) > 0,
        })

    except (KeyError, IndexError, TypeError, ValueError, InvalidOperation) as e:
        logger.exception("Error parsing import file: %s", e)
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Error reading file. Please ensure it is a valid score import template.'
        })
    finally:
        wb.close()


@login_required
@teacher_or_admin_required
def score_import_confirm(request, class_id, subject_id):
    """Confirm and execute score import."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    current_term = Term.get_current()

    # Check authorization
    if not can_edit_scores(request.user, class_obj, subject):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'You are not authorized to import scores for this subject.'
        })

    # Check if grades are locked
    if current_term and current_term.grades_locked:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Grades are locked for this term.'
        })

    # Get data from cache
    cache_key = request.session.get('score_import_cache_key', '')
    cached = django_cache.get(cache_key) if cache_key else None
    if not cached:
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'No import data found. Please upload the file again.'
        })

    try:
        cached_data = json.loads(cached)
        import_data = cached_data['data']
    except (json.JSONDecodeError, KeyError):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Invalid import data. Please upload the file again.'
        })

    # Validate cached data matches current request
    if (cached_data.get('class_id') != str(class_id) or
        cached_data.get('subject_id') != str(subject_id)):
        return render(request, 'gradebook/partials/import_error.html', {
            'error': 'Import data mismatch. Please upload the file again.'
        })

    # Ownership validation: Verify all students and assignments belong to this class/subject
    valid_student_ids = set(
        Student.objects.filter(
            current_class=class_obj,
            status='active'
        ).values_list('pk', flat=True)
    )
    valid_assignment_ids = set(
        Assignment.objects.filter(
            subject=subject,
            term=current_term
        ).values_list('pk', flat=True)
    )

    # Validate each item in import data
    invalid_items = []
    for idx, item in enumerate(import_data):
        student_id = item.get('student_id')
        assignment_id = item.get('assignment_id')

        # Convert to appropriate types for comparison
        try:
            if isinstance(student_id, str):
                student_id = int(student_id)
        except (ValueError, TypeError):
            pass
        try:
            if isinstance(assignment_id, str):
                from uuid import UUID
                assignment_id = UUID(assignment_id)
        except (ValueError, TypeError):
            pass

        if student_id not in valid_student_ids:
            invalid_items.append(f"Row {idx + 1}: Student ID {student_id} not in this class")
        if assignment_id not in valid_assignment_ids:
            invalid_items.append(f"Row {idx + 1}: Assignment ID {assignment_id} not for this subject/term")

    if invalid_items:
        logger.warning(f"Bulk import validation failed: {invalid_items[:5]}")
        return render(request, 'gradebook/partials/import_error.html', {
            'error': f'Data validation failed. {len(invalid_items)} items have invalid student or assignment references.',
            'details': invalid_items[:10]  # Show first 10 errors
        })

    # Get audit context
    client_ip = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')[:255]

    # Import scores
    created_count = 0
    updated_count = 0
    affected_combos = set()  # Track (student_id, subject_id, term_id) for batch recalc

    # Pre-fetch existing scores to avoid per-row queries
    import_keys = [(item['student_id'], item['assignment_id']) for item in import_data]
    existing_scores = {}
    if import_keys:
        student_ids = {k[0] for k in import_keys}
        assignment_ids = {k[1] for k in import_keys}
        for s in Score.objects.filter(student_id__in=student_ids, assignment_id__in=assignment_ids):
            existing_scores[(s.student_id, s.assignment_id)] = s.points

    with signals_disabled(), transaction.atomic():
        for item in import_data:
            student_id = item['student_id']
            assignment_id = item['assignment_id']
            points = Decimal(item['points'])

            old_value = existing_scores.get((student_id, assignment_id))

            score, created = Score.objects.update_or_create(
                student_id=student_id,
                assignment_id=assignment_id,
                defaults={'points': points}
            )

            # Track affected combos for batch recalculation
            affected_combos.add((student_id, str(subject.id), str(current_term.id)))

            # Audit log
            ScoreAuditLog.objects.create(
                score=score,
                student_id=student_id,
                assignment_id=assignment_id,
                user=request.user,
                action='CREATE' if created else 'UPDATE',
                old_value=old_value,
                new_value=points,
                ip_address=client_ip,
                user_agent=f"BULK_IMPORT: {user_agent[:240]}"
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

    # Batch recalculate grades for all affected students after signals re-enabled
    from ..signals import recalculate_subject_grade, recalculate_term_report
    affected_student_ids = {combo[0] for combo in affected_combos}
    student_map = {s.pk: s for s in Student.objects.filter(pk__in=affected_student_ids)}

    for student_id, subj_id, term_id in affected_combos:
        student_obj = student_map.get(student_id)
        if student_obj:
            recalculate_subject_grade(
                student=student_obj,
                subject=subject,
                term=current_term
            )

    for student_id in affected_student_ids:
        student_obj = student_map.get(student_id)
        if student_obj:
            recalculate_term_report(
                student=student_obj,
                term=current_term
            )

    # Clear cached import data
    cache_key = request.session.pop('score_import_cache_key', '')
    if cache_key:
        django_cache.delete(cache_key)

    return render(request, 'gradebook/partials/import_success.html', {
        'created_count': created_count,
        'updated_count': updated_count,
        'total_count': created_count + updated_count,
        'class_obj': class_obj,
        'subject': subject,
    })


# ============ Grade Export ============

@login_required
@admin_required
def export_class_grades(request, class_id):
    """Export all grades for a class to Excel with subject breakdown."""
    current_term = Term.get_current()
    if not current_term:
        return HttpResponse('No current term', status=400)

    class_obj = get_object_or_404(Class, pk=class_id)

    # Get term reports with student info
    reports = TermReport.objects.filter(
        student__current_class=class_obj,
        term=current_term,
    ).select_related('student').order_by('student__last_name', 'student__first_name')

    # Get subject grades
    subject_grades = SubjectTermGrade.objects.filter(
        student__current_class=class_obj,
        term=current_term,
    ).select_related('subject')

    # Build subject grades lookup: {student_id: {subject_name: total_score}}
    grades_by_student = {}
    subject_names = set()
    for sg in subject_grades:
        if sg.student_id not in grades_by_student:
            grades_by_student[sg.student_id] = {}
        subj_name = sg.subject.short_name or sg.subject.name
        grades_by_student[sg.student_id][subj_name] = {
            'score': sg.total_score,
            'grade': sg.grade,
            'position': sg.position,
        }
        subject_names.add(subj_name)

    sorted_subjects = sorted(subject_names)

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{class_obj.name} Grades"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(sorted_subjects))
    title_cell = ws.cell(row=1, column=1, value=f"{class_obj.name} — {current_term.name} Grade Report")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Headers (row 3)
    headers = ['#', 'Adm. No.', 'Student Name'] + sorted_subjects + ['Average', 'Position', 'Aggregate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    for row_idx, report in enumerate(reports, 4):
        student = report.student
        student_grades = grades_by_student.get(student.id, {})

        ws.cell(row=row_idx, column=1, value=row_idx - 3).border = thin_border
        ws.cell(row=row_idx, column=2, value=student.admission_number).border = thin_border
        name_cell = ws.cell(row=row_idx, column=3, value=student.full_name)
        name_cell.border = thin_border
        name_cell.font = Font(bold=True)

        for col_offset, subj in enumerate(sorted_subjects):
            cell = ws.cell(row=row_idx, column=4 + col_offset)
            sg = student_grades.get(subj)
            if sg and sg['score'] is not None:
                cell.value = float(sg['score'])
            cell.alignment = center
            cell.border = thin_border

        # Summary columns
        avg_col = 4 + len(sorted_subjects)
        ws.cell(row=row_idx, column=avg_col, value=float(report.average) if report.average else None).border = thin_border
        ws.cell(row=row_idx, column=avg_col, value=float(report.average) if report.average else None).alignment = center
        ws.cell(row=row_idx, column=avg_col + 1, value=report.position).border = thin_border
        ws.cell(row=row_idx, column=avg_col + 1).alignment = center
        ws.cell(row=row_idx, column=avg_col + 2, value=report.aggregate).border = thin_border
        ws.cell(row=row_idx, column=avg_col + 2).alignment = center

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or '')) for r in range(3, ws.max_row + 1)),
            default=8
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 30)

    ws.column_dimensions['C'].width = 25  # Student name column

    # Return as download
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = DjangoHttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{class_obj.name}_{current_term.name}_grades.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response