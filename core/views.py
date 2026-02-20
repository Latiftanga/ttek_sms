import html
import logging
from decimal import Decimal
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, FileResponse, JsonResponse
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_GET
from django.db import connection

from .models import SchoolSettings, AcademicYear, Term
from .email_backend import get_from_email
from .utils import ratelimit, admin_required, htmx_render
from .forms import (
    SchoolBasicInfoForm,
    SchoolBrandingForm,
    SchoolContactForm,
    SchoolAdminForm,
    AcademicSettingsForm,
    AcademicYearForm,
    TermForm,
)
from finance.models import PaymentGateway, PaymentGatewayConfig
from teachers.models import Teacher

logger = logging.getLogger(__name__)


# ============================================
# PWA / Offline Support Views
# ============================================

def offline(request):
    """Offline fallback page for service worker."""
    return render(request, 'core/offline.html')


@require_GET
@cache_control(max_age=0, no_cache=True, no_store=True, must_revalidate=True)
def service_worker(request):
    """Serve service worker from root scope."""
    import os
    sw_path = os.path.join(settings.BASE_DIR, 'core', 'static', 'core', 'sw.js')
    return FileResponse(
        open(sw_path, 'rb'),
        content_type='application/javascript',
        headers={'Service-Worker-Allowed': '/'}
    )


@require_GET
def manifest(request):
    """Generate PWA manifest with school-specific branding."""
    # Get school branding from tenant (School model)
    try:
        school = connection.tenant
        school_name = school.display_name if school else 'School'
        short_name = getattr(school, 'short_name', None) or school_name[:12]
        theme_color = school.primary_color if school and school.primary_color else '#570df8'
        background_color = '#f2f2f2'

        # Build icon URLs
        icons = []
        if school and school.logo:
            icons.append({
                'src': school.logo.url,
                'sizes': '192x192',
                'type': 'image/png',
                'purpose': 'any maskable'
            })
            icons.append({
                'src': school.logo.url,
                'sizes': '512x512',
                'type': 'image/png',
                'purpose': 'any maskable'
            })
        else:
            # Default icons if no school logo
            icons = [
                {
                    'src': '/static/icons/icon-192.svg',
                    'sizes': '192x192',
                    'type': 'image/svg+xml',
                    'purpose': 'any'
                },
                {
                    'src': '/static/icons/icon-512.svg',
                    'sizes': '512x512',
                    'type': 'image/svg+xml',
                    'purpose': 'any'
                }
            ]
    except Exception as e:
        # Log the error and use defaults - manifest should always return valid JSON
        logger.warning(f"Could not load school settings for manifest: {e}")
        school_name = 'School Management'
        short_name = 'School'
        theme_color = '#570df8'
        background_color = '#f2f2f2'
        icons = [
            {
                'src': '/static/icons/icon-192.svg',
                'sizes': '192x192',
                'type': 'image/svg+xml',
                'purpose': 'any'
            },
            {
                'src': '/static/icons/icon-512.svg',
                'sizes': '512x512',
                'type': 'image/svg+xml',
                'purpose': 'any'
            }
        ]

    manifest_data = {
        'name': school_name,
        'short_name': short_name,
        'description': f'{school_name} - School Management System',
        'id': '/',
        'start_url': '/',
        'scope': '/',
        'display': 'standalone',
        'orientation': 'portrait-primary',
        'theme_color': theme_color,
        'background_color': background_color,
        'icons': icons,
        'categories': ['education', 'productivity'],
        'screenshots': [
            {
                'src': '/static/screenshots/dashboard-wide.png',
                'sizes': '1280x720',
                'type': 'image/png',
                'form_factor': 'wide',
                'label': 'Dashboard overview'
            },
            {
                'src': '/static/screenshots/dashboard-mobile.png',
                'sizes': '390x844',
                'type': 'image/png',
                'form_factor': 'narrow',
                'label': 'Mobile dashboard'
            }
        ],
        'shortcuts': [
            {
                'name': 'Dashboard',
                'short_name': 'Home',
                'url': '/',
                'icons': [{'src': '/static/icons/home.svg', 'sizes': '96x96', 'type': 'image/svg+xml'}]
            },
            {
                'name': 'Take Attendance',
                'short_name': 'Attendance',
                'url': '/my-attendance/',
                'icons': [{'src': '/static/icons/attendance.svg', 'sizes': '96x96', 'type': 'image/svg+xml'}]
            }
        ],
        'prefer_related_applications': False,
        'handle_links': 'preferred'
    }

    return JsonResponse(manifest_data, content_type='application/manifest+json')




# ============================================
# Setup Wizard Views
# ============================================

@admin_required
def setup_wizard(request):
    """
    Setup wizard to guide new schools through initial configuration.
    Steps: Academic Year -> Terms -> Classes -> Houses -> Seed Data
    Education system is configured at tenant level by superadmin.
    """
    from academics.models import Programme, Subject, Class
    from students.models import House

    # Get tenant for education system configuration
    tenant = request.tenant

    # Ensure programmes exist for SHS class creation (seed if needed) - only if SHS is enabled
    programmes = Programme.objects.none()
    if tenant.has_programmes:
        programmes = Programme.objects.filter(is_active=True)
        if not programmes.exists():
            # Create default Ghana SHS programmes
            default_programmes = [
                ('General Arts', 'ART'),
                ('General Science', 'SCI'),
                ('Business', 'BUS'),
                ('Visual Arts', 'VIS'),
                ('Home Economics', 'HEC'),
                ('Agricultural Science', 'AGR'),
                ('Technical', 'TEC'),
            ]
            for name, code in default_programmes:
                Programme.objects.get_or_create(code=code, defaults={'name': name, 'is_active': True})
            programmes = Programme.objects.filter(is_active=True)

    school = SchoolSettings.load()
    step = request.GET.get('step', '1')

    # Handle skip_houses parameter - store in session
    if request.GET.get('skip_houses') == '1':
        request.session['wizard_houses_skipped'] = True

    # Check what's already set up
    has_academic_year = AcademicYear.objects.exists()
    has_terms = Term.objects.exists()
    has_classes = Class.objects.exists()
    # Houses are optional - check if skipped, if any exist, or if not applicable for this school type
    has_houses = House.objects.exists() or request.session.get('wizard_houses_skipped', False) or not tenant.has_houses
    has_subjects = Subject.objects.exists()

    # Get education system and allowed level types from tenant
    education_system = tenant.education_system
    allowed_level_types = tenant.get_allowed_level_types()
    allowed_level_type_values = [lt[0] for lt in allowed_level_types]

    # Get period type from school settings
    period_type = school.academic_period_type
    period_label = school.period_label
    period_count = 2 if period_type == 'semester' else 3

    # Class level options for the select input (filtered by education system)
    all_class_level_options = [
        ('creche-1', 'Creche 1', 'creche'),
        ('creche-2', 'Creche 2', 'creche'),
        ('nursery-1', 'Nursery 1', 'nursery'),
        ('nursery-2', 'Nursery 2', 'nursery'),
        ('kg-1', 'KG 1', 'kg'),
        ('kg-2', 'KG 2', 'kg'),
        ('primary-1', 'Basic 1', 'basic'),
        ('primary-2', 'Basic 2', 'basic'),
        ('primary-3', 'Basic 3', 'basic'),
        ('primary-4', 'Basic 4', 'basic'),
        ('primary-5', 'Basic 5', 'basic'),
        ('primary-6', 'Basic 6', 'basic'),
        ('jhs-1', 'Basic 7', 'basic'),
        ('jhs-2', 'Basic 8', 'basic'),
        ('jhs-3', 'Basic 9', 'basic'),
        ('shs-1', 'SHS 1', 'shs'),
        ('shs-2', 'SHS 2', 'shs'),
        ('shs-3', 'SHS 3', 'shs'),
    ]

    # Filter class level options based on allowed level types
    class_level_options = [
        (value, label) for value, label, level_type in all_class_level_options
        if level_type in allowed_level_type_values
    ]

    context = {
        'school': school,
        'tenant': tenant,
        'step': step,
        'has_academic_year': has_academic_year,
        'has_terms': has_terms,
        'has_classes': has_classes,
        'has_houses': has_houses,
        'has_subjects': has_subjects,
        'academic_years': AcademicYear.objects.all(),
        'terms': Term.objects.select_related('academic_year').all(),
        'classes': Class.objects.select_related('programme').all(),
        'houses': House.objects.all(),
        'programmes': programmes,
        'period_type': period_type,
        'period_label': period_label,
        'period_count': period_count,
        'class_level_options': class_level_options,
        'education_system': education_system,
        'allowed_level_types': allowed_level_types,
    }

    return htmx_render(
        request,
        'core/setup_wizard.html',
        'core/partials/setup_wizard_content.html',
        context
    )


@admin_required
def setup_wizard_academic_year(request):
    """Create academic year in setup wizard."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if name and start_date and end_date:
            try:
                academic_year = AcademicYear(
                    name=name,
                    start_date=start_date,
                    end_date=end_date,
                    is_current=True
                )
                academic_year.full_clean()
                academic_year.save()
                messages.success(request, f'Academic year "{name}" created.')
            except Exception as e:
                # Extract validation error messages
                if hasattr(e, 'message_dict'):
                    for field, errors in e.message_dict.items():
                        for error in errors:
                            messages.error(request, f'{error}')
                else:
                    messages.error(request, str(e))
        else:
            messages.error(request, 'Please fill in all required fields.')

        return setup_wizard(request)

    return redirect('core:setup_wizard')


@admin_required
def setup_wizard_education_system(request):
    """Set education system (basic, shs, or both) in setup wizard."""
    if request.method == 'POST':
        education_system = request.POST.get('education_system', 'both')
        school = SchoolSettings.load()
        school.education_system = education_system

        # Auto-set academic period type based on education system
        if education_system == 'basic':
            school.academic_period_type = 'term'
        elif education_system == 'shs':
            school.academic_period_type = 'semester'
        # For 'both', keep existing or default to 'term'

        school.save()

        # Mark education system as explicitly set in session
        request.session['wizard_education_system_set'] = True

        # Return the wizard content via HTMX
        return setup_wizard(request)

    return redirect('core:setup_wizard')


@admin_required
def setup_wizard_session_type(request):
    """Set session type (terms vs semesters) in setup wizard."""
    if request.method == 'POST':
        session_type = request.POST.get('session_type', 'term')
        school = SchoolSettings.load()
        school.academic_period_type = session_type
        school.save()

        # Return the wizard content via HTMX
        return setup_wizard(request)

    return redirect('core:setup_wizard')


@admin_required
def setup_wizard_terms(request):
    """Create terms in setup wizard."""
    if request.method == 'POST':
        academic_year = AcademicYear.get_current()
        if not academic_year:
            messages.error(request, 'Please create an academic year first.')
            return setup_wizard(request)

        # Get term data from form
        term_count = int(request.POST.get('term_count', 3))
        created_count = 0

        for i in range(1, term_count + 1):
            name = request.POST.get(f'term_{i}_name', f'Term {i}')
            start_date = request.POST.get(f'term_{i}_start')
            end_date = request.POST.get(f'term_{i}_end')

            if start_date and end_date:
                try:
                    term = Term(
                        academic_year=academic_year,
                        name=name,
                        term_number=i,
                        start_date=start_date,
                        end_date=end_date,
                        is_current=(i == 1)
                    )
                    term.full_clean()
                    term.save()
                    created_count += 1
                except Exception as e:
                    if hasattr(e, 'message_dict'):
                        for field, field_errors in e.message_dict.items():
                            for error in field_errors:
                                messages.error(request, f'{name}: {error}')
                    else:
                        messages.error(request, f'{name}: {str(e)}')
            else:
                messages.error(request, f'{name}: Please provide both start and end dates.')

        if created_count > 0:
            messages.success(request, f'{created_count} term(s) created.')

        return setup_wizard(request)

    return redirect('core:setup_wizard')


@admin_required
def setup_wizard_clear_academic_year(request):
    """Clear academic year to go back to step 1."""
    if request.method == 'DELETE':
        # Delete all academic years (and cascades to terms)
        AcademicYear.objects.all().delete()
        # Also clear education system session flag since we're going back
        request.session.pop('wizard_education_system_set', None)
        messages.info(request, 'Academic year cleared.')

    return setup_wizard(request)


@admin_required
def setup_wizard_clear_education_system(request):
    """Clear education system setting to go back to step 2."""
    if request.method == 'DELETE':
        # Reset education system to default
        school = SchoolSettings.load()
        school.education_system = 'both'
        school.save()
        # Clear the session flag
        request.session.pop('wizard_education_system_set', None)
        messages.info(request, 'Education system reset.')

    return setup_wizard(request)


@admin_required
def setup_wizard_clear_terms(request):
    """Clear terms to go back to step 2."""
    if request.method == 'DELETE':
        Term.objects.all().delete()
        messages.info(request, 'Terms cleared.')

    return setup_wizard(request)


@admin_required
def setup_wizard_clear_houses(request):
    """Clear all houses via HTMX in setup wizard."""
    from students.models import House

    if request.method == 'DELETE':
        House.objects.all().delete()
        # Also clear houses skip flag
        request.session.pop('wizard_houses_skipped', None)
        messages.info(request, 'Houses cleared.')

    return setup_wizard(request)


@admin_required
def setup_wizard_add_class(request):
    """Add a single class via HTMX in setup wizard."""
    from academics.models import Class as AcademicClass, Programme
    from django.http import HttpResponse

    if request.method == 'POST':
        # Get separate level_type and level_number fields
        level_type = request.POST.get('level_type', 'basic')
        level_number = int(request.POST.get('level_number', 1))
        section = request.POST.get('section', '').strip()
        programme_id = request.POST.get('programme', '').strip()

        # Validate level type is allowed based on tenant's education system
        tenant = request.tenant
        allowed_level_types = [lt[0] for lt in tenant.get_allowed_level_types()]
        if level_type not in allowed_level_types:
            return HttpResponse(f'<tr><td colspan="5" class="text-error text-center text-xs py-2">{level_type.upper()} is not allowed for this school</td></tr>')

        # Map "basic" to primary (1-6) or jhs (7-9) for database
        db_level_type = level_type
        db_level_number = level_number
        if level_type == 'basic':
            if level_number <= 6:
                db_level_type = 'primary'
                db_level_number = level_number
            else:
                db_level_type = 'jhs'
                db_level_number = level_number - 6  # 7->1, 8->2, 9->3

        # Section is optional - use empty string if not provided
        section = section if section else ''

        # Get programme for SHS classes
        programme = None
        if level_type == 'shs' and programme_id:
            try:
                programme = Programme.objects.get(pk=programme_id)
            except Programme.DoesNotExist:
                return HttpResponse('<span class="text-error text-xs">Invalid programme selected</span>')

        cls, created = AcademicClass.objects.get_or_create(
            level_type=db_level_type,
            level_number=db_level_number,
            section=section,
            programme=programme,
            defaults={'capacity': 35}
        )

        if created:
            # Return the table row HTML for the new class
            # Display user-friendly level type (Basic instead of Primary/JHS)
            if level_type == 'basic':
                display_type = 'Basic'
                display_level = level_number  # Original 1-9
            else:
                display_type = cls.get_level_type_display()
                display_level = cls.level_number

            prog_code = cls.programme.code if cls.programme else '-'
            prog_hidden = '' if level_type == 'shs' else 'hidden'
            html = f'''<tr class="hover" data-level-type="{level_type}" data-level-number="{level_number}">
                <td>{display_type}</td>
                <td>{display_level}</td>
                <td class="programme-cell {prog_hidden}">{prog_code}</td>
                <td>{cls.section or '-'}</td>
                <td class="text-center">
                    <button hx-delete="{reverse('core:setup_wizard_remove_class', args=[cls.pk])}"
                            hx-target="closest tr"
                            hx-swap="outerHTML"
                            class="btn btn-ghost btn-xs text-error">
                        <i class="fa-solid fa-trash-can"></i>
                    </button>
                </td>
            </tr>'''
            return HttpResponse(html)
        else:
            return HttpResponse(f'<tr><td colspan="5" class="text-warning text-center text-xs py-2">{cls.name} already exists</td></tr>')

    return HttpResponse('')


@admin_required
def setup_wizard_remove_class(request, pk):
    """Remove a class via HTMX in setup wizard."""
    from academics.models import Class as AcademicClass
    from django.http import HttpResponse

    if request.method == 'DELETE':
        try:
            cls = AcademicClass.objects.get(pk=pk)
            cls.delete()
        except AcademicClass.DoesNotExist:
            pass

    return HttpResponse('')


@admin_required
def setup_wizard_bulk_classes(request):
    """Bulk add classes via HTMX in setup wizard."""
    from academics.models import Class as AcademicClass
    from django.http import HttpResponse

    if request.method == 'POST':
        bulk_type = request.GET.get('type', '')

        # Define level ranges for each type
        level_configs = {
            'creche': [('creche', 1, 2)],
            'nursery': [('nursery', 1, 2)],
            'kg': [('kg', 1, 2)],
            'basic': [('basic', 1, 9)],
            'shs': [('shs', 1, 3)],
        }

        if bulk_type not in level_configs:
            return HttpResponse('')

        # Validate level type is allowed based on tenant's education system
        tenant = request.tenant
        allowed_level_types = [lt[0] for lt in tenant.get_allowed_level_types()]
        if bulk_type not in allowed_level_types:
            return HttpResponse(f'<tr><td colspan="5" class="text-error text-center text-xs py-2">{bulk_type.upper()} is not allowed for this school</td></tr>')

        html_parts = []
        for db_level_type, start, end in level_configs[bulk_type]:
            for level in range(start, end + 1):
                cls, created = AcademicClass.objects.get_or_create(
                    level_type=db_level_type,
                    level_number=level,
                    section='',  # No section by default
                    programme=None,
                    defaults={'capacity': 35}
                )
                if created:
                    # Display user-friendly level type
                    display_type = cls.get_level_type_display()
                    display_level = cls.level_number

                    prog_hidden = '' if bulk_type == 'shs' else 'hidden'
                    html_parts.append(f'''<tr class="hover" data-level-type="{bulk_type}">
                        <td>{display_type}</td>
                        <td>{display_level}</td>
                        <td class="programme-cell {prog_hidden}">-</td>
                        <td>-</td>
                        <td class="text-right">
                            <button hx-delete="{reverse('core:setup_wizard_remove_class', args=[cls.pk])}"
                                    hx-target="closest tr"
                                    hx-swap="outerHTML"
                                    class="btn btn-ghost btn-xs text-error">
                                <i class="fa-solid fa-trash-can"></i>
                            </button>
                        </td>
                    </tr>''')

        return HttpResponse(''.join(html_parts))

    return HttpResponse('')


@admin_required
def setup_wizard_clear_classes(request):
    """Clear all classes via HTMX in setup wizard."""
    from academics.models import Class as AcademicClass

    if request.method == 'DELETE':
        AcademicClass.objects.all().delete()
        # Also clear houses skip flag since we're going back
        request.session.pop('wizard_houses_skipped', None)
        messages.info(request, 'Classes cleared.')

    return setup_wizard(request)


@admin_required
def setup_wizard_add_house(request):
    """Add a single house via HTMX in setup wizard."""
    from students.models import House
    from django.http import HttpResponse

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        color_code = request.POST.get('color_code', '#000000')

        if name:
            house, created = House.objects.get_or_create(
                name=name,
                defaults={
                    'color_code': color_code,
                    'is_active': True
                }
            )

            if created:
                html = f'''<tr class="hover">
                    <td>
                        <span class="w-6 h-6 rounded-full inline-block border border-base-300" style="background-color: {color_code};"></span>
                    </td>
                    <td>{house.name}</td>
                    <td class="text-center">
                        <button hx-delete="{reverse('core:setup_wizard_remove_house', args=[house.pk])}"
                            hx-target="closest tr"
                            hx-swap="outerHTML"
                            class="btn btn-ghost btn-xs text-error">
                            <i class="fa-solid fa-trash-can"></i>
                        </button>
                    </td>
                </tr>'''
                return HttpResponse(html)
            else:
                return HttpResponse(f'<tr><td colspan="3" class="text-warning text-center text-xs py-2">{house.name} already exists</td></tr>')

    return HttpResponse('')


@admin_required
def setup_wizard_remove_house(request, pk):
    """Remove a house via HTMX in setup wizard."""
    from students.models import House
    from django.http import HttpResponse

    if request.method == 'DELETE':
        try:
            house = House.objects.get(pk=pk)
            house.delete()
        except House.DoesNotExist:
            pass

    return HttpResponse('')


@admin_required
def setup_wizard_classes(request):
    """Legacy view - redirects to wizard."""
    return redirect('core:setup_wizard')


@admin_required
def setup_wizard_houses(request):
    """Create houses in setup wizard."""
    from students.models import House

    if request.method == 'POST':
        house_count = int(request.POST.get('house_count', 4))

        for i in range(1, house_count + 1):
            name = request.POST.get(f'house_{i}_name', '').strip()
            color = request.POST.get(f'house_{i}_color', '').strip()
            color_code = request.POST.get(f'house_{i}_color_code', '#000000')

            if name:
                House.objects.get_or_create(
                    name=name,
                    defaults={
                        'color': color,
                        'color_code': color_code,
                        'is_active': True
                    }
                )

        messages.success(request, 'Houses created.')
        return setup_wizard(request)

    return redirect('core:setup_wizard')


@admin_required
def setup_wizard_seed(request):
    """Run seed commands in setup wizard."""
    from django.core.management import call_command
    from io import StringIO

    if request.method == 'POST':
        seed_type = request.POST.get('seed_type', 'all')
        output = StringIO()

        try:
            if seed_type == 'academics' or seed_type == 'all':
                call_command('seed_academics', stdout=output)

            if seed_type == 'grading' or seed_type == 'all':
                call_command('seed_grading_data', stdout=output)

            messages.success(request, 'Seed data imported successfully.')
        except Exception as e:
            messages.error(request, f'Error importing seed data: {str(e)}')

        return setup_wizard(request)

    return redirect('core:setup_wizard')


@admin_required
def setup_wizard_complete(request):
    """Mark setup as complete."""
    if request.method == 'POST':
        from .models import Notification

        school = SchoolSettings.load()
        school.setup_completed = True
        school.setup_completed_at = timezone.now()
        school.save()

        # Create welcome notification
        Notification.create_notification(
            user=request.user,
            title='Welcome to your School System!',
            message='Setup is complete. You can now add students, teachers, and start managing your school.',
            notification_type='success',
            category='system',
            icon='fa-solid fa-party-horn',
        )

        messages.success(request, 'Setup completed! Your school is ready to use.')
        return index(request)

    return redirect('core:setup_wizard')


# ============================================
# Notification Views
# ============================================

@login_required
def notifications_dropdown(request):
    """Get notifications for the dropdown menu."""
    from .models import Notification

    notifications = Notification.objects.filter(
        user=request.user
    ).order_by('-created_at')[:10]

    unread_count = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).count()

    context = {
        'notifications': notifications,
        'unread_count': unread_count,
    }
    return render(request, 'core/partials/notifications_dropdown.html', context)


@login_required
def notifications_badge(request):
    """Get just the notification count badge (for polling)."""
    from .models import Notification

    unread_count = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).count()

    return render(request, 'core/partials/notifications_badge.html', {
        'unread_count': unread_count
    })


@login_required
def notification_mark_read(request, pk):
    """Mark a single notification as read."""
    from .models import Notification

    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    notification.mark_as_read()

    # If there's a link, redirect to it
    if notification.link and request.htmx:
        return HttpResponse(
            status=200,
            headers={'HX-Redirect': notification.link}
        )

    return notifications_dropdown(request)


@login_required
def notifications_mark_all_read(request):
    """Mark all notifications as read."""
    from .models import Notification

    if request.method == 'POST':
        Notification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True, read_at=timezone.now())

    return notifications_dropdown(request)


@login_required
def profile(request):
    """Show profile based on user role."""
    user = request.user

    # Teacher profile
    if getattr(user, 'is_teacher', False):
        from django.db.models import Count, Q
        from academics.models import Class, ClassSubject
        from students.models import Student
        from teachers.models import Promotion, Qualification

        teacher = getattr(user, 'teacher_profile', None)
        if not teacher:
            return render(request, 'core/profile_error.html', {
                'error': 'No teacher profile linked to your account.'
            })

        homeroom_classes = Class.objects.filter(
            class_teacher=teacher,
            is_active=True
        ).annotate(
            student_count=Count(
                'students', filter=Q(students__status='active')
            )
        ).order_by('name')

        subject_assignments = ClassSubject.objects.filter(
            teacher=teacher
        ).select_related('class_assigned', 'subject').order_by(
            'class_assigned__level_number', 'class_assigned__name'
        )

        class_ids_taught = list(
            subject_assignments.values_list(
                'class_assigned_id', flat=True
            ).distinct()
        )
        total_students = Student.objects.filter(
            current_class_id__in=class_ids_taught,
            status='active'
        ).count() if class_ids_taught else 0

        promotions = Promotion.objects.filter(
            teacher=teacher
        ).order_by('-date_promoted')

        qualifications = Qualification.objects.filter(
            teacher=teacher
        ).order_by('-date_ended')

        context = {
            'teacher': teacher,
            'homeroom_classes': homeroom_classes,
            'subject_assignments': subject_assignments,
            'promotions': promotions,
            'qualifications': qualifications,
            'workload': {
                'classes_taught': len(class_ids_taught),
                'subjects_taught': subject_assignments.count(),
                'total_students': total_students,
                'homeroom_classes': homeroom_classes.count(),
            }
        }
        return htmx_render(
            request,
            'teachers/profile.html',
            'teachers/partials/profile_content.html',
            context
        )

    # Parent/Guardian profile
    if getattr(user, 'is_parent', False):
        from students.models import Student

        guardian = getattr(user, 'guardian_profile', None)
        if not guardian:
            messages.error(request, 'No guardian profile linked to your account.')
            return redirect('core:index')

        # Get all wards (students) for this guardian
        wards = Student.objects.filter(
            guardians__guardian=guardian,
            status='active'
        ).select_related('current_class').order_by('first_name')

        context = {
            'guardian': guardian,
            'wards': wards,
            'ward_count': wards.count(),
        }
        return htmx_render(
            request,
            'students/guardian_profile.html',
            'students/partials/guardian_profile_content.html',
            context
        )

    # Admin profile (placeholder)
    if user.is_superuser or getattr(user, 'is_school_admin', False):
        context = {
            'user': user,
            # Navigation
            'breadcrumbs': [
                {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
                {'label': 'Profile'},
            ],
        }
        return htmx_render(
            request,
            'core/profile.html',
            'core/partials/profile_content.html',
            context
        )

    # Default - redirect to index
    return redirect('core:index')


@login_required
def profile_edit(request):
    """Edit profile based on user role."""
    user = request.user

    # Teacher profile edit
    if getattr(user, 'is_teacher', False):
        teacher = getattr(user, 'teacher_profile', None)
        if not teacher:
            messages.error(request, 'No teacher profile linked to your account.')
            return redirect('core:profile')

        if request.method == 'POST':
            # Update editable fields
            teacher.phone_number = request.POST.get('phone_number', '').strip()
            teacher.address = request.POST.get('address', '').strip()

            dob = request.POST.get('date_of_birth', '').strip()
            if dob:
                from datetime import datetime
                try:
                    teacher.date_of_birth = datetime.strptime(dob, '%Y-%m-%d').date()
                except ValueError:
                    pass

            # Handle photo upload
            if 'photo' in request.FILES:
                teacher.photo = request.FILES['photo']

            try:
                teacher.save()
                messages.success(request, 'Profile updated successfully.')
                return redirect('core:profile')
            except Exception as e:
                messages.error(request, f'Failed to update profile: {str(e)}')

        context = {'teacher': teacher}
        return htmx_render(
            request,
            'teachers/profile_edit.html',
            'teachers/partials/profile_edit_content.html',
            context
        )

    # Parent profile edit
    if getattr(user, 'is_parent', False):
        guardian = getattr(user, 'guardian_profile', None)
        if not guardian:
            messages.error(request, 'No guardian profile linked to your account.')
            return redirect('core:index')

        if request.method == 'POST':
            guardian.phone_number = request.POST.get('phone_number', '').strip()
            guardian.address = request.POST.get('address', '').strip()
            guardian.occupation = request.POST.get('occupation', '').strip()

            try:
                guardian.save()
                messages.success(request, 'Profile updated successfully.')
                return redirect('core:profile')
            except Exception as e:
                messages.error(request, f'Failed to update profile: {str(e)}')

        context = {'guardian': guardian}
        return htmx_render(
            request,
            'students/guardian_profile_edit.html',
            'students/partials/guardian_profile_edit_content.html',
            context
        )

    # Admin profile edit
    if user.is_superuser or getattr(user, 'is_school_admin', False):
        if request.method == 'POST':
            user.first_name = request.POST.get('first_name', '').strip()
            user.last_name = request.POST.get('last_name', '').strip()

            try:
                user.save(update_fields=['first_name', 'last_name'])
                messages.success(request, 'Profile updated successfully.')
                return redirect('core:profile')
            except Exception as e:
                messages.error(request, f'Failed to update profile: {str(e)}')

        context = {
            'user': user,
            # Navigation
            'breadcrumbs': [
                {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
                {'label': 'Profile', 'url': '/profile/'},
                {'label': 'Edit'},
            ],
            'back_url': '/profile/',
        }
        return htmx_render(
            request,
            'core/profile_edit.html',
            'core/partials/profile_edit_content.html',
            context
        )

    return redirect('core:index')


@login_required
def schedule(request):
    """Schedule view - redirects to appropriate schedule based on user role."""
    user = request.user

    # Teacher schedule
    if getattr(user, 'is_teacher', False):
        from django.utils import timezone
        from academics.models import Period, TimetableEntry

        teacher = getattr(user, 'teacher_profile', None)
        if not teacher:
            messages.warning(request, "No teacher profile linked to your account.")
            return redirect('core:index')

        today = timezone.now()
        weekday = today.isoweekday()

        periods = Period.objects.filter(is_active=True).order_by('order')
        entries = TimetableEntry.objects.filter(
            class_subject__teacher=teacher
        ).select_related(
            'class_subject__class_assigned',
            'class_subject__subject',
            'period'
        ).order_by('weekday', 'period__order')

        schedule_grid = {}
        for period in periods:
            schedule_grid[period.id] = {
                'period': period,
                'days': {1: None, 2: None, 3: None, 4: None, 5: None}
            }

        for entry in entries:
            if entry.period_id in schedule_grid:
                schedule_grid[entry.period_id]['days'][entry.weekday] = entry

        total_periods = entries.count()
        classes_taught = entries.values('class_subject__class_assigned').distinct().count()

        context = {
            'teacher': teacher,
            'periods': periods,
            'schedule_grid': schedule_grid,
            'weekdays': TimetableEntry.Weekday.choices,
            'weekday': weekday,
            'today': today,
            'stats': {
                'total_periods': total_periods,
                'classes_taught': classes_taught,
            }
        }

        return htmx_render(
            request,
            'teachers/schedule.html',
            'teachers/partials/schedule_content.html',
            context
        )

    # Default - redirect to index
    return redirect('core:index')


def teacher_dashboard(request):
    """Dashboard for logged-in teachers."""
    from django.db.models import Count, Q
    from django.utils import timezone
    from academics.models import Class, ClassSubject, Period, TimetableEntry
    from students.models import Student

    teacher = getattr(request.user, 'teacher_profile', None)

    if not teacher:
        # Fallback if no teacher profile linked
        context = {
            'error': 'No teacher profile linked to your account.',
        }
        return htmx_render(request, 'core/index.html', 'core/partials/index_content.html', context)

    current_term = Term.get_current()
    today = timezone.now()
    weekday = today.isoweekday()  # 1=Monday, 7=Sunday

    # Homeroom classes (where teacher is class teacher)
    homeroom_classes = Class.objects.filter(
        class_teacher=teacher,
        is_active=True
    ).annotate(
        student_count=Count('students', filter=Q(students__status='active'))
    ).order_by('name')

    # Subject assignments
    subject_assignments = ClassSubject.objects.filter(
        teacher=teacher
    ).select_related('class_assigned', 'subject').order_by(
        'class_assigned__level_number', 'class_assigned__name'
    )

    # Get unique classes taught (preserving order from queryset)
    seen_class_ids = set()
    classes_taught = []
    for assignment in subject_assignments:
        if assignment.class_assigned_id not in seen_class_ids:
            classes_taught.append(assignment.class_assigned)
            seen_class_ids.add(assignment.class_assigned_id)

    # Calculate stats
    total_students = Student.objects.filter(
        current_class_id__in=seen_class_ids, status='active'
    ).count() if seen_class_ids else 0

    homeroom_students = sum(cls.student_count for cls in homeroom_classes)

    # Batch query for per-class student counts
    class_student_counts = dict(
        Student.objects.filter(
            current_class_id__in=seen_class_ids, status='active'
        ).values('current_class_id').annotate(
            count=Count('id')
        ).values_list('current_class_id', 'count')
    ) if seen_class_ids else {}

    # Group assignments by class for easy display
    assignments_by_class = {}
    for assignment in subject_assignments:
        class_name = assignment.class_assigned.name
        if class_name not in assignments_by_class:
            assignments_by_class[class_name] = {
                'class': assignment.class_assigned,
                'subjects': [],
                'student_count': class_student_counts.get(assignment.class_assigned_id, 0),
            }
        assignments_by_class[class_name]['subjects'].append(assignment.subject)

    # Today's schedule
    todays_schedule = []
    if weekday <= 5:  # Only weekdays
        todays_entries = TimetableEntry.objects.filter(
            class_subject__teacher=teacher,
            weekday=weekday
        ).select_related(
            'class_subject__class_assigned',
            'class_subject__subject',
            'period'
        ).order_by('period__start_time')

        for entry in todays_entries:
            class_obj = entry.class_subject.class_assigned
            todays_schedule.append({
                'entry': entry,
                'period': entry.period,
                'subject': entry.class_subject.subject,
                'class': class_obj,
                'is_current': entry.period.start_time <= today.time() <= entry.period.end_time,
                'is_past': entry.period.end_time < today.time(),
                'uses_lesson_attendance': class_obj.attendance_type == Class.AttendanceType.PER_LESSON,
            })

    # Get all periods for reference
    periods = Period.objects.filter(is_active=True).order_by('start_time')

    context = {
        'teacher': teacher,
        'current_term': current_term,
        'homeroom_classes': homeroom_classes,
        'classes_taught': classes_taught,
        'assignments_by_class': assignments_by_class,
        'todays_schedule': todays_schedule,
        'today': today,
        'weekday': weekday,
        'is_weekend': weekday > 5,
        'periods': periods,
        'stats': {
            'classes_count': len(classes_taught),
            'subjects_count': subject_assignments.count(),
            'total_students': total_students,
            'homeroom_students': homeroom_students,
            'periods_today': len(todays_schedule),
        }
    }

    return htmx_render(
        request,
        'teachers/dashboard.html',
        'teachers/partials/dashboard_content.html',
        context
    )


@login_required
def index(request):
    """Dashboard/index view - routes to appropriate dashboard based on user role."""
    from django.db.models import Count, Q
    from django.utils import timezone
    from students.models import Student, Enrollment
    from academics.models import Class, AttendanceSession, AttendanceRecord

    # Check if user is a teacher - show teacher dashboard
    if getattr(request.user, 'is_teacher', False):
        return teacher_dashboard(request)

    # Check if user is a parent/guardian - show guardian dashboard
    if getattr(request.user, 'is_parent', False):
        return guardian_dashboard(request)

    # Admin/other roles - show admin dashboard
    # Get current academic year and term
    current_year = AcademicYear.get_current()
    current_term = Term.get_current()
    today = timezone.now().date()

    # Get tenant's enabled levels for filtering dashboard display
    from django.db import connection
    from schools.models import School
    try:
        tenant = School.objects.get(schema_name=connection.schema_name)
        enabled_levels = tenant.get_allowed_level_types()
        enabled_level_values = [lt[0] for lt in enabled_levels]
    except School.DoesNotExist:
        enabled_level_values = ['creche', 'nursery', 'kg', 'basic', 'shs']

    # Get counts using aggregation (single query instead of multiple)
    # Student counts - single query with aggregation
    student_stats = Student.objects.filter(status='active').aggregate(
        total=Count('id'),
        male=Count('id', filter=Q(gender='M')),
        female=Count('id', filter=Q(gender='F')),
        creche=Count('id', filter=Q(current_class__level_type='creche')),
        nursery=Count('id', filter=Q(current_class__level_type='nursery')),
        kg=Count('id', filter=Q(current_class__level_type='kg')),
        primary=Count('id', filter=Q(current_class__level_type='primary')),
        jhs=Count('id', filter=Q(current_class__level_type='jhs')),
        shs=Count('id', filter=Q(current_class__level_type='shs')),
        unassigned=Count('id', filter=Q(current_class__isnull=True)),
    )

    student_count = student_stats['total']
    male_count = student_stats['male']
    female_count = student_stats['female']

    teacher_count = Teacher.objects.filter(status='active').count()
    class_count = Class.objects.filter(is_active=True).count()

    # Get recent students (last 5 added)
    recent_students = Student.objects.select_related('current_class').order_by('-created_at')[:5]

    # Get active enrollments for current year
    active_enrollments = 0
    if current_year:
        active_enrollments = Enrollment.objects.filter(
            academic_year=current_year,
            status='active'
        ).count()

    # Students by level (filtered by enabled levels)
    # Map enabled level codes to their display names and database keys
    level_display_map = {
        'creche': ('Creche', 'creche'),
        'nursery': ('Nursery', 'nursery'),
        'kg': ('KG', 'kg'),
        'basic': ('Basic', ['primary', 'jhs']),  # Basic combines Primary and JHS
        'shs': ('SHS', 'shs'),
    }

    students_by_level = {}
    for level_code in enabled_level_values:
        if level_code in level_display_map:
            display_name, db_keys = level_display_map[level_code]
            if isinstance(db_keys, list):
                # For 'basic', combine primary and jhs counts
                count = sum(student_stats.get(k, 0) for k in db_keys)
            else:
                count = student_stats.get(db_keys, 0)
            students_by_level[display_name] = count

    # Always show unassigned
    students_by_level['Unassigned'] = student_stats['unassigned']

    # Today's attendance summary - optimized with single aggregation
    attendance_stats = AttendanceRecord.objects.filter(
        session__date=today
    ).aggregate(
        present=Count('id', filter=Q(status__in=['P', 'L'])),
        absent=Count('id', filter=Q(status='A')),
    )

    today_attendance = {
        'sessions_taken': AttendanceSession.objects.filter(date=today).count(),
        'total_classes': class_count,
        'present': attendance_stats['present'],
        'absent': attendance_stats['absent'],
    }

    # Classes needing attention (no attendance today)
    classes_without_attendance = Class.objects.filter(
        is_active=True
    ).exclude(
        attendance_sessions__date=today
    ).select_related('class_teacher')[:5]

    # Recent activity (enrollments, new students)
    recent_enrollments = []
    if current_year:
        recent_enrollments = Enrollment.objects.filter(
            academic_year=current_year
        ).select_related(
            'student', 'class_assigned'
        ).order_by('-created_at')[:5]

    context = {
        'student_count': student_count,
        'male_count': male_count,
        'female_count': female_count,
        'teacher_count': teacher_count,
        'class_count': class_count,
        'current_year': current_year,
        'current_term': current_term,
        'active_enrollments': active_enrollments,
        'recent_students': recent_students,
        'students_by_level': students_by_level,
        'enabled_level_values': enabled_level_values,
        'today_attendance': today_attendance,
        'classes_without_attendance': classes_without_attendance,
        'recent_enrollments': recent_enrollments,
        'today': today,
    }
    return htmx_render(request, 'core/index.html', 'core/partials/index_content.html', context)


# School Admin views
@admin_required
def students_list(request):
    context = {}
    return htmx_render(request, 'core/students/list.html', 'core/students/partials/list_content.html', context)


@admin_required
def teachers_list(request):
    context = {}
    return htmx_render(request, 'core/teachers/list.html', 'core/teachers/partials/list_content.html', context)


@admin_required
def finance_overview(request):
    context = {}
    return htmx_render(request, 'core/finance/overview.html', 'core/finance/partials/overview_content.html', context)


@admin_required
def invoices(request):
    context = {}
    return htmx_render(request, 'core/finance/invoices.html', 'core/finance/partials/invoices_content.html', context)


@admin_required
def payments(request):
    context = {}
    return htmx_render(request, 'core/finance/payments.html', 'core/finance/partials/payments_content.html', context)


@admin_required
def settings_page(request):
    """School settings page with all configuration options."""
    tenant = request.tenant
    school_settings = SchoolSettings.load()
    period_type = school_settings.academic_period_type

    # Initialize forms with current data
    contact_form = SchoolContactForm(initial={
        'email': tenant.email,
        'phone': tenant.phone,
        'address': tenant.address,
        'digital_address': tenant.digital_address,
        'city': tenant.city,
        'region': tenant.location_region,
    })

    admin_form = SchoolAdminForm(initial={
        'headmaster_name': tenant.headmaster_name,
        'headmaster_title': tenant.headmaster_title,
    })

    # Academic settings and data
    academic_settings_form = AcademicSettingsForm(instance=school_settings)
    academic_years = AcademicYear.objects.prefetch_related('terms').all()
    academic_year_form = AcademicYearForm()
    term_form = TermForm(period_type=period_type)

    # SMS settings - derive sender ID from school name
    derived_sender_id = ''
    if tenant.short_name:
        derived_sender_id = ''.join(c for c in tenant.short_name if c.isalnum())[:11]
    elif tenant.name:
        derived_sender_id = ''.join(c for c in tenant.name if c.isalnum())[:11]
    derived_sender_id = derived_sender_id or 'SchoolSMS'

    # Payment gateway settings
    available_gateways = PaymentGateway.objects.filter(is_active=True)
    gateway_configs = PaymentGatewayConfig.objects.select_related('gateway').all()
    primary_gateway = gateway_configs.filter(is_primary=True, is_active=True).first()

    context = {
        'tenant': tenant,
        'school_settings': school_settings,
        'contact_form': contact_form,
        'admin_form': admin_form,
        'academic_settings_form': academic_settings_form,
        'academic_years': academic_years,
        'academic_year_form': academic_year_form,
        'term_form': term_form,
        'period_type': period_type,
        'period_label': school_settings.period_label,
        'period_label_plural': school_settings.period_label_plural,
        'derived_sender_id': derived_sender_id,
        # Education system context (read from tenant)
        'education_system_display': tenant.education_system_display,
        # Payment gateway context
        'available_gateways': available_gateways,
        'gateway_configs': gateway_configs,
        'primary_gateway': primary_gateway,
        # Navigation
        'breadcrumbs': [
            {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
            {'label': 'Settings'},
        ],
    }
    return htmx_render(request, 'core/settings/index.html', 'core/settings/partials/index_content.html', context)


@admin_required
def settings_update_basic(request):
    """Update basic school information."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    tenant = request.tenant
    school_settings = SchoolSettings.load()
    form = SchoolBasicInfoForm(request.POST)

    if form.is_valid():
        tenant.name = form.cleaned_data['name']
        tenant.short_name = form.cleaned_data['short_name']
        tenant.motto = form.cleaned_data['motto']
        tenant.save()

        # For non-HTMX requests, redirect back to settings
        if not request.htmx:
            return redirect('core:settings')

        context = {'tenant': tenant, 'school_settings': school_settings, 'success': True}
    else:
        context = {'tenant': tenant, 'school_settings': school_settings, 'errors': form.errors}

    return render(request, 'core/settings/partials/card_basic.html', context)


@admin_required
def settings_update_branding(request):
    """Update branding settings (logo, favicon, colors)."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    tenant = request.tenant
    school_settings = SchoolSettings.load()
    form = SchoolBrandingForm(request.POST, request.FILES)

    if form.is_valid():
        # Update tenant branding fields
        if form.cleaned_data.get('logo'):
            tenant.logo = form.cleaned_data['logo']
        if form.cleaned_data.get('favicon'):
            tenant.favicon = form.cleaned_data['favicon']
        if form.cleaned_data.get('primary_color'):
            tenant.primary_color = form.cleaned_data['primary_color']
        if form.cleaned_data.get('secondary_color'):
            tenant.secondary_color = form.cleaned_data['secondary_color']
        if form.cleaned_data.get('accent_color'):
            tenant.accent_color = form.cleaned_data['accent_color']
        tenant.save()

        # Always redirect/refresh for branding changes since colors affect entire UI
        if request.htmx:
            # Trigger full page refresh so new colors apply globally
            response = HttpResponse(status=200)
            response['HX-Refresh'] = 'true'
            return response

        return redirect('core:settings')

    # On error, return the form with errors
    context = {'tenant': tenant, 'school_settings': school_settings, 'errors': form.errors}
    return render(request, 'core/settings/partials/card_branding.html', context)


@admin_required
def settings_update_contact(request):
    """Update contact information."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    tenant = request.tenant
    form = SchoolContactForm(request.POST)

    if form.is_valid():
        tenant.email = form.cleaned_data['email']
        tenant.phone = form.cleaned_data['phone']
        tenant.address = form.cleaned_data['address']
        tenant.digital_address = form.cleaned_data['digital_address']
        tenant.city = form.cleaned_data['city']
        tenant.location_region = form.cleaned_data['region']
        tenant.save()

        if not request.htmx:
            return redirect('core:settings')

        context = {'tenant': tenant, 'success': True}
    else:
        context = {'tenant': tenant, 'errors': form.errors}

    return render(request, 'core/settings/partials/card_contact.html', context)


@admin_required
def settings_update_admin(request):
    """Update administration details."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    tenant = request.tenant
    form = SchoolAdminForm(request.POST)

    if form.is_valid():
        tenant.headmaster_name = form.cleaned_data['headmaster_name']
        tenant.headmaster_title = form.cleaned_data['headmaster_title']
        tenant.save()

        if not request.htmx:
            return redirect('core:settings')

        context = {'tenant': tenant, 'success': True}
    else:
        context = {'tenant': tenant, 'errors': form.errors}

    return render(request, 'core/settings/partials/card_admin.html', context)


@admin_required
def settings_update_sms(request):
    """Update SMS configuration settings."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    school_settings = SchoolSettings.load()
    errors = []

    # Handle checkbox - 'on' (HTML default) or 'true' (input_tags)
    sms_enabled = request.POST.get('sms_enabled') in ('on', 'true')
    sms_backend = request.POST.get('sms_backend', 'console')
    sms_api_key = request.POST.get('sms_api_key', '').strip()
    sms_sender_id = request.POST.get('sms_sender_id', '').strip()

    # Validate sender ID length (GSM standard max 11 characters)
    if sms_sender_id and len(sms_sender_id) > 11:
        errors.append(
            f'Sender ID must be 11 characters or fewer (currently {len(sms_sender_id)}). '
            'This is a GSM standard limit.'
        )

    # Validate if SMS is enabled and not console mode
    if sms_enabled and sms_backend != 'console':
        if not sms_api_key and not school_settings.sms_api_key:
            errors.append('API key is required for SMS provider')

        # Validate API key format for specific providers
        if sms_api_key and sms_backend in ('hubtel', 'africastalking'):
            if ':' not in sms_api_key:
                provider_name = 'Hubtel' if sms_backend == 'hubtel' else "Africa's Talking"
                errors.append(f'{provider_name} API key must be in format "id:secret"')

    if errors:
        context = {
            'school_settings': school_settings,
            'errors': errors,
        }
        response = render(request, 'core/settings/partials/card_sms.html', context)
        response['HX-Retarget'] = '#card-sms'
        response['HX-Reswap'] = 'outerHTML'
        return response

    # Update settings
    school_settings.sms_enabled = sms_enabled
    school_settings.sms_backend = sms_backend
    school_settings.sms_sender_id = sms_sender_id

    # Only update API key if a new one was provided (not placeholder)
    if sms_api_key and not sms_api_key.startswith('••'):
        school_settings.sms_api_key = sms_api_key

    school_settings.save()

    if not request.htmx:
        return redirect('core:settings')

    context = {
        'school_settings': school_settings,
        'success': 'SMS settings updated successfully',
    }
    response = render(request, 'core/settings/partials/card_sms.html', context)
    response['HX-Trigger'] = 'closeSmsModal'
    return response


@admin_required
def settings_update_email(request):
    """Update email configuration settings."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    school_settings = SchoolSettings.load()
    errors = []

    # Checkboxes send 'on' (HTML default) or 'true' (input_tags) when checked
    email_enabled = request.POST.get('email_enabled') in ('on', 'true')
    email_host = request.POST.get('email_host', '').strip()
    email_host_user = request.POST.get('email_host_user', '').strip()
    password = request.POST.get('email_host_password', '').strip()

    # Validate if email is enabled
    if email_enabled:
        if not email_host:
            errors.append('SMTP Host is required')
        if not email_host_user:
            errors.append('Email address is required')
        # Password required only if not already set
        if not password and not school_settings.email_host_password:
            errors.append('App Password is required')

    if errors:
        context = {
            'school_settings': school_settings,
            'errors': errors,
        }
        response = render(request, 'core/settings/partials/card_email.html', context)
        response['HX-Retarget'] = '#card-email'
        response['HX-Reswap'] = 'outerHTML'
        return response

    # Update email settings
    school_settings.email_enabled = email_enabled
    school_settings.email_backend = request.POST.get('email_backend', 'smtp')
    school_settings.email_host = email_host

    # Handle port with default
    try:
        school_settings.email_port = int(request.POST.get('email_port', 587))
    except (ValueError, TypeError):
        school_settings.email_port = 587

    school_settings.email_use_tls = request.POST.get('email_use_tls') in ('on', 'true')
    school_settings.email_use_ssl = request.POST.get('email_use_ssl') in ('on', 'true')
    school_settings.email_host_user = email_host_user
    school_settings.email_from_address = request.POST.get('email_from_address', '').strip()
    school_settings.email_from_name = request.POST.get('email_from_name', '').strip()

    # Only update password if a new one was provided (not placeholder)
    if password and not password.startswith('••'):
        school_settings.email_host_password = password

    school_settings.save()

    if not request.htmx:
        return redirect('core:settings')

    context = {
        'school_settings': school_settings,
        'success': 'Email settings updated successfully',
    }
    response = render(request, 'core/settings/partials/card_email.html', context)
    response['HX-Trigger'] = 'closeEmailModal'
    return response


@admin_required
@ratelimit(key='user', rate='5/h')
def settings_test_email(request):
    """Send a test email to verify email configuration. Rate limited to 5/hour."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    from django.core.mail import send_mail

    # Auto-dismiss script for alerts (resets opacity first, then fades out)
    auto_dismiss = '<script>(() => { const el = document.getElementById("test-email-result"); if(el) { el.style.opacity = "1"; el.style.transition = "none"; setTimeout(() => { el.style.transition = "opacity 0.5s"; el.style.opacity = "0"; setTimeout(() => el.innerHTML = "", 500); }, 4000); } })();</script>'

    recipient = request.POST.get('test_email', '').strip()
    if not recipient:
        recipient = request.user.email

    if not recipient:
        return HttpResponse(
            '<div class="alert alert-error text-sm py-2">'
            '<i class="fa-solid fa-circle-xmark"></i> No recipient email address'
            '</div>' + auto_dismiss
        )

    school_settings = SchoolSettings.load()
    from_email = school_settings.email_from_address or None

    try:
        send_mail(
            subject='Test Email - School Management System',
            message=(
                'This is a test email to verify your email configuration is working correctly.\n\n'
                'If you received this email, your email settings are configured properly.'
            ),
            from_email=from_email,
            recipient_list=[recipient],
            fail_silently=False,
        )
        return HttpResponse(
            '<div class="alert alert-success text-sm py-2">'
            f'<i class="fa-solid fa-circle-check"></i> Test email sent to {recipient}'
            '</div>' + auto_dismiss
        )
    except Exception as e:
        error_msg = str(e)
        # Truncate long error messages
        if len(error_msg) > 100:
            error_msg = error_msg[:100] + '...'
        return HttpResponse(
            f'<div class="alert alert-error text-sm py-2">'
            f'<i class="fa-solid fa-circle-xmark"></i> Failed: {error_msg}'
            f'</div>' + auto_dismiss
        )


@admin_required
@ratelimit(key='user', rate='5/h')
def settings_test_sms(request):
    """Send a test SMS to verify SMS configuration using form values. Rate limited to 5/hour."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    from communications.tasks import send_via_arkesel, send_via_hubtel, send_via_africastalking

    # Auto-dismiss script for alerts (resets opacity first, then fades out)
    auto_dismiss = '<script>(() => { const el = document.getElementById("test-sms-result"); if(el) { el.style.opacity = "1"; el.style.transition = "none"; setTimeout(() => { el.style.transition = "opacity 0.5s"; el.style.opacity = "0"; setTimeout(() => el.innerHTML = "", 500); }, 4000); } })();</script>'

    recipient = request.POST.get('test_phone', '').strip()
    if not recipient:
        return HttpResponse(
            '<div class="alert alert-error text-sm py-2">'
            '<i class="fa-solid fa-circle-xmark"></i> Phone number is required'
            '</div>' + auto_dismiss
        )

    # Get form values (allows testing before saving)
    school_settings = SchoolSettings.load()
    sms_backend = request.POST.get('sms_backend', school_settings.sms_backend or 'console')
    sms_api_key = request.POST.get('sms_api_key', '').strip()
    sms_sender_id = request.POST.get('sms_sender_id', '').strip()

    # Use saved API key if form field is empty or placeholder
    if not sms_api_key or sms_api_key.startswith('••'):
        sms_api_key = school_settings.sms_api_key

    # Use saved or derived sender ID
    tenant = request.tenant
    if not sms_sender_id:
        sms_sender_id = school_settings.sms_sender_id
    if not sms_sender_id and tenant.short_name:
        sms_sender_id = ''.join(c for c in tenant.short_name if c.isalnum())[:11]
    elif not sms_sender_id and tenant.name:
        sms_sender_id = ''.join(c for c in tenant.name if c.isalnum())[:11]
    if not sms_sender_id:
        sms_sender_id = 'SchoolSMS'

    if sms_backend == 'console':
        return HttpResponse(
            '<div class="alert alert-info text-sm py-2">'
            '<i class="fa-solid fa-circle-info"></i> Console mode: SMS would be logged only, not sent'
            '</div>' + auto_dismiss
        )

    if not sms_api_key:
        return HttpResponse(
            '<div class="alert alert-warning text-sm py-2">'
            '<i class="fa-solid fa-triangle-exclamation"></i> Enter API key to test'
            '</div>' + auto_dismiss
        )

    # Normalize phone number
    phone = recipient.strip()
    if phone.startswith('0') and len(phone) == 10:
        phone = '233' + phone[1:]
    elif phone.startswith('+'):
        phone = phone[1:]

    message = f"Test SMS from {tenant.short_name or tenant.name or 'School Management System'}. Your SMS configuration is working!"

    try:
        if sms_backend == 'arkesel':
            send_via_arkesel(phone, message, sender_id=sms_sender_id, api_key=sms_api_key)
        elif sms_backend == 'hubtel':
            if ':' not in sms_api_key:
                return HttpResponse(
                    '<div class="alert alert-error text-sm py-2">'
                    '<i class="fa-solid fa-circle-xmark"></i> Hubtel API key must be in format "client_id:client_secret"'
                    '</div>' + auto_dismiss
                )
            send_via_hubtel(phone, message, sender_id=sms_sender_id, api_key=sms_api_key)
        elif sms_backend == 'africastalking':
            if ':' not in sms_api_key:
                return HttpResponse(
                    '<div class="alert alert-error text-sm py-2">'
                    '<i class="fa-solid fa-circle-xmark"></i> Africa\'s Talking API key must be in format "username:api_key"'
                    '</div>' + auto_dismiss
                )
            send_via_africastalking(phone, message, sender_id=sms_sender_id, api_key=sms_api_key)
        else:
            return HttpResponse(
                '<div class="alert alert-warning text-sm py-2">'
                '<i class="fa-solid fa-triangle-exclamation"></i> Unknown SMS provider'
                '</div>' + auto_dismiss
            )

        return HttpResponse(
            f'<div class="alert alert-success text-sm py-2">'
            f'<i class="fa-solid fa-circle-check"></i> Test SMS sent to {recipient} via {sms_backend}'
            f'</div>' + auto_dismiss
        )

    except Exception as e:
        error_msg = str(e)
        if len(error_msg) > 100:
            error_msg = error_msg[:100] + '...'
        return HttpResponse(
            f'<div class="alert alert-error text-sm py-2">'
            f'<i class="fa-solid fa-circle-xmark"></i> Failed: {error_msg}'
            f'</div>' + auto_dismiss
        )


@admin_required
@ratelimit(key='user', rate='5/h')
def settings_test_payment(request):
    """Test payment gateway credentials using form values (without saving). Rate limited to 5/hour."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    # Auto-dismiss script
    auto_dismiss = '''<script>(() => {
        const el = document.getElementById("test-payment-result");
        if(el) {
            el.style.opacity = "1";
            el.style.transition = "none";
            setTimeout(() => {
                el.style.transition = "opacity 0.5s";
                el.style.opacity = "0";
                setTimeout(() => el.innerHTML = "", 500);
            }, 5000);
        }
    })();</script>'''

    gateway_id = request.POST.get('gateway_id', '').strip()
    secret_key = request.POST.get('secret_key', '').strip()
    public_key = request.POST.get('public_key', '').strip()
    is_test_mode = request.POST.get('is_test_mode') == 'on'

    if not gateway_id:
        return HttpResponse(
            f'<div class="alert alert-error text-sm py-1">'
            f'<i class="fa-solid fa-times-circle text-xs"></i> Select a gateway'
            f'</div>{auto_dismiss}'
        )

    if not secret_key or secret_key.startswith('••'):
        return HttpResponse(
            f'<div class="alert alert-error text-sm py-1">'
            f'<i class="fa-solid fa-times-circle text-xs"></i> Secret key required'
            f'</div>{auto_dismiss}'
        )

    try:
        gateway = PaymentGateway.objects.get(pk=gateway_id)
    except PaymentGateway.DoesNotExist:
        return HttpResponse(
            f'<div class="alert alert-error text-sm py-1">'
            f'<i class="fa-solid fa-times-circle text-xs"></i> Gateway not found'
            f'</div>{auto_dismiss}'
        )

    # Get or create config to test with
    config, _ = PaymentGatewayConfig.objects.get_or_create(
        gateway=gateway,
        defaults={'configured_by': request.user}
    )

    # Temporarily set credentials for testing
    old_secret = config.secret_key
    old_public = config.public_key
    old_test_mode = config.is_test_mode

    config.secret_key = secret_key
    config.public_key = public_key if public_key and not public_key.startswith('••') else config.public_key
    config.is_test_mode = is_test_mode

    # Test credentials
    from finance.gateways import get_gateway_adapter

    try:
        adapter = get_gateway_adapter(config)
        is_valid, message = adapter.verify_credentials()

        # Restore original values
        config.secret_key = old_secret
        config.public_key = old_public
        config.is_test_mode = old_test_mode

        if is_valid:
            return HttpResponse(
                f'<div class="alert alert-success text-sm py-1">'
                f'<i class="fa-solid fa-check-circle text-xs"></i> Credentials verified!'
                f'</div>{auto_dismiss}'
            )
        else:
            return HttpResponse(
                f'<div class="alert alert-error text-sm py-1">'
                f'<i class="fa-solid fa-times-circle text-xs"></i> {message}'
                f'</div>{auto_dismiss}'
            )

    except Exception as e:
        # Restore original values
        config.secret_key = old_secret
        config.public_key = old_public
        config.is_test_mode = old_test_mode

        error_msg = str(e)[:80]
        return HttpResponse(
            f'<div class="alert alert-error text-sm py-1">'
            f'<i class="fa-solid fa-times-circle text-xs"></i> {error_msg}'
            f'</div>{auto_dismiss}'
        )


@admin_required
def settings_update_payment(request):
    """Update payment gateway configuration."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    gateway_id = request.POST.get('gateway_id')
    if not gateway_id:
        return HttpResponse("Gateway ID required", status=400)

    gateway = get_object_or_404(PaymentGateway, pk=gateway_id)

    # Get or create the configuration for this gateway
    config, created = PaymentGatewayConfig.objects.get_or_create(
        gateway=gateway,
        defaults={'secret_key': '', 'configured_by': request.user}
    )

    # Update credentials (only if new values provided)
    secret_key = request.POST.get('secret_key', '').strip()
    public_key = request.POST.get('public_key', '').strip()
    webhook_secret = request.POST.get('webhook_secret', '').strip()
    merchant_id = request.POST.get('merchant_id', '').strip()
    encryption_key = request.POST.get('encryption_key', '').strip()
    merchant_account = request.POST.get('merchant_account', '').strip()

    # Only update if new value provided (not placeholder)
    if secret_key and not secret_key.startswith('••'):
        config.secret_key = secret_key
    if public_key and not public_key.startswith('••'):
        config.public_key = public_key
    if webhook_secret and not webhook_secret.startswith('••'):
        config.webhook_secret = webhook_secret
    if merchant_id:
        config.merchant_id = merchant_id
    if encryption_key and not encryption_key.startswith('••'):
        config.encryption_key = encryption_key
    if merchant_account:
        config.merchant_account = merchant_account

    # Update settings
    config.is_active = request.POST.get('is_active') == 'on'
    config.is_test_mode = request.POST.get('is_test_mode') == 'on'
    config.is_primary = request.POST.get('is_primary') == 'on'

    # Update transaction charges
    charge_percentage = request.POST.get('charge_percentage', '').strip()
    charge_fixed = request.POST.get('charge_fixed', '').strip()
    who_bears_charge = request.POST.get('who_bears_charge', 'SCHOOL')

    if charge_percentage:
        config.transaction_charge_percentage = Decimal(charge_percentage)
    if charge_fixed:
        config.transaction_charge_fixed = Decimal(charge_fixed)
    config.who_bears_charge = who_bears_charge

    config.configured_by = request.user
    config.verification_status = 'PENDING'  # Reset verification status
    config.save()

    if not request.htmx:
        return redirect('core:settings')

    # Get updated context for the card
    gateway_configs = PaymentGatewayConfig.objects.select_related('gateway').all()
    primary_gateway = gateway_configs.filter(is_primary=True, is_active=True).first()

    context = {
        'gateway_configs': gateway_configs,
        'primary_gateway': primary_gateway,
        'payment_success': f'{gateway.display_name} configured successfully',
    }
    response = render(request, 'core/settings/partials/card_payment.html', context)
    response['HX-Trigger'] = 'closePaymentModal'
    return response


def get_academic_card_context(request=None, success=None, errors=None):
    """Helper to get common context for academic card."""
    from django.db import connection
    from schools.models import School

    school_settings = SchoolSettings.load()
    period_type = school_settings.academic_period_type

    # Get education system from tenant
    try:
        tenant = School.objects.get(schema_name=connection.schema_name)
        education_system_display = tenant.education_system_display
    except School.DoesNotExist:
        education_system_display = 'Both Basic and SHS'

    return {
        'academic_years': AcademicYear.objects.prefetch_related('terms').all(),
        'academic_year_form': AcademicYearForm(),
        'term_form': TermForm(period_type=period_type),
        'period_type': period_type,
        'period_label': school_settings.period_label,
        'period_label_plural': school_settings.period_label_plural,
        'school_settings': school_settings,
        'education_system_display': education_system_display,
        'success': success,
        'errors': errors,
    }


@admin_required
def settings_update_academic(request):
    """Update academic period settings."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    school_settings = SchoolSettings.load()
    form = AcademicSettingsForm(request.POST, instance=school_settings)

    if form.is_valid():
        form.save()
        if not request.htmx:
            return redirect('core:settings')
        return render(request, 'core/settings/partials/card_academic.html',
                      get_academic_card_context(success='Academic settings updated.'))

    return render(request, 'core/settings/partials/card_academic.html',
                  get_academic_card_context(errors=form.errors))


@admin_required
def settings_update_education_system(request):
    """Update education system setting."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    school_settings = SchoolSettings.load()
    education_system = request.POST.get('education_system', 'both')

    # Validate the choice
    valid_choices = [choice[0] for choice in SchoolSettings.EDUCATION_SYSTEM_CHOICES]
    if education_system in valid_choices:
        school_settings.education_system = education_system

        # Auto-update academic period type based on education system
        if education_system == 'basic':
            school_settings.academic_period_type = 'term'
        elif education_system == 'shs':
            school_settings.academic_period_type = 'semester'
        # For 'both', keep existing period type

        school_settings.save()

        if not request.htmx:
            return redirect('core:settings')
        return render(request, 'core/settings/partials/card_academic.html',
                      get_academic_card_context(success='Education system updated.'))

    return render(request, 'core/settings/partials/card_academic.html',
                  get_academic_card_context(errors={'education_system': ['Invalid choice']}))


# Academic Year views
@admin_required
def academic_year_create(request):
    """Create a new academic year."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    form = AcademicYearForm(request.POST)
    if form.is_valid():
        form.save()
        # Trigger full page refresh so navbar updates
        if request.htmx:
            response = HttpResponse(status=200)
            response['HX-Refresh'] = 'true'
            return response
        return redirect('core:settings')

    # Return form with errors - use 422 so modal doesn't close
    context = {
        'form': form,
        'is_create': True,
    }
    response = render(request, 'core/settings/partials/modal_academic_year_form.html', context)
    response.status_code = 422
    response['HX-Retarget'] = '#modal-academic-year-form'
    response['HX-Reswap'] = 'outerHTML'
    return response


@admin_required
def academic_year_edit(request, pk):
    """Edit an academic year."""
    academic_year = get_object_or_404(AcademicYear, pk=pk)

    if request.method == 'GET':
        form = AcademicYearForm(instance=academic_year)
        return render(request, 'core/settings/partials/modal_academic_year_form.html', {
            'form': form,
            'is_create': False,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    form = AcademicYearForm(request.POST, instance=academic_year)
    if form.is_valid():
        form.save()
        if not request.htmx:
            return redirect('core:settings')
        # Trigger full page refresh so navbar updates
        response = HttpResponse(status=200)
        response['HX-Refresh'] = 'true'
        return response

    # Return form with errors - use 422 so modal doesn't close
    context = {
        'form': form,
        'is_create': False,
    }
    response = render(request, 'core/settings/partials/modal_academic_year_form.html', context)
    response.status_code = 422
    response['HX-Retarget'] = '#modal-academic-year-form'
    response['HX-Reswap'] = 'outerHTML'
    return response


@admin_required
def academic_year_delete(request, pk):
    """Delete an academic year."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    academic_year = get_object_or_404(AcademicYear, pk=pk)
    academic_year.delete()

    if not request.htmx:
        return redirect('core:settings')

    return render(request, 'core/settings/partials/card_academic.html',
                  get_academic_card_context(success='Academic year deleted successfully.'))


@admin_required
def academic_year_set_current(request, pk):
    """Set an academic year as current."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    academic_year = get_object_or_404(AcademicYear, pk=pk)
    academic_year.is_current = True
    academic_year.save()

    # Trigger full page refresh so navbar updates
    if request.htmx:
        response = HttpResponse(status=200)
        response['HX-Refresh'] = 'true'
        return response
    return redirect('core:settings')


# Term views
@admin_required
def term_create(request):
    """Create a new term/semester."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    school_settings = SchoolSettings.load()
    period_type = school_settings.academic_period_type

    form = TermForm(request.POST, period_type=period_type)
    if form.is_valid():
        form.save()
        # Trigger full page refresh so navbar updates
        if request.htmx:
            response = HttpResponse(status=200)
            response['HX-Refresh'] = 'true'
            return response
        return redirect('core:settings')

    # Return form with errors - use 422 so modal doesn't close
    context = {
        'form': form,
        'is_create': True,
        'period_label': school_settings.period_label,
    }
    response = render(request, 'core/settings/partials/modal_term_form.html', context)
    response.status_code = 422
    response['HX-Retarget'] = '#modal-term-form'
    response['HX-Reswap'] = 'outerHTML'
    return response


@admin_required
def term_edit(request, pk):
    """Edit a term/semester."""
    term = get_object_or_404(Term, pk=pk)
    school_settings = SchoolSettings.load()
    period_type = school_settings.academic_period_type
    period_label = school_settings.period_label

    if request.method == 'GET':
        form = TermForm(instance=term, period_type=period_type)
        return render(request, 'core/settings/partials/modal_term_form.html', {
            'form': form,
            'is_create': False,
            'period_label': period_label,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    form = TermForm(request.POST, instance=term, period_type=period_type)
    if form.is_valid():
        form.save()
        if not request.htmx:
            return redirect('core:settings')
        # Trigger full page refresh so navbar updates
        response = HttpResponse(status=200)
        response['HX-Refresh'] = 'true'
        return response

    # Return form with errors - use 422 so modal doesn't close
    context = {
        'form': form,
        'is_create': False,
        'period_label': period_label,
    }
    response = render(request, 'core/settings/partials/modal_term_form.html', context)
    response.status_code = 422
    response['HX-Retarget'] = '#modal-term-form'
    response['HX-Reswap'] = 'outerHTML'
    return response


@admin_required
def term_delete(request, pk):
    """Delete a term/semester."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    term = get_object_or_404(Term, pk=pk)
    term.delete()

    if not request.htmx:
        return redirect('core:settings')

    school_settings = SchoolSettings.load()
    return render(request, 'core/settings/partials/card_academic.html',
                  get_academic_card_context(success=f'{school_settings.period_label} deleted successfully.'))


@admin_required
def term_set_current(request, pk):
    """Set a term/semester as current."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    term = get_object_or_404(Term, pk=pk)
    term.is_current = True
    term.save()

    # Trigger full page refresh so navbar updates
    if request.htmx:
        response = HttpResponse(status=200)
        response['HX-Refresh'] = 'true'
        return response
    return redirect('core:settings')


# Teacher views
@login_required
def my_classes(request):
    """Teacher view of their assigned classes."""
    from academics.models import Class, ClassSubject
    from students.models import Student
    from django.db.models import Count, Q, Prefetch

    user = request.user

    # Verify user is a teacher
    if not getattr(user, 'is_teacher', False):
        messages.error(request, 'This page is only accessible to teachers.')
        return redirect('core:index')

    teacher = getattr(user, 'teacher_profile', None)
    if not teacher:
        messages.error(request, 'No teacher profile linked to your account.')
        return redirect('core:index')

    # Get current term
    current_term = Term.get_current()

    # Get homeroom classes with annotated student counts and prefetched subjects
    homeroom_classes = Class.objects.filter(
        class_teacher=teacher,
        is_active=True
    ).annotate(
        active_student_count=Count('students', filter=Q(students__status='active'))
    ).prefetch_related(
        Prefetch(
            'subjects',
            queryset=ClassSubject.objects.select_related('subject', 'teacher').order_by('-subject__is_core', 'subject__name'),
            to_attr='prefetched_subjects'
        )
    ).order_by('level_number', 'name')

    # Get subject assignments
    subject_assignments = ClassSubject.objects.filter(
        teacher=teacher
    ).select_related('class_assigned', 'subject').order_by(
        'class_assigned__level_number', 'class_assigned__name', 'subject__name'
    )

    # Build class data with student counts
    classes_data = []
    seen_classes = set()

    # Add homeroom classes (counts and subjects already prefetched)
    for cls in homeroom_classes:
        classes_data.append({
            'class': cls,
            'is_homeroom': True,
            'subjects': [],  # Subjects this teacher teaches
            'class_subjects': cls.prefetched_subjects,
            'student_count': cls.active_student_count,
        })
        seen_classes.add(cls.id)

    # Batch-fetch student counts for all non-homeroom classes
    non_homeroom_class_ids = {
        a.class_assigned_id for a in subject_assignments
        if a.class_assigned_id not in seen_classes
    }
    student_counts = {}
    if non_homeroom_class_ids:
        student_counts = dict(
            Student.objects.filter(
                current_class_id__in=non_homeroom_class_ids, status='active'
            ).values('current_class_id').annotate(
                count=Count('id')
            ).values_list('current_class_id', 'count')
        )

    # Batch-prefetch ClassSubjects for non-homeroom classes
    class_subjects_by_class = {}
    if non_homeroom_class_ids:
        for cs in ClassSubject.objects.filter(
            class_assigned_id__in=non_homeroom_class_ids
        ).select_related('subject', 'teacher').order_by('-subject__is_core', 'subject__name'):
            class_subjects_by_class.setdefault(cs.class_assigned_id, []).append(cs)

    # Add classes from subject assignments
    for assignment in subject_assignments:
        cls = assignment.class_assigned
        if cls.id not in seen_classes:
            classes_data.append({
                'class': cls,
                'is_homeroom': False,
                'subjects': [assignment.subject],  # Subjects this teacher teaches
                'class_subjects': class_subjects_by_class.get(cls.id, []),
                'student_count': student_counts.get(cls.id, 0),
            })
            seen_classes.add(cls.id)
        else:
            # Add subject to existing class entry
            for data in classes_data:
                if data['class'].id == cls.id:
                    if assignment.subject not in data['subjects']:
                        data['subjects'].append(assignment.subject)
                    break

    # Separate homeroom and subject-only classes
    homeroom_classes_list = [c for c in classes_data if c['is_homeroom']]
    subject_classes_list = [c for c in classes_data if not c['is_homeroom']]

    # Calculate totals
    total_students = sum(c['student_count'] for c in classes_data)

    # Get all unique subjects teacher teaches
    all_subjects = list({subj for c in classes_data for subj in c['subjects']})
    all_subjects.sort(key=lambda s: s.name)

    context = {
        'teacher': teacher,
        'current_term': current_term,
        'classes_data': classes_data,
        'homeroom_classes': homeroom_classes_list,
        'subject_classes': subject_classes_list,
        'total_classes': len(classes_data),
        'homeroom_count': len(homeroom_classes_list),
        'total_students': total_students,
        'total_subjects': len(all_subjects),
        'all_subjects': all_subjects,
    }

    return htmx_render(request, 'core/teacher/my_classes.html', 'core/teacher/partials/my_classes_content.html', context)


@login_required
def my_timetable(request):
    """Teacher's weekly timetable showing all their scheduled classes."""
    from academics.models import TimetableEntry, Period

    user = request.user

    # Verify user is a teacher
    if not getattr(user, 'is_teacher', False):
        messages.error(request, 'This page is only accessible to teachers.')
        return redirect('core:index')

    teacher = getattr(user, 'teacher_profile', None)
    if not teacher:
        messages.error(request, 'No teacher profile linked to your account.')
        return redirect('core:index')

    # Get all periods ordered by start time
    periods = Period.objects.filter(is_active=True).order_by('order', 'start_time')

    # Get all timetable entries for this teacher
    entries = TimetableEntry.objects.filter(
        class_subject__teacher=teacher
    ).select_related(
        'class_subject__class_assigned',
        'class_subject__subject',
        'period'
    ).order_by('weekday', 'period__order')

    # Build timetable grid: {weekday: {period_id: entry}}
    timetable = {day: {} for day in range(1, 6)}  # Monday=1 to Friday=5
    for entry in entries:
        timetable[entry.weekday][entry.period_id] = entry

    # Weekday labels
    weekdays = [
        (1, 'Monday'),
        (2, 'Tuesday'),
        (3, 'Wednesday'),
        (4, 'Thursday'),
        (5, 'Friday'),
    ]

    # Calculate stats
    total_lessons = entries.count()
    classes_taught = len(set(e.class_subject.class_assigned_id for e in entries))
    subjects_taught = len(set(e.class_subject.subject_id for e in entries))

    context = {
        'periods': periods,
        'timetable': timetable,
        'weekdays': weekdays,
        'entries': entries,
        'total_lessons': total_lessons,
        'classes_taught': classes_taught,
        'subjects_taught': subjects_taught,
    }

    return htmx_render(request, 'core/teacher/my_timetable.html', 'core/teacher/partials/my_timetable_content.html', context)


@login_required
def my_workload(request):
    """Teacher's workload analytics - imported from teachers app."""
    from teachers.views.analytics import my_workload as teacher_my_workload
    return teacher_my_workload(request)


@login_required
def my_promotions(request):
    """Teacher manages their own promotions."""
    from teachers.models import Promotion
    user = request.user
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('core:index')
    teacher = user.teacher_profile
    promotions = Promotion.objects.filter(teacher=teacher).order_by('-date_promoted')
    context = {'teacher': teacher, 'promotions': promotions}
    template = 'teachers/partials/my_promotions_content.html'
    if request.htmx:
        return render(request, template, context)
    return render(request, 'teachers/my_promotions.html', context)


@login_required
def my_qualifications(request):
    """Teacher manages their own qualifications."""
    from teachers.models import Qualification
    user = request.user
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('core:index')
    teacher = user.teacher_profile
    qualifications = Qualification.objects.filter(teacher=teacher).order_by('-date_ended')
    context = {'teacher': teacher, 'qualifications': qualifications}
    template = 'teachers/partials/my_qualifications_content.html'
    if request.htmx:
        return render(request, template, context)
    return render(request, 'teachers/my_qualifications.html', context)


@login_required
def my_attendance(request):
    """Teacher's attendance dashboard - view and take attendance for assigned classes."""
    from django.db.models import Count, Q
    from datetime import timedelta
    from academics.models import (
        Class, ClassSubject, AttendanceSession, AttendanceRecord,
        TimetableEntry
    )

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('core:index')

    teacher = user.teacher_profile

    # Get teacher's classes (homeroom + assigned)
    homeroom_classes = Class.objects.filter(class_teacher=teacher, is_active=True)
    assigned_class_ids = ClassSubject.objects.filter(teacher=teacher).values_list('class_assigned_id', flat=True)
    all_class_ids = set(homeroom_classes.values_list('id', flat=True)) | set(assigned_class_ids)
    homeroom_ids = set(homeroom_classes.values_list('id', flat=True))

    # Get classes with student counts in single query (optimized)
    classes = Class.objects.filter(id__in=all_class_ids, is_active=True).annotate(
        active_student_count=Count('students', filter=Q(students__status='active'))
    ).order_by('level_number', 'name')

    # Get filter parameters
    class_filter = request.GET.get('class', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Default date range: last 30 days
    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    # Base queryset for records (filtered by date/class later)
    records_qs = AttendanceRecord.objects.filter(
        session__class_assigned_id__in=all_class_ids
    )

    # Apply date filter
    if date_from:
        records_qs = records_qs.filter(session__date__gte=date_from)
    if date_to:
        records_qs = records_qs.filter(session__date__lte=date_to)

    # Apply class filter
    if class_filter:
        records_qs = records_qs.filter(session__class_assigned_id=class_filter)

    # Calculate all summary stats in single aggregated query (optimized)
    overall_stats = records_qs.aggregate(
        total=Count('id'),
        present=Count('id', filter=Q(status__in=['P', 'L'])),
        absent=Count('id', filter=Q(status='A')),
        late=Count('id', filter=Q(status='L'))
    )

    total_records = overall_stats['total']
    present_count = overall_stats['present']
    absent_count = overall_stats['absent']
    late_count = overall_stats['late']

    attendance_rate = 0
    if total_records > 0:
        attendance_rate = round((present_count / total_records) * 100, 1)

    # Pre-compute class stats in one aggregated query (eliminates N+1)
    class_stats = records_qs.values('session__class_assigned_id').annotate(
        total=Count('id'),
        present=Count('id', filter=Q(status__in=['P', 'L'])),
        absent=Count('id', filter=Q(status='A'))
    )
    class_stats_dict = {item['session__class_assigned_id']: item for item in class_stats}

    # Pre-compute which classes have today's attendance in one query
    today_sessions = set(
        AttendanceSession.objects.filter(
            class_assigned_id__in=all_class_ids,
            date=today
        ).values_list('class_assigned_id', flat=True)
    )

    # Get today's timetable entries for the teacher (for per-lesson attendance)
    today_weekday = today.weekday()
    teacher_timetable_entries = TimetableEntry.objects.filter(
        class_subject__teacher=teacher,
        weekday=today_weekday,
    ).select_related('class_subject__class_assigned', 'class_subject__subject', 'period').order_by('period__start_time')

    # Get which lessons have attendance taken today (session_type='Lesson')
    lessons_with_attendance = set(
        AttendanceSession.objects.filter(
            timetable_entry__in=teacher_timetable_entries,
            date=today,
            session_type='Lesson'
        ).values_list('timetable_entry_id', flat=True)
    )

    # Build lessons list for per-lesson attendance classes
    today_lessons = []
    for entry in teacher_timetable_entries:
        class_obj = entry.class_subject.class_assigned
        if class_obj.attendance_type == Class.AttendanceType.PER_LESSON:
            today_lessons.append({
                'entry': entry,
                'class': class_obj,
                'subject': entry.class_subject.subject,
                'period': entry.period,
                'has_attendance': entry.id in lessons_with_attendance,
            })

    # Get form class for per-lesson attendance (where teacher is form master)
    # This allows form masters to view reports even if they don't teach lessons
    form_class = None
    for cls in homeroom_classes:
        if cls.attendance_type == Class.AttendanceType.PER_LESSON:
            # Get stats for this class
            cls_stats = class_stats_dict.get(cls.id, {'total': 0, 'present': 0, 'absent': 0})
            cls_total = cls_stats['total']
            cls_present = cls_stats['present']
            cls_rate = round((cls_present / cls_total) * 100, 1) if cls_total > 0 else 0
            form_class = {
                'class': cls,
                'student_count': cls.students.filter(status='active').count(),
                'rate': cls_rate,
            }
            break  # Only one form class per teacher

    # Build class summary without additional queries
    class_summary = []
    for cls in classes:
        stats = class_stats_dict.get(cls.id, {'total': 0, 'present': 0, 'absent': 0})
        cls_total = stats['total']
        cls_present = stats['present']
        cls_absent = stats['absent']
        cls_rate = round((cls_present / cls_total) * 100, 1) if cls_total > 0 else 0

        # Check if this is a per-lesson class
        is_per_lesson = cls.attendance_type == Class.AttendanceType.PER_LESSON

        class_summary.append({
            'class': cls,
            'total': cls_total,
            'present': cls_present,
            'absent': cls_absent,
            'rate': cls_rate,
            'is_homeroom': cls.id in homeroom_ids,
            'has_today': cls.id in today_sessions,
            'student_count': cls.active_student_count,
            'is_per_lesson': is_per_lesson,
        })

    # Base sessions queryset with filters
    sessions_qs = AttendanceSession.objects.filter(
        class_assigned_id__in=all_class_ids
    ).select_related('class_assigned')

    if date_from:
        sessions_qs = sessions_qs.filter(date__gte=date_from)
    if date_to:
        sessions_qs = sessions_qs.filter(date__lte=date_to)
    if class_filter:
        sessions_qs = sessions_qs.filter(class_assigned_id=class_filter)

    total_sessions = sessions_qs.count()

    # Recent sessions with annotated counts (eliminates N+1)
    recent_sessions = sessions_qs.annotate(
        total_records=Count('records'),
        present_records=Count('records', filter=Q(records__status__in=['P', 'L'])),
        absent_records=Count('records', filter=Q(records__status='A'))
    ).order_by('-date')[:10]

    recent_data = []
    for session in recent_sessions:
        recent_data.append({
            'session': session,
            'total': session.total_records,
            'present': session.present_records,
            'absent': session.absent_records,
            'rate': round((session.present_records / session.total_records) * 100, 1) if session.total_records > 0 else 0,
        })

    # Calculate today's completion stats
    # For daily classes, count completed sessions
    daily_classes = [c for c in class_summary if not c['is_per_lesson']]
    daily_done = len([c for c in daily_classes if c['has_today']])
    daily_pending = len(daily_classes) - daily_done

    # For per-lesson classes, count completed lessons
    lessons_done = len([lesson for lesson in today_lessons if lesson['has_attendance']])
    lessons_pending = len(today_lessons) - lessons_done

    # Combined stats
    classes_done_today = daily_done
    classes_pending_today = daily_pending

    context = {
        'teacher': teacher,
        'classes': classes,
        'class_filter': class_filter,
        'date_from': date_from,
        'date_to': date_to,
        'today': today,
        'classes_done_today': classes_done_today,
        'classes_pending_today': classes_pending_today,
        'lessons_done_today': lessons_done,
        'lessons_pending_today': lessons_pending,
        'today_lessons': today_lessons,
        'form_class': form_class,
        'stats': {
            'total_sessions': total_sessions,
            'total_records': total_records,
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
            'rate': attendance_rate,
        },
        'class_summary': class_summary,
        'recent_data': recent_data,
    }

    return htmx_render(request, 'core/teacher/my_attendance.html', 'core/teacher/partials/my_attendance_content.html', context)


@login_required
def take_attendance(request, class_id):
    """Teacher takes attendance for a specific class."""
    from academics.models import Class, ClassSubject, AttendanceSession, AttendanceRecord
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'You do not have permission to take attendance.')
        return redirect('core:index')

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)

    # Check permission: must be class teacher or assigned to teach this class
    is_class_teacher = class_obj.class_teacher == teacher
    is_subject_teacher = ClassSubject.objects.filter(
        class_assigned=class_obj,
        teacher=teacher
    ).exists()

    if not is_class_teacher and not is_subject_teacher:
        messages.error(request, 'You are not assigned to this class.')
        return redirect('core:my_attendance')

    target_date = timezone.now().date()

    # Get or create session with explicit session_type
    session, created = AttendanceSession.objects.get_or_create(
        class_assigned=class_obj,
        date=target_date,
        session_type=AttendanceSession.SessionType.DAILY,
        defaults={'created_by': teacher}
    )

    if request.method == 'POST':
        students = list(Student.objects.filter(current_class=class_obj, status='active'))

        # Get existing records for bulk update vs create
        existing_records = {
            r.student_id: r for r in AttendanceRecord.objects.filter(
                session=session,
                student__in=students
            ).select_for_update()
        }

        records_to_create = []
        records_to_update = []

        for student in students:
            status_key = f"status_{student.id}"
            new_status = request.POST.get(status_key, AttendanceRecord.Status.PRESENT)

            if student.id in existing_records:
                # Update existing record
                record = existing_records[student.id]
                if record.status != new_status:
                    record.status = new_status
                    records_to_update.append(record)
            else:
                # Create new record
                records_to_create.append(AttendanceRecord(
                    session=session,
                    student=student,
                    status=new_status
                ))

        # Bulk operations with error handling
        try:
            if records_to_create:
                AttendanceRecord.objects.bulk_create(records_to_create)
            if records_to_update:
                AttendanceRecord.objects.bulk_update(records_to_update, ['status'])

            total_saved = len(records_to_create) + len(records_to_update)
            messages.success(request, f'Attendance saved for {class_obj.name} ({total_saved} records).')
        except Exception as e:
            messages.error(request, f'Failed to save attendance: {str(e)}')
            if request.htmx:
                response = HttpResponse(status=500)
                response['HX-Reswap'] = 'none'
                return response
            return redirect('core:take_attendance', class_id=class_id)

        if request.htmx:
            response = HttpResponse(status=204)
            response['HX-Redirect'] = reverse('core:my_attendance')
            return response

        return redirect('core:my_attendance')

    # GET: Prepare form data
    students = Student.objects.filter(current_class=class_obj, status='active').order_by('first_name', 'last_name')
    records = {r.student_id: r.status for r in session.records.only('student_id', 'status')}

    student_list = []
    for student in students:
        student_list.append({
            'obj': student,
            'status': records.get(student.id, 'P')
        })

    context = {
        'class': class_obj,
        'session': session,
        'student_list': student_list,
        'date': target_date,
        'is_homeroom': class_obj.class_teacher == teacher,
    }

    return htmx_render(request, 'core/teacher/take_attendance.html', 'core/teacher/partials/take_attendance_content.html', context)


@login_required
def my_grading(request):
    """Teacher's grading dashboard - view and enter scores for assigned classes."""
    from django.db.models import Count, Q
    from academics.models import ClassSubject
    from gradebook.models import Assignment, Score, AssessmentCategory

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('core:index')

    teacher = user.teacher_profile
    current_term = Term.get_current()

    # Get teacher's class-subject assignments with student counts (optimized - single query)
    assignments = ClassSubject.objects.filter(
        teacher=teacher,
        class_assigned__is_active=True
    ).select_related('class_assigned', 'subject').annotate(
        student_count=Count(
            'class_assigned__students',
            filter=Q(class_assigned__students__status='active')
        )
    ).order_by(
        'class_assigned__level_number', 'class_assigned__name', 'subject__name'
    )

    # Pre-compute assignment and score counts efficiently (eliminates N+1 queries)
    assignment_counts = {}
    score_counts = {}

    if current_term and assignments.exists():
        subject_ids = list(assignments.values_list('subject_id', flat=True).distinct())
        class_ids = list(assignments.values_list('class_assigned_id', flat=True).distinct())

        # Get assignment counts per subject in one query
        assignment_counts = dict(
            Assignment.objects.filter(
                subject_id__in=subject_ids,
                term=current_term
            ).values('subject_id').annotate(
                count=Count('id')
            ).values_list('subject_id', 'count')
        )

        # Get score counts per (class, subject) combination in one query
        score_data = Score.objects.filter(
            assignment__term=current_term,
            assignment__subject_id__in=subject_ids,
            student__current_class_id__in=class_ids
        ).values(
            'student__current_class_id',
            'assignment__subject_id'
        ).annotate(count=Count('id'))

        for item in score_data:
            key = (item['student__current_class_id'], item['assignment__subject_id'])
            score_counts[key] = item['count']

    # Build class data using pre-computed values (no additional queries)
    class_data = []
    for assignment in assignments:
        cls_id = assignment.class_assigned_id
        subject_id = assignment.subject_id
        student_count = assignment.student_count

        term_assignments = assignment_counts.get(subject_id, 0)
        scores_entered = score_counts.get((cls_id, subject_id), 0)

        if term_assignments > 0 and student_count > 0:
            total_possible = term_assignments * student_count
            progress = round((scores_entered / total_possible) * 100)
        else:
            progress = 0

        class_data.append({
            'class': assignment.class_assigned,
            'subject': assignment.subject,
            'student_count': student_count,
            'assignments': term_assignments,
            'scores_entered': scores_entered,
            'progress': progress,
        })

    # Get categories for reference
    categories = AssessmentCategory.objects.filter(is_active=True).order_by('order')

    context = {
        'teacher': teacher,
        'current_term': current_term,
        'class_data': class_data,
        'categories': categories,
        'total_classes': len(class_data),
    }

    return htmx_render(request, 'core/teacher/my_grading.html', 'core/teacher/partials/my_grading_content.html', context)


@login_required
def enter_scores(request, class_id, subject_id):
    """Teacher enters scores for a specific class/subject."""
    from academics.models import Class, ClassSubject, Subject
    from gradebook.models import Assignment, Score, AssessmentCategory
    from students.models import Student
    from collections import defaultdict

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'You do not have permission to enter scores.')
        return redirect('core:index')

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    current_term = Term.get_current()

    # Check permission: must be assigned to teach this subject in this class
    is_assigned = ClassSubject.objects.filter(
        class_assigned=class_obj,
        subject=subject,
        teacher=teacher
    ).exists()

    if not is_assigned:
        messages.error(request, 'You are not assigned to teach this subject in this class.')
        return redirect('core:my_grading')

    # Check if grades are locked
    grades_locked = current_term.grades_locked if current_term else True

    # Get students - only fetch needed fields
    students = list(Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).only('id', 'first_name', 'last_name', 'admission_number', 'photo').order_by('last_name', 'first_name'))

    # Get assignments for this subject/term
    assignments_list = list(Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name'))

    # Get existing scores - build nested dict for O(1) lookup
    scores_dict = defaultdict(dict)
    student_totals = {}  # Server-side totals calculation
    if students and assignments_list:
        student_ids = [s.id for s in students]
        assignment_ids = [a.id for a in assignments_list]
        for score in Score.objects.filter(
            student_id__in=student_ids,
            assignment_id__in=assignment_ids
        ).only('student_id', 'assignment_id', 'points'):
            scores_dict[score.student_id][score.assignment_id] = score.points

        # Calculate totals server-side
        for student in students:
            total = sum(
                scores_dict[student.id].get(a.id, 0) or 0
                for a in assignments_list
            )
            student_totals[student.id] = round(total, 1) if total > 0 else None

    # Get categories
    categories = AssessmentCategory.objects.filter(is_active=True).order_by('order')

    # Check for view mode (table or student)
    view_mode = request.GET.get('view', 'table')
    current_student = None
    current_student_index = 0
    prev_student = None
    next_student = None
    current_student_scores = {}
    assignments_by_category = {}
    filled_scores = 0

    if view_mode == 'student' and students:
        # Get specific student or default to first
        student_id = request.GET.get('student_id')
        if student_id:
            try:
                student_id = int(student_id)
                current_student = next((s for s in students if s.id == student_id), students[0])
            except (ValueError, TypeError):
                current_student = students[0]
        else:
            current_student = students[0]

        # Find index for prev/next navigation
        current_student_index = next((i for i, s in enumerate(students) if s.id == current_student.id), 0)
        prev_student = students[current_student_index - 1] if current_student_index > 0 else None
        next_student = students[current_student_index + 1] if current_student_index < len(students) - 1 else None

        # Get scores for current student
        current_student_scores = scores_dict.get(current_student.id, {})

        # Group assignments by category for student view
        assignments_by_category = {}
        for assignment in assignments_list:
            cat = assignment.assessment_category
            if cat not in assignments_by_category:
                assignments_by_category[cat] = []
            assignments_by_category[cat].append(assignment)

        # Score progress
        filled_scores = sum(
            1 for a in assignments_list if a.id in current_student_scores
        )

    # Check if teacher is class teacher for this class
    is_class_teacher = class_obj.class_teacher == teacher

    context = {
        'class_obj': class_obj,
        'subject': subject,
        'current_term': current_term,
        'students': students,
        'assignments': assignments_list,
        'categories': categories,
        'scores_dict': dict(scores_dict),
        'student_totals': student_totals,
        'grades_locked': grades_locked,
        'can_edit': not grades_locked,
        'is_class_teacher': is_class_teacher,
        'editing_allowed': not grades_locked,
        # Student view context
        'view_mode': view_mode,
        'current_student': current_student,
        'current_student_index': current_student_index,
        'prev_student': prev_student,
        'next_student': next_student,
        'current_student_scores': current_student_scores,
        'assignments_by_category': assignments_by_category if view_mode == 'student' else {},
        'filled_scores': filled_scores if view_mode == 'student' else 0,
        'total_assignments': len(assignments_list),
    }

    return htmx_render(
        request, 'core/teacher/enter_scores.html',
        'core/teacher/partials/enter_scores_content.html', context
    )


@login_required
def enter_scores_student(request, class_id, subject_id, student_id):
    """HTMX endpoint for per-student score entry (mobile-optimized)."""
    from academics.models import Class, ClassSubject, Subject
    from gradebook.models import Assignment, Score
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        return HttpResponse('Unauthorized', status=403)

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    current_term = Term.get_current()

    # Check permission
    is_assigned = ClassSubject.objects.filter(
        class_assigned=class_obj,
        subject=subject,
        teacher=teacher
    ).exists()

    if not is_assigned:
        return HttpResponse('Not authorized', status=403)

    grades_locked = current_term.grades_locked if current_term else True

    # Get all students for navigation
    students = list(Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).only('id', 'first_name', 'last_name', 'admission_number', 'photo').order_by('last_name', 'first_name'))

    # Get specific student
    student = get_object_or_404(Student, pk=student_id, current_class=class_obj)

    # Find index for prev/next
    current_index = next((i for i, s in enumerate(students) if s.id == student.id), 0)
    prev_student = students[current_index - 1] if current_index > 0 else None
    next_student = students[current_index + 1] if current_index < len(students) - 1 else None

    # Get assignments (only if current_term exists)
    assignments = []
    if current_term:
        assignments = list(Assignment.objects.filter(
            subject=subject,
            term=current_term
        ).select_related('assessment_category').order_by('assessment_category__order', 'name'))

    # Get scores for this student
    scores_dict = {}
    for score in Score.objects.filter(
        student=student,
        assignment__in=assignments
    ).only('assignment_id', 'points'):
        scores_dict[score.assignment_id] = score.points

    context = {
        'class_obj': class_obj,
        'subject': subject,
        'student': student,
        'students': students,
        'current_index': current_index,
        'prev_student': prev_student,
        'next_student': next_student,
        'assignments': assignments,
        'scores_dict': scores_dict,
        'grades_locked': grades_locked,
    }

    return render(request, 'core/teacher/partials/enter_scores_student.html', context)


@login_required
def export_scores(request, class_id, subject_id):
    """Export scores for a class/subject as CSV or Excel."""
    import csv
    from io import BytesIO
    from django.http import HttpResponse
    from academics.models import Class, ClassSubject, Subject
    from gradebook.models import Assignment, Score
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'You do not have permission to export scores.')
        return redirect('core:index')

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    current_term = Term.get_current()

    # Check permission
    is_assigned = ClassSubject.objects.filter(
        class_assigned=class_obj,
        subject=subject,
        teacher=teacher
    ).exists()

    if not is_assigned:
        messages.error(request, 'You are not assigned to teach this subject.')
        return redirect('core:my_grading')

    # Get students
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('last_name', 'first_name')

    # Get assignments
    assignments = Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name')

    # Get existing scores
    scores_dict = {}
    if students and assignments:
        student_ids = [s.id for s in students]
        assignment_ids = [a.id for a in assignments]
        for score in Score.objects.filter(
            student_id__in=student_ids,
            assignment_id__in=assignment_ids
        ):
            if score.student_id not in scores_dict:
                scores_dict[score.student_id] = {}
            scores_dict[score.student_id][score.assignment_id] = score.points

    # Check format
    export_format = request.GET.get('format', 'csv')

    # Build headers
    headers = ['Admission No', 'Student Name']
    for assignment in assignments:
        headers.append(f"{assignment.name} ({assignment.assessment_category.short_name}) /{assignment.points_possible}")

    # Build rows
    rows = []
    for student in students:
        row = [student.admission_number, student.full_name]
        student_scores = scores_dict.get(student.id, {})
        for assignment in assignments:
            score = student_scores.get(assignment.id, '')
            row.append(score if score != '' else '')
        rows.append(row)

    filename = f"{class_obj.name}_{subject.name}_{current_term.name}_scores".replace(' ', '_')

    if export_format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scores"

            # Header style
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Write data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            for col_idx in range(3, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

            # Create response
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response

        except ImportError:
            export_format = 'csv'

    # CSV export
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    return response


@login_required
def import_scores(request, class_id, subject_id):
    """Import scores from CSV/Excel file - preview step."""
    import csv
    import json
    from io import TextIOWrapper
    from django.http import HttpResponse
    from academics.models import Class, ClassSubject, Subject
    from gradebook.models import Assignment
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        return HttpResponse('<div class="alert alert-error">Permission denied.</div>')

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    current_term = Term.get_current()

    # Check permission
    is_assigned = ClassSubject.objects.filter(
        class_assigned=class_obj,
        subject=subject,
        teacher=teacher
    ).exists()

    if not is_assigned:
        return HttpResponse('<div class="alert alert-error">You are not assigned to teach this subject.</div>')

    # Check if grades locked
    if current_term and current_term.grades_locked:
        return HttpResponse('<div class="alert alert-warning">Grades are locked for this term.</div>')

    if request.method != 'POST':
        return HttpResponse('<div class="alert alert-error">Invalid request.</div>')

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return HttpResponse('<div class="alert alert-error">No file uploaded.</div>')

    # File size validation (max 5MB)
    MAX_UPLOAD_SIZE = 5 * 1024 * 1024  # 5MB
    if uploaded_file.size > MAX_UPLOAD_SIZE:
        return HttpResponse(
            '<div class="alert alert-error">File too large. Maximum size is 5MB.</div>'
        )

    # Get assignments for validation
    assignments = list(Assignment.objects.filter(
        subject=subject,
        term=current_term
    ).select_related('assessment_category').order_by('assessment_category__order', 'name'))

    # Get students for validation
    students = {s.admission_number: s for s in Student.objects.filter(
        current_class=class_obj,
        status='active'
    )}

    # Parse file
    rows = []
    filename = uploaded_file.name.lower()

    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                import openpyxl
                from io import BytesIO

                wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()))
                ws = wb.active

                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == 0:
                        continue
                    if row and row[0]:
                        rows.append(list(row))
            except ImportError:
                return HttpResponse('<div class="alert alert-error">Excel support not available. Please use CSV format.</div>')
        else:
            decoded_file = TextIOWrapper(uploaded_file.file, encoding='utf-8-sig')
            reader = csv.reader(decoded_file)
            next(reader)
            for row in reader:
                if row and row[0]:
                    rows.append(row)
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-error">Error reading file: {str(e)}</div>')

    # Validate and prepare preview data
    preview_data = []
    valid_count = 0
    error_count = 0

    for row_idx, row in enumerate(rows):
        if len(row) < 2:
            continue

        admission_no = str(row[0]).strip()
        student = students.get(admission_no)

        row_data = {
            'row_num': row_idx + 2,
            'admission_no': admission_no,
            'student_name': row[1] if len(row) > 1 else '',
            'student': student,
            'scores': [],
            'errors': []
        }

        if not student:
            row_data['errors'].append(f"Student not found: {admission_no}")
            error_count += 1
        else:
            for idx, assignment in enumerate(assignments):
                col_idx = idx + 2
                score_value = row[col_idx] if len(row) > col_idx else ''

                score_data = {
                    'assignment': assignment,
                    'value': score_value,
                    'valid': True,
                    'error': None
                }

                if score_value != '' and score_value is not None:
                    try:
                        score_float = float(score_value)
                        if score_float < 0:
                            score_data['valid'] = False
                            score_data['error'] = 'Negative value'
                        elif score_float > assignment.points_possible:
                            score_data['valid'] = False
                            score_data['error'] = f'Exceeds max ({assignment.points_possible})'
                    except (ValueError, TypeError):
                        score_data['valid'] = False
                        score_data['error'] = 'Invalid number'

                row_data['scores'].append(score_data)

                if not score_data['valid']:
                    error_count += 1

            if not row_data['errors'] and all(s['valid'] for s in row_data['scores']):
                valid_count += 1

        preview_data.append(row_data)

    # Store in session for confirmation
    session_data = []
    for row in preview_data:
        if row['student']:
            scores_to_save = []
            for score_data in row['scores']:
                if score_data['valid'] and score_data['value'] != '' and score_data['value'] is not None:
                    scores_to_save.append({
                        'assignment_id': str(score_data['assignment'].id),
                        'value': float(score_data['value'])
                    })
            session_data.append({
                'student_id': row['student'].id,
                'scores': scores_to_save
            })

    request.session[f'import_scores_{class_id}_{subject_id}'] = json.dumps(session_data)

    # Render preview
    from django.middleware.csrf import get_token
    csrf_token = get_token(request)

    html_content = f'''
    <div class="space-y-4">
        <div class="flex gap-4">
            <div class="stat bg-success/10 rounded-lg p-3">
                <div class="stat-title text-xs">Valid Rows</div>
                <div class="stat-value text-lg text-success">{valid_count}</div>
            </div>
            <div class="stat bg-error/10 rounded-lg p-3">
                <div class="stat-title text-xs">Errors</div>
                <div class="stat-value text-lg text-error">{error_count}</div>
            </div>
        </div>

        <div class="overflow-x-auto max-h-64">
            <table class="table table-xs">
                <thead>
                    <tr>
                        <th>Row</th>
                        <th>Admission No</th>
                        <th>Name</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
    '''

    for row in preview_data[:20]:
        status_class = 'text-success' if row['student'] and not row['errors'] else 'text-error'
        status_icon = 'fa-check' if row['student'] and not row['errors'] else 'fa-xmark'
        status_text = 'Valid' if row['student'] and not row['errors'] else (row['errors'][0] if row['errors'] else 'Has errors')

        # Escape user-supplied data to prevent XSS
        escaped_admission_no = html.escape(str(row['admission_no']))
        escaped_student_name = html.escape(str(row['student_name']))
        escaped_status_text = html.escape(str(status_text))

        html_content += f'''
            <tr>
                <td>{row['row_num']}</td>
                <td class="font-mono">{escaped_admission_no}</td>
                <td>{escaped_student_name}</td>
                <td class="{status_class}">
                    <i class="fa-solid {status_icon} mr-1"></i>{escaped_status_text}
                </td>
            </tr>
        '''

    if len(preview_data) > 20:
        html_content += f'<tr><td colspan="4" class="text-center text-base-content/60">... and {len(preview_data) - 20} more rows</td></tr>'

    html_content += '''
                </tbody>
            </table>
        </div>
    '''

    from django.urls import reverse
    confirm_url = reverse('core:import_scores_confirm', args=[class_id, subject_id])

    if valid_count > 0:
        html_content += f'''
        <form hx-post="{confirm_url}"
              hx-target="#import-content"
              hx-swap="innerHTML">
            <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
            <div class="modal-action">
                <button type="button" class="btn btn-ghost" onclick="modal_import.close()">Cancel</button>
                <button type="submit" class="btn btn-primary gap-2">
                    <i class="fa-solid fa-check"></i>
                    Import {valid_count} Rows
                </button>
            </div>
        </form>
        '''
    else:
        html_content += '''
        <div class="modal-action">
            <button type="button" class="btn btn-ghost" onclick="modal_import.close()">Close</button>
        </div>
        '''

    html_content += '</div>'

    return HttpResponse(html_content)


@login_required
def import_scores_confirm(request, class_id, subject_id):
    """Confirm and save imported scores."""
    import json
    from django.http import HttpResponse
    from academics.models import Class, Subject, ClassSubject
    from gradebook.models import Score

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        return HttpResponse('<div class="alert alert-error">Permission denied.</div>')

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)
    subject = get_object_or_404(Subject, pk=subject_id)
    current_term = Term.get_current()

    # Check permission
    is_assigned = ClassSubject.objects.filter(
        class_assigned=class_obj,
        subject=subject,
        teacher=teacher
    ).exists()

    if not is_assigned:
        return HttpResponse('<div class="alert alert-error">You are not assigned to teach this subject.</div>')

    # Check if grades locked
    if current_term and current_term.grades_locked:
        return HttpResponse('<div class="alert alert-warning">Grades are locked for this term.</div>')

    # Get session data
    session_key = f'import_scores_{class_id}_{subject_id}'
    session_data = request.session.get(session_key)

    if not session_data:
        return HttpResponse('<div class="alert alert-error">Session expired. Please upload the file again.</div>')

    try:
        import_data = json.loads(session_data)
    except json.JSONDecodeError:
        return HttpResponse('<div class="alert alert-error">Invalid session data.</div>')

    # Import scores
    saved_count = 0
    updated_count = 0

    for row in import_data:
        student_id = row['student_id']
        for score_data in row['scores']:
            assignment_id = score_data['assignment_id']
            value = score_data['value']

            score, created = Score.objects.update_or_create(
                student_id=student_id,
                assignment_id=assignment_id,
                defaults={'points': value}
            )

            if created:
                saved_count += 1
            else:
                updated_count += 1

    # Clear session
    del request.session[session_key]

    # Return success message
    html = f'''
    <div class="text-center py-8">
        <div class="w-16 h-16 mx-auto mb-4 rounded-full bg-success/20 flex items-center justify-center">
            <i class="fa-solid fa-check text-success text-3xl"></i>
        </div>
        <h3 class="font-bold text-lg mb-2">Import Complete!</h3>
        <p class="text-base-content/60 mb-4">
            {saved_count} new scores added, {updated_count} scores updated.
        </p>
        <button type="button" class="btn btn-primary" onclick="modal_import.close(); location.reload();">
            Done
        </button>
    </div>
    '''

    return HttpResponse(html)


@login_required
def class_students(request, class_id):
    """Form teacher view to manage students in their homeroom class."""
    from academics.models import Class, ClassSubject, StudentSubjectEnrollment
    from students.models import Student
    from django.db.models import Prefetch

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        messages.error(request, 'This page is only accessible to teachers.')
        return redirect('core:index')

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)

    # Must be the form teacher (class teacher) of this class
    if class_obj.class_teacher != teacher:
        messages.error(request, 'You are not the form teacher for this class.')
        return redirect('core:my_classes')

    current_term = Term.get_current()

    # Get students with prefetched enrollments for this class
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).prefetch_related(
        Prefetch(
            'subject_enrollments',
            queryset=StudentSubjectEnrollment.objects.filter(
                class_subject__class_assigned=class_obj, is_active=True
            ).select_related('class_subject__subject'),
            to_attr='active_enrollments'
        )
    ).order_by('last_name', 'first_name')

    # Get class subjects (for elective enrollment)
    class_subjects = ClassSubject.objects.filter(
        class_assigned=class_obj
    ).select_related('subject', 'teacher').order_by('-subject__is_core', 'subject__name')

    core_subjects = [cs for cs in class_subjects if cs.subject.is_core]

    # Filter elective subjects by programme if class has a programme
    # This ensures students only see electives relevant to their programme
    if class_obj.programme:
        programme_subject_ids = set(
            class_obj.programme.subjects.values_list('id', flat=True)
        )
        elective_subjects = [
            cs for cs in class_subjects
            if not cs.subject.is_core and (
                # Include if subject is in programme's electives OR has no programme restrictions
                cs.subject.id in programme_subject_ids or
                not cs.subject.programmes.exists()
            )
        ]
        required_electives = class_obj.programme.required_electives
    else:
        elective_subjects = [cs for cs in class_subjects if not cs.subject.is_core]
        required_electives = 3  # Default minimum for SHS

    # Build enrollment data from prefetched results
    students_data = []
    students_missing_electives = 0

    for student in students:
        enrollments = student.active_enrollments

        enrolled_subjects = [e.class_subject.subject for e in enrollments]
        enrolled_electives = [e for e in enrollments if not e.class_subject.subject.is_core]
        enrolled_elective_ids = [e.class_subject_id for e in enrolled_electives]
        elective_count = len(enrolled_electives)

        # Check if student has enough electives (only for SHS with electives)
        electives_complete = True
        electives_missing = 0
        if elective_subjects and required_electives > 0:
            if elective_count < required_electives:
                electives_complete = False
                electives_missing = required_electives - elective_count
                students_missing_electives += 1

        students_data.append({
            'student': student,
            'enrolled_subjects_count': len(enrolled_subjects),
            'enrolled_electives': [e.class_subject.subject for e in enrolled_electives],
            'enrolled_elective_ids': enrolled_elective_ids,
            'elective_count': elective_count,
            'electives_complete': electives_complete,
            'electives_missing': electives_missing,
        })

    # Get students available for enrollment (not in any class)
    available_students = Student.objects.filter(
        status='active',
        current_class__isnull=True
    ).order_by('last_name', 'first_name')[:50]

    context = {
        'teacher': teacher,
        'class_obj': class_obj,
        'current_term': current_term,
        'students_data': students_data,
        'student_count': len(students_data),
        'core_subjects': core_subjects,
        'elective_subjects': elective_subjects,
        'available_students': available_students,
        'required_electives': required_electives if elective_subjects else 0,
        'students_missing_electives': students_missing_electives,
    }

    return htmx_render(
        request,
        'core/teacher/class_students.html',
        'core/teacher/partials/class_students_content.html',
        context
    )


@login_required
def enroll_student(request, class_id):
    """Enroll a student into a class (form teacher only)."""
    from academics.models import Class, StudentSubjectEnrollment
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        return HttpResponse('<div class="alert alert-error">Permission denied.</div>', status=403)

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)

    # Must be the form teacher
    if class_obj.class_teacher != teacher:
        return HttpResponse('<div class="alert alert-error">You are not the form teacher for this class.</div>', status=403)

    if request.method != 'POST':
        return HttpResponse(status=405)

    student_id = request.POST.get('student_id')
    if not student_id:
        return HttpResponse('<div class="alert alert-error">No student selected.</div>', status=400)

    student = get_object_or_404(Student, pk=student_id)

    # Check if student already in a class
    if student.current_class:
        return HttpResponse(f'<div class="alert alert-warning">Student is already in {student.current_class.name}.</div>', status=400)

    # Enroll student in the class
    student.current_class = class_obj
    student.save()

    # Auto-enroll in core subjects
    StudentSubjectEnrollment.enroll_student_in_core_subjects(student, class_obj, enrolled_by=teacher)

    # Return success response with redirect trigger
    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'studentEnrolled'
    return response


@login_required
def remove_student(request, class_id, student_id):
    """Remove a student from a class (form teacher only)."""
    from academics.models import Class, StudentSubjectEnrollment
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        return HttpResponse('<div class="alert alert-error">Permission denied.</div>', status=403)

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)

    # Must be the form teacher
    if class_obj.class_teacher != teacher:
        return HttpResponse('<div class="alert alert-error">You are not the form teacher for this class.</div>', status=403)

    if request.method != 'POST':
        return HttpResponse(status=405)

    student = get_object_or_404(Student, pk=student_id)

    # Check if student is in this class
    if student.current_class != class_obj:
        return HttpResponse('<div class="alert alert-error">Student is not in this class.</div>', status=400)

    # Remove subject enrollments
    StudentSubjectEnrollment.objects.filter(
        student=student,
        class_subject__class_assigned=class_obj
    ).update(is_active=False)

    # Remove from class
    student.current_class = None
    student.save()

    # Return success response with refresh trigger
    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'studentRemoved'
    return response


@login_required
def update_student_electives(request, class_id, student_id):
    """Update elective subject enrollments for a student."""
    from academics.models import Class, ClassSubject, StudentSubjectEnrollment
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        return HttpResponse('<div class="alert alert-error">Permission denied.</div>', status=403)

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)

    # Must be the form teacher
    if class_obj.class_teacher != teacher:
        return HttpResponse('<div class="alert alert-error">You are not the form teacher for this class.</div>', status=403)

    if request.method != 'POST':
        return HttpResponse(status=405)

    student = get_object_or_404(Student, pk=student_id)

    # Check if student is in this class
    if student.current_class != class_obj:
        return HttpResponse('<div class="alert alert-error">Student is not in this class.</div>', status=400)

    # Get selected elective IDs
    selected_elective_ids = request.POST.getlist('electives')

    # Get all elective class subjects for this class
    elective_class_subjects = ClassSubject.objects.filter(
        class_assigned=class_obj,
        subject__is_core=False
    )

    # Update enrollments
    for class_subject in elective_class_subjects:
        should_be_enrolled = str(class_subject.id) in selected_elective_ids

        if should_be_enrolled:
            # Create or reactivate enrollment
            enrollment, created = StudentSubjectEnrollment.objects.get_or_create(
                student=student,
                class_subject=class_subject,
                defaults={
                    'enrolled_by': teacher,
                    'is_active': True
                }
            )
            if not created and not enrollment.is_active:
                enrollment.is_active = True
                enrollment.save()
        else:
            # Deactivate enrollment
            StudentSubjectEnrollment.objects.filter(
                student=student,
                class_subject=class_subject
            ).update(is_active=False)

    # Return success response
    response = HttpResponse(status=204)
    response['HX-Trigger'] = 'electivesUpdated'
    return response


@login_required
def bulk_assign_electives(request, class_id):
    """Bulk assign elective subjects to all students missing electives."""
    import json
    from django.db.models import Q
    from academics.models import Class, ClassSubject, StudentSubjectEnrollment
    from students.models import Student

    user = request.user

    # Must be a teacher
    if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
        return HttpResponse('<div class="alert alert-error">Permission denied.</div>', status=403)

    teacher = user.teacher_profile
    class_obj = get_object_or_404(Class, pk=class_id)

    # Must be the form teacher
    if class_obj.class_teacher != teacher:
        return HttpResponse('<div class="alert alert-error">You are not the form teacher for this class.</div>', status=403)

    # Get elective subjects filtered by programme
    if class_obj.programme:
        programme_subject_ids = set(
            class_obj.programme.subjects.values_list('id', flat=True)
        )
        elective_class_subjects = ClassSubject.objects.filter(
            class_assigned=class_obj,
            subject__is_core=False
        ).filter(
            Q(subject_id__in=programme_subject_ids) |
            Q(subject__programmes__isnull=True)
        ).select_related('subject', 'teacher').distinct()
        required_electives = class_obj.programme.required_electives
    else:
        elective_class_subjects = ClassSubject.objects.filter(
            class_assigned=class_obj,
            subject__is_core=False
        ).select_related('subject', 'teacher')
        required_electives = 3

    # Find students who need electives
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('last_name', 'first_name')

    students_needing_electives = []
    for student in students:
        enrolled_elective_count = StudentSubjectEnrollment.objects.filter(
            student=student,
            class_subject__class_assigned=class_obj,
            class_subject__subject__is_core=False,
            is_active=True
        ).count()

        if enrolled_elective_count < required_electives:
            students_needing_electives.append({
                'student': student,
                'current_count': enrolled_elective_count,
                'needed': required_electives - enrolled_elective_count
            })

    context = {
        'class_obj': class_obj,
        'elective_subjects': elective_class_subjects,
        'students_needing_electives': students_needing_electives,
        'required_electives': required_electives,
    }

    if request.method == 'POST':
        # Get the selected electives to apply
        selected_elective_ids = request.POST.getlist('electives')
        student_ids = request.POST.getlist('students')

        if not selected_elective_ids:
            context['error'] = 'Please select at least one elective subject.'
            return render(request, 'core/teacher/partials/bulk_electives_modal.html', context)

        if not student_ids:
            context['error'] = 'Please select at least one student.'
            return render(request, 'core/teacher/partials/bulk_electives_modal.html', context)

        # Get students to update
        students_to_update = Student.objects.filter(pk__in=student_ids, current_class=class_obj)

        # Get the class subjects
        class_subjects = ClassSubject.objects.filter(
            id__in=selected_elective_ids,
            class_assigned=class_obj,
            subject__is_core=False
        )

        # Apply enrollments
        for student in students_to_update:
            for class_subject in class_subjects:
                enrollment, created = StudentSubjectEnrollment.objects.get_or_create(
                    student=student,
                    class_subject=class_subject,
                    defaults={
                        'enrolled_by': teacher,
                        'is_active': True
                    }
                )
                if not created and not enrollment.is_active:
                    enrollment.is_active = True
                    enrollment.save()

        # Close modal and trigger refresh
        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'electivesUpdated': True,
            'closeModal': True,
        })
        return response

    return render(request, 'core/teacher/partials/bulk_electives_modal.html', context)


# Student views
@login_required
def my_results(request):
    """Student view of their own grades and results."""
    from gradebook.models import SubjectTermGrade, TermReport, GradingSystem

    user = request.user
    student = getattr(user, 'student_profile', None)

    if not student:
        return redirect('core:index')

    current_term = Term.get_current()

    # Get all terms this student has results for
    available_terms = Term.objects.filter(
        subject_grades__student=student
    ).distinct().order_by('-academic_year__start_date', '-term_number')

    # Get selected term (default to current)
    selected_term_id = request.GET.get('term')
    if selected_term_id:
        try:
            selected_term = Term.objects.get(pk=selected_term_id)
        except Term.DoesNotExist:
            selected_term = current_term
    else:
        selected_term = current_term

    # Get subject grades for selected term
    subject_grades = []
    term_report = None
    grading_system = None
    grade_scales = []

    if selected_term:
        subject_grades = SubjectTermGrade.objects.filter(
            student=student,
            term=selected_term,
            total_score__isnull=False
        ).select_related('subject').order_by('subject__name')

        term_report = TermReport.objects.filter(
            student=student,
            term=selected_term
        ).first()

        # Get grading system for display
        if student.current_class:
            level_type = student.current_class.level_type
            grading_level = 'SHS' if level_type == 'shs' else 'BASIC'
            grading_system = GradingSystem.objects.filter(
                level=grading_level,
                is_active=True
            ).first()
            if grading_system:
                grade_scales = grading_system.scales.all().order_by('order')

    # Separate core and elective subjects
    core_grades = [g for g in subject_grades if g.subject.is_core]
    elective_grades = [g for g in subject_grades if not g.subject.is_core]

    context = {
        'student': student,
        'current_term': current_term,
        'selected_term': selected_term,
        'available_terms': available_terms,
        'subject_grades': subject_grades,
        'core_grades': core_grades,
        'elective_grades': elective_grades,
        'term_report': term_report,
        'grading_system': grading_system,
        'grade_scales': grade_scales,
    }
    return htmx_render(request, 'core/student/my_results.html', 'core/student/partials/my_results_content.html', context)


@login_required
def timetable(request):
    """Student's class timetable view."""
    from academics.models import TimetableEntry, Period

    user = request.user

    # Verify user is a student
    if not getattr(user, 'is_student', False):
        messages.error(request, 'This page is only accessible to students.')
        return redirect('core:index')

    student = getattr(user, 'student_profile', None)
    if not student:
        messages.error(request, 'No student profile linked to your account.')
        return redirect('core:index')

    # Get student's current class
    current_class = student.current_class
    if not current_class:
        context = {
            'no_class': True,
        }
        return htmx_render(request, 'core/student/timetable.html', 'core/student/partials/timetable_content.html', context)

    # Get all periods ordered by start time
    periods = Period.objects.filter(is_active=True).order_by('order', 'start_time')

    # Get all timetable entries for this class
    entries = TimetableEntry.objects.filter(
        class_subject__class_assigned=current_class
    ).select_related(
        'class_subject__subject',
        'class_subject__teacher__user',
        'period'
    ).order_by('weekday', 'period__order')

    # Build timetable grid: {weekday: {period_id: entry}}
    timetable = {day: {} for day in range(1, 6)}  # Monday=1 to Friday=5
    for entry in entries:
        timetable[entry.weekday][entry.period_id] = entry

    # Weekday labels
    weekdays = [
        (1, 'Monday'),
        (2, 'Tuesday'),
        (3, 'Wednesday'),
        (4, 'Thursday'),
        (5, 'Friday'),
    ]

    # Calculate stats
    total_lessons = entries.count()
    total_subjects = len(set(e.class_subject.subject_id for e in entries))

    context = {
        'student': student,
        'current_class': current_class,
        'periods': periods,
        'timetable': timetable,
        'weekdays': weekdays,
        'entries': entries,
        'total_lessons': total_lessons,
        'total_subjects': total_subjects,
    }

    return htmx_render(request, 'core/student/timetable.html', 'core/student/partials/timetable_content.html', context)


@login_required
def my_fees(request):
    context = {}
    return htmx_render(request, 'core/student/my_fees.html', 'core/student/partials/my_fees_content.html', context)


# Parent/Guardian views
@login_required
def guardian_dashboard(request):
    """Dashboard for logged-in guardians/parents showing their wards."""
    from gradebook.models import SubjectTermGrade, TermReport
    from students.models import StudentGuardian
    from academics.models import AttendanceSession, AttendanceRecord

    user = request.user
    current_term = Term.get_current()
    current_year = AcademicYear.get_current()

    # Get guardian profile linked to this user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.warning(request, "No guardian profile linked to your account.")
        context = {'guardian': None, 'wards': []}
        return htmx_render(request, 'core/parent/dashboard.html', 'core/parent/partials/dashboard_content.html', context)

    # Get wards (students) linked to this guardian
    student_guardians = StudentGuardian.objects.filter(
        guardian=guardian
    ).select_related(
        'student__current_class'
    ).order_by('-is_primary', 'student__first_name')

    # Collect active ward IDs for batch queries
    active_guardians = [(sg, sg.student) for sg in student_guardians if sg.student.status == 'active']
    active_student_ids = [s.id for _, s in active_guardians]

    # Batch-fetch term reports, subject counts, and attendance data
    term_reports = {}
    subject_counts = {}
    session_counts = {}
    present_counts = {}

    if current_term and active_student_ids:
        from django.db.models import Count

        for tr in TermReport.objects.filter(student_id__in=active_student_ids, term=current_term):
            term_reports[tr.student_id] = tr

        for row in SubjectTermGrade.objects.filter(
            student_id__in=active_student_ids, term=current_term, total_score__isnull=False
        ).values('student_id').annotate(count=Count('id')):
            subject_counts[row['student_id']] = row['count']

        # Batch attendance: session counts by class, present counts by student
        class_ids = list({s.current_class_id for _, s in active_guardians if s.current_class_id})
        if class_ids:
            session_counts = dict(
                AttendanceSession.objects.filter(
                    class_assigned_id__in=class_ids,
                    date__gte=current_term.start_date,
                    date__lte=current_term.end_date
                ).values('class_assigned_id').annotate(count=Count('id')).values_list('class_assigned_id', 'count')
            )
            present_counts = dict(
                AttendanceRecord.objects.filter(
                    student_id__in=active_student_ids,
                    status__in=['P', 'L'],
                    session__date__gte=current_term.start_date,
                    session__date__lte=current_term.end_date
                ).values('student_id').annotate(count=Count('id')).values_list('student_id', 'count')
            )

    # Build ward data from batch results
    wards_data = []
    total_average = 0
    avg_count = 0

    for sg, student in active_guardians:
        ward_data = {
            'student': student,
            'relationship': sg.get_relationship_display(),
            'is_primary': sg.is_primary,
            'term_report': term_reports.get(student.id),
            'subject_count': subject_counts.get(student.id, 0),
            'attendance_rate': None,
        }

        if current_term:
            if ward_data['term_report'] and ward_data['term_report'].average:
                total_average += ward_data['term_report'].average
                avg_count += 1

            if student.current_class_id:
                total_sessions = session_counts.get(student.current_class_id, 0)
                if total_sessions > 0:
                    present_count = present_counts.get(student.id, 0)
                    ward_data['attendance_rate'] = round((present_count / total_sessions) * 100)

        wards_data.append(ward_data)

    # Calculate overall stats
    stats = {
        'total_wards': len(wards_data),
        'average_score': round(total_average / avg_count, 1) if avg_count > 0 else None,
    }

    context = {
        'guardian': guardian,
        'wards': wards_data,
        'current_term': current_term,
        'current_year': current_year,
        'stats': stats,
    }
    return htmx_render(request, 'core/parent/dashboard.html', 'core/parent/partials/dashboard_content.html', context)


@login_required
def my_wards(request):
    """Parent view of their children (wards) with grades summary."""
    from gradebook.models import SubjectTermGrade, TermReport
    from students.models import StudentGuardian

    user = request.user
    current_term = Term.get_current()

    # Get guardian profile linked to this user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        context = {'wards': [], 'current_term': current_term}
        return htmx_render(request, 'core/parent/my_wards.html', 'core/parent/partials/my_wards_content.html', context)

    # Get wards (students) linked to this guardian
    student_guardians = StudentGuardian.objects.filter(
        guardian=guardian
    ).select_related(
        'student__current_class'
    ).order_by('-is_primary', 'student__first_name')

    # Collect active wards for batch queries
    active_guardians = [(sg, sg.student) for sg in student_guardians if sg.student.status == 'active']
    active_student_ids = [s.id for _, s in active_guardians]

    # Batch-fetch term reports and subject counts
    term_reports = {}
    subject_counts = {}
    if current_term and active_student_ids:
        from django.db.models import Count

        for tr in TermReport.objects.filter(student_id__in=active_student_ids, term=current_term):
            term_reports[tr.student_id] = tr

        for row in SubjectTermGrade.objects.filter(
            student_id__in=active_student_ids, term=current_term, total_score__isnull=False
        ).values('student_id').annotate(count=Count('id')):
            subject_counts[row['student_id']] = row['count']

    # Build ward data from batch results
    wards_data = []
    for sg, student in active_guardians:
        wards_data.append({
            'student': student,
            'relationship': sg.get_relationship_display(),
            'is_primary': sg.is_primary,
            'term_report': term_reports.get(student.id),
            'subject_count': subject_counts.get(student.id, 0),
        })

    context = {
        'guardian': guardian,
        'wards': wards_data,
        'current_term': current_term,
    }
    return htmx_render(request, 'core/parent/my_wards.html', 'core/parent/partials/my_wards_content.html', context)


@login_required
def ward_detail(request, pk):
    """Detailed view of a specific ward for guardians."""
    from gradebook.models import SubjectTermGrade, TermReport
    from students.models import StudentGuardian
    from academics.models import AttendanceSession, AttendanceRecord

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.error(request, "No guardian profile linked to your account.")
        return redirect('core:index')

    # Verify this student is a ward of the logged-in guardian
    student_guardian = StudentGuardian.objects.filter(
        guardian=guardian,
        student_id=pk
    ).select_related('student__current_class').first()

    if not student_guardian:
        messages.error(request, "You are not authorized to view this student.")
        return redirect('core:my_wards')

    student = student_guardian.student
    current_term = Term.get_current()
    current_year = AcademicYear.get_current()

    # Get term report and grades
    term_report = None
    subject_grades = []
    if current_term:
        term_report = TermReport.objects.filter(
            student=student,
            term=current_term
        ).first()
        subject_grades = SubjectTermGrade.objects.filter(
            student=student,
            term=current_term
        ).select_related('subject').order_by('subject__name')

    # Get attendance records for current term
    attendance_records = []
    attendance_stats = {'present': 0, 'absent': 0, 'late': 0, 'excused': 0, 'total': 0}
    if current_term and student.current_class:
        from django.db.models import Count, Q

        sessions = AttendanceSession.objects.filter(
            class_assigned=student.current_class,
            date__gte=current_term.start_date,
            date__lte=current_term.end_date
        ).order_by('-date')[:10]

        # Batch-fetch attendance records for all sessions at once
        session_ids = [s.id for s in sessions]
        records_by_session = {}
        for record in AttendanceRecord.objects.filter(student=student, session_id__in=session_ids):
            records_by_session[record.session_id] = record

        for session in sessions:
            record = records_by_session.get(session.id)
            attendance_records.append({
                'date': session.date,
                'status': record.status if record else None,
                'status_display': record.get_status_display() if record else 'Not Recorded'
            })

        # Calculate attendance stats with a single aggregate query
        attendance_stats = AttendanceRecord.objects.filter(
            student=student,
            session__class_assigned=student.current_class,
            session__date__gte=current_term.start_date,
            session__date__lte=current_term.end_date
        ).aggregate(
            present=Count('id', filter=Q(status='P')),
            late=Count('id', filter=Q(status='L')),
            absent=Count('id', filter=Q(status='A')),
            excused=Count('id', filter=Q(status='E')),
            total=Count('id'),
        )

    context = {
        'student': student,
        'relationship': student_guardian.get_relationship_display(),
        'is_primary': student_guardian.is_primary,
        'term_report': term_report,
        'subject_grades': subject_grades,
        'attendance_records': attendance_records,
        'attendance_stats': attendance_stats,
        'current_term': current_term,
        'current_year': current_year,
    }
    return htmx_render(request, 'core/parent/ward_detail.html', 'core/parent/partials/ward_detail_content.html', context)


@login_required
def fee_payments(request):
    """Guardian view for viewing fee payments for their wards."""
    from students.models import StudentGuardian
    from finance.models import Invoice, Payment, PaymentGatewayConfig
    from django.db.models import Sum

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)
    current_year = AcademicYear.get_current()
    current_term = Term.get_current()

    # Check if online payments are available
    online_payments_enabled = PaymentGatewayConfig.objects.filter(
        is_active=True,
        is_primary=True,
        verification_status='VERIFIED'
    ).exists()

    # Get selected ward filter from query params
    selected_ward_id = request.GET.get('ward')

    wards_fees = []
    all_invoices = []
    all_payments = []
    total_outstanding = 0

    if guardian:
        student_guardians = StudentGuardian.objects.filter(
            guardian=guardian
        ).select_related('student', 'student__current_class')

        ward_students = []
        ward_sgs = []
        for sg in student_guardians:
            if sg.student.status == 'active':
                ward_students.append(sg.student)
                ward_sgs.append(sg)

        ward_ids = [s.id for s in ward_students]

        # Batch-fetch invoice aggregates for all wards
        year_filter = {'academic_year': current_year} if current_year else {}
        invoice_aggregates = {}
        if ward_ids:
            for row in Invoice.objects.filter(
                student_id__in=ward_ids, **year_filter
            ).exclude(status='CANCELLED').values('student_id').annotate(
                total_fees=Sum('total_amount'),
                total_paid=Sum('amount_paid'),
                total_balance=Sum('balance')
            ):
                invoice_aggregates[row['student_id']] = row

        # Batch-fetch recent payments grouped by student
        payments_by_student = {}
        if ward_ids:
            for payment in Payment.objects.filter(
                invoice__student_id__in=ward_ids,
                status='COMPLETED'
            ).select_related('invoice').order_by('-transaction_date'):
                sid = payment.invoice.student_id
                if sid not in payments_by_student:
                    payments_by_student[sid] = []
                if len(payments_by_student[sid]) < 5:
                    payments_by_student[sid].append(payment)

        for sg, student in zip(ward_sgs, ward_students):
            agg = invoice_aggregates.get(student.id, {})
            total_fees = agg.get('total_fees') or 0
            total_paid = agg.get('total_paid') or 0
            balance = agg.get('total_balance') or 0

            # Per-student queries for sliced results (can't batch slices)
            invoices = Invoice.objects.filter(
                student=student, **year_filter
            ).exclude(status='CANCELLED').order_by('-created_at')

            current_invoice = invoices.filter(term=current_term).first() if current_term else None

            wards_fees.append({
                'student': student,
                'relationship': sg.get_relationship_display(),
                'total_fees': total_fees,
                'total_paid': total_paid,
                'balance': balance,
                'current_invoice': current_invoice,
                'invoices': invoices[:3],
                'recent_payments': payments_by_student.get(student.id, []),
            })

            total_outstanding += balance

        # For detailed view - get all invoices and payments (optionally filtered by ward)
        if selected_ward_id:
            try:
                selected_student = next(
                    (s for s in ward_students if str(s.pk) == selected_ward_id),
                    None
                )
                if selected_student:
                    all_invoices = Invoice.objects.filter(
                        student=selected_student
                    ).exclude(status='CANCELLED').select_related(
                        'student', 'term', 'academic_year'
                    ).prefetch_related('items').order_by('-created_at')

                    all_payments = Payment.objects.filter(
                        invoice__student=selected_student,
                        status='COMPLETED'
                    ).select_related('invoice').order_by('-transaction_date')
            except (ValueError, StopIteration):
                pass
        else:
            # All wards' invoices and payments
            all_invoices = Invoice.objects.filter(
                student__in=ward_students
            ).exclude(status='CANCELLED').select_related(
                'student', 'term', 'academic_year'
            ).prefetch_related('items').order_by('-created_at')[:20]

            all_payments = Payment.objects.filter(
                invoice__student__in=ward_students,
                status='COMPLETED'
            ).select_related('invoice', 'invoice__student').order_by('-transaction_date')[:10]

    context = {
        'guardian': guardian,
        'wards_fees': wards_fees,
        'all_invoices': all_invoices,
        'all_payments': all_payments,
        'total_outstanding': total_outstanding,
        'selected_ward_id': selected_ward_id,
        'current_year': current_year,
        'current_term': current_term,
        'online_payments_enabled': online_payments_enabled,
    }
    return htmx_render(request, 'core/parent/fee_payments.html', 'core/parent/partials/fee_payments_content.html', context)


@login_required
def guardian_pay_invoice(request, invoice_id):
    """
    Guardian view to initiate online payment for an invoice.
    Verifies the guardian has access to this invoice's student.
    """
    from students.models import StudentGuardian
    from finance.models import Invoice, Payment, PaymentGatewayConfig, PaymentGatewayTransaction
    from finance.gateways import get_gateway_adapter
    import uuid as uuid_module

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.error(request, 'No guardian profile found.')
        return redirect('core:fee_payments')

    # Get the invoice
    invoice = get_object_or_404(
        Invoice.objects.select_related('student'),
        pk=invoice_id
    )

    # Verify guardian has access to this student
    has_access = StudentGuardian.objects.filter(
        guardian=guardian,
        student=invoice.student
    ).exists()

    if not has_access:
        messages.error(request, 'You do not have access to this invoice.')
        return redirect('core:fee_payments')

    # Check invoice can be paid
    if invoice.status in ['PAID', 'CANCELLED']:
        messages.error(request, 'This invoice cannot be paid.')
        return redirect('core:fee_payments')

    if invoice.balance <= 0:
        messages.error(request, 'This invoice has no outstanding balance.')
        return redirect('core:fee_payments')

    # Get primary gateway config
    gateway_config = PaymentGatewayConfig.objects.filter(
        is_active=True,
        is_primary=True
    ).select_related('gateway').first()

    if not gateway_config:
        messages.error(request, 'Online payments are not available. Please contact the school.')
        return redirect('core:fee_payments')

    if gateway_config.verification_status != 'VERIFIED':
        messages.error(request, 'Payment gateway is not configured. Please contact the school.')
        return redirect('core:fee_payments')

    # Get gateway adapter
    adapter = get_gateway_adapter(gateway_config)

    # Generate unique reference
    reference = f"GP-{invoice.invoice_number}-{uuid_module.uuid4().hex[:8].upper()}"

    # Get payer email from guardian
    payer_email = guardian.email or user.email or 'noreply@school.com'
    payer_name = guardian.full_name
    payer_phone = guardian.phone or ''

    # Build callback URL - guardian specific
    callback_url = request.build_absolute_uri(
        reverse('core:guardian_payment_callback')
    ) + f'?reference={reference}'

    # Metadata for tracking
    metadata = {
        'invoice_id': str(invoice.pk),
        'invoice_number': invoice.invoice_number,
        'student_id': str(invoice.student.pk),
        'student_name': invoice.student.full_name,
        'guardian_id': str(guardian.pk),
        'guardian_name': guardian.full_name,
        'source': 'guardian_portal',
    }

    # Initialize payment with gateway
    response = adapter.initialize_payment(
        amount=invoice.balance,
        email=payer_email,
        reference=reference,
        callback_url=callback_url,
        metadata=metadata
    )

    if response.success:
        # Create pending payment record
        payment = Payment.objects.create(
            invoice=invoice,
            amount=invoice.balance,
            method='ONLINE',
            status='PENDING',
            reference=reference,
            payer_email=payer_email,
            payer_name=payer_name,
            payer_phone=payer_phone,
        )

        # Create gateway transaction record
        PaymentGatewayTransaction.objects.create(
            payment=payment,
            gateway_config=gateway_config,
            gateway_reference=response.gateway_reference or '',
            amount_charged=invoice.balance,
            net_amount=invoice.balance,
            full_response=response.raw_response,
        )

        # Redirect to payment gateway
        return redirect(response.authorization_url)
    else:
        messages.error(request, f'Could not initiate payment: {response.message}')
        return redirect('core:fee_payments')


def send_payment_receipt_email(payment, guardian):
    """
    Send payment receipt email to guardian.
    Returns True if email was sent successfully, False otherwise.
    """
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    from datetime import datetime

    # Get recipient email
    recipient_email = payment.payer_email or guardian.email
    if not recipient_email:
        logger.warning(f"No email address for payment receipt {payment.receipt_number}")
        return False

    # Get school from tenant
    school = getattr(connection, 'tenant', None)

    # Build context
    context = {
        'payment': payment,
        'invoice': payment.invoice,
        'student': payment.invoice.student,
        'guardian': guardian,
        'school': school,
        'current_year': datetime.now().year,
    }

    # Render email
    subject = f"Payment Receipt - {payment.receipt_number}"
    html_message = render_to_string('core/emails/payment_receipt_email.html', context)
    plain_message = strip_tags(html_message)

    try:
        send_mail(
            subject,
            plain_message,
            get_from_email(),
            [recipient_email],
            html_message=html_message,
            fail_silently=False,
        )
        logger.info(f"Payment receipt email sent for {payment.receipt_number} to {recipient_email}")
        return True
    except Exception as e:
        logger.error(f"Failed to send payment receipt email for {payment.receipt_number}: {str(e)}")
        return False


@login_required
def guardian_payment_callback(request):
    """
    Handle return from payment gateway for guardian payments.
    Verifies the payment and redirects to success page or fee payments.
    """
    from finance.models import Payment, PaymentGatewayTransaction
    from finance.gateways import get_gateway_adapter

    reference = request.GET.get('reference', '')

    if not reference:
        messages.error(request, 'Invalid payment callback.')
        return redirect('core:fee_payments')

    # Find the payment
    try:
        payment = Payment.objects.select_related(
            'invoice__student'
        ).get(reference=reference)
    except Payment.DoesNotExist:
        messages.error(request, 'Payment not found.')
        return redirect('core:fee_payments')

    # If already processed, redirect to appropriate page
    if payment.status == 'COMPLETED':
        return redirect('core:guardian_payment_success', payment_id=payment.pk)
    elif payment.status in ['FAILED', 'CANCELLED']:
        return redirect('core:guardian_payment_failed', payment_id=payment.pk)

    # Get gateway transaction
    try:
        gateway_tx = payment.gateway_transaction
        gateway_config = gateway_tx.gateway_config
    except PaymentGatewayTransaction.DoesNotExist:
        messages.error(request, 'Payment configuration error.')
        return redirect('core:fee_payments')

    # Verify with gateway
    adapter = get_gateway_adapter(gateway_config)
    response = adapter.verify_payment(reference)

    if response.success:
        # Update payment
        payment.status = 'COMPLETED'
        payment.transaction_date = timezone.now()
        payment.save()

        # Update gateway transaction
        gateway_tx.gateway_transaction_id = response.transaction_id or ''
        gateway_tx.gateway_fee = response.gateway_fee or 0
        gateway_tx.net_amount = response.amount - (response.gateway_fee or 0)
        gateway_tx.full_response = response.raw_response
        gateway_tx.save()

        # Invoice totals are updated automatically via Payment.save()

        # Send receipt email to guardian
        guardian = getattr(request.user, 'guardian_profile', None)
        if guardian:
            # Refresh payment with updated invoice data
            payment.refresh_from_db()
            payment.invoice.refresh_from_db()
            send_payment_receipt_email(payment, guardian)

        # Redirect to success page
        return redirect('core:guardian_payment_success', payment_id=payment.pk)
    else:
        payment.status = 'FAILED'
        payment.save()
        # Store error message in session for display on failed page
        request.session['payment_error_message'] = response.message
        return redirect('core:guardian_payment_failed', payment_id=payment.pk)


@login_required
def guardian_payment_success(request, payment_id):
    """
    Display payment success confirmation page for guardians.
    Shows receipt details and allows printing.
    """
    from students.models import StudentGuardian
    from finance.models import Payment

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.error(request, 'No guardian profile found.')
        return redirect('core:fee_payments')

    # Get the payment with related data
    payment = get_object_or_404(
        Payment.objects.select_related(
            'invoice__student',
            'invoice__term',
            'invoice__academic_year'
        ).prefetch_related('invoice__items'),
        pk=payment_id,
        status='COMPLETED'
    )

    # Verify guardian has access to this student
    has_access = StudentGuardian.objects.filter(
        guardian=guardian,
        student=payment.invoice.student
    ).exists()

    if not has_access:
        messages.error(request, 'You do not have access to this payment.')
        return redirect('core:fee_payments')

    # Get school from tenant for branding
    school = getattr(connection, 'tenant', None)

    context = {
        'payment': payment,
        'invoice': payment.invoice,
        'student': payment.invoice.student,
        'school': school,
        'guardian': guardian,
    }

    # For HTMX requests, return partial content
    if request.htmx:
        return render(request, 'core/parent/partials/payment_success_content.html', context)
    return render(request, 'core/parent/payment_success.html', context)


@login_required
def guardian_payment_failed(request, payment_id):
    """
    Display payment failed page for guardians.
    Shows error details and allows retry.
    """
    from students.models import StudentGuardian
    from finance.models import Payment

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.error(request, 'No guardian profile found.')
        return redirect('core:fee_payments')

    # Get the payment with related data
    payment = get_object_or_404(
        Payment.objects.select_related(
            'invoice__student',
            'invoice__term',
            'invoice__academic_year'
        ),
        pk=payment_id
    )

    # Verify guardian has access to this student
    has_access = StudentGuardian.objects.filter(
        guardian=guardian,
        student=payment.invoice.student
    ).exists()

    if not has_access:
        messages.error(request, 'You do not have access to this payment.')
        return redirect('core:fee_payments')

    # Get error message from session if available
    error_message = request.session.pop('payment_error_message', None)

    context = {
        'payment': payment,
        'invoice': payment.invoice,
        'student': payment.invoice.student,
        'guardian': guardian,
        'error_message': error_message,
    }

    # For HTMX requests, return partial content
    if request.htmx:
        return render(request, 'core/parent/partials/payment_failed_content.html', context)
    return render(request, 'core/parent/payment_failed.html', context)


def verify_document(request, code):
    """
    Public view to verify document authenticity.
    No login required - anyone with the code can verify.
    """
    from .models import DocumentVerification

    verification = None
    try:
        verification = DocumentVerification.objects.get(verification_code=code.upper())
        verification.record_verification()
    except DocumentVerification.DoesNotExist:
        pass

    context = {
        'code': code,
        'verification': verification,
        'is_valid': verification is not None,
    }
    return render(request, 'core/verify_document.html', context)