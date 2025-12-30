from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import models
from django.contrib import messages

from .models import Programme, Class, Subject, ClassSubject, AttendanceSession, AttendanceRecord, Period, TimetableEntry
from .forms import ProgrammeForm, ClassForm, SubjectForm, StudentEnrollmentForm, ClassSubjectForm
from students.models import Student


def is_school_admin(user):
    """Check if user is a school admin or superuser."""
    return user.is_superuser or getattr(user, 'is_school_admin', False)


def admin_required(view_func):
    """Decorator to require school admin or superuser access."""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        if not is_school_admin(request.user):
            messages.error(request, "You don't have permission to access this page.")
            return redirect('core:index')
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def htmx_render(request, full_template, partial_template, context=None):
    """Render full template for regular requests, partial for HTMX requests."""
    context = context or {}
    template = partial_template if request.htmx else full_template
    return render(request, template, context)




def get_academics_context():
    """Get common context for academics page."""
    from django.db.models import Count, Q

    # Get classes with student counts, grouped by level
    classes = Class.objects.select_related('programme', 'class_teacher').annotate(
        student_count=Count('students', filter=models.Q(students__status='active'))
    ).order_by('level_number', 'section')

    # Group classes by level type with stats
    classes_by_level = {
        'kg': {'classes': [], 'student_count': 0},
        'primary': {'classes': [], 'student_count': 0},
        'jhs': {'classes': [], 'student_count': 0},
        'shs': {'classes': [], 'student_count': 0},
    }
    for cls in classes:
        if cls.level_type in classes_by_level:
            classes_by_level[cls.level_type]['classes'].append(cls)
            classes_by_level[cls.level_type]['student_count'] += cls.student_count

    # Programmes with stats
    programmes = Programme.objects.annotate(
        class_count=Count('classes', filter=Q(classes__is_active=True)),
        student_count=Count(
            'classes__students',
            filter=Q(classes__is_active=True, classes__students__status='active')
        )
    ).order_by('name')

    # Subjects with stats
    subjects = Subject.objects.prefetch_related('programmes').annotate(
        class_count=Count(
            'class_allocations',
            filter=Q(class_allocations__class_assigned__is_active=True)
        ),
        student_count=Count(
            'class_allocations__class_assigned__students',
            filter=Q(
                class_allocations__class_assigned__is_active=True,
                class_allocations__class_assigned__students__status='active'
            )
        )
    ).order_by('-is_core', 'name')

    total_students = Student.objects.filter(status='active').count()

    return {
        'programmes': programmes,
        'classes': classes,
        'classes_by_level': classes_by_level,
        'subjects': subjects,
        'stats': {
            'total_classes': classes.count(),
            'total_subjects': subjects.count(),
            'total_programmes': programmes.count(),
            'total_students': total_students,
        },
        'programme_form': ProgrammeForm(),
        'class_form': ClassForm(),
        'subject_form': SubjectForm(),
    }


@admin_required
def index(request):
    """Academics dashboard page - Admin only."""
    from django.db.models import Count, Q, Avg
    from core.models import Term

    current_term = Term.get_current()

    # Get classes with student counts grouped by level
    classes = Class.objects.select_related('programme', 'class_teacher').annotate(
        student_count=Count('students', filter=models.Q(students__status='active'))
    ).filter(is_active=True).order_by('level_number', 'section')

    # Group classes by level type with stats
    classes_by_level = {
        'kg': {'label': 'KG', 'classes': [], 'student_count': 0, 'icon': 'fa-baby', 'color': 'info'},
        'primary': {'label': 'Primary', 'classes': [], 'student_count': 0, 'icon': 'fa-child', 'color': 'success'},
        'jhs': {'label': 'JHS', 'classes': [], 'student_count': 0, 'icon': 'fa-user-graduate', 'color': 'warning'},
        'shs': {'label': 'SHS', 'classes': [], 'student_count': 0, 'icon': 'fa-graduation-cap', 'color': 'primary'},
    }
    for cls in classes:
        if cls.level_type in classes_by_level:
            classes_by_level[cls.level_type]['classes'].append(cls)
            classes_by_level[cls.level_type]['student_count'] += cls.student_count

    # Programmes with stats
    programmes = Programme.objects.annotate(
        class_count=Count('classes', filter=Q(classes__is_active=True)),
        student_count=Count(
            'classes__students',
            filter=Q(classes__is_active=True, classes__students__status='active')
        )
    ).order_by('name')

    # Subjects with stats
    subjects = Subject.objects.prefetch_related('programmes').annotate(
        class_count=Count(
            'class_allocations',
            filter=Q(class_allocations__class_assigned__is_active=True)
        )
    ).order_by('-is_core', 'name')

    total_students = Student.objects.filter(status='active').count()
    total_capacity = sum(cls.capacity for cls in classes)

    # Recent attendance stats (last 7 days)
    from datetime import timedelta
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    recent_attendance = AttendanceRecord.objects.filter(
        session__date__gte=week_ago
    )
    total_records = recent_attendance.count()
    present_count = recent_attendance.filter(status__in=['P', 'L']).count()
    attendance_rate = round((present_count / total_records) * 100, 1) if total_records > 0 else 0

    context = {
        'current_term': current_term,
        'programmes': programmes,
        'subjects': subjects,
        'classes_by_level': classes_by_level,
        'stats': {
            'total_classes': classes.count(),
            'total_subjects': subjects.count(),
            'total_programmes': programmes.count(),
            'total_students': total_students,
            'total_capacity': total_capacity,
            'attendance_rate': attendance_rate,
        },
        'programme_form': ProgrammeForm(),
        'subject_form': SubjectForm(),
    }

    return htmx_render(
        request,
        'academics/index.html',
        'academics/partials/index_content.html',
        context
    )


@admin_required
def classes_list(request):
    """Classes list page with search and filters."""
    from django.db.models import Count, Q

    # Get filter parameters
    search = request.GET.get('search', '').strip()
    level_filter = request.GET.get('level', '')

    # Get classes with student counts
    classes = Class.objects.select_related('programme', 'class_teacher').annotate(
        student_count=Count('students', filter=models.Q(students__status='active'))
    ).order_by('level_number', 'section')

    # Apply search filter
    if search:
        classes = classes.filter(
            Q(name__icontains=search) |
            Q(class_teacher__first_name__icontains=search) |
            Q(class_teacher__last_name__icontains=search) |
            Q(programme__name__icontains=search)
        )

    # Apply level filter
    if level_filter:
        classes = classes.filter(level_type=level_filter)

    context = {
        'classes': classes,
        'search': search,
        'level_filter': level_filter,
        'stats': {
            'total_classes': Class.objects.count(),
            'total_students': Student.objects.filter(status='active').count(),
        },
        'class_form': ClassForm(),
    }

    return htmx_render(
        request,
        'academics/classes.html',
        'academics/partials/classes_content.html',
        context
    )


# ============ PROGRAMME VIEWS ============

def get_programmes_list_context():
    """Get context for programmes list with stats."""
    from django.db.models import Count, Q

    programmes = Programme.objects.annotate(
        class_count=Count('classes', filter=Q(classes__is_active=True)),
        student_count=Count(
            'classes__students',
            filter=Q(classes__is_active=True, classes__students__status='active')
        )
    ).order_by('name')

    return {'programmes': programmes}


@admin_required
def programme_create(request):
    """Create a new programme."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    form = ProgrammeForm(request.POST)
    if form.is_valid():
        form.save()
        if request.htmx:
            response = render(request, 'academics/partials/programmes_list.html', get_programmes_list_context())
            response['HX-Trigger'] = 'closeModal'
            response['HX-Reswap'] = 'outerHTML'
            response['HX-Retarget'] = '#programmes-container'
            return response
        return redirect('academics:index')

    # Validation error - show form with errors in modal
    if request.htmx:
        response = render(request, 'academics/partials/modal_programme_form.html', {
            'form': form,
            'is_create': True,
        })
        response.status_code = 422
        return response
    return redirect('academics:index')


@admin_required
def programme_edit(request, pk):
    """Edit a programme."""
    programme = get_object_or_404(Programme, pk=pk)

    if request.method == 'GET':
        form = ProgrammeForm(instance=programme)
        return render(request, 'academics/partials/modal_programme_form.html', {
            'form': form,
            'programme': programme,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    form = ProgrammeForm(request.POST, instance=programme)
    if form.is_valid():
        form.save()
        if request.htmx:
            response = render(request, 'academics/partials/programmes_list.html', get_programmes_list_context())
            response['HX-Trigger'] = 'closeModal'
            response['HX-Reswap'] = 'outerHTML'
            response['HX-Retarget'] = '#programmes-container'
            return response
        return redirect('academics:index')

    # Validation error - show form with errors in modal
    if request.htmx:
        response = render(request, 'academics/partials/modal_programme_form.html', {
            'form': form,
            'programme': programme,
        })
        response.status_code = 422
        return response
    return redirect('academics:index')


@login_required
def programme_delete(request, pk):
    """Delete a programme."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    programme = get_object_or_404(Programme, pk=pk)
    programme.delete()

    if request.htmx:
        return render(request, 'academics/partials/programmes_list.html', get_programmes_list_context())
    return redirect('academics:index')


# ============ CLASS VIEWS ============

def get_classes_list_context():
    """Get context for classes list."""
    from django.db.models import Count
    classes = Class.objects.select_related('programme', 'class_teacher').annotate(
        student_count=Count('students', filter=models.Q(students__status='active'))
    ).order_by('level_number', 'section')
    return {'classes': classes}


@login_required
def class_create(request):
    """Create a new class."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    form = ClassForm(request.POST)
    if form.is_valid():
        form.save()
        if request.htmx:
            # Refresh the classes page
            response = HttpResponse(status=204)
            response['HX-Redirect'] = '/academics/classes/'
            return response
        return redirect('academics:classes')

    # Validation error (422 keeps modal open)
    if request.htmx:
        response = render(request, 'academics/partials/modal_class_form.html', {
            'form': form,
            'is_create': True,
        })
        response.status_code = 422
        return response
    return redirect('academics:classes')


@login_required
def class_edit(request, pk):
    """Edit a class."""
    cls = get_object_or_404(Class, pk=pk)

    # Check if editing from class detail page (not from classes list)
    current_url = request.headers.get('HX-Current-URL', '')
    is_detail_page = f'/academics/classes/{pk}/' in current_url and not current_url.endswith('/academics/classes/')

    if request.method == 'GET':
        form = ClassForm(instance=cls)
        return render(request, 'academics/partials/modal_class_form.html', {
            'form': form,
            'class': cls,
            'is_detail_page': is_detail_page,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    form = ClassForm(request.POST, instance=cls)
    if form.is_valid():
        form.save()
        if request.htmx:
            # Always refresh the page to show updated data
            response = HttpResponse(status=204)
            response['HX-Refresh'] = 'true'
            return response
        return redirect('academics:classes')

    # Validation error (422 keeps modal open)
    if request.htmx:
        response = render(request, 'academics/partials/modal_class_form.html', {
            'form': form,
            'class': cls,
            'is_detail_page': is_detail_page,
        })
        response.status_code = 422
        return response
    return redirect('academics:classes')


@login_required
def class_delete(request, pk):
    """Delete a class."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    cls = get_object_or_404(Class, pk=pk)
    cls.delete()

    if request.htmx:
        response = HttpResponse(status=204)
        response['HX-Refresh'] = 'true'
        return response
    return redirect('academics:classes')


# ============ SUBJECT VIEWS ============

def get_subjects_list_context():
    """Get context for subjects list with stats."""
    from django.db.models import Count, Q

    subjects = Subject.objects.prefetch_related('programmes').annotate(
        class_count=Count(
            'class_allocations',
            filter=Q(class_allocations__class_assigned__is_active=True)
        ),
        student_count=Count(
            'class_allocations__class_assigned__students',
            filter=Q(
                class_allocations__class_assigned__is_active=True,
                class_allocations__class_assigned__students__status='active'
            )
        )
    ).order_by('-is_core', 'name')

    return {'subjects': subjects}


@login_required
def subject_create(request):
    """Create a new subject."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    form = SubjectForm(request.POST)
    if form.is_valid():
        form.save()
        if request.htmx:
            response = render(request, 'academics/partials/subjects_list.html', get_subjects_list_context())
            response['HX-Trigger'] = 'closeModal'
            response['HX-Reswap'] = 'outerHTML'
            response['HX-Retarget'] = '#subjects-container'
            return response
        return redirect('academics:index')

    # Validation error - show form with errors in modal
    if request.htmx:
        response = render(request, 'academics/partials/modal_subject_form.html', {
            'form': form,
            'is_create': True,
        })
        response.status_code = 422
        return response
    return redirect('academics:index')


@login_required
def subject_edit(request, pk):
    """Edit a subject."""
    subject = get_object_or_404(Subject, pk=pk)

    if request.method == 'GET':
        form = SubjectForm(instance=subject)
        return render(request, 'academics/partials/modal_subject_form.html', {
            'form': form,
            'subject': subject,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    form = SubjectForm(request.POST, instance=subject)
    if form.is_valid():
        form.save()
        if request.htmx:
            response = render(request, 'academics/partials/subjects_list.html', get_subjects_list_context())
            response['HX-Trigger'] = 'closeModal'
            response['HX-Reswap'] = 'outerHTML'
            response['HX-Retarget'] = '#subjects-container'
            return response
        return redirect('academics:index')

    # Validation error - show form with errors in modal
    if request.htmx:
        response = render(request, 'academics/partials/modal_subject_form.html', {
            'form': form,
            'subject': subject,
        })
        response.status_code = 422
        return response
    return redirect('academics:index')


@login_required
def subject_delete(request, pk):
    """Delete a subject."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    subject = get_object_or_404(Subject, pk=pk)
    subject.delete()

    if request.htmx:
        return render(request, 'academics/partials/subjects_list.html', get_subjects_list_context())
    return redirect('academics:index')



def get_register_tab_context(class_obj):
    """Context for the Students/Register tab."""
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('first_name')

    # Gender breakdown
    male_count = students.filter(gender='M').count()
    female_count = students.filter(gender='F').count()

    return {
        'class': class_obj,
        'students': students,
        'gender_stats': {
            'male': male_count,
            'female': female_count,
        }
    }

def get_teachers_tab_context(class_obj):
    """Context for the Teachers/Subjects tab."""
    subject_allocations = ClassSubject.objects.filter(
        class_assigned=class_obj
    ).select_related('subject', 'teacher').order_by('subject__name')
    return {
        'class': class_obj,
        'subject_allocations': subject_allocations
    }

def get_attendance_tab_context(class_obj):
    """Context for the Attendance tab with real data."""
    from django.db.models import Count, Q

    sessions = AttendanceSession.objects.filter(
        class_assigned=class_obj
    ).annotate(
        present_count=Count('records', filter=Q(records__status__in=['P', 'L'])),
        absent_count=Count('records', filter=Q(records__status='A')),
        total_count=Count('records')
    ).order_by('-date')

    # Calculate overall attendance percentage
    total_records = AttendanceRecord.objects.filter(
        session__class_assigned=class_obj
    ).count()
    present_count = AttendanceRecord.objects.filter(
        session__class_assigned=class_obj,
        status__in=['P', 'L']  # Present or Late counts as present
    ).count()

    if total_records > 0:
        attendance_percentage = int((present_count / total_records) * 100)
    else:
        attendance_percentage = "--"

    return {
        'class': class_obj,
        'attendance_sessions': sessions,
        'attendance_percentage': attendance_percentage,
    }


def get_promotion_history_context(class_obj):
    """Context for showing recent enrollment/promotion activity for this class."""
    from students.models import Enrollment
    from core.models import AcademicYear

    current_year = AcademicYear.get_current()

    # Get recent enrollments for this class (incoming students)
    incoming_enrollments = Enrollment.objects.filter(
        class_assigned=class_obj
    ).select_related('student', 'academic_year', 'promoted_from').order_by(
        '-created_at'
    )[:5]

    # Get promotions/graduations out of this class
    outgoing_enrollments = Enrollment.objects.filter(
        promoted_from__class_assigned=class_obj
    ).select_related('student', 'academic_year', 'class_assigned').order_by(
        '-created_at'
    )[:5]

    # Summary stats
    if current_year:
        active_in_current_year = Enrollment.objects.filter(
            class_assigned=class_obj,
            academic_year=current_year,
            status=Enrollment.Status.ACTIVE
        ).count()

        promoted_count = Enrollment.objects.filter(
            promoted_from__class_assigned=class_obj,
            promoted_from__academic_year=current_year
        ).count()

        graduated_count = Enrollment.objects.filter(
            class_assigned=class_obj,
            academic_year=current_year,
            status=Enrollment.Status.GRADUATED
        ).count()
    else:
        active_in_current_year = 0
        promoted_count = 0
        graduated_count = 0

    return {
        'incoming_enrollments': incoming_enrollments,
        'outgoing_enrollments': outgoing_enrollments,
        'promotion_stats': {
            'active': active_in_current_year,
            'promoted': promoted_count,
            'graduated': graduated_count,
        },
        'current_year': current_year,
    }

# --- Main Detail View ---

@login_required
def class_detail(request, pk):
    """Detailed view of a specific class."""
    class_obj = get_object_or_404(Class, pk=pk)

    # Base context
    context = {
        'class': class_obj,
    }

    # Combine all tab contexts for the initial page load
    context.update(get_register_tab_context(class_obj))
    context.update(get_teachers_tab_context(class_obj))
    context.update(get_attendance_tab_context(class_obj))
    context.update(get_promotion_history_context(class_obj))

    return htmx_render(
        request,
        'academics/class_detail.html',
        'academics/partials/class_detail_content.html',
        context
    )

# --- Action Views (Using Helpers) ---
@login_required
def class_subject_create(request, pk):
    class_obj = get_object_or_404(Class, pk=pk)

    if request.method == 'POST':
        form = ClassSubjectForm(request.POST, class_instance=class_obj)
        if form.is_valid():
            allocation = form.save(commit=False)
            allocation.class_assigned = class_obj
            allocation.save()

            if request.htmx:
                response = HttpResponse(status=204)
                response['HX-Refresh'] = 'true'
                return response

            return redirect('academics:class_detail', pk=pk)
    else:
        form = ClassSubjectForm(class_instance=class_obj)

    return render(request, 'academics/partials/modal_subject_allocation.html', {
        'form': form, 'class': class_obj
    })


@login_required
def class_subject_delete(request, class_pk, pk):
    allocation = get_object_or_404(ClassSubject, pk=pk, class_assigned_id=class_pk)
    allocation.delete()

    if request.htmx:
        response = HttpResponse(status=204)
        response['HX-Refresh'] = 'true'
        return response

    return redirect('academics:class_detail', pk=class_pk)


@login_required
def class_student_enroll(request, pk):
    """Enroll existing students into a class."""
    class_obj = get_object_or_404(Class, pk=pk)

    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST, class_instance=class_obj)
        if form.is_valid():
            students_to_add = form.cleaned_data['students']
            for student in students_to_add:
                student.current_class = class_obj
                student.save()

            if request.htmx:
                response = HttpResponse(status=204)
                response['HX-Refresh'] = 'true'
                return response

            return redirect('academics:class_detail', pk=pk)
    else:
        form = StudentEnrollmentForm(class_instance=class_obj)

    return render(request, 'academics/partials/modal_student_enroll.html', {
        'form': form, 'class': class_obj
    })


@login_required
def class_student_remove(request, class_pk, student_pk):
    """Remove a student from a class."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    student = get_object_or_404(Student, pk=student_pk, current_class_id=class_pk)
    student.current_class = None
    student.save()

    if request.htmx:
        response = HttpResponse(status=204)
        response['HX-Refresh'] = 'true'
        return response

    return redirect('academics:class_detail', pk=class_pk)


# --- 2. Take Attendance Action ---
@login_required
def class_attendance_take(request, pk):
    """
    Opens the attendance sheet for a specific date (defaults to today).
    Handles saving the records.
    """
    from teachers.models import Teacher
    from django.contrib import messages

    class_obj = get_object_or_404(Class, pk=pk)
    user = request.user
    is_admin = user.is_superuser or getattr(user, 'is_school_admin', False)

    # Check permission: must be admin, class teacher, or subject teacher for this class
    if not is_admin:
        if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
            messages.error(request, 'You do not have permission to take attendance.')
            return redirect('core:index')

        teacher = user.teacher_profile
        is_class_teacher = class_obj.class_teacher == teacher
        is_subject_teacher = ClassSubject.objects.filter(
            class_assigned=class_obj,
            teacher=teacher
        ).exists()

        if not is_class_teacher and not is_subject_teacher:
            messages.error(request, 'You are not assigned to this class.')
            return redirect('academics:attendance_reports')

    target_date = timezone.now().date()  # For now, default to today

    # Check if session exists
    session, created = AttendanceSession.objects.get_or_create(
        class_assigned=class_obj,
        date=target_date
    )

    if request.method == 'POST':
        # Process the form submission manually for grid data
        # Data format: "status_STUDENTID" : "STATUS_CODE"
        
        students = Student.objects.filter(current_class=class_obj, status='active')
        
        for student in students:
            status_key = f"status_{student.id}"
            new_status = request.POST.get(status_key, AttendanceRecord.Status.PRESENT)
            
            AttendanceRecord.objects.update_or_create(
                session=session,
                student=student,
                defaults={'status': new_status}
            )
            
        # HTMX Success: Refresh page to show updated data
        if request.htmx:
            response = HttpResponse(status=204)
            response['HX-Refresh'] = 'true'
            return response

        return redirect('academics:class_detail', pk=pk)

    # GET Request: Prepare data for the form
    students = Student.objects.filter(current_class=class_obj, status='active').order_by('first_name')
    records = {r.student_id: r.status for r in session.records.all()}
    
    # Combine student + their status
    student_list = []
    for student in students:
        student_list.append({
            'obj': student,
            'status': records.get(student.id, 'P') # Default to Present if new
        })

    return render(request, 'academics/partials/modal_attendance_take.html', {
        'class': class_obj,
        'session': session,
        'student_list': student_list,
        'date': target_date
    })


@login_required
def class_promote(request, pk):
    """Promote students from a specific class (modal-based)."""
    from students.models import Enrollment
    from core.models import AcademicYear

    class_obj = get_object_or_404(Class, pk=pk)
    current_year = AcademicYear.get_current()

    if not current_year:
        return render(request, 'academics/partials/modal_class_promote.html', {
            'class': class_obj,
            'error': 'No current academic year set.',
        })

    # Get next academic year
    next_year = AcademicYear.objects.filter(
        start_date__gt=current_year.end_date
    ).order_by('start_date').first()

    if not next_year:
        return render(request, 'academics/partials/modal_class_promote.html', {
            'class': class_obj,
            'error': 'No next academic year configured. Please create one first.',
        })

    # Determine if this is a final-year class
    is_final_year = (
        class_obj.level_type == Class.LevelType.SHS and class_obj.level_number == 3
    )

    # Get students with active enrollments in this class for current year
    students = Student.objects.filter(
        enrollments__class_assigned=class_obj,
        enrollments__academic_year=current_year,
        enrollments__status=Enrollment.Status.ACTIVE,
        status=Student.Status.ACTIVE
    ).select_related('current_class').order_by('last_name', 'first_name')

    if request.method == 'POST':
        from django.contrib import messages

        promoted_count = 0
        repeated_count = 0
        graduated_count = 0
        errors = []

        for key, value in request.POST.items():
            if key.startswith('action_'):
                student_id = key.replace('action_', '')
                action = value

                if action == 'skip':
                    continue

                try:
                    student = Student.objects.get(pk=student_id)
                    current_enrollment = student.enrollments.filter(
                        academic_year=current_year,
                        class_assigned=class_obj,
                        status=Enrollment.Status.ACTIVE
                    ).first()

                    if not current_enrollment:
                        continue

                    if action == 'promote':
                        target_class_id = request.POST.get(f'target_class_{student_id}')
                        if not target_class_id:
                            errors.append(f'{student.full_name}: No target class selected')
                            continue

                        try:
                            target_class = Class.objects.get(pk=target_class_id)
                        except Class.DoesNotExist:
                            errors.append(f'{student.full_name}: Invalid target class')
                            continue

                        current_enrollment.status = Enrollment.Status.PROMOTED
                        current_enrollment.save()

                        Enrollment.objects.create(
                            student=student,
                            academic_year=next_year,
                            class_assigned=target_class,
                            status=Enrollment.Status.ACTIVE,
                            promoted_from=current_enrollment,
                        )

                        student.current_class = target_class
                        student.save()
                        promoted_count += 1

                    elif action == 'repeat':
                        current_enrollment.status = Enrollment.Status.REPEATED
                        current_enrollment.save()

                        Enrollment.objects.create(
                            student=student,
                            academic_year=next_year,
                            class_assigned=class_obj,
                            status=Enrollment.Status.ACTIVE,
                            promoted_from=current_enrollment,
                            remarks='Repeated year',
                        )
                        repeated_count += 1

                    elif action == 'graduate':
                        current_enrollment.status = Enrollment.Status.GRADUATED
                        current_enrollment.save()

                        student.status = Student.Status.GRADUATED
                        student.current_class = None
                        student.save()
                        graduated_count += 1

                except Student.DoesNotExist:
                    errors.append(f'Student ID {student_id}: Not found')
                except Exception as e:
                    errors.append(f'Error: {str(e)}')

        # HTMX Response: Refresh page to show updated data
        if request.htmx:
            response = HttpResponse(status=204)
            response['HX-Refresh'] = 'true'
            return response

        return redirect('academics:class_detail', pk=pk)

    # GET: Prepare form data
    all_target_classes = Class.objects.filter(is_active=True).order_by(
        'programme__name', 'level_number', 'name'
    )

    return render(request, 'academics/partials/modal_class_promote.html', {
        'class': class_obj,
        'students': students,
        'is_final_year': is_final_year,
        'current_year': current_year,
        'next_year': next_year,
        'all_classes': all_target_classes,
    })


@login_required
def class_attendance_edit(request, pk, session_pk):
    """Edit an existing attendance session."""
    class_obj = get_object_or_404(Class, pk=pk)
    session = get_object_or_404(AttendanceSession, pk=session_pk, class_assigned=class_obj)

    if request.method == 'POST':
        students = Student.objects.filter(current_class=class_obj, status='active')

        for student in students:
            status_key = f"status_{student.id}"
            new_status = request.POST.get(status_key, AttendanceRecord.Status.PRESENT)

            AttendanceRecord.objects.update_or_create(
                session=session,
                student=student,
                defaults={'status': new_status}
            )

        if request.htmx:
            response = HttpResponse(status=204)
            response['HX-Refresh'] = 'true'
            return response

        return redirect('academics:class_detail', pk=pk)

    # GET Request: Load existing records
    students = Student.objects.filter(current_class=class_obj, status='active').order_by('first_name')
    records = {r.student_id: r.status for r in session.records.all()}

    student_list = []
    for student in students:
        student_list.append({
            'obj': student,
            'status': records.get(student.id, 'P')
        })

    return render(request, 'academics/partials/modal_attendance_take.html', {
        'class': class_obj,
        'session': session,
        'student_list': student_list,
        'date': session.date,
        'is_edit': True
    })


@login_required
def class_export(request, pk):
    """Export class register to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as DjangoHttpResponse
    from core.models import SchoolSettings

    class_obj = get_object_or_404(Class, pk=pk)
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('last_name', 'first_name')

    school = SchoolSettings.load()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_obj.name} Register"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    table_header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # School Header
    ws.merge_cells('A1:F1')
    ws['A1'] = school.display_name or request.tenant.name
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Class Info
    ws.merge_cells('A2:F2')
    ws['A2'] = f"{class_obj.name} - {class_obj.level_display}"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Date
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Generated: {timezone.now().strftime('%B %d, %Y')}"
    ws['A3'].alignment = Alignment(horizontal='center')

    # Empty row
    ws.append([])

    # Table Headers
    headers = ['#', 'Admission No.', 'Full Name', 'Gender', 'Guardian', 'Phone']
    ws.append(headers)
    header_row = 5

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Student Data
    for idx, student in enumerate(students, 1):
        row_data = [
            idx,
            student.admission_number,
            student.full_name,
            student.get_gender_display(),
            student.guardian_name or '-',
            student.guardian_phone or '-',
        ]
        ws.append(row_data)

        # Apply borders
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=header_row + idx, column=col_num)
            cell.border = border
            if col_num == 1:
                cell.alignment = Alignment(horizontal='center')

    # Set column widths
    column_widths = [5, 15, 30, 10, 25, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Summary row
    summary_row = header_row + len(students) + 2
    ws.cell(row=summary_row, column=1, value=f"Total Students: {students.count()}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    male_count = students.filter(gender='M').count()
    female_count = students.filter(gender='F').count()
    ws.cell(row=summary_row + 1, column=1, value=f"Male: {male_count} | Female: {female_count}")

    # Create response
    response = DjangoHttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{class_obj.name.replace(' ', '_')}_Register.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# ============ ATTENDANCE REPORTS ============

@login_required
def attendance_reports(request):
    """Attendance reports with filters."""
    from django.db.models import Count, Q, F
    from datetime import timedelta
    from teachers.models import Teacher

    user = request.user
    is_admin = user.is_superuser or getattr(user, 'is_school_admin', False)

    # Get filter parameters
    class_filter = request.GET.get('class', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    view_mode = request.GET.get('view', 'summary')  # summary, daily, students

    # Default date range: last 30 days
    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    # Filter classes based on user role
    if is_admin:
        classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')
    elif getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile'):
        teacher = user.teacher_profile
        # Teachers see classes they're class teacher for OR assigned to teach
        homeroom_ids = Class.objects.filter(class_teacher=teacher).values_list('id', flat=True)
        assigned_ids = ClassSubject.objects.filter(teacher=teacher).values_list('class_assigned_id', flat=True)
        all_class_ids = set(homeroom_ids) | set(assigned_ids)
        classes = Class.objects.filter(id__in=all_class_ids, is_active=True).order_by('level_number', 'name')
    else:
        classes = Class.objects.none()

    allowed_class_ids = list(classes.values_list('id', flat=True))

    # Base querysets - filter by allowed classes for teachers
    sessions = AttendanceSession.objects.select_related('class_assigned')
    records = AttendanceRecord.objects.select_related('session', 'student', 'session__class_assigned')

    if not is_admin:
        sessions = sessions.filter(class_assigned_id__in=allowed_class_ids)
        records = records.filter(session__class_assigned_id__in=allowed_class_ids)

    # Apply date filter
    if date_from:
        sessions = sessions.filter(date__gte=date_from)
        records = records.filter(session__date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__lte=date_to)
        records = records.filter(session__date__lte=date_to)

    # Apply class filter
    if class_filter:
        sessions = sessions.filter(class_assigned_id=class_filter)
        records = records.filter(session__class_assigned_id=class_filter)

    # Calculate summary stats
    total_sessions = sessions.count()
    total_records = records.count()
    present_count = records.filter(status__in=['P', 'L']).count()
    absent_count = records.filter(status='A').count()
    late_count = records.filter(status='L').count()

    attendance_rate = 0
    if total_records > 0:
        attendance_rate = round((present_count / total_records) * 100, 1)

    # Summary by class (only for allowed classes)
    class_summary = []
    for cls in classes:
        cls_records = records.filter(session__class_assigned=cls)
        cls_total = cls_records.count()
        if cls_total > 0:
            cls_present = cls_records.filter(status__in=['P', 'L']).count()
            cls_absent = cls_records.filter(status='A').count()
            cls_rate = round((cls_present / cls_total) * 100, 1)
            class_summary.append({
                'class': cls,
                'total': cls_total,
                'present': cls_present,
                'absent': cls_absent,
                'rate': cls_rate,
            })

    # Daily breakdown (for daily view)
    daily_data = []
    if view_mode == 'daily':
        daily_sessions = sessions.order_by('-date')[:30]
        for session in daily_sessions:
            session_records = session.records.all()
            s_total = session_records.count()
            s_present = session_records.filter(status__in=['P', 'L']).count()
            s_absent = session_records.filter(status='A').count()
            daily_data.append({
                'session': session,
                'total': s_total,
                'present': s_present,
                'absent': s_absent,
                'rate': round((s_present / s_total) * 100, 1) if s_total > 0 else 0,
            })

    # Students with low attendance (for students view)
    low_attendance_students = []
    student_stats_all = {}  # Track all student stats for consecutive absences

    # Build student stats from records
    for record in records:
        sid = record.student_id
        if sid not in student_stats_all:
            student_stats_all[sid] = {
                'student': record.student,
                'total': 0,
                'present': 0,
                'records': []
            }
        student_stats_all[sid]['total'] += 1
        if record.status in ['P', 'L']:
            student_stats_all[sid]['present'] += 1
        student_stats_all[sid]['records'].append({
            'date': record.session.date,
            'status': record.status
        })

    if view_mode == 'students':
        for sid, stats in student_stats_all.items():
            if stats['total'] > 0:
                rate = round((stats['present'] / stats['total']) * 100, 1)
                if rate < 80:  # Low attendance threshold
                    low_attendance_students.append({
                        'student': stats['student'],
                        'total': stats['total'],
                        'present': stats['present'],
                        'absent': stats['total'] - stats['present'],
                        'rate': rate,
                    })

        # Sort by attendance rate (lowest first)
        low_attendance_students.sort(key=lambda x: x['rate'])

    # Trends data for chart (attendance rate over time)
    trend_data = []
    if view_mode == 'trends':
        from collections import defaultdict
        from datetime import datetime

        # Group records by date
        daily_rates = defaultdict(lambda: {'present': 0, 'total': 0})
        for record in records:
            date_str = record.session.date.isoformat()
            daily_rates[date_str]['total'] += 1
            if record.status in ['P', 'L']:
                daily_rates[date_str]['present'] += 1

        # Sort by date and calculate rates
        for date_str in sorted(daily_rates.keys()):
            data = daily_rates[date_str]
            rate = round((data['present'] / data['total']) * 100, 1) if data['total'] > 0 else 0
            trend_data.append({
                'date': date_str,
                'rate': rate,
                'present': data['present'],
                'absent': data['total'] - data['present'],
                'total': data['total']
            })

    # Calculate consecutive absences for alert
    students_with_consecutive_absences = []
    CONSECUTIVE_THRESHOLD = 3

    for sid, stats in student_stats_all.items():
        # Sort records by date
        sorted_records = sorted(stats['records'], key=lambda x: x['date'], reverse=True)
        consecutive = 0
        for rec in sorted_records:
            if rec['status'] == 'A':
                consecutive += 1
            else:
                break

        if consecutive >= CONSECUTIVE_THRESHOLD:
            students_with_consecutive_absences.append({
                'student': stats['student'],
                'consecutive_days': consecutive,
                'last_present': None
            })
            # Find last present date
            for rec in sorted_records:
                if rec['status'] in ['P', 'L']:
                    students_with_consecutive_absences[-1]['last_present'] = rec['date']
                    break

    # Sort by consecutive days (highest first)
    students_with_consecutive_absences.sort(key=lambda x: x['consecutive_days'], reverse=True)

    # Calculate trend indicators (compare to previous period)
    prev_period_rate = None
    rate_change = None
    if date_from and date_to:
        try:
            from datetime import datetime
            df = datetime.fromisoformat(date_from)
            dt = datetime.fromisoformat(date_to)
            period_length = (dt - df).days
            prev_start = (df - timedelta(days=period_length)).date()
            prev_end = (df - timedelta(days=1)).date()

            # Get previous period records
            prev_records = AttendanceRecord.objects.filter(
                session__date__gte=prev_start,
                session__date__lte=prev_end
            )
            if not is_admin:
                prev_records = prev_records.filter(session__class_assigned_id__in=allowed_class_ids)
            if class_filter:
                prev_records = prev_records.filter(session__class_assigned_id=class_filter)

            prev_total = prev_records.count()
            if prev_total > 0:
                prev_present = prev_records.filter(status__in=['P', 'L']).count()
                prev_period_rate = round((prev_present / prev_total) * 100, 1)
                rate_change = round(attendance_rate - prev_period_rate, 1)
        except (ValueError, TypeError):
            pass

    context = {
        'classes': classes,
        'class_filter': class_filter,
        'date_from': date_from,
        'date_to': date_to,
        'view_mode': view_mode,
        'is_admin': is_admin,
        'stats': {
            'total_sessions': total_sessions,
            'total_records': total_records,
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
            'rate': attendance_rate,
        },
        'class_summary': class_summary,
        'daily_data': daily_data,
        'low_attendance_students': low_attendance_students,
        'trend_data': trend_data,
        'students_with_consecutive_absences': students_with_consecutive_absences,
        'prev_period_rate': prev_period_rate,
        'rate_change': rate_change,
    }

    return htmx_render(
        request,
        'academics/attendance_reports.html',
        'academics/partials/attendance_reports_content.html',
        context
    )


@login_required
def attendance_export(request):
    """Export attendance data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as DjangoHttpResponse
    from core.models import SchoolSettings
    from datetime import timedelta

    # Get filter parameters
    class_filter = request.GET.get('class', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Default date range
    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    school = SchoolSettings.load()

    # Get records
    records = AttendanceRecord.objects.select_related(
        'session', 'student', 'session__class_assigned'
    ).filter(
        session__date__gte=date_from,
        session__date__lte=date_to
    ).order_by('session__date', 'session__class_assigned__name', 'student__last_name')

    if class_filter:
        records = records.filter(session__class_assigned_id=class_filter)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    table_header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    late_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = school.display_name or request.tenant.name
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Attendance Report: {date_from} to {date_to}"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:F3')
    ws['A3'] = f"Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}"
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Table headers
    headers = ['Date', 'Class', 'Student Name', 'Admission No.', 'Status', 'Remarks']
    ws.append(headers)
    header_row = 5

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data rows
    status_map = {'P': 'Present', 'A': 'Absent', 'L': 'Late', 'E': 'Excused'}
    for idx, record in enumerate(records, 1):
        row_data = [
            record.session.date.strftime('%Y-%m-%d'),
            record.session.class_assigned.name,
            record.student.full_name,
            record.student.admission_number,
            status_map.get(record.status, record.status),
            record.remarks or '',
        ]
        ws.append(row_data)

        row_num = header_row + idx
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

        # Color-code status
        status_cell = ws.cell(row=row_num, column=5)
        if record.status == 'P':
            status_cell.fill = present_fill
        elif record.status == 'A':
            status_cell.fill = absent_fill
        elif record.status == 'L':
            status_cell.fill = late_fill

    # Column widths
    column_widths = [12, 15, 30, 15, 12, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Summary
    summary_row = header_row + records.count() + 2
    ws.cell(row=summary_row, column=1, value=f"Total Records: {records.count()}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    present = records.filter(status__in=['P', 'L']).count()
    absent = records.filter(status='A').count()
    rate = round((present / records.count()) * 100, 1) if records.count() > 0 else 0
    ws.cell(row=summary_row + 1, column=1, value=f"Present: {present} | Absent: {absent} | Rate: {rate}%")

    # Response
    response = DjangoHttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Attendance_Report_{date_from}_to_{date_to}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def student_attendance_detail(request, student_id):
    """Get detailed attendance for a single student."""
    from students.models import Student
    from django.http import JsonResponse
    from datetime import timedelta

    student = get_object_or_404(Student, pk=student_id)

    # Get filter parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Default date range: last 30 days
    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    # Get attendance records
    records = AttendanceRecord.objects.filter(
        student=student,
        session__date__gte=date_from,
        session__date__lte=date_to
    ).select_related('session').order_by('-session__date')

    total = records.count()
    present = records.filter(status__in=['P', 'L']).count()
    absent = records.filter(status='A').count()
    late = records.filter(status='L').count()
    rate = round((present / total) * 100, 1) if total > 0 else 0

    # Build attendance calendar data
    attendance_data = []
    for record in records:
        attendance_data.append({
            'date': record.session.date.isoformat(),
            'status': record.status,
            'status_display': record.get_status_display(),
            'remarks': record.remarks or ''
        })

    context = {
        'student': student,
        'stats': {
            'total': total,
            'present': present,
            'absent': absent,
            'late': late,
            'rate': rate,
        },
        'attendance_data': attendance_data,
        'date_from': date_from,
        'date_to': date_to,
    }

    # Return as partial template for HTMX
    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/student_attendance_detail.html', context)

    return JsonResponse({
        'student': {
            'id': str(student.pk),
            'name': str(student),
            'admission_number': student.admission_number,
            'class': student.current_class.name if student.current_class else None,
        },
        'stats': context['stats'],
        'records': attendance_data,
    })


@login_required
def notify_absent_parents(request):
    """Send SMS notifications to parents of students with consecutive absences."""
    from django.contrib import messages
    from students.models import Student
    from core.models import SchoolSettings

    if request.method != 'POST':
        return redirect('academics:attendance_reports')

    student_ids = request.POST.getlist('student_ids')
    message_template = request.POST.get('message', '')

    if not student_ids:
        messages.warning(request, "No students selected for notification.")
        return redirect('academics:attendance_reports')

    if not message_template:
        messages.error(request, "Message template is required.")
        return redirect('academics:attendance_reports')

    # Get school settings for school name
    school_settings = SchoolSettings.load()
    school_name = school_settings.display_name if school_settings else ''

    # Get students with their attendance stats
    students = Student.objects.filter(pk__in=student_ids, guardian_phone__isnull=False)
    sent_count = 0
    failed_count = 0

    for student in students:
        # Calculate consecutive absences for this student
        recent_records = AttendanceRecord.objects.filter(
            student=student
        ).select_related('session').order_by('-session__date')[:30]

        consecutive_days = 0
        for record in recent_records:
            if record.status == 'A':
                consecutive_days += 1
            else:
                break

        # Render the message with placeholders
        message = message_template.replace('{student_name}', student.first_name)
        message = message.replace('{full_name}', str(student))
        message = message.replace('{days}', str(consecutive_days))
        message = message.replace('{class_name}', student.current_class.name if student.current_class else '')
        message = message.replace('{school_name}', school_name)

        # Send SMS
        try:
            from communications.utils import send_sms
            result = send_sms(
                to_phone=student.guardian_phone,
                message=message,
                student=student,
                message_type='attendance',
                created_by=request.user
            )
            if result.get('success'):
                sent_count += 1
            else:
                failed_count += 1
        except Exception as e:
            failed_count += 1

    if sent_count > 0:
        messages.success(request, f"Successfully sent {sent_count} notification(s) to parents.")
    if failed_count > 0:
        messages.warning(request, f"{failed_count} notification(s) failed to send.")

    return redirect('academics:attendance_reports')


# ============ API ENDPOINTS ============

@login_required
def api_class_subjects(request, pk):
    """API endpoint to get subjects for a class.

    For admins: returns all subjects assigned to the class.
    For teachers: returns only subjects they are assigned to teach.
    """
    class_obj = get_object_or_404(Class, pk=pk)
    user = request.user

    # Build base query
    class_subjects = ClassSubject.objects.filter(
        class_assigned=class_obj
    ).select_related('subject', 'teacher')

    # Filter by teacher assignment for non-admins
    if not (user.is_superuser or getattr(user, 'is_school_admin', False)):
        # Check if user has a teacher profile
        if hasattr(user, 'teacher_profile') and user.teacher_profile:
            class_subjects = class_subjects.filter(teacher=user.teacher_profile)
        else:
            class_subjects = class_subjects.none()

    subjects = [
        {
            'id': cs.subject.pk,
            'name': cs.subject.name,
            'is_assigned': cs.teacher_id == getattr(getattr(user, 'teacher_profile', None), 'id', None)
        }
        for cs in class_subjects
    ]

    return JsonResponse({'subjects': subjects})


# ============ PERIOD MANAGEMENT ============

@login_required
@admin_required
def periods(request):
    """List all school periods."""
    periods_list = Period.objects.all().order_by('order', 'start_time')

    context = {
        'periods': periods_list,
        'active_tab': 'periods',
    }

    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/periods_content.html', context)
    return render(request, 'academics/periods.html', context)


@login_required
@admin_required
def period_create(request):
    """Create a new period."""
    from .forms import PeriodForm

    if request.method == 'POST':
        form = PeriodForm(request.POST)
        if form.is_valid():
            period = form.save()
            messages.success(request, f'Period "{period.name}" created successfully.')

            if request.headers.get('HX-Request'):
                response = HttpResponse(status=204)
                response['HX-Trigger'] = 'periodChanged'
                return response
            return redirect('academics:periods')
    else:
        # Set default order to max + 1
        max_order = Period.objects.aggregate(max_order=models.Max('order'))['max_order'] or 0
        form = PeriodForm(initial={'order': max_order + 1})

    context = {'form': form, 'action': 'Create'}

    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/modal_period_form.html', context)
    return render(request, 'academics/period_form.html', context)


@login_required
@admin_required
def period_edit(request, pk):
    """Edit an existing period."""
    from .forms import PeriodForm

    period = get_object_or_404(Period, pk=pk)

    if request.method == 'POST':
        form = PeriodForm(request.POST, instance=period)
        if form.is_valid():
            form.save()
            messages.success(request, f'Period "{period.name}" updated successfully.')

            if request.headers.get('HX-Request'):
                response = HttpResponse(status=204)
                response['HX-Trigger'] = 'periodChanged'
                return response
            return redirect('academics:periods')
    else:
        form = PeriodForm(instance=period)

    context = {'form': form, 'period': period, 'action': 'Edit'}

    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/modal_period_form.html', context)
    return render(request, 'academics/period_form.html', context)


@login_required
@admin_required
def period_delete(request, pk):
    """Delete a period."""
    period = get_object_or_404(Period, pk=pk)

    if request.method == 'POST':
        name = period.name
        # Check if period has timetable entries
        if period.timetable_entries.exists():
            messages.error(request, f'Cannot delete "{name}" - it has timetable entries. Remove them first.')
        else:
            period.delete()
            messages.success(request, f'Period "{name}" deleted successfully.')

        if request.headers.get('HX-Request'):
            response = HttpResponse(status=204)
            response['HX-Trigger'] = 'periodChanged'
            return response
        return redirect('academics:periods')

    return HttpResponse(status=405)


# ============ TIMETABLE MANAGEMENT ============

@login_required
@admin_required
def timetable_index(request):
    """Timetable overview - select a class to view/edit."""
    classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')
    periods_count = Period.objects.filter(is_active=True).count()

    context = {
        'classes': classes,
        'periods_count': periods_count,
        'active_tab': 'timetable',
    }

    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/timetable_index_content.html', context)
    return render(request, 'academics/timetable_index.html', context)


@login_required
@admin_required
def class_timetable(request, class_id):
    """View and manage timetable for a specific class."""
    class_obj = get_object_or_404(Class, pk=class_id)
    periods_list = list(Period.objects.filter(is_active=True).order_by('order'))
    weekdays = TimetableEntry.Weekday.choices

    # Get all timetable entries for this class
    entries = TimetableEntry.objects.filter(
        class_subject__class_assigned=class_obj
    ).select_related('class_subject__subject', 'class_subject__teacher', 'period')

    # Build timetable grid: {weekday: {period_id: entry}}
    # Also track which slots are occupied by double periods from previous period
    timetable_grid = {}
    double_period_slots = {}  # {weekday: {period_id: entry}} - slots occupied by double periods

    # Create period order lookup for finding next period
    period_order_map = {p.pk: i for i, p in enumerate(periods_list)}

    for entry in entries:
        if entry.weekday not in timetable_grid:
            timetable_grid[entry.weekday] = {}
            double_period_slots[entry.weekday] = {}
        timetable_grid[entry.weekday][entry.period_id] = entry

        # If it's a double period, mark the next period slot as occupied
        if entry.is_double:
            current_idx = period_order_map.get(entry.period_id, -1)
            if current_idx >= 0 and current_idx + 1 < len(periods_list):
                next_period = periods_list[current_idx + 1]
                # Only mark if next period is not a break
                if not next_period.is_break:
                    double_period_slots[entry.weekday][next_period.pk] = entry

    # Get class subjects for the add entry form
    class_subjects = ClassSubject.objects.filter(
        class_assigned=class_obj
    ).select_related('subject', 'teacher')

    context = {
        'class_obj': class_obj,
        'periods': periods_list,
        'weekdays': weekdays,
        'timetable_grid': timetable_grid,
        'double_period_slots': double_period_slots,
        'class_subjects': class_subjects,
        'active_tab': 'timetable',
    }

    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/class_timetable_content.html', context)
    return render(request, 'academics/class_timetable.html', context)


@login_required
@admin_required
def timetable_entry_create(request, class_id):
    """Create a new timetable entry for a class."""
    from .forms import TimetableEntryForm

    class_obj = get_object_or_404(Class, pk=class_id)

    if request.method == 'POST':
        form = TimetableEntryForm(request.POST, class_instance=class_obj)
        if form.is_valid():
            entry = form.save()
            messages.success(request, f'Timetable entry added: {entry.class_subject.subject.name} on {entry.get_weekday_display()}')

            if request.headers.get('HX-Request'):
                response = HttpResponse(status=204)
                response['HX-Trigger'] = 'timetableChanged'
                return response
            return redirect('academics:class_timetable', class_id=class_id)
        else:
            # Return form with errors
            if request.headers.get('HX-Request'):
                context = {'form': form, 'class_obj': class_obj, 'weekdays': TimetableEntry.Weekday.choices}
                return render(request, 'academics/partials/modal_timetable_entry_form.html', context)
    else:
        # Pre-fill from query params if provided
        initial = {}
        if 'weekday' in request.GET:
            initial['weekday'] = request.GET['weekday']
        if 'period' in request.GET:
            initial['period'] = request.GET['period']

        form = TimetableEntryForm(class_instance=class_obj, initial=initial)

    context = {
        'form': form,
        'class_obj': class_obj,
        'weekdays': TimetableEntry.Weekday.choices,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/modal_timetable_entry_form.html', context)
    return render(request, 'academics/timetable_entry_form.html', context)


@login_required
@admin_required
def timetable_entry_delete(request, pk):
    """Delete a timetable entry."""
    entry = get_object_or_404(TimetableEntry, pk=pk)
    class_id = entry.class_subject.class_assigned_id

    if request.method == 'POST':
        subject_name = entry.class_subject.subject.name
        day_name = entry.get_weekday_display()
        entry.delete()
        messages.success(request, f'Removed {subject_name} from {day_name}')

        if request.headers.get('HX-Request'):
            response = HttpResponse(status=204)
            response['HX-Trigger'] = 'timetableChanged'
            return response
        return redirect('academics:class_timetable', class_id=class_id)

    return HttpResponse(status=405)