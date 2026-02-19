"""Class management views including CRUD, detail, and student enrollment."""
import io
import json
import logging
from datetime import datetime
from io import BytesIO

import pandas as pd

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from django.db import connection, models, transaction
from django.db.models import Count, Q, Sum, Case, When, IntegerField
from django.contrib import messages
from django.core.paginator import Paginator

from students.models import Student

from ..models import (
    Class, ClassSubject, AttendanceSession, AttendanceRecord,
    TimetableEntry, StudentSubjectEnrollment, Subject
)
from ..forms import ClassForm, StudentEnrollmentForm, ClassSubjectForm, CopySubjectsForm
from .base import admin_required, teacher_or_admin_required, htmx_render


def get_classes_list_context():
    """Get context for classes list."""
    classes = Class.objects.select_related('programme', 'class_teacher').annotate(
        student_count=Count('students', filter=models.Q(students__status='active'))
    ).order_by('level_number', 'section')
    return {'classes': classes}


@admin_required
def classes_list(request):
    """Classes list page with search and filters."""
    # Get filter parameters
    search = request.GET.get('search', '').strip()
    level_filter = request.GET.get('level', '')  # Combined filter: "shs_1", "basic_3", etc.

    # Get classes with student counts
    classes = Class.objects.select_related('programme', 'class_teacher').annotate(
        student_count=Count('students', filter=models.Q(students__status='active'))
    ).order_by('level_type', 'level_number', 'section')

    # Apply search filter
    if search:
        classes = classes.filter(
            Q(name__icontains=search) |
            Q(class_teacher__first_name__icontains=search) |
            Q(class_teacher__last_name__icontains=search) |
            Q(programme__name__icontains=search)
        )

    # Apply level filter (combined level_type and level_number)
    if level_filter and '_' in level_filter:
        level_type, level_number = level_filter.split('_', 1)
        classes = classes.filter(level_type=level_type, level_number=level_number)

    # Pagination
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100]:
            per_page = 25
    except ValueError:
        per_page = 25

    paginator = Paginator(classes, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Build subtitle - cache total_students to avoid querying twice
    total_classes = paginator.count
    all_classes_count = Class.objects.count()

    # Calculate comprehensive stats in a single aggregate query
    stats_agg = Class.objects.aggregate(
        total_students=Count('students', filter=models.Q(students__status='active')),
        total_capacity=Sum('capacity'),
        classes_with_teacher=Count('id', filter=models.Q(class_teacher__isnull=False))
    )
    total_students = stats_agg['total_students'] or 0
    total_capacity = stats_agg['total_capacity'] or 0
    classes_with_teacher = stats_agg['classes_with_teacher'] or 0
    fill_rate = round((total_students / total_capacity * 100), 1) if total_capacity > 0 else 0

    subtitle = f"{total_classes} class{'es' if total_classes != 1 else ''} • {total_students} student{'s' if total_students != 1 else ''}"
    if search or level_filter:
        subtitle += " (filtered)"

    # Get distinct level combinations for filter dropdown
    level_choices = []
    distinct_levels = Class.objects.values('level_type', 'level_number').distinct().order_by('level_type', 'level_number')
    for item in distinct_levels:
        lt, ln = item['level_type'], item['level_number']
        # Create display name based on level type
        if lt == 'shs':
            display = f"Form {ln}"
        elif lt == 'creche':
            display = f"Creche {ln}"
        elif lt == 'nursery':
            display = f"Nursery {ln}"
        elif lt == 'kg':
            display = f"KG {ln}"
        elif lt == 'basic':
            display = f"Basic {ln}"
        else:
            display = f"{lt.title()} {ln}"
        level_choices.append((f"{lt}_{ln}", display))

    context = {
        'classes': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'search': search,
        'level_filter': level_filter,
        'level_choices': level_choices,
        'stats': {
            'total_classes': all_classes_count,
            'total_students': total_students,
            'total_capacity': total_capacity,
            'fill_rate': fill_rate,
            'classes_with_teacher': classes_with_teacher,
        },
        'class_form': ClassForm(),
        # Navigation
        'back_url': '/academics/',
        'breadcrumbs': [
            {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
            {'label': 'Academics', 'url': '/academics/'},
            {'label': 'Classes'},
        ],
        'subtitle': subtitle,
    }

    return htmx_render(
        request,
        'academics/classes.html',
        'academics/partials/classes_content.html',
        context
    )


@admin_required
def class_create(request):
    """Create a new class."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    form = ClassForm(request.POST)
    if form.is_valid():
        form.save()
        if request.htmx:
            # Trigger event for page to refresh itself
            response = HttpResponse(status=204)
            response['HX-Trigger'] = 'classChanged'
            return response
        return redirect('academics:classes')

    # Validation error (422 keeps modal open)
    if request.htmx:
        response = render(request, 'academics/partials/modal_class_form.html', {
            'form': form,
            'is_create': True,
        })
        response.status_code = 422
        return response
    return redirect('academics:classes')


@admin_required
def class_edit(request, pk):
    """Edit a class (HTMX modal endpoint)."""
    cls = get_object_or_404(Class, pk=pk)

    if request.method == 'GET':
        # Non-HTMX GET (direct URL access/refresh) - redirect to classes list
        if not request.htmx:
            return redirect('academics:classes')
        # HTMX GET - return modal form partial
        form = ClassForm(instance=cls)
        return render(request, 'academics/partials/modal_class_form.html', {
            'form': form,
            'class': cls,
        })

    if request.method != 'POST':
        return HttpResponse(status=405)

    form = ClassForm(request.POST, instance=cls)
    if form.is_valid():
        form.save()
        if request.htmx:
            # Trigger event for page to refresh itself
            response = HttpResponse(status=204)
            response['HX-Trigger'] = 'classChanged'
            return response
        return redirect('academics:classes')

    # Validation error (422 keeps modal open)
    if request.htmx:
        response = render(request, 'academics/partials/modal_class_form.html', {
            'form': form,
            'class': cls,
        })
        response.status_code = 422
        return response
    return redirect('academics:classes')


@admin_required
def class_delete(request, pk):
    """Delete a class."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    cls = get_object_or_404(Class, pk=pk)
    class_name = cls.name
    cls.delete()

    messages.success(request, f'Class "{class_name}" has been deleted.')

    if request.htmx:
        response = HttpResponse(status=200)
        response['HX-Redirect'] = reverse('academics:classes')
        return response
    return redirect('academics:classes')


# ============ CLASS DETAIL HELPERS ============

def get_register_tab_context(class_obj, request=None, page=1, search='', gender='', sort='name'):
    """Context for the Students/Register tab with pagination, search, filter and sort."""
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    )

    # Gender breakdown (before filtering) - single aggregate query
    gender_stats = students.aggregate(
        total=Count('id'),
        male_count=Count('id', filter=models.Q(gender='M')),
        female_count=Count('id', filter=models.Q(gender='F'))
    )
    total_students = gender_stats['total'] or 0
    male_count = gender_stats['male_count'] or 0
    female_count = gender_stats['female_count'] or 0

    # Apply gender filter if provided
    if gender in ('M', 'F'):
        students = students.filter(gender=gender)

    # Apply search filter if provided
    if search:
        students = students.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(middle_name__icontains=search) |
            Q(admission_number__icontains=search) |
            Q(guardians__full_name__icontains=search)
        ).distinct()

    # Apply sorting
    sort_options = {
        'name': ('first_name', 'last_name'),
        '-name': ('-first_name', '-last_name'),
        'admission': ('admission_number',),
        '-admission': ('-admission_number',),
        'gender': ('gender', 'first_name'),
    }
    order_by = sort_options.get(sort, ('first_name', 'last_name'))
    students = students.order_by(*order_by)

    # Paginate
    paginator = Paginator(students, 20)  # 20 students per page
    students_page = paginator.get_page(page)

    # Get subjects that require manual assignment (auto_enroll=False)
    manual_subjects = ClassSubject.objects.filter(
        class_assigned=class_obj,
        auto_enroll=False
    ).select_related('subject', 'teacher')

    return {
        'class': class_obj,
        'students': students_page,
        'total_students': total_students,
        'search_query': search,
        'gender_filter': gender,
        'sort_by': sort,
        'gender_stats': {
            'male': male_count,
            'female': female_count,
        },
        'manual_subjects': manual_subjects,
    }


def get_teachers_tab_context(class_obj):
    """Context for the Teachers/Subjects tab with periods calculated from timetable."""
    # Get subject allocations with period counts from timetable
    subject_allocations = ClassSubject.objects.filter(
        class_assigned=class_obj
    ).select_related('subject', 'teacher').annotate(
        # Count periods: single = 1, double = 2
        timetable_periods=Sum(
            Case(
                When(timetable_entries__is_double=True, then=2),
                default=1,
                output_field=IntegerField()
            )
        )
    ).order_by('subject__name')

    return {
        'class': class_obj,
        'subject_allocations': subject_allocations
    }


def get_attendance_tab_context(class_obj):
    """Context for the Attendance tab with real data."""
    sessions = AttendanceSession.objects.filter(
        class_assigned=class_obj
    ).annotate(
        present_count=Count('records', filter=Q(records__status__in=['P', 'L'])),
        absent_count=Count('records', filter=Q(records__status='A')),
        total_count=Count('records')
    ).order_by('-date')

    # Calculate overall attendance percentage - single aggregate query
    stats = AttendanceRecord.objects.filter(
        session__class_assigned=class_obj
    ).aggregate(
        total_records=Count('id'),
        present_count=Count('id', filter=Q(status__in=['P', 'L']))
    )
    total_records = stats['total_records'] or 0
    present_count = stats['present_count'] or 0

    if total_records > 0:
        attendance_percentage = int((present_count / total_records) * 100)
    else:
        attendance_percentage = "--"

    return {
        'class': class_obj,
        'attendance_sessions': sessions,
        'attendance_percentage': attendance_percentage,
    }


def get_timetable_stats(class_obj):
    """Get timetable statistics for a class."""
    timetable_entries = TimetableEntry.objects.filter(
        class_subject__class_assigned=class_obj
    )
    return {
        'total_entries': timetable_entries.count(),
        'subjects_count': timetable_entries.values('class_subject__subject').distinct().count(),
        'teachers_count': timetable_entries.exclude(
            class_subject__teacher__isnull=True
        ).values('class_subject__teacher').distinct().count(),
    }


def get_class_detail_base_context(class_obj):
    """Get common context for class detail page."""
    # Check if class has any timetable entries (for per-lesson attendance warning)
    has_timetable = TimetableEntry.objects.filter(
        class_subject__class_assigned=class_obj
    ).exists()

    return {
        'class': class_obj,
        'timetable_stats': get_timetable_stats(class_obj),
        'has_timetable': has_timetable,
        'back_url': '/academics/classes/',
        'breadcrumbs': [
            {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
            {'label': 'Academics', 'url': '/academics/'},
            {'label': 'Classes', 'url': '/academics/classes/'},
            {'label': class_obj.name},
        ],
    }


def get_promotion_history_context(class_obj):
    """Context for showing recent enrollment/promotion activity for this class."""
    from students.models import Enrollment
    from core.models import AcademicYear

    current_year = AcademicYear.get_current()

    # Get recent enrollments for this class (incoming students)
    incoming_enrollments = Enrollment.objects.filter(
        class_assigned=class_obj
    ).select_related('student', 'academic_year', 'promoted_from').order_by(
        '-created_at'
    )[:5]

    # Get promotions/graduations out of this class
    outgoing_enrollments = Enrollment.objects.filter(
        promoted_from__class_assigned=class_obj
    ).select_related('student', 'academic_year', 'class_assigned').order_by(
        '-created_at'
    )[:5]

    # Summary stats
    if current_year:
        active_in_current_year = Enrollment.objects.filter(
            class_assigned=class_obj,
            academic_year=current_year,
            status=Enrollment.Status.ACTIVE
        ).count()

        promoted_count = Enrollment.objects.filter(
            promoted_from__class_assigned=class_obj,
            promoted_from__academic_year=current_year
        ).count()

        graduated_count = Enrollment.objects.filter(
            class_assigned=class_obj,
            academic_year=current_year,
            status=Enrollment.Status.GRADUATED
        ).count()
    else:
        active_in_current_year = 0
        promoted_count = 0
        graduated_count = 0

    return {
        'incoming_enrollments': incoming_enrollments,
        'outgoing_enrollments': outgoing_enrollments,
        'promotion_stats': {
            'active': active_in_current_year,
            'promoted': promoted_count,
            'graduated': graduated_count,
        },
        'current_year': current_year,
    }


# ============ CLASS DETAIL VIEW ============

@login_required
@teacher_or_admin_required
def class_detail(request, pk):
    """Detailed view of a specific class."""
    class_obj = get_object_or_404(Class, pk=pk)

    # Base context with timetable stats and navigation
    context = get_class_detail_base_context(class_obj)

    # Combine all tab contexts for the initial page load
    page = request.GET.get('page', 1)
    search = request.GET.get('q', '')
    gender = request.GET.get('gender', '')
    sort = request.GET.get('sort', 'name')
    context.update(get_register_tab_context(class_obj, request=request, page=page, search=search, gender=gender, sort=sort))
    context.update(get_teachers_tab_context(class_obj))
    context.update(get_attendance_tab_context(class_obj))
    context.update(get_promotion_history_context(class_obj))

    return htmx_render(
        request,
        'academics/class_detail.html',
        'academics/partials/class_detail_content.html',
        context
    )


@login_required
@teacher_or_admin_required
def class_register(request, pk):
    """Paginated student register for a class (HTMX endpoint)."""
    class_obj = get_object_or_404(Class, pk=pk)
    page = request.GET.get('page', 1)
    search = request.GET.get('q', '')
    gender = request.GET.get('gender', '')
    sort = request.GET.get('sort', 'name')

    context = get_register_tab_context(class_obj, request=request, page=page, search=search, gender=gender, sort=sort)

    return render(request, 'academics/includes/tab_register_content.html', context)


# ============ CLASS SUBJECT MANAGEMENT ============

def _get_class_subjects_queryset(class_obj):
    """Helper to get annotated class subjects queryset."""
    return ClassSubject.objects.filter(
        class_assigned=class_obj
    ).select_related('subject', 'teacher').annotate(
        timetable_periods=Sum(
            Case(
                When(timetable_entries__is_double=True, then=2),
                default=1,
                output_field=IntegerField()
            )
        )
    ).order_by('subject__name')


def _enroll_class_students_in_subject(class_obj, class_subject):
    """
    Enroll all active students in a class into a specific subject.
    Called when a new ClassSubject with auto_enroll=True is created.
    """
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).only('id')

    enrollments_to_create = []
    existing_student_ids = set(
        StudentSubjectEnrollment.objects.filter(
            class_subject=class_subject
        ).values_list('student_id', flat=True)
    )

    for student in students:
        if student.id not in existing_student_ids:
            enrollments_to_create.append(
                StudentSubjectEnrollment(
                    student=student,
                    class_subject=class_subject,
                    is_active=True
                )
            )

    if enrollments_to_create:
        StudentSubjectEnrollment.objects.bulk_create(
            enrollments_to_create,
            ignore_conflicts=True
        )

    return len(enrollments_to_create)


@admin_required
def class_subjects(request, pk):
    """View and manage subjects for a specific class."""
    class_obj = get_object_or_404(Class, pk=pk)
    subject_allocations = _get_class_subjects_queryset(class_obj)

    # Calculate stats using database aggregation (avoids N+1)
    stats = ClassSubject.objects.filter(class_assigned=class_obj).aggregate(
        teachers_assigned=Count('teacher', filter=Q(teacher__isnull=False)),
        total_periods=Sum(
            Case(
                When(timetable_entries__is_double=True, then=2),
                default=1,
                output_field=IntegerField()
            )
        )
    )

    context = {
        'class': class_obj,
        'subject_allocations': subject_allocations,
        'teachers_assigned': stats['teachers_assigned'] or 0,
        'total_periods': stats['total_periods'] or 0,
    }

    if request.htmx:
        return render(request, 'academics/partials/class_subjects_content.html', context)
    return render(request, 'academics/class_subjects.html', context)


@admin_required
def class_subjects_modal(request, pk):
    """Show class subjects in a modal (legacy, used by timetable page)."""
    class_obj = get_object_or_404(Class, pk=pk)
    subject_allocations = _get_class_subjects_queryset(class_obj)

    context = {
        'class': class_obj,
        'subject_allocations': subject_allocations,
    }

    return render(request, 'academics/includes/modal_subjects_content.html', context)


@admin_required
def class_subject_create(request, pk):
    """Add a subject allocation to a class (HTMX modal endpoint)."""
    class_obj = get_object_or_404(Class, pk=pk)

    # Non-HTMX GET (direct URL access/refresh) - redirect to class subjects page
    if request.method == 'GET' and not request.htmx:
        return redirect('academics:class_subjects', pk=pk)

    # Check if editing existing allocation
    subject_id = request.GET.get('subject_id') or request.POST.get('subject_id')
    allocation = None
    is_new = True
    if subject_id:
        allocation = get_object_or_404(ClassSubject, pk=subject_id, class_assigned=class_obj)
        is_new = False

    if request.method == 'POST':
        form = ClassSubjectForm(request.POST, instance=allocation, class_instance=class_obj)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.class_assigned = class_obj
            obj.save()

            # Auto-enroll existing students if this is a new auto_enroll subject
            if obj.auto_enroll and is_new:
                _enroll_class_students_in_subject(class_obj, obj)

            if request.htmx:
                # Close modal and trigger page refresh
                response = HttpResponse(status=204)
                response['HX-Trigger'] = 'closeModal, refreshSubjects'
                return response

            return redirect('academics:class_subjects', pk=pk)
    else:
        form = ClassSubjectForm(instance=allocation, class_instance=class_obj)

    return render(request, 'academics/partials/modal_subject_allocation.html', {
        'form': form,
        'class': class_obj,
        'allocation': allocation,
        'subject_id': subject_id,
    })


@admin_required
def class_subject_delete(request, class_pk, pk):
    """Delete a subject allocation from a class."""
    allocation = get_object_or_404(ClassSubject, pk=pk, class_assigned_id=class_pk)
    allocation.delete()

    if request.htmx:
        # Trigger page refresh
        response = HttpResponse(status=204)
        response['HX-Trigger'] = 'refreshSubjects'
        return response

    return redirect('academics:class_subjects', pk=class_pk)


@admin_required
def copy_subjects(request, pk):
    """Copy subjects from another class (HTMX modal endpoint)."""
    class_obj = get_object_or_404(Class, pk=pk)

    # Non-HTMX GET (direct URL access/refresh) - redirect to class subjects page
    if request.method == 'GET' and not request.htmx:
        return redirect('academics:class_subjects', pk=pk)

    if request.method == 'POST':
        form = CopySubjectsForm(request.POST, target_class=class_obj)
        if form.is_valid():
            created_count, skipped_count = form.save()

            # Auto-enroll students for subjects with auto_enroll=True
            for class_subject in ClassSubject.objects.filter(
                class_assigned=class_obj,
                auto_enroll=True
            ):
                _enroll_class_students_in_subject(class_obj, class_subject)

            if request.htmx:
                response = HttpResponse(status=204)
                response['HX-Trigger'] = 'closeModal, refreshSubjects'
                return response

            messages.success(
                request,
                f'Subjects copied: {created_count} added, {skipped_count} already existed.'
            )
            return redirect('academics:class_subjects', pk=pk)

        # Validation error
        if request.htmx:
            response = render(request, 'academics/partials/modal_copy_subjects.html', {
                'form': form,
                'class': class_obj,
            })
            response.status_code = 422
            return response
    else:
        form = CopySubjectsForm(target_class=class_obj)

    # Check if any source classes are available
    has_sources = form.fields['source_class'].queryset.exists()

    return render(request, 'academics/partials/modal_copy_subjects.html', {
        'form': form,
        'class': class_obj,
        'has_sources': has_sources,
    })


# ============ STUDENT ENROLLMENT & ELECTIVES ============

@admin_required
def class_student_enroll(request, pk):
    """Enroll existing students into a class (HTMX modal endpoint)."""
    class_obj = get_object_or_404(Class, pk=pk)

    # Non-HTMX GET (direct URL access/refresh) - redirect to class detail
    if request.method == 'GET' and not request.htmx:
        return redirect('academics:class_detail', pk=pk)

    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST, class_instance=class_obj)
        if form.is_valid():
            students_to_add = form.cleaned_data['students']
            for student in students_to_add:
                student.current_class = class_obj
                student.save()

                # Auto-enroll in class subjects
                # SHS: core subjects only | Other levels: all subjects
                StudentSubjectEnrollment.enroll_student_in_class_subjects(
                    student, class_obj
                )

            if request.htmx:
                response = HttpResponse()
                response['HX-Trigger'] = 'studentsEnrolled, closeModal'
                return response

            return redirect('academics:class_detail', pk=pk)
    else:
        form = StudentEnrollmentForm(class_instance=class_obj)

    return render(request, 'academics/partials/modal_student_enroll.html', {
        'form': form, 'class': class_obj
    })


@admin_required
def class_student_remove(request, class_pk, student_pk):
    """Remove a student from a class."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    class_obj = get_object_or_404(Class, pk=class_pk)
    student = get_object_or_404(Student, pk=student_pk, current_class_id=class_pk)
    student.current_class = None
    student.save()

    if request.htmx:
        # Return updated register tab content
        context = get_register_tab_context(class_obj)
        return render(request, 'academics/includes/tab_register_content.html', context)

    return redirect('academics:class_detail', pk=class_pk)


@admin_required
def class_student_electives(request, class_pk, student_pk):
    """Manage manual subject assignments for a student (HTMX modal endpoint)."""
    class_obj = get_object_or_404(Class, pk=class_pk)
    student = get_object_or_404(Student, pk=student_pk, current_class=class_obj)

    # Non-HTMX GET (direct URL access/refresh) - redirect to class detail
    if request.method == 'GET' and not request.htmx:
        return redirect('academics:class_detail', pk=class_pk)

    # Get subjects that require manual assignment (auto_enroll=False)
    manual_subjects = ClassSubject.objects.filter(
        class_assigned=class_obj,
        auto_enroll=False
    ).select_related('subject', 'teacher')

    # Get current enrollments for manual subjects
    enrolled_ids = list(StudentSubjectEnrollment.objects.filter(
        student=student,
        class_subject__class_assigned=class_obj,
        class_subject__auto_enroll=False,
        is_active=True
    ).values_list('class_subject_id', flat=True))

    context = {
        'class': class_obj,
        'student': student,
        'manual_subjects': manual_subjects,
        'enrolled_ids': enrolled_ids,
    }

    if request.method == 'POST':
        # Get selected subject IDs
        selected_ids = request.POST.getlist('subjects')

        # Update enrollments for manual subjects only
        for class_subject in manual_subjects:
            should_be_enrolled = str(class_subject.id) in selected_ids

            if should_be_enrolled:
                enrollment, created = StudentSubjectEnrollment.objects.get_or_create(
                    student=student,
                    class_subject=class_subject,
                    defaults={'is_active': True}
                )
                if not created and not enrollment.is_active:
                    enrollment.is_active = True
                    enrollment.save()
            else:
                StudentSubjectEnrollment.objects.filter(
                    student=student,
                    class_subject=class_subject
                ).update(is_active=False)

        # Return success message with OOB swap for tab content
        tab_context = get_register_tab_context(class_obj)
        tab_context['student'] = student
        tab_context['success'] = True
        return render(request, 'academics/includes/modal_student_subjects_success.html', tab_context)

    return render(request, 'academics/includes/modal_student_subjects.html', context)


@login_required
@teacher_or_admin_required
def class_bulk_subject_assign(request, pk):
    """Bulk assign manual subjects to multiple students (HTMX modal endpoint)."""
    class_obj = get_object_or_404(Class, pk=pk)

    # Non-HTMX GET (direct URL access/refresh) - redirect to class detail
    if request.method == 'GET' and not request.htmx:
        return redirect('academics:class_detail', pk=pk)

    # Get subjects that require manual assignment (auto_enroll=False)
    manual_subjects = ClassSubject.objects.filter(
        class_assigned=class_obj,
        auto_enroll=False
    ).select_related('subject', 'teacher')

    # Get all active students in the class
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('first_name', 'last_name')

    context = {
        'class': class_obj,
        'manual_subjects': manual_subjects,
        'students': students,
    }

    if request.method == 'POST':
        selected_subject_ids = request.POST.getlist('subjects')
        selected_student_ids = request.POST.getlist('students')

        if not selected_subject_ids:
            context['error'] = 'Please select at least one subject.'
            return render(request, 'academics/includes/modal_bulk_subject_assign.html', context)

        if not selected_student_ids:
            context['error'] = 'Please select at least one student.'
            return render(request, 'academics/includes/modal_bulk_subject_assign.html', context)

        # Get the selected class subjects and students
        class_subjects = manual_subjects.filter(id__in=selected_subject_ids)
        students_to_update = students.filter(id__in=selected_student_ids)

        # Create enrollments
        enrolled_count = 0
        for student in students_to_update:
            for class_subject in class_subjects:
                enrollment, created = StudentSubjectEnrollment.objects.get_or_create(
                    student=student,
                    class_subject=class_subject,
                    defaults={'is_active': True}
                )
                if created:
                    enrolled_count += 1
                elif not enrollment.is_active:
                    enrollment.is_active = True
                    enrollment.save()
                    enrolled_count += 1

        context['success'] = True
        context['enrolled_count'] = enrolled_count
        context['student_count'] = students_to_update.count()
        context['subject_count'] = class_subjects.count()
        return render(request, 'academics/includes/modal_bulk_subject_assign_success.html', context)

    return render(request, 'academics/includes/modal_bulk_subject_assign.html', context)


@admin_required
def class_bulk_electives(request, pk):
    """Bulk assign elective subjects to students (HTMX modal endpoint)."""
    class_obj = get_object_or_404(Class, pk=pk)

    # Non-HTMX GET (direct URL access/refresh) - redirect to class detail
    if request.method == 'GET' and not request.htmx:
        return redirect('academics:class_detail', pk=pk)

    # Get elective subjects filtered by programme
    if class_obj.programme:
        programme_subject_ids = set(
            class_obj.programme.subjects.values_list('id', flat=True)
        )
        elective_subjects = ClassSubject.objects.filter(
            class_assigned=class_obj,
            subject__is_core=False
        ).filter(
            models.Q(subject_id__in=programme_subject_ids) |
            models.Q(subject__programmes__isnull=True)
        ).select_related('subject', 'teacher').distinct()
        required_electives = class_obj.programme.required_electives
    else:
        elective_subjects = ClassSubject.objects.filter(
            class_assigned=class_obj,
            subject__is_core=False
        ).select_related('subject', 'teacher')
        required_electives = 3

    # Find students needing electives - use annotation instead of N+1 queries
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).annotate(
        elective_count=Count(
            'subject_enrollments',
            filter=Q(
                subject_enrollments__class_subject__class_assigned=class_obj,
                subject_enrollments__class_subject__subject__is_core=False,
                subject_enrollments__is_active=True
            )
        )
    )

    students_needing_electives = [
        {
            'student': student,
            'current_count': student.elective_count,
            'needed': required_electives - student.elective_count
        }
        for student in students
        if student.elective_count < required_electives
    ]

    context = {
        'class': class_obj,
        'elective_subjects': elective_subjects,
        'students_needing_electives': students_needing_electives,
        'required_electives': required_electives,
    }

    if request.method == 'POST':
        # Get selected electives and students
        selected_ids = request.POST.getlist('electives')
        student_ids = request.POST.getlist('students')

        if not selected_ids:
            context['error'] = 'Please select at least one elective subject.'
            return render(request, 'academics/includes/modal_bulk_electives.html', context)

        if not student_ids:
            context['error'] = 'Please select at least one student.'
            return render(request, 'academics/includes/modal_bulk_electives.html', context)

        # Get students
        students_to_update = Student.objects.filter(pk__in=student_ids, current_class=class_obj)

        # Get class subjects
        class_subjects = ClassSubject.objects.filter(
            id__in=selected_ids,
            class_assigned=class_obj,
            subject__is_core=False
        )

        # Apply enrollments
        enrolled_count = 0
        for student in students_to_update:
            for class_subject in class_subjects:
                enrollment, created = StudentSubjectEnrollment.objects.get_or_create(
                    student=student,
                    class_subject=class_subject,
                    defaults={'is_active': True}
                )
                if created:
                    enrolled_count += 1
                elif not enrollment.is_active:
                    enrollment.is_active = True
                    enrollment.save()
                    enrolled_count += 1

        # Return success message in modal + OOB swap for tab content
        tab_context = get_register_tab_context(class_obj)
        tab_context['enrolled_count'] = enrolled_count
        tab_context['bulk_success'] = True
        # Don't auto-close modal - let user see success message and click Close
        return render(request, 'academics/includes/modal_bulk_electives_success.html', tab_context)

    return render(request, 'academics/includes/modal_bulk_electives.html', context)


# ============ SYNC SUBJECTS ============

@admin_required
def class_sync_subjects(request, pk):
    """
    Sync all students in a class with their subject enrollments.

    For SHS: Enrolls students in core subjects only.
    For other levels: Enrolls students in all class subjects.

    Useful to fix existing data or after adding new subjects to a class.
    """
    if request.method != 'POST':
        return HttpResponse(status=405)

    class_obj = get_object_or_404(Class, pk=pk)

    # Get all active students in this class
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    )

    enrolled_count = 0
    for student in students:
        enrollments = StudentSubjectEnrollment.enroll_student_in_class_subjects(
            student, class_obj
        )
        enrolled_count += len(enrollments)

    if request.htmx:
        if enrolled_count > 0:
            messages.success(
                request,
                f'Synced subjects: {enrolled_count} enrollment(s) created for {students.count()} student(s).'
            )
        else:
            messages.info(request, 'All students already enrolled in required subjects.')

        # Return a redirect trigger
        response = HttpResponse(status=204)
        response['HX-Refresh'] = 'true'
        return response

    return redirect('academics:class_detail', pk=pk)


# ============ PROMOTION ============

@admin_required
def class_promote(request, pk):
    """Redirect to main promotion page (class-level promotion is now in students app)."""
    return redirect('students:promotion')


# ============ EXPORT ============

@login_required
@teacher_or_admin_required
def class_export(request, pk):
    """Export class register to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as DjangoHttpResponse
    from core.models import SchoolSettings

    class_obj = get_object_or_404(Class, pk=pk)
    students_qs = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('last_name', 'first_name')

    # Get stats upfront with single aggregate query
    student_stats = students_qs.aggregate(
        total=Count('id'),
        male_count=Count('id', filter=models.Q(gender='M')),
        female_count=Count('id', filter=models.Q(gender='F'))
    )
    students = list(students_qs)  # Convert to list since we need both iteration and count

    # Prefetch primary guardians to avoid N+1 on guardian_name/guardian_phone
    from students.models import StudentGuardian
    sg_qs = StudentGuardian.objects.filter(
        student__in=students, is_primary=True
    ).select_related('guardian')
    guardian_map = {sg.student_id: sg.guardian for sg in sg_qs}
    for student in students:
        student._cached_primary_guardian = guardian_map.get(student.id)

    school = SchoolSettings.load()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_obj.name} Register"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    table_header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # School Header
    ws.merge_cells('A1:F1')
    ws['A1'] = school.display_name or request.tenant.name
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Class Info
    ws.merge_cells('A2:F2')
    ws['A2'] = f"{class_obj.name} - {class_obj.level_display}"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Date
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Generated: {timezone.now().strftime('%B %d, %Y')}"
    ws['A3'].alignment = Alignment(horizontal='center')

    # Empty row
    ws.append([])

    # Table Headers
    headers = ['#', 'Admission No.', 'Full Name', 'Gender', 'Guardian', 'Phone']
    ws.append(headers)
    header_row = 5

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Student Data
    for idx, student in enumerate(students, 1):
        row_data = [
            idx,
            student.admission_number,
            student.full_name,
            student.get_gender_display(),
            student.guardian_name or '-',
            student.guardian_phone or '-',
        ]
        ws.append(row_data)

        # Apply borders
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=header_row + idx, column=col_num)
            cell.border = border
            if col_num == 1:
                cell.alignment = Alignment(horizontal='center')

    # Set column widths
    column_widths = [5, 15, 30, 10, 25, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Summary row - use pre-computed stats
    summary_row = header_row + len(students) + 2
    ws.cell(row=summary_row, column=1, value=f"Total Students: {student_stats['total'] or 0}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    ws.cell(row=summary_row + 1, column=1, value=f"Male: {student_stats['male_count'] or 0} | Female: {student_stats['female_count'] or 0}")

    # Create response
    response = DjangoHttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{class_obj.name.replace(' ', '_')}_Register.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@admin_required
def classes_bulk_export(request):
    """Export all classes to Excel with current filters applied."""
    import pandas as pd

    # Get filter parameters (same as classes_list view)
    search = request.GET.get('search', '').strip()
    level_filter = request.GET.get('level', '')

    # Get classes with student counts
    classes = Class.objects.select_related('programme', 'class_teacher').annotate(
        student_count=Count('students', filter=models.Q(students__status='active'))
    ).order_by('level_number', 'section')

    # Apply search filter
    if search:
        classes = classes.filter(
            Q(name__icontains=search) |
            Q(class_teacher__first_name__icontains=search) |
            Q(class_teacher__last_name__icontains=search) |
            Q(programme__name__icontains=search)
        )

    # Apply level filter
    if level_filter:
        classes = classes.filter(level_type=level_filter)

    # Build export data
    export_data = []
    for cls in classes:
        export_data.append({
            'Class Name': cls.name,
            'Level Type': cls.get_level_type_display(),
            'Level Number': cls.level_number,
            'Section': cls.section or '',
            'Programme': cls.programme.name if cls.programme else '',
            'Class Teacher': f"{cls.class_teacher.get_title_display()} {cls.class_teacher.full_name}" if cls.class_teacher else '',
            'Capacity': cls.capacity,
            'Current Students': cls.student_count,
            'Available Seats': cls.capacity - cls.student_count,
            'Status': 'Active' if cls.is_active else 'Inactive',
        })

    # Create Excel file
    df = pd.DataFrame(export_data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Classes')

        # Auto-adjust column widths
        worksheet = writer.sheets['Classes']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(col)
            ) + 2
            worksheet.column_dimensions[chr(65 + idx) if idx < 26 else 'A' + chr(65 + idx - 26)].width = min(max_length, 50)

    output.seek(0)

    # Generate filename with date
    filename = f"classes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============ PDF GENERATION ============

@login_required
@teacher_or_admin_required
def class_detail_pdf(request, pk):
    """Generate a branded PDF class register."""
    logger = logging.getLogger(__name__)

    class_obj = get_object_or_404(Class, pk=pk)

    # Get students (ordered)
    students = Student.objects.filter(
        current_class=class_obj,
        status='active'
    ).order_by('last_name', 'first_name')

    # Get subjects
    subjects = ClassSubject.objects.filter(
        class_assigned=class_obj
    ).select_related('subject', 'teacher').order_by('subject__name')

    # Get school context with logo
    try:
        from gradebook.utils import get_school_context
        school_ctx = get_school_context(include_logo_base64=True)
    except ImportError:
        school_ctx = {
            'school': getattr(connection, 'tenant', None),
            'logo_base64': None
        }

    # Generate PDF using WeasyPrint
    try:
        from weasyprint import HTML
        from django.template.loader import render_to_string
        from django.conf import settings as django_settings

        school = school_ctx.get('school') or getattr(connection, 'tenant', None)
        context = {
            'class': class_obj,
            'students': students,
            'subjects': subjects,
            'school': school,
            'logo_base64': school_ctx.get('logo_base64'),
            'generated_at': timezone.now(),
            'generated_by': request.user,
        }

        html_string = render_to_string('academics/class_detail_pdf.html', context)
        html = HTML(string=html_string, base_url=str(django_settings.BASE_DIR))
        pdf_buffer = BytesIO()
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = f"{class_obj.name.replace(' ', '_')}_Register.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except ImportError:
        logger.error("WeasyPrint not installed")
        messages.error(request, 'PDF generation is not available. WeasyPrint is not installed.')
        return redirect('academics:class_detail', pk=pk)
    except Exception as e:
        logger.error(f"Failed to generate class PDF: {str(e)}")
        messages.error(request, f'Failed to generate PDF: {str(e)}')
        return redirect('academics:class_detail', pk=pk)


# ============ BULK SUBJECT IMPORT ============

SUBJECT_IMPORT_COLUMNS = [
    'class_name', 'subject_name', 'teacher_email', 'periods_per_week', 'auto_enroll'
]


def clean_value(val):
    """Clean a value from pandas, handling NaN and whitespace."""
    if pd.isna(val):
        return ''
    return str(val).strip()


@admin_required
def bulk_subject_import(request):
    """Handle bulk import of subject assignments from Excel/CSV."""
    from teachers.models import Teacher

    if request.method == 'GET':
        return render(request, 'academics/partials/modal_bulk_subject_import.html', {
            'expected_columns': SUBJECT_IMPORT_COLUMNS,
        })

    # POST - process file
    if 'file' not in request.FILES:
        return render(request, 'academics/partials/modal_bulk_subject_import.html', {
            'expected_columns': SUBJECT_IMPORT_COLUMNS,
            'error': 'Please select a file to upload.',
        })

    file = request.FILES['file']
    ext = file.name.split('.')[-1].lower()

    if ext not in ['xlsx', 'csv']:
        return render(request, 'academics/partials/modal_bulk_subject_import.html', {
            'expected_columns': SUBJECT_IMPORT_COLUMNS,
            'error': 'Only .xlsx and .csv files are supported.',
        })

    try:
        # Read file
        if ext == 'xlsx':
            df = pd.read_excel(file, engine='openpyxl')
        else:
            df = pd.read_csv(file)

        if df.empty:
            return render(request, 'academics/partials/modal_bulk_subject_import.html', {
                'expected_columns': SUBJECT_IMPORT_COLUMNS,
                'error': 'The file is empty.',
            })

        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

        # Build lookups
        class_map = {c.name: c for c in Class.objects.filter(is_active=True)}
        subject_map = {s.name.lower(): s for s in Subject.objects.filter(is_active=True)}
        subject_map.update({s.short_name.lower(): s for s in Subject.objects.filter(is_active=True)})
        teacher_map = {t.user.email.lower(): t for t in Teacher.objects.filter(
            status='active', user__isnull=False
        ).select_related('user')}

        # Get existing class-subject combinations
        existing_combinations = set(
            ClassSubject.objects.values_list('class_assigned__name', 'subject__name')
        )

        # Process rows
        all_errors = []
        valid_rows = []

        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row number
            errors = []

            # Extract and clean values
            class_name = clean_value(row.get('class_name', ''))
            subject_name = clean_value(row.get('subject_name', ''))
            teacher_email = clean_value(row.get('teacher_email', '')).lower()
            periods_per_week = row.get('periods_per_week', 0)
            auto_enroll = clean_value(row.get('auto_enroll', 'true')).lower()

            # Parse periods_per_week
            try:
                periods_per_week = int(periods_per_week) if periods_per_week else 0
            except (ValueError, TypeError):
                periods_per_week = 0

            # Parse auto_enroll
            auto_enroll = auto_enroll in ['true', 'yes', '1', 't', 'y']

            # Validate class
            class_obj = None
            if not class_name:
                errors.append('Class name is required')
            elif class_name not in class_map:
                errors.append(f'Class "{class_name}" not found')
            else:
                class_obj = class_map[class_name]

            # Validate subject
            subject_obj = None
            if not subject_name:
                errors.append('Subject name is required')
            elif subject_name.lower() not in subject_map:
                errors.append(f'Subject "{subject_name}" not found')
            else:
                subject_obj = subject_map[subject_name.lower()]

            # Check if combination already exists
            if class_obj and subject_obj:
                if (class_obj.name, subject_obj.name) in existing_combinations:
                    errors.append(f'{subject_obj.name} already assigned to {class_obj.name}')

            # Validate teacher (optional)
            teacher_obj = None
            if teacher_email:
                if teacher_email not in teacher_map:
                    errors.append(f'Teacher with email "{teacher_email}" not found')
                else:
                    teacher_obj = teacher_map[teacher_email]

            if errors:
                all_errors.append({'row': row_num, 'errors': errors})
            else:
                valid_rows.append({
                    'row_num': row_num,
                    'class_name': class_obj.name,
                    'class_pk': class_obj.pk,
                    'subject_name': subject_obj.name,
                    'subject_pk': subject_obj.pk,
                    'teacher_email': teacher_email,
                    'teacher_pk': teacher_obj.pk if teacher_obj else None,
                    'teacher_name': teacher_obj.full_name if teacher_obj else None,
                    'periods_per_week': periods_per_week,
                    'auto_enroll': auto_enroll,
                })
                # Track to prevent duplicates in same import
                existing_combinations.add((class_obj.name, subject_obj.name))

        # Store in session for confirmation
        request.session['bulk_subject_import_data'] = json.dumps(valid_rows)

        return render(request, 'academics/partials/modal_bulk_subject_preview.html', {
            'valid_rows': valid_rows,
            'all_errors': all_errors,
            'total_rows': len(df),
            'valid_count': len(valid_rows),
            'error_count': len(all_errors),
        })

    except Exception as e:
        return render(request, 'academics/partials/modal_bulk_subject_import.html', {
            'expected_columns': SUBJECT_IMPORT_COLUMNS,
            'error': f'Error reading file: {str(e)}',
        })


@admin_required
def bulk_subject_import_confirm(request):
    """Confirm and process the bulk subject import."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    data = request.session.get('bulk_subject_import_data')
    if not data:
        return render(request, 'academics/partials/modal_bulk_subject_import.html', {
            'expected_columns': SUBJECT_IMPORT_COLUMNS,
            'error': 'Session expired. Please upload the file again.',
        })

    try:
        rows = json.loads(data)
    except json.JSONDecodeError:
        return render(request, 'academics/partials/modal_bulk_subject_import.html', {
            'expected_columns': SUBJECT_IMPORT_COLUMNS,
            'error': 'Invalid session data. Please upload the file again.',
        })

    created_count = 0
    errors = []

    try:
        with transaction.atomic():
            for row in rows:
                try:
                    class_subject = ClassSubject.objects.create(
                        class_assigned_id=row['class_pk'],
                        subject_id=row['subject_pk'],
                        teacher_id=row.get('teacher_pk'),
                        periods_per_week=row.get('periods_per_week', 0),
                        auto_enroll=row.get('auto_enroll', True)
                    )

                    # Auto-enroll students if auto_enroll is True
                    if class_subject.auto_enroll:
                        class_obj = Class.objects.get(pk=row['class_pk'])
                        _enroll_class_students_in_subject(class_obj, class_subject)

                    created_count += 1

                except Exception as e:
                    errors.append(f"Row {row.get('row_num', '?')}: {str(e)}")

    except Exception as e:
        errors.append(f"Error during import: {str(e)}")

    # Clear session
    request.session.pop('bulk_subject_import_data', None)

    if errors:
        return render(request, 'academics/partials/modal_bulk_subject_success.html', {
            'created_count': created_count,
            'errors': errors,
            'has_errors': True,
        })

    return render(request, 'academics/partials/modal_bulk_subject_success.html', {
        'created_count': created_count,
        'has_errors': False,
    })


@admin_required
def bulk_subject_import_template(request):
    """Download a sample import template for subject assignments."""
    sample_data = {
        'class_name': ['B1-A', 'B1-A', 'B1-B', 'B1-B'],
        'subject_name': ['Mathematics', 'English', 'Mathematics', 'English'],
        'teacher_email': ['john@school.com', 'jane@school.com', 'john@school.com', 'jane@school.com'],
        'periods_per_week': [5, 5, 5, 5],
        'auto_enroll': ['true', 'true', 'true', 'true'],
    }

    df = pd.DataFrame(sample_data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Subject Assignments')

    output.seek(0)
    return FileResponse(
        output,
        as_attachment=True,
        filename='subject_import_template.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============ ASSIGNMENT DASHBOARD ============

@admin_required
def assignment_dashboard(request):
    """
    Dashboard showing incomplete subject/teacher assignments at a glance.
    Helps admins identify what needs attention.
    """
    # Get active classes
    active_classes = Class.objects.filter(is_active=True)

    # 1. Classes without subjects
    classes_without_subjects = active_classes.annotate(
        subject_count=Count('subjects')
    ).filter(subject_count=0).order_by('level_type', 'level_number', 'section')

    # 2. Subjects without teachers (ClassSubject where teacher is null)
    subjects_without_teachers = ClassSubject.objects.filter(
        class_assigned__is_active=True,
        teacher__isnull=True
    ).select_related(
        'class_assigned', 'subject'
    ).order_by('class_assigned__level_type', 'class_assigned__level_number', 'subject__name')

    # 3. Students missing enrollments in manual subjects (auto_enroll=False)
    # Find students in classes with manual subjects they're not enrolled in
    manual_subjects = list(ClassSubject.objects.filter(
        class_assigned__is_active=True,
        auto_enroll=False
    ).select_related('class_assigned', 'subject'))

    students_missing_enrollments = []
    if manual_subjects:
        from collections import defaultdict

        # Batch-fetch all enrollments for manual subjects in one query
        enrolled_by_cs = defaultdict(set)
        for cs_id, student_id in StudentSubjectEnrollment.objects.filter(
            class_subject__in=manual_subjects
        ).values_list('class_subject_id', 'student_id'):
            enrolled_by_cs[cs_id].add(student_id)

        # Batch-fetch all active students for relevant classes
        class_ids = {cs.class_assigned_id for cs in manual_subjects}
        students_by_class = defaultdict(list)
        for s in Student.objects.filter(
            current_class_id__in=class_ids, status='active'
        ).only('id', 'current_class_id', 'first_name', 'last_name'):
            students_by_class[s.current_class_id].append(s)

        for cs in manual_subjects:
            enrolled = enrolled_by_cs.get(cs.id, set())
            missing = [s for s in students_by_class.get(cs.class_assigned_id, []) if s.id not in enrolled]
            if missing:
                students_missing_enrollments.append({
                    'class_subject': cs,
                    'missing_count': len(missing),
                    'students': missing[:5],
                    'has_more': len(missing) > 5,
                })

    # Summary stats
    total_classes = active_classes.count()
    classes_with_subjects = total_classes - classes_without_subjects.count()
    total_class_subjects = ClassSubject.objects.filter(class_assigned__is_active=True).count()
    subjects_with_teachers = total_class_subjects - subjects_without_teachers.count()

    context = {
        'classes_without_subjects': classes_without_subjects,
        'subjects_without_teachers': subjects_without_teachers,
        'students_missing_enrollments': students_missing_enrollments,
        'total_classes': total_classes,
        'classes_with_subjects': classes_with_subjects,
        'total_class_subjects': total_class_subjects,
        'subjects_with_teachers': subjects_with_teachers,
        'completion_percentage': round(
            (classes_with_subjects / total_classes * 100) if total_classes > 0 else 0
        ),
        'teacher_assignment_percentage': round(
            (subjects_with_teachers / total_class_subjects * 100) if total_class_subjects > 0 else 0
        ),
    }

    return htmx_render(
        request,
        'academics/assignment_dashboard.html',
        'academics/partials/assignment_dashboard_content.html',
        context
    )
