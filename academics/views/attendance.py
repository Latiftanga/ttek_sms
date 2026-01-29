"""Attendance management views including taking attendance, reports, and exports."""
from collections import defaultdict
from datetime import timedelta, datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import models
from django.db.models import Count, Q
from django.contrib import messages

from students.models import Student

from ..models import (
    Class, ClassSubject, AttendanceSession, AttendanceRecord, TimetableEntry
)
from ..utils import (
    should_use_lesson_attendance, get_students_for_lesson, get_lesson_attendance_stats
)
from .base import admin_required, teacher_or_admin_required, htmx_render, is_school_admin


# ============ DAILY ATTENDANCE ============

@login_required
@teacher_or_admin_required
def class_attendance_take(request, pk):
    """
    Opens the attendance sheet for a specific date (defaults to today).
    For per-lesson classes, redirects to lesson selection.
    Handles saving the records for daily attendance.
    """
    class_obj = get_object_or_404(Class, pk=pk)
    user = request.user
    is_admin = user.is_superuser or getattr(user, 'is_school_admin', False)

    # Check permission: must be admin, class teacher, or subject teacher for this class
    teacher = None
    if not is_admin:
        if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
            messages.error(request, 'You do not have permission to take attendance.')
            return redirect('core:index')

        teacher = user.teacher_profile
        is_class_teacher = class_obj.class_teacher == teacher
        is_subject_teacher = ClassSubject.objects.filter(
            class_assigned=class_obj,
            teacher=teacher
        ).exists()

        if not is_class_teacher and not is_subject_teacher:
            messages.error(request, 'You are not assigned to this class.')
            return redirect('academics:attendance_reports')
    else:
        # Admin user - check if they have a teacher profile
        teacher = getattr(user, 'teacher_profile', None)

    # Check if class uses per-lesson attendance
    if should_use_lesson_attendance(class_obj):
        # For per-lesson classes, redirect to lesson selection
        return redirect('academics:lesson_attendance_list', pk=pk)

    target_date = timezone.now().date()  # For now, default to today

    # Check if session exists (for daily attendance)
    session, created = AttendanceSession.objects.get_or_create(
        class_assigned=class_obj,
        date=target_date,
        session_type=AttendanceSession.SessionType.DAILY,
        defaults={'created_by': teacher}
    )

    if request.method == 'POST':
        # Process the form submission manually for grid data
        # Data format: "status_STUDENTID" : "STATUS_CODE"

        students = list(Student.objects.filter(current_class=class_obj, status='active'))
        student_ids = [s.id for s in students]

        # Fetch existing records in a single query
        existing_records = {
            r.student_id: r
            for r in AttendanceRecord.objects.filter(session=session, student_id__in=student_ids)
        }

        records_to_create = []
        records_to_update = []

        for student in students:
            status_key = f"status_{student.id}"
            new_status = request.POST.get(status_key, AttendanceRecord.Status.PRESENT)

            if student.id in existing_records:
                # Update existing record
                record = existing_records[student.id]
                if record.status != new_status:
                    record.status = new_status
                    records_to_update.append(record)
            else:
                # Create new record
                records_to_create.append(AttendanceRecord(
                    session=session,
                    student=student,
                    status=new_status
                ))

        # Bulk operations
        if records_to_create:
            AttendanceRecord.objects.bulk_create(records_to_create)
        if records_to_update:
            AttendanceRecord.objects.bulk_update(records_to_update, ['status'])

        # HTMX Success: Close modal or redirect
        if request.htmx:
            from django.urls import reverse
            url = reverse('academics:class_detail', args=[pk])
            # Return script that closes modal if exists, otherwise redirects
            script = '''<script>
                var dialog = document.querySelector('dialog[open]');
                if (dialog) {
                    dialog.close();
                    // Refresh the stats bar by triggering a refresh of class detail content
                    htmx.ajax('GET', '%s', {target: '#main-content', swap: 'innerHTML'});
                } else {
                    window.location.href = '%s';
                }
            </script>''' % (url, url)
            return HttpResponse(script)

        return redirect('academics:class_detail', pk=pk)

    # GET Request: Prepare data for the form
    students = Student.objects.filter(current_class=class_obj, status='active').order_by('first_name')
    records = {r.student_id: r.status for r in session.records.all()}

    # Combine student + their status
    student_list = []
    for student in students:
        student_list.append({
            'obj': student,
            'status': records.get(student.id, 'P')  # Default to Present if new
        })

    return render(request, 'academics/partials/modal_attendance_take.html', {
        'class': class_obj,
        'session': session,
        'student_list': student_list,
        'date': target_date,
        'is_lesson': False,
    })


@login_required
@teacher_or_admin_required
def class_attendance_edit(request, pk, session_pk):
    """Edit an existing attendance session."""
    class_obj = get_object_or_404(Class, pk=pk)
    session = get_object_or_404(AttendanceSession, pk=session_pk, class_assigned=class_obj)

    if request.method == 'POST':
        students = list(Student.objects.filter(current_class=class_obj, status='active'))
        student_ids = [s.id for s in students]

        # Fetch existing records in a single query
        existing_records = {
            r.student_id: r
            for r in AttendanceRecord.objects.filter(session=session, student_id__in=student_ids)
        }

        records_to_create = []
        records_to_update = []

        for student in students:
            status_key = f"status_{student.id}"
            new_status = request.POST.get(status_key, AttendanceRecord.Status.PRESENT)

            if student.id in existing_records:
                # Update existing record
                record = existing_records[student.id]
                if record.status != new_status:
                    record.status = new_status
                    records_to_update.append(record)
            else:
                # Create new record
                records_to_create.append(AttendanceRecord(
                    session=session,
                    student=student,
                    status=new_status
                ))

        # Bulk operations
        if records_to_create:
            AttendanceRecord.objects.bulk_create(records_to_create)
        if records_to_update:
            AttendanceRecord.objects.bulk_update(records_to_update, ['status'])

        if request.htmx:
            from django.urls import reverse
            url = reverse('academics:class_detail', args=[pk])
            # Return script that closes modal if exists, otherwise redirects
            script = '''<script>
                var dialog = document.querySelector('dialog[open]');
                if (dialog) {
                    dialog.close();
                    htmx.ajax('GET', '%s', {target: '#main-content', swap: 'innerHTML'});
                } else {
                    window.location.href = '%s';
                }
            </script>''' % (url, url)
            return HttpResponse(script)

        return redirect('academics:class_detail', pk=pk)

    # GET Request: Load existing records
    students = Student.objects.filter(current_class=class_obj, status='active').order_by('first_name')
    records = {r.student_id: r.status for r in session.records.all()}

    student_list = []
    for student in students:
        student_list.append({
            'obj': student,
            'status': records.get(student.id, 'P')
        })

    return render(request, 'academics/partials/modal_attendance_take.html', {
        'class': class_obj,
        'session': session,
        'student_list': student_list,
        'date': session.date,
        'is_edit': True
    })


@admin_required
def class_attendance_history(request, pk):
    """Show class attendance history in a modal."""
    class_obj = get_object_or_404(Class, pk=pk)

    attendance_sessions = AttendanceSession.objects.filter(
        class_assigned=class_obj
    ).order_by('-date')[:20]

    # Add counts
    for session in attendance_sessions:
        session.present_count = session.records.filter(status='present').count()
        session.absent_count = session.records.filter(status='absent').count()
        session.total_count = session.present_count + session.absent_count

    context = {
        'class': class_obj,
        'attendance_sessions': attendance_sessions,
    }

    return render(request, 'academics/includes/modal_attendance_history.html', context)


# ============ PER-LESSON ATTENDANCE ============

@login_required
@teacher_or_admin_required
def lesson_attendance_list(request, pk):
    """
    Shows all lessons for today for a per-lesson attendance class.
    Teachers select which lesson to mark attendance for.
    """
    class_obj = get_object_or_404(Class, pk=pk)
    user = request.user
    is_admin = user.is_superuser or getattr(user, 'is_school_admin', False)

    # Check permission
    teacher = None
    if not is_admin:
        if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
            messages.error(request, 'You do not have permission to view this page.')
            return redirect('core:index')

        teacher = user.teacher_profile
        is_class_teacher = class_obj.class_teacher == teacher
        is_subject_teacher = ClassSubject.objects.filter(
            class_assigned=class_obj,
            teacher=teacher
        ).exists()

        if not is_class_teacher and not is_subject_teacher:
            messages.error(request, 'You are not assigned to this class.')
            return redirect('academics:attendance_reports')
    else:
        teacher = getattr(user, 'teacher_profile', None)

    # If class doesn't use per-lesson attendance, redirect to daily attendance
    if not should_use_lesson_attendance(class_obj):
        return redirect('academics:class_attendance_take', pk=pk)

    today = timezone.now().date()
    today_weekday = timezone.now().isoweekday()

    # Get all timetable entries for this class today
    entries = TimetableEntry.objects.filter(
        class_subject__class_assigned=class_obj,
        weekday=today_weekday
    ).select_related(
        'class_subject__subject',
        'class_subject__teacher',
        'period',
        'classroom'
    ).order_by('period__order')

    # Get existing attendance sessions for today
    existing_sessions = {
        s.timetable_entry_id: s
        for s in AttendanceSession.objects.filter(
            class_assigned=class_obj,
            date=today,
            session_type=AttendanceSession.SessionType.LESSON
        )
    }

    current_time = timezone.localtime().time()
    lessons = []

    for entry in entries:
        is_past = entry.period.end_time < current_time
        is_current = entry.period.start_time <= current_time <= entry.period.end_time
        session = existing_sessions.get(entry.id)
        attendance_taken = session is not None

        # Check if this teacher can mark this lesson
        can_mark = is_admin or (teacher and entry.class_subject.teacher == teacher)

        lessons.append({
            'entry': entry,
            'period': entry.period,
            'subject': entry.class_subject.subject,
            'teacher': entry.class_subject.teacher,
            'classroom': entry.classroom,
            'is_past': is_past,
            'is_current': is_current,
            'attendance_taken': attendance_taken,
            'session': session,
            'can_mark': can_mark,
        })

    # Check if timetable exists at all for this class (any day)
    has_any_timetable = TimetableEntry.objects.filter(
        class_subject__class_assigned=class_obj
    ).exists()

    context = {
        'class': class_obj,
        'lessons': lessons,
        'date': today,
        'has_lessons': len(lessons) > 0,
        'no_timetable': not has_any_timetable,
    }

    return render(request, 'academics/partials/lesson_select.html', context)


@login_required
@teacher_or_admin_required
def take_lesson_attendance(request, timetable_entry_id):
    """
    Takes attendance for a specific timetable entry (lesson).
    Only the assigned teacher or admin can mark attendance.
    """
    entry = get_object_or_404(
        TimetableEntry.objects.select_related(
            'class_subject__class_assigned',
            'class_subject__subject',
            'class_subject__teacher',
            'period'
        ),
        pk=timetable_entry_id
    )

    class_obj = entry.class_subject.class_assigned
    class_subject = entry.class_subject
    user = request.user
    is_admin = user.is_superuser or getattr(user, 'is_school_admin', False)

    # Permission check: only assigned teacher or admin
    teacher = None
    if not is_admin:
        if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
            messages.error(request, 'You do not have permission to take attendance.')
            return redirect('core:index')

        teacher = user.teacher_profile

        # Must be the assigned teacher for this subject
        if class_subject.teacher != teacher:
            messages.error(request, 'You are not assigned to teach this lesson.')
            return redirect('academics:lesson_attendance_list', pk=class_obj.pk)
    else:
        teacher = getattr(user, 'teacher_profile', None)

    target_date = timezone.now().date()

    # Get or create the lesson attendance session
    session, created = AttendanceSession.objects.get_or_create(
        class_assigned=class_obj,
        date=target_date,
        timetable_entry=entry,
        session_type=AttendanceSession.SessionType.LESSON,
        defaults={
            'created_by': teacher,
            'period': entry.period,
            'class_subject': class_subject,
        }
    )

    # Get students for this lesson (considers elective enrollment)
    students = get_students_for_lesson(class_obj, class_subject)

    if request.method == 'POST':
        student_ids = [s.id for s in students]

        # Fetch existing records
        existing_records = {
            r.student_id: r
            for r in AttendanceRecord.objects.filter(session=session, student_id__in=student_ids)
        }

        records_to_create = []
        records_to_update = []

        for student in students:
            status_key = f"status_{student.id}"
            new_status = request.POST.get(status_key, AttendanceRecord.Status.PRESENT)

            if student.id in existing_records:
                record = existing_records[student.id]
                if record.status != new_status:
                    record.status = new_status
                    records_to_update.append(record)
            else:
                records_to_create.append(AttendanceRecord(
                    session=session,
                    student=student,
                    status=new_status
                ))

        # Bulk operations
        if records_to_create:
            AttendanceRecord.objects.bulk_create(records_to_create)
        if records_to_update:
            AttendanceRecord.objects.bulk_update(records_to_update, ['status'])

        # HTMX Success
        if request.htmx:
            from django.urls import reverse
            url = reverse('academics:lesson_attendance_list', args=[class_obj.pk])
            script = '''<script>
                var dialog = document.querySelector('dialog[open]');
                if (dialog) {
                    dialog.close();
                    htmx.ajax('GET', '%s', {target: '#main-content', swap: 'innerHTML'});
                } else {
                    window.location.href = '%s';
                }
            </script>''' % (url, url)
            return HttpResponse(script)

        return redirect('academics:lesson_attendance_list', pk=class_obj.pk)

    # GET Request: Prepare data for the form
    records = {r.student_id: r.status for r in session.records.all()}

    student_list = []
    for student in students:
        student_list.append({
            'obj': student,
            'status': records.get(student.id, 'P')
        })

    is_edit = not created and session.records.exists()

    return render(request, 'academics/partials/modal_attendance_take.html', {
        'class': class_obj,
        'session': session,
        'student_list': student_list,
        'date': target_date,
        'is_lesson': True,
        'is_edit': is_edit,
        'entry': entry,
        'subject': class_subject.subject,
        'period': entry.period,
        'timetable_entry_id': timetable_entry_id,
    })


@login_required
@teacher_or_admin_required
def class_weekly_attendance_report(request, pk):
    """
    Weekly attendance report by subject for form masters.
    Shows attendance rate per subject for per-lesson classes.
    """
    class_obj = get_object_or_404(Class, pk=pk)
    user = request.user
    is_admin = user.is_superuser or getattr(user, 'is_school_admin', False)

    # Permission check: form master or admin
    if not is_admin:
        if not getattr(user, 'is_teacher', False) or not hasattr(user, 'teacher_profile'):
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('core:index')

        teacher = user.teacher_profile
        if class_obj.class_teacher != teacher:
            messages.error(request, 'Only the form teacher can view this report.')
            return redirect('academics:class_detail', pk=pk)

    # Get date range from query params or default to current week
    today = timezone.now().date()
    week_offset = int(request.GET.get('week', 0))

    # Calculate start of week (Monday)
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    end_of_week = start_of_week + timedelta(days=4)  # Friday

    # Get attendance stats
    stats = get_lesson_attendance_stats(class_obj, start_of_week, end_of_week)

    # Calculate overall stats
    total_sessions = sum(s['sessions'] for s in stats)
    total_present = sum(s['present'] for s in stats)
    total_absent = sum(s['absent'] for s in stats)
    total_late = sum(s['late'] for s in stats)
    total_records = total_present + total_absent + total_late
    overall_rate = round((total_present + total_late) / total_records * 100, 1) if total_records > 0 else 0

    context = {
        'class': class_obj,
        'stats': stats,
        'start_date': start_of_week,
        'end_date': end_of_week,
        'week_offset': week_offset,
        'prev_week': week_offset - 1,
        'next_week': week_offset + 1 if week_offset < 0 else None,
        'is_current_week': week_offset == 0,
        'uses_lesson_attendance': should_use_lesson_attendance(class_obj),
        'overall': {
            'sessions': total_sessions,
            'present': total_present,
            'absent': total_absent,
            'late': total_late,
            'rate': overall_rate,
        },
    }

    return render(request, 'academics/partials/weekly_attendance_report.html', context)


# ============ ATTENDANCE REPORTS ============

@login_required
@teacher_or_admin_required
def attendance_reports(request):
    """Attendance reports with filters."""
    user = request.user
    is_admin = user.is_superuser or getattr(user, 'is_school_admin', False)

    # Get filter parameters
    class_filter = request.GET.get('class', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    view_mode = request.GET.get('view', 'summary')  # summary, daily, students

    # Default date range: last 30 days
    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    # Filter classes based on user role
    if is_admin:
        classes = Class.objects.filter(is_active=True).order_by('level_number', 'name')
    elif getattr(user, 'is_teacher', False) and hasattr(user, 'teacher_profile'):
        teacher = user.teacher_profile
        # Teachers see classes they're class teacher for OR assigned to teach
        homeroom_ids = Class.objects.filter(class_teacher=teacher).values_list('id', flat=True)
        assigned_ids = ClassSubject.objects.filter(teacher=teacher).values_list('class_assigned_id', flat=True)
        all_class_ids = set(homeroom_ids) | set(assigned_ids)
        classes = Class.objects.filter(id__in=all_class_ids, is_active=True).order_by('level_number', 'name')
    else:
        classes = Class.objects.none()

    allowed_class_ids = list(classes.values_list('id', flat=True))

    # Base querysets - filter by allowed classes for teachers
    sessions = AttendanceSession.objects.select_related('class_assigned')
    records = AttendanceRecord.objects.select_related('session', 'student', 'session__class_assigned')

    if not is_admin:
        sessions = sessions.filter(class_assigned_id__in=allowed_class_ids)
        records = records.filter(session__class_assigned_id__in=allowed_class_ids)

    # Apply date filter
    if date_from:
        sessions = sessions.filter(date__gte=date_from)
        records = records.filter(session__date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__lte=date_to)
        records = records.filter(session__date__lte=date_to)

    # Apply class filter
    if class_filter:
        sessions = sessions.filter(class_assigned_id=class_filter)
        records = records.filter(session__class_assigned_id=class_filter)

    # Calculate summary stats
    total_sessions = sessions.count()
    total_records = records.count()
    present_count = records.filter(status__in=['P', 'L']).count()
    absent_count = records.filter(status='A').count()
    late_count = records.filter(status='L').count()

    attendance_rate = 0
    if total_records > 0:
        attendance_rate = round((present_count / total_records) * 100, 1)

    # Summary by class - use aggregated queries instead of N+1
    # Get attendance stats per class in a single query
    class_stats = records.values('session__class_assigned_id').annotate(
        total=Count('id'),
        present=Count('id', filter=Q(status__in=['P', 'L'])),
        absent=Count('id', filter=Q(status='A'))
    )
    class_stats_dict = {
        s['session__class_assigned_id']: s for s in class_stats
    }

    # Get today's sessions in a single query
    today_sessions = set(
        sessions.filter(date=today).values_list('class_assigned_id', flat=True)
    )

    # Get student counts per class in a single query
    student_counts = dict(
        Student.objects.filter(
            current_class__in=classes,
            status='active'
        ).values('current_class_id').annotate(
            count=Count('id')
        ).values_list('current_class_id', 'count')
    )

    class_summary = []
    for cls in classes:
        stats = class_stats_dict.get(cls.id, {'total': 0, 'present': 0, 'absent': 0})
        cls_total = stats['total']
        cls_present = stats['present']
        cls_absent = stats['absent']
        cls_rate = round((cls_present / cls_total) * 100, 1) if cls_total > 0 else 0

        class_summary.append({
            'class': cls,
            'total': student_counts.get(cls.id, 0),
            'present': cls_present,
            'absent': cls_absent,
            'rate': cls_rate,
            'has_today': cls.id in today_sessions,
        })

    # Daily breakdown - use annotated queryset instead of N+1
    daily_sessions = sessions.order_by('-date')[:20]
    session_ids = [s.id for s in daily_sessions]

    # Get stats per session in a single query
    session_stats = AttendanceRecord.objects.filter(
        session_id__in=session_ids
    ).values('session_id').annotate(
        total=Count('id'),
        present=Count('id', filter=Q(status__in=['P', 'L'])),
        absent=Count('id', filter=Q(status='A'))
    )
    session_stats_dict = {s['session_id']: s for s in session_stats}

    daily_data = []
    for session in daily_sessions:
        stats = session_stats_dict.get(session.id, {'total': 0, 'present': 0, 'absent': 0})
        s_total = stats['total']
        s_present = stats['present']
        s_absent = stats['absent']
        daily_data.append({
            'session': session,
            'total': s_total,
            'present': s_present,
            'absent': s_absent,
            'rate': round((s_present / s_total) * 100, 1) if s_total > 0 else 0,
        })

    # Students with low attendance (for students view)
    low_attendance_students = []
    student_stats_all = {}  # Track all student stats for consecutive absences

    # Build student stats from records
    for record in records:
        sid = record.student_id
        if sid not in student_stats_all:
            student_stats_all[sid] = {
                'student': record.student,
                'total': 0,
                'present': 0,
                'records': []
            }
        student_stats_all[sid]['total'] += 1
        if record.status in ['P', 'L']:
            student_stats_all[sid]['present'] += 1
        student_stats_all[sid]['records'].append({
            'date': record.session.date,
            'status': record.status
        })

    # Always calculate low attendance students
    for sid, stats in student_stats_all.items():
        if stats['total'] > 0:
            rate = round((stats['present'] / stats['total']) * 100, 1)
            if rate < 80:  # Low attendance threshold
                low_attendance_students.append({
                    'student': stats['student'],
                    'total': stats['total'],
                    'present': stats['present'],
                    'absent': stats['total'] - stats['present'],
                    'rate': rate,
                })

    # Sort by attendance rate (lowest first)
    low_attendance_students.sort(key=lambda x: x['rate'])

    # Trends data for chart (always calculate for sidebar chart)
    trend_data = []

    # Group records by date
    daily_rates = defaultdict(lambda: {'present': 0, 'total': 0})
    for record in records:
        date_str = record.session.date.isoformat()
        daily_rates[date_str]['total'] += 1
        if record.status in ['P', 'L']:
            daily_rates[date_str]['present'] += 1

    # Sort by date and calculate rates
    for date_str in sorted(daily_rates.keys()):
        data = daily_rates[date_str]
        rate = round((data['present'] / data['total']) * 100, 1) if data['total'] > 0 else 0
        trend_data.append({
            'date': date_str,
            'rate': rate,
            'present': data['present'],
            'absent': data['total'] - data['present'],
            'total': data['total']
        })

    # Calculate consecutive absences for alert
    students_with_consecutive_absences = []
    CONSECUTIVE_THRESHOLD = 3

    for sid, stats in student_stats_all.items():
        # Sort records by date
        sorted_records = sorted(stats['records'], key=lambda x: x['date'], reverse=True)
        consecutive = 0
        for rec in sorted_records:
            if rec['status'] == 'A':
                consecutive += 1
            else:
                break

        if consecutive >= CONSECUTIVE_THRESHOLD:
            students_with_consecutive_absences.append({
                'student': stats['student'],
                'consecutive_days': consecutive,
                'last_present': None
            })
            # Find last present date
            for rec in sorted_records:
                if rec['status'] in ['P', 'L']:
                    students_with_consecutive_absences[-1]['last_present'] = rec['date']
                    break

    # Sort by consecutive days (highest first)
    students_with_consecutive_absences.sort(key=lambda x: x['consecutive_days'], reverse=True)

    # Calculate trend indicators (compare to previous period)
    prev_period_rate = None
    rate_change = None
    if date_from and date_to:
        try:
            df = datetime.fromisoformat(date_from)
            dt = datetime.fromisoformat(date_to)
            period_length = (dt - df).days
            prev_start = (df - timedelta(days=period_length)).date()
            prev_end = (df - timedelta(days=1)).date()

            # Get previous period records
            prev_records = AttendanceRecord.objects.filter(
                session__date__gte=prev_start,
                session__date__lte=prev_end
            )
            if not is_admin:
                prev_records = prev_records.filter(session__class_assigned_id__in=allowed_class_ids)
            if class_filter:
                prev_records = prev_records.filter(session__class_assigned_id=class_filter)

            prev_total = prev_records.count()
            if prev_total > 0:
                prev_present = prev_records.filter(status__in=['P', 'L']).count()
                prev_period_rate = round((prev_present / prev_total) * 100, 1)
                rate_change = round(attendance_rate - prev_period_rate, 1)
        except (ValueError, TypeError):
            pass

    context = {
        'classes': classes,
        'class_filter': class_filter,
        'date_from': date_from,
        'date_to': date_to,
        'view_mode': view_mode,
        'is_admin': is_admin,
        'today': today,
        'stats': {
            'total_sessions': total_sessions,
            'total_records': total_records,
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
            'rate': attendance_rate,
        },
        'class_summary': class_summary,
        'daily_data': daily_data,
        'low_attendance_students': low_attendance_students,
        'trend_data': trend_data,
        'students_with_consecutive_absences': students_with_consecutive_absences,
        'prev_period_rate': prev_period_rate,
        'rate_change': rate_change,
    }

    return htmx_render(
        request,
        'academics/attendance_reports.html',
        'academics/partials/attendance_reports_content.html',
        context
    )


@login_required
@teacher_or_admin_required
def attendance_export(request):
    """Export attendance data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as DjangoHttpResponse
    from core.models import SchoolSettings

    # Get filter parameters
    class_filter = request.GET.get('class', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Default date range
    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    school = SchoolSettings.load()

    # Get records
    records = AttendanceRecord.objects.select_related(
        'session', 'student', 'session__class_assigned'
    ).filter(
        session__date__gte=date_from,
        session__date__lte=date_to
    ).order_by('session__date', 'session__class_assigned__name', 'student__last_name')

    if class_filter:
        records = records.filter(session__class_assigned_id=class_filter)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    table_header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    late_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = school.display_name or request.tenant.name
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Attendance Report: {date_from} to {date_to}"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:F3')
    ws['A3'] = f"Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}"
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Table headers
    headers = ['Date', 'Class', 'Student Name', 'Admission No.', 'Status', 'Remarks']
    ws.append(headers)
    header_row = 5

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data rows
    status_map = {'P': 'Present', 'A': 'Absent', 'L': 'Late', 'E': 'Excused'}
    for idx, record in enumerate(records, 1):
        row_data = [
            record.session.date.strftime('%Y-%m-%d'),
            record.session.class_assigned.name,
            record.student.full_name,
            record.student.admission_number,
            status_map.get(record.status, record.status),
            record.remarks or '',
        ]
        ws.append(row_data)

        row_num = header_row + idx
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

        # Color-code status
        status_cell = ws.cell(row=row_num, column=5)
        if record.status == 'P':
            status_cell.fill = present_fill
        elif record.status == 'A':
            status_cell.fill = absent_fill
        elif record.status == 'L':
            status_cell.fill = late_fill

    # Column widths
    column_widths = [12, 15, 30, 15, 12, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Summary - use aggregate to get all counts in one query
    stats = records.aggregate(
        total=Count('id'),
        present=Count('id', filter=models.Q(status__in=['P', 'L'])),
        absent=Count('id', filter=models.Q(status='A'))
    )
    total_records = stats['total'] or 0
    present = stats['present'] or 0
    absent = stats['absent'] or 0

    summary_row = header_row + total_records + 2
    ws.cell(row=summary_row, column=1, value=f"Total Records: {total_records}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    rate = round((present / total_records) * 100, 1) if total_records > 0 else 0
    ws.cell(row=summary_row + 1, column=1, value=f"Present: {present} | Absent: {absent} | Rate: {rate}%")

    # Response
    response = DjangoHttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Attendance_Report_{date_from}_to_{date_to}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@teacher_or_admin_required
def student_attendance_detail(request, student_id):
    """Get detailed attendance for a single student."""
    student = get_object_or_404(Student, pk=student_id)

    # Get filter parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Default date range: last 30 days
    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=30)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    # Get attendance records
    records = AttendanceRecord.objects.filter(
        student=student,
        session__date__gte=date_from,
        session__date__lte=date_to
    ).select_related('session').order_by('-session__date')

    total = records.count()
    present = records.filter(status__in=['P', 'L']).count()
    absent = records.filter(status='A').count()
    late = records.filter(status='L').count()
    rate = round((present / total) * 100, 1) if total > 0 else 0

    # Build attendance calendar data
    attendance_data = []
    for record in records:
        attendance_data.append({
            'date': record.session.date.isoformat(),
            'status': record.status,
            'status_display': record.get_status_display(),
            'remarks': record.remarks or ''
        })

    context = {
        'student': student,
        'stats': {
            'total': total,
            'present': present,
            'absent': absent,
            'late': late,
            'rate': rate,
        },
        'attendance_data': attendance_data,
        'date_from': date_from,
        'date_to': date_to,
    }

    # Return as partial template for HTMX
    if request.headers.get('HX-Request'):
        return render(request, 'academics/partials/student_attendance_detail.html', context)

    return JsonResponse({
        'student': {
            'id': str(student.pk),
            'name': str(student),
            'admission_number': student.admission_number,
            'class': student.current_class.name if student.current_class else None,
        },
        'stats': context['stats'],
        'records': attendance_data,
    })


@login_required
@teacher_or_admin_required
def notify_absent_parents(request):
    """Send SMS notifications to parents of students with consecutive absences."""
    from core.models import SchoolSettings

    if request.method != 'POST':
        return redirect('academics:attendance_reports')

    student_ids = request.POST.getlist('student_ids')
    message_template = request.POST.get('message', '')

    if not student_ids:
        messages.warning(request, "No students selected for notification.")
        return redirect('academics:attendance_reports')

    if not message_template:
        messages.error(request, "Message template is required.")
        return redirect('academics:attendance_reports')

    # Get school settings for school name
    school_settings = SchoolSettings.load()
    school_name = school_settings.display_name if school_settings else ''

    # Get students with their attendance stats
    students = Student.objects.filter(pk__in=student_ids, guardian_phone__isnull=False)
    sent_count = 0
    failed_count = 0

    for student in students:
        # Calculate consecutive absences for this student
        recent_records = AttendanceRecord.objects.filter(
            student=student
        ).select_related('session').order_by('-session__date')[:30]

        consecutive_days = 0
        for record in recent_records:
            if record.status == 'A':
                consecutive_days += 1
            else:
                break

        # Render the message with placeholders
        message = message_template.replace('{student_name}', student.first_name)
        message = message.replace('{full_name}', str(student))
        message = message.replace('{days}', str(consecutive_days))
        message = message.replace('{class_name}', student.current_class.name if student.current_class else '')
        message = message.replace('{school_name}', school_name)

        # Send SMS
        try:
            from communications.utils import send_sms
            result = send_sms(
                to_phone=student.guardian_phone,
                message=message,
                student=student,
                message_type='attendance',
                created_by=request.user
            )
            if result.get('success'):
                sent_count += 1
            else:
                failed_count += 1
        except Exception:
            failed_count += 1

    if sent_count > 0:
        messages.success(request, f"Successfully sent {sent_count} notification(s) to parents.")
    if failed_count > 0:
        messages.warning(request, f"{failed_count} notification(s) failed to send.")

    return redirect('academics:attendance_reports')
