import logging
import json
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import connection, transaction
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate
from django.contrib import messages
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils.html import escape
from core.email_backend import get_from_email
from core.utils import cache_page_per_tenant, admin_required, htmx_render

from .models import (
    PaymentGateway, PaymentGatewayConfig, PaymentGatewayTransaction,
    FeeStructure, CATEGORY_CHOICES,
    Scholarship, StudentScholarship, Invoice, InvoiceItem, Payment
)
from .forms import (
    FeeStructureForm, ScholarshipForm, StudentScholarshipForm,
    InvoiceGenerateForm, PaymentForm, GatewayConfigForm
)
from students.models import Student
from academics.models import Class
from core.models import AcademicYear, Term

logger = logging.getLogger(__name__)


# =============================================================================
# DASHBOARD
# =============================================================================

@admin_required
@cache_page_per_tenant(timeout=300)  # Cache for 5 minutes
def index(request):
    """Finance dashboard with summary statistics. Cached for 5 minutes."""
    current_year = AcademicYear.get_current()
    current_term = Term.get_current() if hasattr(Term, 'get_current') else None

    # Calculate statistics
    total_invoiced = Invoice.objects.filter(
        academic_year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')

    total_collected = Payment.objects.filter(
        invoice__academic_year=current_year,
        status='COMPLETED'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_outstanding = Invoice.objects.filter(
        academic_year=current_year,
        status__in=['ISSUED', 'PARTIALLY_PAID', 'OVERDUE']
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00')

    # Recent payments
    recent_payments = Payment.objects.filter(
        status='COMPLETED'
    ).select_related('invoice__student').order_by('-created_at')[:10]

    # Overdue invoices
    overdue_invoices = Invoice.objects.filter(
        status='OVERDUE'
    ).select_related('student').order_by('due_date')[:10]

    # Collection by method
    collection_by_method = Payment.objects.filter(
        invoice__academic_year=current_year,
        status='COMPLETED'
    ).values('method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    # Students with outstanding balance
    students_with_balance = Invoice.objects.filter(
        academic_year=current_year,
        balance__gt=0
    ).values('student').annotate(
        total_balance=Sum('balance')
    ).count()

    context = {
        'current_year': current_year,
        'current_term': current_term,
        'total_invoiced': total_invoiced,
        'total_collected': total_collected,
        'total_outstanding': total_outstanding,
        'collection_rate': (total_collected / total_invoiced * 100) if total_invoiced > 0 else 0,
        'recent_payments': recent_payments,
        'overdue_invoices': overdue_invoices,
        'collection_by_method': collection_by_method,
        'students_with_balance': students_with_balance,
        # Navigation
        'breadcrumbs': [
            {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
            {'label': 'Finance'},
        ],
    }

    return htmx_render(
        request,
        'finance/index.html',
        'finance/partials/index_content.html',
        context
    )


# =============================================================================
# FEE STRUCTURES
# =============================================================================

@admin_required
def fee_structures(request):
    """List fee structures with filtering."""
    current_year = AcademicYear.get_current()

    structures = FeeStructure.objects.select_related(
        'class_assigned', 'programme', 'academic_year', 'term'
    ).order_by('-academic_year__start_date', 'category')

    # Filters
    year_filter = request.GET.get('year')
    if year_filter:
        structures = structures.filter(academic_year_id=year_filter)
    else:
        structures = structures.filter(academic_year=current_year)

    category_filter = request.GET.get('category')
    if category_filter:
        structures = structures.filter(category=category_filter)

    # Stats
    total_count = structures.count()
    active_count = structures.filter(is_active=True).count()
    mandatory_count = structures.filter(is_mandatory=True).count()
    category_count = structures.values('category').distinct().count()

    context = {
        'structures': structures,
        'current_year': current_year,
        'academic_years': AcademicYear.objects.all().order_by('-start_date'),
        'categories': CATEGORY_CHOICES,
        'form': FeeStructureForm(),
        'total_count': total_count,
        'active_count': active_count,
        'mandatory_count': mandatory_count,
        'category_count': category_count,
        # Navigation
        'breadcrumbs': [
            {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
            {'label': 'Finance', 'url': '/finance/'},
            {'label': 'Fee Structures'},
        ],
        'back_url': '/finance/',
    }

    return htmx_render(
        request,
        'finance/fee_structures.html',
        'finance/partials/fee_structures_content.html',
        context
    )


@admin_required
def fee_structure_create(request):
    """Create a new fee structure."""
    if request.method == 'POST':
        form = FeeStructureForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fee structure created successfully.')
            # Return the refreshed list for HTMX
            if request.htmx:
                return redirect('finance:fee_structures')
            return redirect('finance:fee_structures')
    else:
        form = FeeStructureForm()

    # For modal, just return the partial
    if request.htmx:
        return render(request, 'finance/partials/fee_structure_form_content.html', {'form': form})

    return render(request, 'finance/fee_structure_form.html', {'form': form})


@admin_required
def fee_structure_edit(request, pk):
    """Edit a fee structure."""
    structure = get_object_or_404(FeeStructure, pk=pk)

    if request.method == 'POST':
        form = FeeStructureForm(request.POST, instance=structure)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fee structure updated successfully.')
            return redirect('finance:fee_structures')
    else:
        form = FeeStructureForm(instance=structure)

    # For modal, just return the partial
    if request.htmx:
        return render(request, 'finance/partials/fee_structure_form_content.html', {'form': form, 'structure': structure})

    return render(request, 'finance/fee_structure_form.html', {'form': form, 'structure': structure})


@admin_required
def fee_structure_delete(request, pk):
    """Delete a fee structure."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    structure = get_object_or_404(FeeStructure, pk=pk)
    structure.delete()
    messages.success(request, 'Fee structure deleted successfully.')

    if request.htmx:
        # Return refreshed fee structures list
        current_year = AcademicYear.get_current()
        structures = FeeStructure.objects.select_related(
            'class_assigned', 'programme', 'academic_year', 'term'
        ).filter(academic_year=current_year).order_by('-academic_year__start_date', 'category')

        return render(request, 'finance/partials/fee_structures_content.html', {
            'structures': structures,
            'current_year': current_year,
            'academic_years': AcademicYear.objects.all().order_by('-start_date'),
            'categories': CATEGORY_CHOICES,
        })

    return redirect('finance:fee_structures')


# =============================================================================
# SCHOLARSHIPS
# =============================================================================

@admin_required
def scholarships(request):
    """List all scholarships."""
    scholarships_list = Scholarship.objects.annotate(
        recipient_count=Count('recipients', filter=Q(recipients__is_active=True))
    ).order_by('name')

    # Stats for the dashboard
    total_recipients = StudentScholarship.objects.filter(is_active=True).count()
    percentage_count = Scholarship.objects.filter(discount_type='PERCENTAGE', is_active=True).count()
    full_count = Scholarship.objects.filter(discount_type='FULL', is_active=True).count()

    context = {
        'scholarships': scholarships_list,
        'form': ScholarshipForm(),
        'total_recipients': total_recipients,
        'percentage_count': percentage_count,
        'full_count': full_count,
    }

    return htmx_render(
        request,
        'finance/scholarships.html',
        'finance/partials/scholarships_content.html',
        context
    )


@admin_required
def scholarship_create(request):
    """Create a new scholarship."""
    if request.method == 'POST':
        form = ScholarshipForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Scholarship created successfully.')
            return redirect('finance:scholarships')
    else:
        form = ScholarshipForm()

    # For modal, just return the partial
    if request.htmx:
        return render(request, 'finance/partials/scholarship_form_content.html', {'form': form})

    return render(request, 'finance/scholarship_form.html', {'form': form})


@admin_required
def scholarship_edit(request, pk):
    """Edit a scholarship."""
    scholarship = get_object_or_404(Scholarship, pk=pk)

    if request.method == 'POST':
        form = ScholarshipForm(request.POST, instance=scholarship)
        if form.is_valid():
            form.save()
            messages.success(request, 'Scholarship updated successfully.')
            return redirect('finance:scholarships')
    else:
        form = ScholarshipForm(instance=scholarship)

    # For modal, just return the partial
    if request.htmx:
        return render(request, 'finance/partials/scholarship_form_content.html', {'form': form, 'scholarship': scholarship})

    return render(request, 'finance/scholarship_form.html', {'form': form, 'scholarship': scholarship})


@admin_required
def scholarship_delete(request, pk):
    """Delete a scholarship."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    scholarship = get_object_or_404(Scholarship, pk=pk)

    # Check for active recipients
    active_recipients = scholarship.recipients.filter(is_active=True).count()
    if active_recipients > 0:
        messages.error(
            request,
            f'Cannot delete "{scholarship.name}": It has {active_recipients} active recipient(s). '
            'Remove or deactivate those assignments first.'
        )
    else:
        scholarship.delete()
        messages.success(request, f'Scholarship "{scholarship.name}" deleted successfully.')

    if request.htmx:
        # Return refreshed scholarships list
        scholarships_list = Scholarship.objects.annotate(
            recipient_count=Count('recipients', filter=Q(recipients__is_active=True))
        ).order_by('name')

        return render(request, 'finance/partials/scholarships_content.html', {
            'scholarships': scholarships_list,
        })

    return redirect('finance:scholarships')


@admin_required
def scholarship_assign(request, pk):
    """Assign scholarship to students."""
    scholarship = get_object_or_404(Scholarship, pk=pk)
    current_year = AcademicYear.get_current()

    if request.method == 'POST':
        form = StudentScholarshipForm(request.POST)
        student_ids = request.POST.getlist('student_ids')
        if form.is_valid() and student_ids:
            from django.db import IntegrityError
            students = Student.objects.filter(pk__in=student_ids, status='active')
            assigned = []
            skipped = []
            for student in students:
                try:
                    StudentScholarship.objects.create(
                        student=student,
                        scholarship=scholarship,
                        academic_year=form.cleaned_data['academic_year'],
                        reason=form.cleaned_data['reason'],
                        start_date=form.cleaned_data['start_date'],
                        end_date=form.cleaned_data['end_date'],
                        approved_by=request.user,
                    )
                    assigned.append(student.full_name)
                except IntegrityError:
                    skipped.append(student.full_name)
            if assigned:
                messages.success(request, f'Scholarship assigned to {", ".join(assigned)}.')
            if skipped:
                messages.warning(request, f'Already assigned: {", ".join(skipped)}.')
            return redirect('finance:scholarship_assign', pk=scholarship.pk)
        elif not student_ids:
            messages.error(request, 'Please select at least one student.')
    else:
        form = StudentScholarshipForm(initial={
            'scholarship': scholarship,
            'academic_year': current_year
        })

    # Get current recipients
    recipients = StudentScholarship.objects.filter(
        scholarship=scholarship,
        is_active=True
    ).select_related('student', 'academic_year')

    context = {
        'scholarship': scholarship,
        'form': form,
        'recipients': recipients,
    }

    return htmx_render(
        request,
        'finance/scholarship_assign.html',
        'finance/partials/scholarship_assign_content.html',
        context
    )


# =============================================================================
# INVOICES
# =============================================================================

def _build_invoices_context(request):
    """Build the context dict for the invoices list page."""
    current_year = AcademicYear.get_current()
    has_fee_structures = FeeStructure.objects.filter(
        academic_year=current_year, is_active=True
    ).exists() if current_year else False

    invoices_list = Invoice.objects.select_related(
        'student', 'academic_year', 'term'
    ).order_by('-created_at')

    # Filters
    status_filter = request.GET.get('status')
    if status_filter:
        invoices_list = invoices_list.filter(status=status_filter)

    class_filter = request.GET.get('class')
    if class_filter:
        invoices_list = invoices_list.filter(student__current_class_id=class_filter)

    search = request.GET.get('search', '').strip()
    if search:
        invoices_list = invoices_list.filter(
            Q(invoice_number__icontains=search) |
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(student__admission_number__icontains=search)
        )

    # Pagination with selectable page size
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100]:
            per_page = 25
    except ValueError:
        per_page = 25

    paginator = Paginator(invoices_list, per_page)
    page = request.GET.get('page', 1)
    invoices_page = paginator.get_page(page)

    # Stats for dashboard
    all_invoices = Invoice.objects.all()
    total_count = all_invoices.count()
    paid_count = all_invoices.filter(status='PAID').count()
    pending_count = all_invoices.filter(status__in=['ISSUED', 'PARTIALLY_PAID']).count()
    overdue_count = all_invoices.filter(status='OVERDUE').count()
    draft_count = all_invoices.filter(status='DRAFT').count()

    return {
        'invoices': invoices_page,
        'page_obj': invoices_page,
        'paginator': paginator,
        'per_page': per_page,
        'current_year': current_year,
        'has_fee_structures': has_fee_structures,
        'status_choices': Invoice.STATUS_CHOICES,
        'classes': Class.objects.filter(is_active=True),
        'status_filter': status_filter,
        'class_filter': class_filter,
        'search': search,
        'total_count': total_count,
        'paid_count': paid_count,
        'pending_count': pending_count,
        'overdue_count': overdue_count,
        'draft_count': draft_count,
        # Navigation
        'breadcrumbs': [
            {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
            {'label': 'Finance', 'url': '/finance/'},
            {'label': 'Invoices'},
        ],
        'back_url': '/finance/',
    }


@admin_required
def invoices(request):
    """List all invoices with filtering."""
    context = _build_invoices_context(request)

    return htmx_render(
        request,
        'finance/invoices.html',
        'finance/partials/invoices_content.html',
        context
    )


@admin_required
def invoice_generate(request):
    """Generate invoices for students."""
    current_year = AcademicYear.get_current()
    current_term = Term.get_current() if hasattr(Term, 'get_current') else None

    if request.method == 'POST':
        form = InvoiceGenerateForm(request.POST)
        if form.is_valid():
            class_obj = form.cleaned_data.get('class_assigned')
            student = form.cleaned_data.get('student')
            term = form.cleaned_data['term']
            due_date = form.cleaned_data['due_date']

            # Get students to invoice
            if student:
                students = [student]
            elif class_obj:
                students = Student.objects.filter(
                    current_class=class_obj,
                    status='active'
                )
            else:
                if request.htmx:
                    response = HttpResponse(status=204)
                    response['HX-Trigger'] = json.dumps({
                        'showToast': {'message': 'Please select a class or student.', 'type': 'error'}
                    })
                    return response
                messages.error(request, 'Please select a class or student.')
                return redirect('finance:invoice_generate')

            # Check if fee structures exist before looping
            fee_structure_count = FeeStructure.objects.filter(
                academic_year=current_year,
                is_active=True,
            ).filter(
                Q(term=term) | Q(term__isnull=True)
            ).count()

            if fee_structure_count == 0:
                err = 'No active fee structures found for this academic year and term. Please create fee structures first.'
                if request.htmx:
                    response = HttpResponse(status=204)
                    response['HX-Trigger'] = json.dumps({
                        'showToast': {'message': err, 'type': 'error'},
                    })
                    return response
                messages.error(request, err)
                return redirect('finance:invoice_generate')

            invoices_created = 0
            notified_students = []
            for student in students:
                invoice = create_student_invoice(
                    student=student,
                    academic_year=current_year,
                    term=term,
                    due_date=due_date,
                    created_by=request.user
                )
                if invoice:
                    invoices_created += 1
                    notified_students.append(student)

            # Bell notifications for guardians
            if notified_students:
                from core.notifications import notify_guardian
                for s in notified_students:
                    notify_guardian(
                        s,
                        title='New Invoice Generated',
                        message=f'A new fee invoice has been generated for {s.full_name}.',
                        category='finance',
                        notification_type='info',
                        icon='fa-solid fa-file-invoice-dollar',
                        link=reverse('finance:fee_payments'),
                    )

            if invoices_created == 0:
                msg = 'No new invoices generated. Students may already have invoices for this term, or no fee structures match their class.'
                toast_type = 'warning'
            else:
                msg = f'{invoices_created} invoice(s) generated successfully.'
                toast_type = 'success'
            if request.htmx:
                invoices_ctx = _build_invoices_context(request)
                response = render(
                    request,
                    'finance/partials/invoices_content.html',
                    invoices_ctx,
                )
                response['HX-Trigger'] = json.dumps({
                    'closeModal': True,
                    'showToast': {'message': msg, 'type': toast_type},
                })
                response['HX-Push-Url'] = reverse('finance:invoices')
                return response
            messages.success(request, msg)
            return redirect('finance:invoices')
    else:
        form = InvoiceGenerateForm(initial={
            'academic_year': current_year,
            'term': current_term
        })

    # For modal, just return the partial
    if request.htmx:
        return render(request, 'finance/partials/invoice_generate_content.html', {'form': form})

    return render(request, 'finance/invoice_generate.html', {'form': form})


def create_student_invoice(student, academic_year, term, due_date, created_by):
    """Create an invoice for a student based on applicable fee structures."""
    # Get applicable fee structures (read-only, safe outside transaction)
    fee_structures = FeeStructure.objects.filter(
        academic_year=academic_year,
        is_active=True
    ).filter(
        Q(term=term) | Q(term__isnull=True)
    )

    # Filter by class or level
    applicable_structures = []
    for structure in fee_structures:
        if structure.class_assigned:
            if structure.class_assigned == student.current_class:
                applicable_structures.append(structure)
        elif structure.level_type:
            if student.current_class and student.current_class.level_type == structure.level_type:
                applicable_structures.append(structure)
        else:
            applicable_structures.append(structure)

    if not applicable_structures:
        return None

    with transaction.atomic():
        # Atomic duplicate check — prevents race condition
        existing = Invoice.objects.select_for_update().filter(
            student=student,
            academic_year=academic_year,
            term=term
        ).exists()

        if existing:
            return None

        # Create invoice
        invoice = Invoice.objects.create(
            student=student,
            academic_year=academic_year,
            term=term,
            due_date=due_date,
            created_by=created_by,
            status='DRAFT'
        )

        # Add line items
        subtotal = Decimal('0.00')
        for structure in applicable_structures:
            # Check if fee applies to student type (boarding/day)
            if hasattr(student, 'is_boarding'):
                if student.is_boarding and not structure.applies_to_boarding:
                    continue
                if not student.is_boarding and not structure.applies_to_day:
                    continue

            InvoiceItem.objects.create(
                invoice=invoice,
                category=structure.category,
                description=structure.get_description(),
                amount=structure.amount
            )
            subtotal += structure.amount

        # Apply scholarships
        discount = Decimal('0.00')
        student_scholarships = StudentScholarship.objects.filter(
            student=student,
            academic_year=academic_year,
            is_active=True
        ).select_related('scholarship')

        # Build per-category subtotals for category-specific scholarships
        category_subtotals = {}
        for structure in applicable_structures:
            cat = structure.category
            category_subtotals[cat] = category_subtotals.get(cat, Decimal('0.00')) + structure.amount

        for ss in student_scholarships:
            scholarship = ss.scholarship
            # Determine the base amount this scholarship applies to
            if scholarship.applies_to_categories:
                applicable_amount = sum(
                    category_subtotals.get(cat, Decimal('0.00'))
                    for cat in scholarship.applies_to_categories
                )
            else:
                applicable_amount = subtotal

            if scholarship.discount_type == 'FULL':
                discount += applicable_amount
            elif scholarship.discount_type == 'PERCENTAGE':
                discount += applicable_amount * (scholarship.discount_value / Decimal('100'))
            elif scholarship.discount_type == 'FIXED':
                discount += scholarship.discount_value

        # Update invoice totals
        invoice.subtotal = subtotal
        invoice.discount = min(discount, subtotal)  # Can't discount more than subtotal
        invoice.total_amount = subtotal - invoice.discount
        invoice.balance = invoice.total_amount
        invoice.save()

    return invoice


@admin_required
def invoice_detail(request, pk):
    """View invoice details."""
    invoice = get_object_or_404(
        Invoice.objects.select_related(
            'student', 'academic_year', 'term', 'created_by'
        ).prefetch_related('items', 'payments'),
        pk=pk
    )

    items = invoice.items.all()
    payments = invoice.payments.all().order_by('-created_at')

    # Check if online payment gateway is available
    gateway_available = PaymentGatewayConfig.objects.filter(
        is_active=True,
        is_primary=True,
        verification_status='VERIFIED'
    ).exists()

    # Get last notification
    last_notification = invoice.notification_logs.first()

    context = {
        'invoice': invoice,
        'items': items,
        'payments': payments,
        'gateway_available': gateway_available,
        'last_notification': last_notification,
    }

    return htmx_render(
        request,
        'finance/invoice_detail.html',
        'finance/partials/invoice_detail_content.html',
        context
    )


@admin_required
def invoice_edit(request, pk):
    """Edit invoice (only draft invoices) - handles issue action."""
    if request.method != 'POST':
        return redirect('finance:invoice_detail', pk=pk)

    invoice = get_object_or_404(Invoice, pk=pk)

    if invoice.status != 'DRAFT':
        messages.error(request, 'Only draft invoices can be edited.')
        return redirect('finance:invoice_detail', pk=pk)

    action = request.POST.get('action')
    if action == 'issue':
        invoice.status = 'ISSUED'
        invoice.issue_date = timezone.now().date()
        invoice.save()
        messages.success(request, 'Invoice issued successfully.')

    return redirect('finance:invoice_detail', pk=pk)


@admin_required
def invoice_bulk_issue(request):
    """Issue draft invoices with optional filters."""
    if request.method == 'POST':
        scope = request.POST.get('scope', 'all')
        drafts = Invoice.objects.filter(status='DRAFT')

        if scope == 'class':
            class_id = request.POST.get('class_id')
            if class_id:
                drafts = drafts.filter(student__current_class_id=class_id)
        elif scope == 'student':
            student_id = request.POST.get('student_id')
            if student_id:
                drafts = drafts.filter(student_id=student_id)

        updated = drafts.update(
            status='ISSUED',
            issue_date=timezone.now().date(),
        )
        messages.success(request, f'{updated} invoice{"s" if updated != 1 else ""} issued successfully.')
        return redirect('finance:invoices')

    # GET: render the modal content
    draft_count = Invoice.objects.filter(status='DRAFT').count()
    classes = Class.objects.filter(is_active=True).order_by('name')
    # Count drafts per class
    class_draft_counts = dict(
        Invoice.objects.filter(status='DRAFT')
        .values_list('student__current_class_id')
        .annotate(count=Count('id'))
        .values_list('student__current_class_id', 'count')
    )
    for cls in classes:
        cls.draft_count = class_draft_counts.get(cls.pk, 0)

    return render(request, 'finance/partials/bulk_issue_content.html', {
        'draft_count': draft_count,
        'classes': classes,
    })


@admin_required
def invoice_cancel(request, pk):
    """Cancel an invoice."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    invoice = get_object_or_404(Invoice, pk=pk)

    if invoice.status == 'PAID':
        messages.error(request, 'Cannot cancel a paid invoice.')
        return redirect('finance:invoice_detail', pk=pk)

    # Check for completed payments — can't cancel with payments made
    completed_payments = invoice.payments.filter(status='COMPLETED').exists()
    if completed_payments:
        messages.error(
            request,
            'Cannot cancel an invoice with completed payments. Refund payments first.'
        )
        return redirect('finance:invoice_detail', pk=pk)

    # Cancel any pending payments associated with this invoice
    invoice.payments.filter(status='PENDING').update(status='CANCELLED')

    invoice.status = 'CANCELLED'
    invoice.save()
    messages.success(request, 'Invoice cancelled successfully.')

    if request.htmx:
        # Return to invoices list
        return redirect('finance:invoices')

    return redirect('finance:invoices')


@admin_required
def invoice_print(request, pk):
    """Generate invoice PDF using WeasyPrint."""
    invoice = get_object_or_404(
        Invoice.objects.select_related('student', 'academic_year', 'term'),
        pk=pk
    )

    school = connection.tenant

    # Encode logo as base64 for PDF
    logo_base64 = None
    try:
        from gradebook.utils import encode_logo_base64
        if school and school.logo:
            logo_base64 = encode_logo_base64(school.logo, connection.schema_name)
    except Exception:
        logger.debug("Could not encode logo for invoice %s", pk)

    # Create verification record and generate QR code
    from core.models import DocumentVerification
    from core.utils import generate_verification_qr
    verification = None
    qr_code_base64 = None
    try:
        verification = DocumentVerification.create_for_document(
            document_type=DocumentVerification.DocumentType.INVOICE,
            student=invoice.student,
            title=f"Invoice #{invoice.invoice_number}",
            user=request.user,
            term=invoice.term,
            academic_year=invoice.academic_year.name if invoice.academic_year else '',
        )
        qr_code_base64 = generate_verification_qr(verification.verification_code, request=request)
    except Exception:
        logger.warning("Could not create verification record for invoice %s", pk)

    context = {
        'invoice': invoice,
        'items': invoice.items.all(),
        'school': school,
        'logo_base64': logo_base64,
        'verification': verification,
        'qr_code_base64': qr_code_base64,
    }

    try:
        from weasyprint import HTML
        from django.template.loader import render_to_string
        from django.conf import settings as django_settings
        from io import BytesIO

        html_string = render_to_string('finance/invoice_print.html', context)
        html = HTML(string=html_string, base_url=str(django_settings.BASE_DIR))
        pdf_buffer = BytesIO()
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="invoice_{invoice.invoice_number}.pdf"'
        return response

    except ImportError:
        logger.error("WeasyPrint not installed")
        messages.error(request, 'PDF generation is not available. WeasyPrint is not installed.')
        return redirect('finance:invoice_detail', pk=pk)
    except Exception as e:
        logger.error("Failed to generate invoice PDF: %s", e, exc_info=True)
        messages.error(request, 'Failed to generate PDF. Please try again.')
        return redirect('finance:invoice_detail', pk=pk)


# =============================================================================
# PAYMENTS
# =============================================================================

@admin_required
def payments(request):
    """List all payments with filtering."""
    payments_list = Payment.objects.select_related(
        'invoice__student', 'received_by'
    ).order_by('-created_at')

    # Filters
    status_filter = request.GET.get('status')
    if status_filter:
        payments_list = payments_list.filter(status=status_filter)

    method_filter = request.GET.get('method')
    if method_filter:
        payments_list = payments_list.filter(method=method_filter)

    search = request.GET.get('search', '').strip()
    if search:
        payments_list = payments_list.filter(
            Q(receipt_number__icontains=search) |
            Q(invoice__invoice_number__icontains=search) |
            Q(invoice__student__first_name__icontains=search) |
            Q(invoice__student__last_name__icontains=search)
        )

    # Date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        payments_list = payments_list.filter(transaction_date__date__gte=date_from)
    if date_to:
        payments_list = payments_list.filter(transaction_date__date__lte=date_to)

    # Pagination with selectable page size
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100]:
            per_page = 25
    except ValueError:
        per_page = 25

    paginator = Paginator(payments_list, per_page)
    page = request.GET.get('page', 1)
    payments_page = paginator.get_page(page)

    # Stats for dashboard
    all_payments = Payment.objects.filter(status='COMPLETED')
    total_count = all_payments.count()
    total_amount = all_payments.aggregate(total=Sum('amount'))['total'] or 0
    momo_count = all_payments.filter(method='MOBILE_MONEY').count()
    cash_count = all_payments.filter(method='CASH').count()

    context = {
        'payments': payments_page,
        'page_obj': payments_page,
        'paginator': paginator,
        'per_page': per_page,
        'status_choices': Payment.STATUS_CHOICES,
        'method_choices': Payment.METHOD_CHOICES,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'search': search,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'total_count': total_count,
        'total_amount': total_amount,
        'momo_count': momo_count,
        'cash_count': cash_count,
        # Navigation
        'breadcrumbs': [
            {'label': 'Home', 'url': '/', 'icon': 'fa-solid fa-home'},
            {'label': 'Finance', 'url': '/finance/'},
            {'label': 'Payments'},
        ],
        'back_url': '/finance/',
    }

    return htmx_render(
        request,
        'finance/payments.html',
        'finance/partials/payments_content.html',
        context
    )


@admin_required
def payment_record(request):
    """Record a manual payment."""
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                payment = form.save(commit=False)
                payment.received_by = request.user
                payment.status = 'COMPLETED'
                payment.save()

            # Bell notification for guardian
            if payment.invoice and payment.invoice.student:
                from core.notifications import notify_guardian
                student = payment.invoice.student
                notify_guardian(
                    student,
                    title='Payment Confirmed',
                    message=f'Payment of {payment.amount} received for {student.full_name}.',
                    category='finance',
                    notification_type='success',
                    icon='fa-solid fa-circle-check',
                )

            # Send payment confirmation SMS to guardian
            from .tasks import send_payment_confirmation_sms
            send_payment_confirmation_sms.delay(str(payment.pk), connection.schema_name)

            messages.success(request, f'Payment recorded successfully. Receipt: {payment.receipt_number}')
            return redirect('finance:payments')
    else:
        form = PaymentForm()

    # For modal, just return the partial
    if request.htmx:
        return render(request, 'finance/partials/payment_record_content.html', {'form': form})

    return render(request, 'finance/payment_record.html', {'form': form})


@admin_required
def payment_detail(request, pk):
    """View payment details."""
    payment = get_object_or_404(
        Payment.objects.select_related(
            'invoice__student', 'invoice__academic_year', 'invoice__term', 'received_by'
        ),
        pk=pk
    )

    context = {
        'payment': payment,
    }

    return htmx_render(
        request,
        'finance/payment_detail.html',
        'finance/partials/payment_detail_content.html',
        context
    )


@admin_required
def payment_receipt(request, pk):
    """Generate payment receipt PDF using WeasyPrint."""
    payment = get_object_or_404(
        Payment.objects.select_related('invoice__student', 'invoice__academic_year', 'invoice__term', 'received_by'),
        pk=pk
    )

    school = connection.tenant

    # Encode logo as base64 for PDF
    logo_base64 = None
    try:
        from gradebook.utils import encode_logo_base64
        if school and school.logo:
            logo_base64 = encode_logo_base64(school.logo, connection.schema_name)
    except Exception:
        logger.debug("Could not encode logo for receipt %s", pk)

    # Create verification record and generate QR code
    from core.models import DocumentVerification
    from core.utils import generate_verification_qr
    verification = None
    qr_code_base64 = None
    try:
        verification = DocumentVerification.create_for_document(
            document_type=DocumentVerification.DocumentType.RECEIPT,
            student=payment.invoice.student,
            title=f"Receipt #{payment.receipt_number}",
            user=request.user,
            term=payment.invoice.term,
            academic_year=payment.invoice.academic_year.name if payment.invoice.academic_year else '',
        )
        qr_code_base64 = generate_verification_qr(verification.verification_code, request=request)
    except Exception:
        logger.warning("Could not create verification record for receipt %s", pk)

    context = {
        'payment': payment,
        'school': school,
        'logo_base64': logo_base64,
        'verification': verification,
        'qr_code_base64': qr_code_base64,
    }

    try:
        from weasyprint import HTML
        from django.template.loader import render_to_string
        from django.conf import settings as django_settings
        from io import BytesIO

        html_string = render_to_string('finance/payment_receipt.html', context)
        html = HTML(string=html_string, base_url=str(django_settings.BASE_DIR))
        pdf_buffer = BytesIO()
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)

        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="receipt_{payment.receipt_number}.pdf"'
        return response

    except ImportError:
        logger.error("WeasyPrint not installed")
        messages.error(request, 'PDF generation is not available. WeasyPrint is not installed.')
        return redirect('finance:payment_detail', pk=pk)
    except Exception as e:
        logger.error("Failed to generate receipt PDF: %s", e, exc_info=True)
        messages.error(request, 'Failed to generate PDF. Please try again.')
        return redirect('finance:payment_detail', pk=pk)


# =============================================================================
# STUDENT FEES
# =============================================================================

@admin_required
def student_fees(request, student_id):
    """View a student's fee summary."""
    student = get_object_or_404(Student, pk=student_id)

    invoices = Invoice.objects.filter(
        student=student
    ).select_related('academic_year', 'term').order_by('-created_at')

    payments = Payment.objects.filter(
        invoice__student=student,
        status='COMPLETED'
    ).select_related('invoice').order_by('-created_at')

    # Calculate totals
    total_invoiced = invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_paid = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_balance = invoices.filter(
        status__in=['ISSUED', 'PARTIALLY_PAID', 'OVERDUE']
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00')

    # Get scholarships
    scholarships = StudentScholarship.objects.filter(
        student=student,
        is_active=True
    ).select_related('scholarship', 'academic_year')

    context = {
        'student': student,
        'invoices': invoices,
        'payments': payments,
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'total_balance': total_balance,
        'scholarships': scholarships,
    }

    return htmx_render(
        request,
        'finance/student_fees.html',
        'finance/partials/student_fees_content.html',
        context
    )


@admin_required
def student_statement(request, student_id):
    """Generate a fee statement for a student."""
    student = get_object_or_404(Student, pk=student_id)

    invoices = Invoice.objects.filter(
        student=student
    ).select_related('academic_year', 'term').order_by('created_at')

    payments = Payment.objects.filter(
        invoice__student=student,
        status='COMPLETED'
    ).select_related('invoice').order_by('transaction_date')

    # Build statement entries
    entries = []
    balance = Decimal('0.00')

    for invoice in invoices:
        balance += invoice.total_amount
        entries.append({
            'date': invoice.issue_date,
            'description': f'Invoice {invoice.invoice_number}',
            'type': 'invoice',
            'debit': invoice.total_amount,
            'credit': None,
            'balance': balance,
        })

    for payment in payments:
        balance -= payment.amount
        entries.append({
            'date': payment.transaction_date.date(),
            'description': f'Payment {payment.receipt_number}',
            'type': 'payment',
            'debit': None,
            'credit': payment.amount,
            'balance': balance,
        })

    # Sort by date
    entries.sort(key=lambda x: x['date'])

    context = {
        'student': student,
        'entries': entries,
        'final_balance': balance,
    }

    return render(request, 'finance/student_statement.html', context)


# =============================================================================
# REPORTS
# =============================================================================

@admin_required
def reports(request):
    """Finance reports dashboard."""
    context = {}
    return htmx_render(
        request,
        'finance/reports.html',
        'finance/partials/reports_content.html',
        context
    )


@admin_required
def collection_report(request):
    """Fee collection report."""
    # Get date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    payments = Payment.objects.filter(status='COMPLETED')

    if date_from:
        payments = payments.filter(transaction_date__date__gte=date_from)
    if date_to:
        payments = payments.filter(transaction_date__date__lte=date_to)

    # Summary by method
    by_method = payments.values('method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    # Summary by day (using TruncDate for database-agnostic date grouping)
    by_day = payments.annotate(
        date=TruncDate('transaction_date')
    ).values('date').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-date')[:30]

    # Summary by class
    by_class = payments.values(
        'invoice__student__current_class__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    total_collected = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    context = {
        'by_method': by_method,
        'by_day': by_day,
        'by_class': by_class,
        'total_collected': total_collected,
        'date_from': date_from,
        'date_to': date_to,
    }

    return htmx_render(
        request,
        'finance/collection_report.html',
        'finance/partials/collection_report_content.html',
        context
    )


@admin_required
def outstanding_report(request):
    """Outstanding fees report."""
    current_year = AcademicYear.get_current()

    # Get students with outstanding balance
    students_with_balance = Invoice.objects.filter(
        academic_year=current_year,
        balance__gt=0
    ).values(
        'student__id',
        'student__first_name',
        'student__last_name',
        'student__admission_number',
        'student__current_class__name'
    ).annotate(
        total_invoiced=Sum('total_amount'),
        total_paid=Sum('amount_paid'),
        total_balance=Sum('balance')
    ).order_by('-total_balance')

    # Summary by class
    by_class = Invoice.objects.filter(
        academic_year=current_year,
        balance__gt=0
    ).values('student__current_class__name').annotate(
        student_count=Count('student', distinct=True),
        total_balance=Sum('balance')
    ).order_by('-total_balance')

    total_outstanding = Invoice.objects.filter(
        academic_year=current_year,
        balance__gt=0
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00')

    context = {
        'students_with_balance': students_with_balance,
        'by_class': by_class,
        'total_outstanding': total_outstanding,
        'current_year': current_year,
    }

    return htmx_render(
        request,
        'finance/outstanding_report.html',
        'finance/partials/outstanding_report_content.html',
        context
    )


@admin_required
def export_report(request):
    """Export report data to CSV/Excel."""
    import csv
    from django.http import HttpResponse

    report_type = request.GET.get('type', 'collection')

    MAX_EXPORT_ROWS = 10_000

    if report_type == 'collection':
        payments = Payment.objects.filter(
            status='COMPLETED'
        ).select_related('invoice__student').order_by('-transaction_date')[:MAX_EXPORT_ROWS]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="collection_report.csv"'

        writer = csv.writer(response)
        writer.writerow(['Receipt', 'Date', 'Student', 'Invoice', 'Amount', 'Method'])

        for payment in payments:
            writer.writerow([
                payment.receipt_number,
                payment.transaction_date.strftime('%Y-%m-%d'),
                payment.invoice.student.full_name,
                payment.invoice.invoice_number,
                payment.amount,
                payment.get_method_display()
            ])

        return response

    elif report_type == 'outstanding':
        invoices = Invoice.objects.filter(
            balance__gt=0
        ).select_related('student', 'student__current_class').order_by('student__last_name')[:MAX_EXPORT_ROWS]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="outstanding_report.csv"'

        writer = csv.writer(response)
        writer.writerow(['Student', 'Admission No', 'Class', 'Invoice', 'Total', 'Paid', 'Balance'])

        for invoice in invoices:
            writer.writerow([
                invoice.student.full_name,
                invoice.student.admission_number,
                invoice.student.current_class.name if invoice.student.current_class else '-',
                invoice.invoice_number,
                invoice.total_amount,
                invoice.amount_paid,
                invoice.balance
            ])

        return response

    return redirect('finance:reports')


@admin_required
def invoices_export(request):
    """Export invoices to Excel with current filters applied."""
    import io
    from datetime import datetime
    from django.http import FileResponse
    import pandas as pd

    # Get filter parameters (same as invoices view)
    status_filter = request.GET.get('status')
    class_filter = request.GET.get('class')
    search = request.GET.get('search', '').strip()

    invoices_list = Invoice.objects.select_related(
        'student', 'student__current_class', 'academic_year', 'term'
    ).order_by('-created_at')

    if status_filter:
        invoices_list = invoices_list.filter(status=status_filter)
    if class_filter:
        invoices_list = invoices_list.filter(student__current_class_id=class_filter)
    if search:
        invoices_list = invoices_list.filter(
            Q(invoice_number__icontains=search) |
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(student__admission_number__icontains=search)
        )

    # Build export data (capped at 10,000 rows to prevent memory issues)
    MAX_EXPORT_ROWS = 10_000
    export_data = []
    for inv in invoices_list[:MAX_EXPORT_ROWS].iterator():
        export_data.append({
            'Invoice Number': inv.invoice_number,
            'Student Name': inv.student.full_name,
            'Admission Number': inv.student.admission_number,
            'Class': inv.student.current_class.name if inv.student.current_class else '',
            'Academic Year': inv.academic_year.name if inv.academic_year else '',
            'Term': inv.term.name if inv.term else '',
            'Issue Date': inv.issue_date.strftime('%Y-%m-%d') if inv.issue_date else '',
            'Due Date': inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '',
            'Subtotal': float(inv.subtotal),
            'Discount': float(inv.discount),
            'Total Amount': float(inv.total_amount),
            'Amount Paid': float(inv.amount_paid),
            'Balance': float(inv.balance),
            'Status': inv.get_status_display(),
        })

    df = pd.DataFrame(export_data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Invoices')

        from openpyxl.utils import get_column_letter
        worksheet = writer.sheets['Invoices']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(col)
            ) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)

    output.seek(0)
    filename = f"invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@admin_required
def payments_export(request):
    """Export payments to Excel with current filters applied."""
    import io
    from datetime import datetime
    from django.http import FileResponse
    import pandas as pd

    # Get filter parameters (same as payments view)
    status_filter = request.GET.get('status')
    method_filter = request.GET.get('method')
    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    payments_list = Payment.objects.select_related(
        'invoice__student', 'invoice__student__current_class', 'received_by'
    ).order_by('-created_at')

    if status_filter:
        payments_list = payments_list.filter(status=status_filter)
    if method_filter:
        payments_list = payments_list.filter(method=method_filter)
    if search:
        payments_list = payments_list.filter(
            Q(receipt_number__icontains=search) |
            Q(invoice__invoice_number__icontains=search) |
            Q(invoice__student__first_name__icontains=search) |
            Q(invoice__student__last_name__icontains=search)
        )
    if date_from:
        payments_list = payments_list.filter(transaction_date__date__gte=date_from)
    if date_to:
        payments_list = payments_list.filter(transaction_date__date__lte=date_to)

    # Build export data (capped at 10,000 rows to prevent memory issues)
    MAX_EXPORT_ROWS = 10_000
    export_data = []
    for pmt in payments_list[:MAX_EXPORT_ROWS].iterator():
        export_data.append({
            'Receipt Number': pmt.receipt_number,
            'Invoice Number': pmt.invoice.invoice_number,
            'Student Name': pmt.invoice.student.full_name,
            'Admission Number': pmt.invoice.student.admission_number,
            'Class': pmt.invoice.student.current_class.name if pmt.invoice.student.current_class else '',
            'Amount': float(pmt.amount),
            'Payment Method': pmt.get_method_display(),
            'Status': pmt.get_status_display(),
            'Transaction Date': pmt.transaction_date.strftime('%Y-%m-%d %H:%M') if pmt.transaction_date else '',
            'Reference': pmt.reference or '',
            'Payer Name': pmt.payer_name or '',
            'Payer Phone': pmt.payer_phone or '',
            'Received By': pmt.received_by.get_full_name() if pmt.received_by else '',
        })

    df = pd.DataFrame(export_data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Payments')

        from openpyxl.utils import get_column_letter
        worksheet = writer.sheets['Payments']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(col)
            ) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)

    output.seek(0)
    filename = f"payments_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# =============================================================================
# PAYMENT GATEWAY SETTINGS
# =============================================================================

@admin_required
def gateway_settings(request):
    """Payment gateway configuration page."""
    gateways = PaymentGateway.objects.all()
    configs = PaymentGatewayConfig.objects.select_related('gateway').all()

    context = {
        'gateways': gateways,
        'configs': configs,
    }

    return htmx_render(
        request,
        'finance/gateway_settings.html',
        'finance/partials/gateway_settings_content.html',
        context
    )


@admin_required
def gateway_configure(request, pk):
    """Configure a payment gateway."""
    gateway = get_object_or_404(PaymentGateway, pk=pk)

    # Get or create config for this gateway
    config, created = PaymentGatewayConfig.objects.get_or_create(
        gateway=gateway,
        defaults={'configured_by': request.user}
    )

    if request.method == 'POST':
        form = GatewayConfigForm(request.POST, instance=config)
        if form.is_valid():
            config = form.save(commit=False)
            config.configured_by = request.user
            config.configured_at = timezone.now()
            config.verification_status = 'PENDING'
            config.save()
            messages.success(request, f'{gateway.display_name} configuration saved.')
            return redirect('finance:gateway_settings')
    else:
        form = GatewayConfigForm(instance=config)

    context = {
        'gateway': gateway,
        'config': config,
        'form': form,
    }

    return htmx_render(
        request,
        'finance/gateway_configure.html',
        'finance/partials/gateway_configure_content.html',
        context
    )


@admin_required
def gateway_verify(request, pk):
    """Verify gateway credentials."""
    config = get_object_or_404(PaymentGatewayConfig, pk=pk)

    # Import gateway adapter
    from .gateways import get_gateway_adapter

    try:
        adapter = get_gateway_adapter(config)
        is_valid, message = adapter.verify_credentials()

        if is_valid:
            config.verification_status = 'VERIFIED'
            config.verification_error = ''
            config.last_verified = timezone.now()
            messages.success(request, f'{config.gateway.display_name} credentials verified successfully.')
        else:
            config.verification_status = 'FAILED'
            config.verification_error = message
            messages.error(request, f'Verification failed: {message}')

        config.save()

    except Exception as e:
        config.verification_status = 'FAILED'
        config.verification_error = str(e)
        config.save()
        messages.error(request, f'Verification error: {str(e)}')

    return redirect('finance:gateway_settings')


@admin_required
def gateway_test_credentials(request, pk):
    """Test gateway credentials using form values (without saving)."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    gateway = get_object_or_404(PaymentGateway, pk=pk)

    # Auto-dismiss script for test results
    auto_dismiss = '''<script>(() => {
        const el = document.getElementById("test-gateway-result");
        if(el) {
            el.style.opacity = "1";
            el.style.transition = "none";
            setTimeout(() => {
                el.style.transition = "opacity 0.5s";
                el.style.opacity = "0";
                setTimeout(() => el.innerHTML = "", 500);
            }, 5000);
        }
    })();</script>'''

    # Get credentials from form
    secret_key = request.POST.get('secret_key', '').strip()
    public_key = request.POST.get('public_key', '').strip()

    if not secret_key:
        return HttpResponse(
            f'<div class="alert alert-error text-sm"><i class="fa-solid fa-times-circle"></i> '
            f'Secret key is required.</div>{auto_dismiss}'
        )

    # Create a mock config object to test with
    from .models import PaymentGatewayConfig

    # Get or create a temporary config
    config, _ = PaymentGatewayConfig.objects.get_or_create(
        gateway=gateway,
        defaults={'configured_by': request.user}
    )

    # Temporarily set credentials for testing
    old_secret = config.secret_key
    old_public = config.public_key
    old_is_test = config.is_test_mode

    config.secret_key = secret_key
    config.public_key = public_key
    config.is_test_mode = request.POST.get('is_test_mode') == 'on'

    # Test the credentials
    from .gateways import get_gateway_adapter

    try:
        adapter = get_gateway_adapter(config)
        is_valid, message = adapter.verify_credentials()

        if is_valid:
            # Restore original values (don't save test values)
            config.secret_key = old_secret
            config.public_key = old_public
            config.is_test_mode = old_is_test
            return HttpResponse(
                f'<div class="alert alert-success text-sm"><i class="fa-solid fa-check-circle"></i> '
                f'Credentials verified successfully! You can now save the configuration.</div>{auto_dismiss}'
            )
        else:
            config.secret_key = old_secret
            config.public_key = old_public
            config.is_test_mode = old_is_test
            return HttpResponse(
                f'<div class="alert alert-error text-sm"><i class="fa-solid fa-times-circle"></i> '
                f'Verification failed: {message}</div>{auto_dismiss}'
            )

    except Exception as e:
        config.secret_key = old_secret
        config.public_key = old_public
        config.is_test_mode = old_is_test
        return HttpResponse(
            f'<div class="alert alert-error text-sm"><i class="fa-solid fa-times-circle"></i> '
            f'Error: {str(e)}</div>{auto_dismiss}'
        )


# =============================================================================
# API ENDPOINTS
# =============================================================================

@admin_required
def student_search(request):
    """Search students for HTMX autocomplete."""
    q = request.GET.get('q', '').strip()

    if len(q) < 2:
        return HttpResponse('')

    students = Student.objects.filter(
        status='active'
    ).filter(
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q) |
        Q(admission_number__icontains=q)
    ).select_related('current_class')[:10]

    if not students:
        return HttpResponse('<div class="p-2 text-sm text-base-content/60">No students found</div>')

    html = '<ul class="menu bg-base-100 shadow-lg rounded-box absolute z-50 w-full mt-1 max-h-48 overflow-y-auto">'
    for student in students:
        class_name = escape(student.current_class.name) if student.current_class else 'No class'
        name = escape(student.full_name)
        adm = escape(student.admission_number)
        # Use data attributes instead of inline JS to avoid XSS via quotes in names
        html += f'''<li><a data-student-id="{student.pk}" data-student-name="{name}"
            onclick="selectStudent(this.dataset.studentId, this.dataset.studentName)" class="text-sm">
            <span class="font-medium">{name}</span>
            <span class="text-xs text-base-content/60">{adm} • {class_name}</span>
        </a></li>'''
    html += '</ul>'

    return HttpResponse(html)


@admin_required
def invoice_search(request):
    """Search invoices for HTMX autocomplete."""
    q = request.GET.get('q', '').strip()

    if len(q) < 2:
        return HttpResponse('')

    payable_statuses = {'ISSUED', 'PARTIALLY_PAID', 'OVERDUE'}

    invoices = Invoice.objects.filter(
        Q(invoice_number__icontains=q) |
        Q(student__first_name__icontains=q) |
        Q(student__last_name__icontains=q) |
        Q(student__admission_number__icontains=q)
    ).select_related('student')[:10]

    if not invoices:
        return HttpResponse('<div class="p-2 text-sm text-base-content/60">No invoices found</div>')

    html = '<ul class="menu bg-base-100 shadow-lg rounded-box absolute z-50 w-full mt-1 max-h-48 overflow-y-auto">'
    for invoice in invoices:
        balance = f"{invoice.balance:.2f}"
        inv_num = escape(invoice.invoice_number)
        name = escape(invoice.student.full_name)
        status_label = escape(invoice.get_status_display())
        payable = invoice.status in payable_statuses

        if payable:
            html += f'''<li><a data-invoice-id="{invoice.pk}" data-invoice-num="{inv_num}"
                data-student-name="{name}" data-balance="{balance}"
                onclick="selectInvoice(this.dataset.invoiceId, this.dataset.invoiceNum, this.dataset.studentName, this.dataset.balance)"
                class="text-sm py-2">
                <div class="flex flex-col">
                    <span class="font-medium">{inv_num}</span>
                    <span class="text-xs text-base-content/60">{name} • Balance: GHS {balance}</span>
                </div>
            </a></li>'''
        else:
            if invoice.status == 'DRAFT':
                hint = 'issue invoice first'
            elif invoice.status == 'PAID':
                hint = 'fully paid'
            elif invoice.status == 'CANCELLED':
                hint = 'cancelled'
            else:
                hint = status_label.lower()
            html += f'''<li class="disabled"><span class="text-sm py-2 opacity-50 cursor-not-allowed">
                <div class="flex flex-col">
                    <span class="font-medium">{inv_num}</span>
                    <span class="text-xs text-error">{name} • {hint}</span>
                </div>
            </span></li>'''
    html += '</ul>'

    return HttpResponse(html)


@admin_required
def api_student_balance(request, student_id):
    """Get student's current balance."""
    student = get_object_or_404(Student, pk=student_id)

    balance = Invoice.objects.filter(
        student=student,
        status__in=['ISSUED', 'PARTIALLY_PAID', 'OVERDUE']
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00')

    return JsonResponse({
        'student_id': str(student_id),
        'student_name': student.full_name,
        'balance': float(balance)
    })


@admin_required
def api_class_fees(request, class_id):
    """Get fee structures for a class."""
    class_obj = get_object_or_404(Class, pk=class_id)
    current_year = AcademicYear.get_current()

    structures = FeeStructure.objects.filter(
        academic_year=current_year,
        is_active=True
    ).filter(
        Q(class_assigned=class_obj) |
        Q(level_type=class_obj.level_type) |
        Q(class_assigned__isnull=True, level_type='')
    )

    fees = []
    for structure in structures:
        fees.append({
            'category': structure.get_category_display(),
            'description': structure.get_description(),
            'amount': float(structure.amount),
            'is_mandatory': structure.is_mandatory,
        })

    return JsonResponse({
        'class_id': class_id,
        'class_name': class_obj.name,
        'fees': fees,
        'total': sum(f['amount'] for f in fees)
    })


# =============================================================================
# ONLINE PAYMENT
# =============================================================================

@login_required
def pay_online(request, invoice_pk):
    """
    Initiate an online payment for an invoice.
    Redirects user to the payment gateway.
    """
    with transaction.atomic():
        invoice = get_object_or_404(
            Invoice.objects.select_for_update().select_related('student'),
            pk=invoice_pk
        )

        # Check invoice can be paid
        if invoice.status in ['PAID', 'CANCELLED']:
            messages.error(request, 'This invoice cannot be paid.')
            return redirect('finance:invoice_detail', pk=invoice_pk)

        if invoice.balance <= 0:
            messages.error(request, 'This invoice has no outstanding balance.')
            return redirect('finance:invoice_detail', pk=invoice_pk)

    # Get primary gateway config
    from .models import PaymentGatewayConfig
    gateway_config = PaymentGatewayConfig.objects.filter(
        is_active=True,
        is_primary=True
    ).select_related('gateway').first()

    if not gateway_config:
        messages.error(request, 'No payment gateway is configured. Please contact the school.')
        return redirect('finance:invoice_detail', pk=invoice_pk)

    if gateway_config.verification_status != 'VERIFIED':
        messages.error(request, 'Payment gateway is not verified. Please contact the school.')
        return redirect('finance:invoice_detail', pk=invoice_pk)

    # Get gateway adapter
    from .gateways import get_gateway_adapter
    adapter = get_gateway_adapter(gateway_config)

    # Generate unique reference
    import uuid
    reference = f"PAY-{invoice.invoice_number}-{uuid.uuid4().hex[:8].upper()}"

    # Get payer email (student's guardian email or use a default)
    payer_email = getattr(invoice.student, 'guardian_email', '') or \
                  getattr(invoice.student, 'email', '') or \
                  'noreply@school.com'

    # Build callback URL
    callback_url = request.build_absolute_uri(
        reverse('finance:payment_callback')
    ) + f'?reference={reference}'

    # Metadata for tracking
    metadata = {
        'invoice_id': str(invoice.pk),
        'invoice_number': invoice.invoice_number,
        'student_id': str(invoice.student.pk),
        'student_name': invoice.student.full_name,
    }

    # Initialize payment
    response = adapter.initialize_payment(
        amount=invoice.balance,
        email=payer_email,
        reference=reference,
        callback_url=callback_url,
        metadata=metadata
    )

    if response.success:
        # Create pending payment record
        from .models import Payment, PaymentGatewayTransaction
        payment = Payment.objects.create(
            invoice=invoice,
            amount=invoice.balance,
            method='ONLINE',
            status='PENDING',
            reference=reference,
            payer_email=payer_email,
        )

        # Create gateway transaction record
        PaymentGatewayTransaction.objects.create(
            payment=payment,
            gateway_config=gateway_config,
            gateway_reference=reference,
            amount_charged=response.amount,
            net_amount=invoice.balance,
            full_response=response.raw_response,
        )

        # Redirect to gateway
        return redirect(response.authorization_url)
    else:
        messages.error(request, f'Payment initialization failed: {response.message}')
        return redirect('finance:invoice_detail', pk=invoice_pk)


@login_required
def payment_callback(request):
    """
    Handle return from payment gateway.
    Verifies the payment and updates records.
    """
    reference = request.GET.get('reference', '')

    if not reference:
        messages.error(request, 'Invalid payment callback.')
        return redirect('finance:payments')

    # Find the payment
    from .models import Payment, PaymentGatewayTransaction
    try:
        payment = Payment.objects.select_related(
            'invoice__student'
        ).get(reference=reference)
    except Payment.DoesNotExist:
        messages.error(request, 'Payment not found.')
        return redirect('finance:payments')

    # If already processed, show result
    if payment.status == 'COMPLETED':
        messages.success(request, 'Payment was successful!')
        return redirect('finance:payment_detail', pk=payment.pk)
    elif payment.status in ['FAILED', 'CANCELLED']:
        messages.error(request, 'Payment was not successful.')
        return redirect('finance:invoice_detail', pk=payment.invoice.pk)

    # Get gateway transaction
    try:
        gateway_tx = payment.gateway_transaction
        gateway_config = gateway_tx.gateway_config
    except PaymentGatewayTransaction.DoesNotExist:
        messages.error(request, 'Payment configuration error.')
        return redirect('finance:invoice_detail', pk=payment.invoice.pk)

    # Verify with gateway
    from .gateways import get_gateway_adapter
    adapter = get_gateway_adapter(gateway_config)
    response = adapter.verify_payment(reference)

    # Atomically update payment to prevent double-processing
    with transaction.atomic():
        payment = Payment.objects.select_for_update().get(pk=payment.pk)

        # Re-check status under lock
        if payment.status == 'COMPLETED':
            messages.success(request, 'Payment was successful!')
            return redirect('finance:payment_detail', pk=payment.pk)

        if response.success:
            payment.status = 'COMPLETED'
            payment.transaction_date = timezone.now()
            payment.save()

            gateway_tx.gateway_transaction_id = response.transaction_id
            gateway_tx.gateway_fee = Decimal(str(response.gateway_fee or 0))
            gateway_tx.net_amount = Decimal(str(response.amount or 0)) - gateway_tx.gateway_fee
            gateway_tx.full_response = response.raw_response
            gateway_tx.save()

            messages.success(request, f'Payment successful! Receipt: {payment.receipt_number}')
            return redirect('finance:payment_detail', pk=payment.pk)
        else:
            payment.status = 'FAILED'
            payment.save()

            gateway_tx.full_response = response.raw_response
            gateway_tx.save()

            messages.error(request, f'Payment verification failed: {response.message}')
            return redirect('finance:invoice_detail', pk=payment.invoice.pk)


@csrf_exempt
@require_POST
def payment_webhook(request):
    """
    Handle webhook notifications from payment gateway.
    This is called server-to-server by the gateway.

    SECURITY: Signature verification happens in the gateway adapter.
    We reject webhooks with invalid signatures before any database changes.
    """
    import logging
    logger = logging.getLogger(__name__)

    # Get signature from headers (each gateway uses a different header)
    signature = request.headers.get('X-Paystack-Signature', '') or \
                request.headers.get('verif-hash', '') or \
                request.headers.get('X-Hubtel-Signature', '')

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        logger.warning("Payment webhook received invalid JSON")
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    # Determine which gateway this is from and get reference
    reference = None
    if 'data' in payload and 'reference' in payload.get('data', {}):
        reference = payload['data']['reference']
    elif 'Data' in payload and 'ClientReference' in payload.get('Data', {}):
        reference = payload['Data']['ClientReference']  # Hubtel format

    if not reference:
        logger.warning("Payment webhook received without reference")
        return JsonResponse({'status': 'error', 'message': 'No reference found'}, status=400)

    # Find the payment
    from .models import Payment, PaymentGatewayTransaction
    try:
        payment = Payment.objects.get(reference=reference)
        gateway_tx = payment.gateway_transaction
        gateway_config = gateway_tx.gateway_config
    except (Payment.DoesNotExist, PaymentGatewayTransaction.DoesNotExist):
        # Don't reveal whether payment exists - use generic message
        logger.warning(f"Payment webhook for unknown reference: {reference[:20]}...")
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

    # Skip if already processed (quick check before signature verification)
    if payment.status == 'COMPLETED':
        return JsonResponse({'status': 'success', 'message': 'Already processed'})

    # Verify signature and process with gateway adapter
    # SECURITY: The adapter verifies the signature using HMAC
    # If signature is invalid, response.success will be False
    from .gateways import get_gateway_adapter
    adapter = get_gateway_adapter(gateway_config)
    response = adapter.handle_webhook(payload, signature, raw_body=request.body)

    # SECURITY: Check for signature verification failure
    # The adapter returns success=False with "signature" in message for invalid signatures
    if not response.success and 'signature' in response.message.lower():
        logger.warning(f"Payment webhook signature verification failed for {reference[:20]}...")
        return JsonResponse({'status': 'error', 'message': 'Signature verification failed'}, status=403)

    # Atomically update payment to prevent double-processing from concurrent webhooks
    with transaction.atomic():
        payment = Payment.objects.select_for_update().get(pk=payment.pk)

        # Re-check status under lock
        if payment.status == 'COMPLETED':
            return JsonResponse({'status': 'success', 'message': 'Already processed'})

        gateway_tx.callback_data = payload
        gateway_tx.save()

        if response.success:
            payment.status = 'COMPLETED'
            payment.transaction_date = timezone.now()
            payment.save()

            gateway_tx.gateway_transaction_id = response.transaction_id
            gateway_tx.gateway_fee = Decimal(str(response.gateway_fee or 0))
            gateway_tx.net_amount = Decimal(str(response.amount or 0)) - gateway_tx.gateway_fee
            gateway_tx.save()

            # Bell notification for guardian
            if payment.invoice and payment.invoice.student:
                from core.notifications import notify_guardian
                student = payment.invoice.student
                notify_guardian(
                    student,
                    title='Payment Confirmed',
                    message=f'Payment of {payment.amount} received for {student.full_name}.',
                    category='finance',
                    notification_type='success',
                    icon='fa-solid fa-circle-check',
                )

            # Send payment confirmation SMS to guardian
            from .tasks import send_payment_confirmation_sms
            send_payment_confirmation_sms.delay(str(payment.pk), connection.schema_name)

            logger.info(f"Payment {reference[:20]}... confirmed via webhook")
            return JsonResponse({'status': 'success', 'message': 'Payment confirmed'})

        # Payment failed (but webhook signature was valid)
        payment.status = 'FAILED'
        payment.save()

        logger.info(f"Payment {reference[:20]}... failed: {response.message}")
        return JsonResponse({'status': 'success', 'message': 'Payment failed recorded'})


# =============================================================================
# FINANCE NOTIFICATIONS
# =============================================================================

@admin_required
def notification_center(request):
    """Finance notification center - manage and send invoice notifications."""
    from datetime import timedelta
    from .models import FinanceNotificationLog

    current_year = AcademicYear.get_current()
    current_term = Term.get_current() if hasattr(Term, 'get_current') else None
    today = timezone.now().date()

    # Get invoices with balance
    from django.db.models import Prefetch
    invoices = Invoice.objects.filter(
        balance__gt=0,
        status__in=['ISSUED', 'PARTIALLY_PAID', 'OVERDUE']
    ).select_related(
        'student', 'student__current_class', 'academic_year', 'term'
    ).prefetch_related(
        Prefetch(
            'notification_logs',
            queryset=FinanceNotificationLog.objects.order_by('-created_at'),
        )
    ).order_by('student__last_name', 'student__first_name')

    # Apply filters
    status_filter = request.GET.get('status', '')
    class_filter = request.GET.get('class_id', '')
    search = request.GET.get('search', '')

    if status_filter:
        invoices = invoices.filter(status=status_filter)
    if class_filter:
        invoices = invoices.filter(student__current_class_id=class_filter)
    if search:
        invoices = invoices.filter(
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(invoice_number__icontains=search)
        )

    # Overall Stats
    all_invoices = Invoice.objects.filter(academic_year=current_year)
    total_invoices = all_invoices.count()
    paid_count = all_invoices.filter(status='PAID').count()

    all_with_balance = Invoice.objects.filter(balance__gt=0, status__in=['ISSUED', 'PARTIALLY_PAID', 'OVERDUE'])
    overdue_count = all_with_balance.filter(status='OVERDUE').count()
    partial_count = all_with_balance.filter(status='PARTIALLY_PAID').count()
    issued_count = all_with_balance.filter(status='ISSUED').count()
    total_balance = all_with_balance.aggregate(total=Sum('balance'))['total'] or Decimal('0.00')
    overdue_balance = all_with_balance.filter(status='OVERDUE').aggregate(total=Sum('balance'))['total'] or Decimal('0.00')

    # Collection rate
    total_invoiced = all_invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_collected = Payment.objects.filter(
        invoice__academic_year=current_year,
        status='COMPLETED'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    collection_rate = int((total_collected / total_invoiced * 100) if total_invoiced > 0 else 0)

    # Notification stats
    sent_today = FinanceNotificationLog.objects.filter(
        created_at__date=today
    ).count()

    week_start = today - timedelta(days=today.weekday())
    sent_week = FinanceNotificationLog.objects.filter(
        created_at__date__gte=week_start
    ).count()

    # Class summary with fee collection stats - single annotated query
    classes = Class.objects.filter(
        students__invoices__academic_year=current_year
    ).distinct().annotate(
        class_total=Count(
            'students__invoices',
            filter=Q(students__invoices__academic_year=current_year)
        ),
        class_paid=Count(
            'students__invoices',
            filter=Q(students__invoices__academic_year=current_year, students__invoices__status='PAID')
        ),
        class_overdue=Count(
            'students__invoices',
            filter=Q(students__invoices__academic_year=current_year, students__invoices__status='OVERDUE')
        ),
        class_pending=Count(
            'students__invoices',
            filter=Q(
                students__invoices__academic_year=current_year,
                students__invoices__status__in=['ISSUED', 'PARTIALLY_PAID']
            )
        ),
        class_balance=Sum(
            'students__invoices__balance',
            filter=Q(
                students__invoices__academic_year=current_year,
                students__invoices__balance__gt=0
            )
        ),
    ).order_by('name')

    class_summary = []
    for cls in classes:
        total = cls.class_total or 0
        paid = cls.class_paid or 0
        class_summary.append({
            'class': cls,
            'total': total,
            'paid': paid,
            'overdue': cls.class_overdue or 0,
            'pending': cls.class_pending or 0,
            'balance': cls.class_balance or Decimal('0.00'),
            'rate': int((paid / total * 100) if total > 0 else 0),
        })

    # Paginate invoices
    per_page = request.GET.get('per_page', '25')
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100]:
            per_page = 25
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(invoices, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    invoice_list = list(page_obj)

    # Prefetch primary guardians in bulk to avoid N+1
    from students.models import StudentGuardian
    student_ids = [inv.student_id for inv in invoice_list]
    guardian_map = {}
    if student_ids:
        sg_qs = StudentGuardian.objects.filter(
            student_id__in=student_ids, is_primary=True
        ).select_related('guardian')
        guardian_map = {sg.student_id: sg.guardian for sg in sg_qs}

    # Get last notification and guardian info for each invoice
    for invoice in invoice_list:
        last_log = invoice.notification_logs.first()
        invoice.last_notification = last_log
        # Calculate days overdue
        if invoice.status == 'OVERDUE' and invoice.due_date:
            invoice.days_overdue = (today - invoice.due_date).days
        else:
            invoice.days_overdue = 0
        # Get guardian contact info from prefetched map
        primary_guardian = guardian_map.get(invoice.student_id)
        invoice.guardian_phone = primary_guardian.phone_number if primary_guardian else None
        invoice.guardian_email = primary_guardian.email if primary_guardian else None

    # Separate overdue invoices for dedicated section
    overdue_invoices = [inv for inv in invoice_list if inv.status == 'OVERDUE']

    context = {
        'invoices': invoice_list,
        'overdue_invoices': overdue_invoices,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'classes': classes,
        'class_summary': class_summary,
        'total_invoices': total_invoices,
        'paid_count': paid_count,
        'overdue_count': overdue_count,
        'partial_count': partial_count,
        'issued_count': issued_count,
        'total_balance': total_balance,
        'overdue_balance': overdue_balance,
        'collection_rate': collection_rate,
        'sent_today': sent_today,
        'sent_week': sent_week,
        'today': today,
        'status_filter': status_filter,
        'class_filter': class_filter,
        'search': search,
        'current_year': current_year,
        'current_term': current_term,
    }

    return htmx_render(
        request,
        'finance/notification_center.html',
        'finance/partials/notification_center_content.html',
        context
    )


@admin_required
def send_invoice_notification_view(request, pk):
    """Send notification for a single invoice."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    invoice = get_object_or_404(Invoice, pk=pk)
    distribution_type = request.POST.get('type', 'SMS')
    notification_type = request.POST.get('notification_type', 'BALANCE_REMINDER')
    custom_sms = request.POST.get('custom_sms', '').strip() or None

    # Get tenant schema
    from django.db import connection
    tenant_schema = connection.tenant.schema_name if hasattr(connection, 'tenant') else 'public'

    # Queue the notification task
    from .tasks import send_invoice_notification
    send_invoice_notification.delay(
        invoice_id=str(invoice.pk),
        notification_type=notification_type,
        distribution_type=distribution_type,
        tenant_schema=tenant_schema,
        sent_by_id=request.user.id,
        custom_sms=custom_sms,
    )

    # Return success response with HTMX trigger for toast
    response = HttpResponse(status=204)
    response['HX-Trigger'] = '{"showToast": {"message": "Notification queued for sending", "type": "success"}}'
    return response


@admin_required
def send_bulk_notifications_view(request):
    """Send bulk notifications for multiple invoices."""
    if request.method != 'POST':
        return HttpResponse(status=405)

    distribution_type = request.POST.get('type', 'SMS')
    notification_type = request.POST.get('notification_type', 'OVERDUE_REMINDER')
    custom_sms = request.POST.get('custom_sms', '').strip() or None

    # Build filters from request
    filters = {}
    if request.POST.get('status'):
        filters['status'] = request.POST.get('status')
    if request.POST.get('class_id'):
        filters['class_id'] = request.POST.get('class_id')

    # Get selected invoice IDs if any
    selected_ids = request.POST.getlist('invoice_ids')
    if selected_ids:
        filters['invoice_ids'] = selected_ids

    # Get tenant schema
    from django.db import connection
    tenant_schema = connection.tenant.schema_name if hasattr(connection, 'tenant') else 'public'

    # Queue the bulk notification task
    from .tasks import send_bulk_notifications
    send_bulk_notifications.delay(
        notification_type=notification_type,
        distribution_type=distribution_type,
        tenant_schema=tenant_schema,
        sent_by_id=request.user.id,
        filters=filters,
        custom_sms=custom_sms,
    )

    # Return success response with HTMX trigger
    response = HttpResponse(status=204)
    response['HX-Trigger'] = '{"showToast": {"message": "Bulk notifications queued for sending", "type": "success"}}'
    return response


@admin_required
def notification_history(request):
    """View notification history with filters."""
    from datetime import timedelta
    from .models import FinanceNotificationLog

    today = timezone.now().date()

    logs = FinanceNotificationLog.objects.select_related(
        'invoice', 'invoice__student', 'sent_by'
    ).order_by('-created_at')

    # Stats for all logs (before filtering)
    all_logs = FinanceNotificationLog.objects.all()
    total_count = all_logs.count()
    success_count = all_logs.filter(Q(email_status='SENT') | Q(sms_status='SENT')).count()
    failed_count = all_logs.filter(email_status='FAILED', sms_status='FAILED').count()
    delivery_rate = int((success_count / total_count * 100) if total_count > 0 else 0)

    today_count = all_logs.filter(created_at__date=today).count()
    week_start = today - timedelta(days=today.weekday())
    week_count = all_logs.filter(created_at__date__gte=week_start).count()

    # Filters
    notification_type = request.GET.get('notification_type', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')

    if notification_type:
        logs = logs.filter(notification_type=notification_type)
    if status:
        if status == 'SUCCESS':
            logs = logs.filter(Q(email_status='SENT') | Q(sms_status='SENT'))
        elif status == 'FAILED':
            logs = logs.filter(email_status='FAILED', sms_status='FAILED')
        elif status == 'PENDING':
            logs = logs.filter(email_status='PENDING', sms_status='PENDING')
    if search:
        logs = logs.filter(
            Q(invoice__invoice_number__icontains=search) |
            Q(invoice__student__first_name__icontains=search) |
            Q(invoice__student__last_name__icontains=search)
        )

    # Paginate
    paginator = Paginator(logs, 50)
    page = request.GET.get('page', 1)
    logs = paginator.get_page(page)

    context = {
        'logs': logs,
        'notification_type': notification_type,
        'status_filter': status,
        'search': search,
        'notification_types': FinanceNotificationLog.NOTIFICATION_TYPE,
        'success_count': success_count,
        'failed_count': failed_count,
        'delivery_rate': delivery_rate,
        'today_count': today_count,
        'week_count': week_count,
    }

    return htmx_render(
        request,
        'finance/notification_history.html',
        'finance/partials/notification_history_content.html',
        context
    )


# =============================================================================
# GUARDIAN / PARENT PAYMENT VIEWS
# =============================================================================

@login_required
def fee_payments(request):
    """Guardian view for viewing fee payments for their wards."""
    from students.models import StudentGuardian

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)
    current_year = AcademicYear.get_current()
    current_term = Term.get_current()

    # Check if online payments are available
    online_payments_enabled = PaymentGatewayConfig.objects.filter(
        is_active=True,
        is_primary=True,
        verification_status='VERIFIED'
    ).exists()

    # Get selected ward filter from query params
    selected_ward_id = request.GET.get('ward')

    wards_fees = []
    all_invoices = []
    all_payments = []
    total_outstanding = 0

    if guardian:
        student_guardians = StudentGuardian.objects.filter(
            guardian=guardian
        ).select_related('student', 'student__current_class')

        ward_students = []
        ward_sgs = []
        for sg in student_guardians:
            if sg.student.status == 'active':
                ward_students.append(sg.student)
                ward_sgs.append(sg)

        ward_ids = [s.id for s in ward_students]

        # Batch-fetch invoice aggregates for all wards
        year_filter = {'academic_year': current_year} if current_year else {}
        invoice_aggregates = {}
        if ward_ids:
            for row in Invoice.objects.filter(
                student_id__in=ward_ids, **year_filter
            ).exclude(status='CANCELLED').values('student_id').annotate(
                total_fees=Sum('total_amount'),
                total_paid=Sum('amount_paid'),
                total_balance=Sum('balance')
            ):
                invoice_aggregates[row['student_id']] = row

        # Batch-fetch recent payments grouped by student
        payments_by_student = {}
        if ward_ids:
            for payment in Payment.objects.filter(
                invoice__student_id__in=ward_ids,
                status='COMPLETED'
            ).select_related('invoice').order_by('-transaction_date'):
                sid = payment.invoice.student_id
                if sid not in payments_by_student:
                    payments_by_student[sid] = []
                if len(payments_by_student[sid]) < 5:
                    payments_by_student[sid].append(payment)

        for sg, student in zip(ward_sgs, ward_students):
            agg = invoice_aggregates.get(student.id, {})
            total_fees = agg.get('total_fees') or 0
            total_paid = agg.get('total_paid') or 0
            balance = agg.get('total_balance') or 0

            # Per-student queries for sliced results (can't batch slices)
            invoices = Invoice.objects.filter(
                student=student, **year_filter
            ).exclude(status='CANCELLED').order_by('-created_at')

            current_invoice = invoices.filter(term=current_term).first() if current_term else None

            wards_fees.append({
                'student': student,
                'relationship': sg.get_relationship_display(),
                'total_fees': total_fees,
                'total_paid': total_paid,
                'balance': balance,
                'current_invoice': current_invoice,
                'invoices': invoices[:3],
                'recent_payments': payments_by_student.get(student.id, []),
            })

            total_outstanding += balance

        # For detailed view - get all invoices and payments (optionally filtered by ward)
        if selected_ward_id:
            try:
                selected_student = next(
                    (s for s in ward_students if str(s.pk) == selected_ward_id),
                    None
                )
                if selected_student:
                    all_invoices = Invoice.objects.filter(
                        student=selected_student
                    ).exclude(status='CANCELLED').select_related(
                        'student', 'term', 'academic_year'
                    ).prefetch_related('items').order_by('-created_at')

                    all_payments = Payment.objects.filter(
                        invoice__student=selected_student,
                        status='COMPLETED'
                    ).select_related('invoice').order_by('-transaction_date')
            except (ValueError, StopIteration):
                pass
        else:
            # All wards' invoices and payments
            all_invoices = Invoice.objects.filter(
                student__in=ward_students
            ).exclude(status='CANCELLED').select_related(
                'student', 'term', 'academic_year'
            ).prefetch_related('items').order_by('-created_at')[:20]

            all_payments = Payment.objects.filter(
                invoice__student__in=ward_students,
                status='COMPLETED'
            ).select_related('invoice', 'invoice__student').order_by('-transaction_date')[:10]

    context = {
        'guardian': guardian,
        'wards_fees': wards_fees,
        'all_invoices': all_invoices,
        'all_payments': all_payments,
        'total_outstanding': total_outstanding,
        'selected_ward_id': selected_ward_id,
        'current_year': current_year,
        'current_term': current_term,
        'online_payments_enabled': online_payments_enabled,
    }
    return htmx_render(request, 'finance/parent/fee_payments.html', 'finance/parent/partials/fee_payments_content.html', context)


@login_required
def guardian_pay_invoice(request, invoice_id):
    """
    Guardian view to initiate online payment for an invoice.
    Verifies the guardian has access to this invoice's student.
    """
    from students.models import StudentGuardian
    from .gateways import get_gateway_adapter
    import uuid as uuid_module

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.error(request, 'No guardian profile found.')
        return redirect('finance:fee_payments')

    # Get the invoice
    invoice = get_object_or_404(
        Invoice.objects.select_related('student'),
        pk=invoice_id
    )

    # Verify guardian has access to this student
    has_access = StudentGuardian.objects.filter(
        guardian=guardian,
        student=invoice.student
    ).exists()

    if not has_access:
        messages.error(request, 'You do not have access to this invoice.')
        return redirect('finance:fee_payments')

    # Check invoice can be paid
    if invoice.status in ['PAID', 'CANCELLED']:
        messages.error(request, 'This invoice cannot be paid.')
        return redirect('finance:fee_payments')

    if invoice.balance <= 0:
        messages.error(request, 'This invoice has no outstanding balance.')
        return redirect('finance:fee_payments')

    # Get primary gateway config
    gateway_config = PaymentGatewayConfig.objects.filter(
        is_active=True,
        is_primary=True
    ).select_related('gateway').first()

    if not gateway_config:
        messages.error(request, 'Online payments are not available. Please contact the school.')
        return redirect('finance:fee_payments')

    if gateway_config.verification_status != 'VERIFIED':
        messages.error(request, 'Payment gateway is not configured. Please contact the school.')
        return redirect('finance:fee_payments')

    # Get gateway adapter
    adapter = get_gateway_adapter(gateway_config)

    # Generate unique reference
    reference = f"GP-{invoice.invoice_number}-{uuid_module.uuid4().hex[:8].upper()}"

    # Get payer email from guardian
    payer_email = guardian.email or user.email or 'noreply@school.com'
    payer_name = guardian.full_name
    payer_phone = guardian.phone or ''

    # Build callback URL - guardian specific
    callback_url = request.build_absolute_uri(
        reverse('finance:guardian_payment_callback')
    ) + f'?reference={reference}'

    # Metadata for tracking
    metadata = {
        'invoice_id': str(invoice.pk),
        'invoice_number': invoice.invoice_number,
        'student_id': str(invoice.student.pk),
        'student_name': invoice.student.full_name,
        'guardian_id': str(guardian.pk),
        'guardian_name': guardian.full_name,
        'source': 'guardian_portal',
    }

    # Initialize payment with gateway
    response = adapter.initialize_payment(
        amount=invoice.balance,
        email=payer_email,
        reference=reference,
        callback_url=callback_url,
        metadata=metadata
    )

    if response.success:
        # Create pending payment record
        payment = Payment.objects.create(
            invoice=invoice,
            amount=invoice.balance,
            method='ONLINE',
            status='PENDING',
            reference=reference,
            payer_email=payer_email,
            payer_name=payer_name,
            payer_phone=payer_phone,
        )

        # Create gateway transaction record
        PaymentGatewayTransaction.objects.create(
            payment=payment,
            gateway_config=gateway_config,
            gateway_reference=response.gateway_reference or '',
            amount_charged=invoice.balance,
            net_amount=invoice.balance,
            full_response=response.raw_response,
        )

        # Redirect to payment gateway
        return redirect(response.authorization_url)
    else:
        messages.error(request, f'Could not initiate payment: {response.message}')
        return redirect('finance:fee_payments')


def _send_payment_receipt_email(payment, guardian):
    """
    Send payment receipt email to guardian.
    Returns True if email was sent successfully, False otherwise.
    """
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    from datetime import datetime

    # Get recipient email
    recipient_email = payment.payer_email or guardian.email
    if not recipient_email:
        logger.warning(f"No email address for payment receipt {payment.receipt_number}")
        return False

    # Get school from tenant
    school = getattr(connection, 'tenant', None)

    # Build context
    context = {
        'payment': payment,
        'invoice': payment.invoice,
        'student': payment.invoice.student,
        'guardian': guardian,
        'school': school,
        'current_year': datetime.now().year,
    }

    # Render email
    subject = f"Payment Receipt - {payment.receipt_number}"
    html_message = render_to_string('finance/emails/payment_receipt_email.html', context)
    plain_message = strip_tags(html_message)

    try:
        send_mail(
            subject,
            plain_message,
            get_from_email(),
            [recipient_email],
            html_message=html_message,
            fail_silently=False,
        )
        logger.info(f"Payment receipt email sent for {payment.receipt_number} to {recipient_email}")
        return True
    except Exception as e:
        logger.error(f"Failed to send payment receipt email for {payment.receipt_number}: {str(e)}")
        return False


@login_required
def guardian_payment_callback(request):
    """
    Handle return from payment gateway for guardian payments.
    Verifies the payment and redirects to success page or fee payments.
    """
    from .gateways import get_gateway_adapter

    reference = request.GET.get('reference', '')

    if not reference:
        messages.error(request, 'Invalid payment callback.')
        return redirect('finance:fee_payments')

    # Find the payment
    try:
        payment = Payment.objects.select_related(
            'invoice__student'
        ).get(reference=reference)
    except Payment.DoesNotExist:
        messages.error(request, 'Payment not found.')
        return redirect('finance:fee_payments')

    # If already processed, redirect to appropriate page
    if payment.status == 'COMPLETED':
        return redirect('finance:guardian_payment_success', payment_id=payment.pk)
    elif payment.status in ['FAILED', 'CANCELLED']:
        return redirect('finance:guardian_payment_failed', payment_id=payment.pk)

    # Get gateway transaction
    try:
        gateway_tx = payment.gateway_transaction
        gateway_config = gateway_tx.gateway_config
    except PaymentGatewayTransaction.DoesNotExist:
        messages.error(request, 'Payment configuration error.')
        return redirect('finance:fee_payments')

    # Verify with gateway
    adapter = get_gateway_adapter(gateway_config)
    response = adapter.verify_payment(reference)

    if response.success:
        # Update payment
        payment.status = 'COMPLETED'
        payment.transaction_date = timezone.now()
        payment.save()

        # Update gateway transaction
        gateway_tx.gateway_transaction_id = response.transaction_id or ''
        gateway_tx.gateway_fee = response.gateway_fee or 0
        gateway_tx.net_amount = response.amount - (response.gateway_fee or 0)
        gateway_tx.full_response = response.raw_response
        gateway_tx.save()

        # Invoice totals are updated automatically via Payment.save()

        # Send receipt email to guardian
        guardian = getattr(request.user, 'guardian_profile', None)
        if guardian:
            # Refresh payment with updated invoice data
            payment.refresh_from_db()
            payment.invoice.refresh_from_db()
            _send_payment_receipt_email(payment, guardian)

        # Send payment confirmation SMS to guardian
        from .tasks import send_payment_confirmation_sms
        send_payment_confirmation_sms.delay(str(payment.pk), connection.schema_name)

        # Redirect to success page
        return redirect('finance:guardian_payment_success', payment_id=payment.pk)
    else:
        payment.status = 'FAILED'
        payment.save()
        # Store error message in session for display on failed page
        request.session['payment_error_message'] = response.message
        return redirect('finance:guardian_payment_failed', payment_id=payment.pk)


@login_required
def guardian_payment_success(request, payment_id):
    """
    Display payment success confirmation page for guardians.
    Shows receipt details and allows printing.
    """
    from students.models import StudentGuardian

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.error(request, 'No guardian profile found.')
        return redirect('finance:fee_payments')

    # Get the payment with related data
    payment = get_object_or_404(
        Payment.objects.select_related(
            'invoice__student',
            'invoice__term',
            'invoice__academic_year'
        ).prefetch_related('invoice__items'),
        pk=payment_id,
        status='COMPLETED'
    )

    # Verify guardian has access to this student
    has_access = StudentGuardian.objects.filter(
        guardian=guardian,
        student=payment.invoice.student
    ).exists()

    if not has_access:
        messages.error(request, 'You do not have access to this payment.')
        return redirect('finance:fee_payments')

    # Get school from tenant for branding
    school = getattr(connection, 'tenant', None)

    context = {
        'payment': payment,
        'invoice': payment.invoice,
        'student': payment.invoice.student,
        'school': school,
        'guardian': guardian,
    }

    # For HTMX requests, return partial content
    if request.htmx:
        return render(request, 'finance/parent/partials/payment_success_content.html', context)
    return render(request, 'finance/parent/payment_success.html', context)


@login_required
def guardian_payment_failed(request, payment_id):
    """
    Display payment failed page for guardians.
    Shows error details and allows retry.
    """
    from students.models import StudentGuardian

    user = request.user
    guardian = getattr(user, 'guardian_profile', None)

    if not guardian:
        messages.error(request, 'No guardian profile found.')
        return redirect('finance:fee_payments')

    # Get the payment with related data
    payment = get_object_or_404(
        Payment.objects.select_related(
            'invoice__student',
            'invoice__term',
            'invoice__academic_year'
        ),
        pk=payment_id
    )

    # Verify guardian has access to this student
    has_access = StudentGuardian.objects.filter(
        guardian=guardian,
        student=payment.invoice.student
    ).exists()

    if not has_access:
        messages.error(request, 'You do not have access to this payment.')
        return redirect('finance:fee_payments')

    # Get error message from session if available
    error_message = request.session.pop('payment_error_message', None)

    context = {
        'payment': payment,
        'invoice': payment.invoice,
        'student': payment.invoice.student,
        'guardian': guardian,
        'error_message': error_message,
    }

    # For HTMX requests, return partial content
    if request.htmx:
        return render(request, 'finance/parent/partials/payment_failed_content.html', context)
    return render(request, 'finance/parent/payment_failed.html', context)
